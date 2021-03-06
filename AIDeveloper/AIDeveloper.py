# -*- coding: utf-8 -*-
"""
AIDeveloper
---------
@author: maikherbig
"""
import os,sys
if not sys.platform.startswith("win"):
    from multiprocessing import freeze_support
    freeze_support()
# Make sure to get the right icon file on win,linux and mac
if sys.platform=="darwin":
    icon_suff = ".icns"
else:
    icon_suff = ".ico"

import pyqtgraph as pg
from pyqtgraph.Qt import QtCore, QtWidgets, QtGui
from pyqtgraph import Qt

import aid_start
dir_root = os.path.dirname(aid_start.__file__)#ask the module for its origin
dir_settings = os.path.join(dir_root,"AIDeveloper_Settings.json")#dir to settings
Default_dict = aid_start.get_default_dict(dir_settings) 

try:
    splashapp = QtWidgets.QApplication(sys.argv)
    #splashapp.setWindowIcon(QtGui.QIcon("."+os.sep+"art"+os.sep+Default_dict["Icon theme"]+os.sep+"main_icon_simple_04_256.ico"))
    # Create and display the splash screen
    splash_pix = os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff)
    splash_pix = QtGui.QPixmap(splash_pix)
    #splash_pix = QtGui.QPixmap("."+os.sep+"art"+os.sep+Default_dict["Icon theme"]+os.sep+"main_icon_simple_04_256"+icon_suff)
    splash = QtWidgets.QSplashScreen(splash_pix, QtCore.Qt.WindowStaysOnTopHint)
    splash.setMask(splash_pix.mask())
    splash.show()
except:
    pass

#BEFORE importing tensorflow or anything from keras: make sure the keras.json has
#certain properties
keras_json_path = os.path.expanduser('~')+os.sep+'.keras'+os.sep+'keras.json'
if not os.path.isdir(os.path.expanduser('~')+os.sep+'.keras'):
    os.mkdir(os.path.expanduser('~')+os.sep+'.keras')

aid_start.banner() #show a fancy banner in console
aid_start.keras_json_check(keras_json_path)

import traceback,shutil,re,ast,io,platform
import h5py,json,time,copy,urllib
from stat import S_IREAD,S_IRGRP,S_IROTH,S_IWRITE,S_IWGRP,S_IWOTH

import tensorflow as tf
from tensorflow.python.client import device_lib
devices = device_lib.list_local_devices()
device_types = [devices[i].device_type for i in range(len(devices))]
config_gpu = tf.ConfigProto()
if 'GPU' in device_types:
    config_gpu.gpu_options.allow_growth = True
    config_gpu.gpu_options.per_process_gpu_memory_fraction = 0.7

print("Found "+str(len(devices))+" device(s):")
print("------------------------")
for i in range(len(devices)):
    print("Device "+str(i)+": "+devices[i].name)
    print("Device type: "+devices[i].device_type)
    print("Device description: "+devices[i].physical_device_desc)
    print("------------------------")

import numpy as np
rand_state = np.random.RandomState(117) #to get the same random number on diff. PCs
from scipy import ndimage,misc
from sklearn import metrics,preprocessing
import PIL
import dclab
import cv2
import pandas as pd
import openpyxl,xlrd 
import psutil

from keras.models import model_from_json,model_from_config,load_model,clone_model
from keras import backend as K
if 'GPU' in device_types:
    keras_gpu_avail = K.tensorflow_backend._get_available_gpus()
    if len(keras_gpu_avail)>0:
        print("Following GPU is used:")
        print(keras_gpu_avail)
        print("------------------------")
    else:
        print("TensorFlow detected GPU, but Keras didn't")
        print("------------------------")
        
from keras.preprocessing.image import load_img
from keras.utils import np_utils
from keras.utils.conv_utils import convert_kernel
import keras_metrics #side package for precision, recall etc during training
global keras_metrics
import model_zoo 
from keras2onnx import convert_keras
from onnx import save_model as save_onnx

import aid_img, aid_dl, aid_bin
import aid_frontend
from partial_trainability import partial_trainability

VERSION = "0.0.8_dev2" #Python 3.5.6 Version
model_zoo_version = model_zoo.__version__()
print("AIDeveloper Version: "+VERSION)
print("model_zoo.py Version: "+model_zoo.__version__())

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtWidgets.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtWidgets.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtWidgets.QApplication.translate(context, text, disambig)

def MyExceptionHook(etype, value, trace):
    """
    Copied from: https://github.com/ZELLMECHANIK-DRESDEN/ShapeOut/blob/07d741db3bb5685790d9f9f6df394cd9577e8236/shapeout/gui/frontend.py
    Handler for all unhandled exceptions.
 
    :param `etype`: the exception type (`SyntaxError`, `ZeroDivisionError`, etc...);
    :type `etype`: `Exception`
    :param string `value`: the exception error message;
    :param string `trace`: the traceback header, if any (otherwise, it prints the
     standard Python header: ``Traceback (most recent call last)``.
    """
    tmp = traceback.format_exception(etype, value, trace)
    exception = "".join(tmp)
    msg = QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Information)       
    msg.setText(exception)
    msg.setWindowTitle("Error")
    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
    msg.exec_()
    return

class MyTable(QtWidgets.QTableWidget):
    dropped = QtCore.pyqtSignal(list)

    def __init__(self,  rows, columns, parent):
        super(MyTable, self).__init__(rows, columns, parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        #self.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.drag_item = None
        self.drag_row = None

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls:
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        self.drag_item = None
        if event.mimeData().hasUrls:
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
            links = []
            for url in event.mimeData().urls():
                links.append(str(url.toLocalFile()))
            self.dropped.emit(links)
        else:
            event.ignore()       
        
    def startDrag(self, supportedActions):
        super(MyTable, self).startDrag(supportedActions)
        self.drag_item = self.currentItem()
        self.drag_row = self.row(self.drag_item)

class MyPopup(QtWidgets.QWidget):
    def __init__(self):
        QtWidgets.QWidget.__init__(self)


#Define some custom metrics which will allow to use precision, recall, etc during training
def get_custom_metrics():
    custom_metrics = []
    custom_metrics.append(keras_metrics.categorical_f1_score())
    custom_metrics.append(keras_metrics.categorical_precision())
    custom_metrics.append(keras_metrics.categorical_recall())
    custom_metrics = {m.__name__: m for m in custom_metrics}
    custom_metrics["sin"] = K.sin
    custom_metrics["abs"] = K.abs
    return custom_metrics

class WorkerSignals(QtCore.QObject):
    '''
    Code inspired from here: https://www.learnpyqt.com/courses/concurrent-execution/multithreading-pyqt-applications-qthreadpool/
    
    Defines the signals available from a running worker thread.
    Supported signals are:
    finished
        No data
    error
        `tuple` (exctype, value, traceback.format_exc() )
    result
        `object` data returned from processing, anything
    progress
        `int` indicating % progress
    history
        `dict` containing keras model history.history resulting from .fit
    '''
    finished = QtCore.pyqtSignal()
    error = QtCore.pyqtSignal(tuple)
    result = QtCore.pyqtSignal(object)
    progress = QtCore.pyqtSignal(int)
    history = QtCore.pyqtSignal(dict)

class Worker(QtCore.QRunnable):
    '''
    Code inspired/copied from: https://www.learnpyqt.com/courses/concurrent-execution/multithreading-pyqt-applications-qthreadpool/
    Worker thread
    Inherits from QRunnable to handler worker thread setup, signals and wrap-up.
    :param callback: The function callback to run on this worker thread. Supplied args and 
                     kwargs will be passed through to the runner.
    :type callback: function
    :param args: Arguments to pass to the callback function
    :param kwargs: Keywords to pass to the callback function
    '''
    def __init__(self, fn, *args, **kwargs):
        super(Worker, self).__init__()

        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

        # Add the callback to our kwargs
        self.kwargs['progress_callback'] = self.signals.progress
        self.kwargs['history_callback'] = self.signals.history

    @QtCore.pyqtSlot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''

        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi()

    def setupUi(self):
        self.setObjectName(_fromUtf8("MainWindow"))
        self.resize(1000, 800)

        sys.excepthook = MyExceptionHook
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName(_fromUtf8("gridLayout_2"))
        self.tabWidget_Modelbuilder = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget_Modelbuilder.setObjectName(_fromUtf8("tabWidget_Modelbuilder"))
        self.tab_Build = QtWidgets.QWidget()
        self.tab_Build.setObjectName(_fromUtf8("tab_Build"))
        self.gridLayout_17 = QtWidgets.QGridLayout(self.tab_Build)
        self.gridLayout_17.setObjectName(_fromUtf8("gridLayout_17"))
        self.splitter_5 = QtWidgets.QSplitter(self.tab_Build)
        self.splitter_5.setOrientation(QtCore.Qt.Vertical)
        self.splitter_5.setObjectName(_fromUtf8("splitter_5"))
        self.splitter_3 = QtWidgets.QSplitter(self.splitter_5)
        self.splitter_3.setOrientation(QtCore.Qt.Vertical)
        self.splitter_3.setObjectName(_fromUtf8("splitter_3"))
        self.groupBox_dragdrop = QtWidgets.QGroupBox(self.splitter_3)
        self.groupBox_dragdrop.setMinimumSize(QtCore.QSize(0, 200))
        self.groupBox_dragdrop.setObjectName(_fromUtf8("groupBox_dragdrop"))
        self.gridLayout_8 = QtWidgets.QGridLayout(self.groupBox_dragdrop)
        self.gridLayout_8.setObjectName(_fromUtf8("gridLayout_8"))
        
        self.table_dragdrop = MyTable(0,10,self.groupBox_dragdrop) #table with 9 columns
        self.table_dragdrop.setObjectName(_fromUtf8("table_dragdrop"))
        header_labels = ["File", "Class" ,"T", "V", "Show","Events total","Events/Epoch","PIX","Shuffle","Zoom"]

        self.table_dragdrop.setHorizontalHeaderLabels(header_labels) 
        header = self.table_dragdrop.horizontalHeader()
        for i in [1,2,3,4,5,6,7,8,9]:#range(len(header_labels)):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        self.table_dragdrop.setAcceptDrops(True)
        self.table_dragdrop.setDragEnabled(True)
        self.table_dragdrop.dropped.connect(self.dataDropped)
        self.table_dragdrop.clicked.connect(self.item_click)
        self.table_dragdrop.itemChanged.connect(self.uncheck_if_zero)
        self.table_dragdrop.doubleClicked.connect(self.item_dclick)
        self.table_dragdrop.itemChanged.connect(self.dataOverviewOn_OnChange)

        self.table_dragdrop.resizeRowsToContents()

        self.table_dragdrop.horizontalHeader().sectionClicked.connect(self.select_all)




        ############################Variables##################################
        #######################################################################
        #Initilaize some variables which are lateron filled in the program
        self.w = None #Initialize a variable for a popup window
        self.threadpool = QtCore.QThreadPool()
        self.threadpool_single = QtCore.QThreadPool()
        self.threadpool_single.setMaxThreadCount(1)
        self.threadpool_single_queue = 0 #count nr. of threads in queue; 

        #self.threadpool_single = QtCore.QThread()
        self.fittingpopups = []  #This app will be designed to allow training of several models ...
        self.fittingpopups_ui = [] #...simultaneously (threading). The info of each model is appended to a list
        self.popupcounter = 0
        self.colorsQt = 10*['red','yellow','blue','cyan','magenta','green','gray','darkRed','darkYellow','darkBlue','darkCyan','darkMagenta','darkGreen','darkGray']    #Some colors which are later used for different subpopulations

        self.model_keras = None #Variable for storing Keras model   
        self.model_keras_path = None
        self.load_model_path = None
        self.loaded_history = None #Variable for storing a loaded history file (for display on History-Tab)
        self.loaded_para = None #Variable for storing a loaded Parameters-file (for display on History-Tab)
        self.plt1 = None #Used for the popup window to display hist and scatter of single experiments
        self.plt2 = None #Used for the history-tab to show accuracy of loaded history files
        self.plt_cm = [] #Used to show images from the interactive Confusion matrix
        self.model_2_convert = None #Variable to store the path to a chosen model (for converting to .nnet)
        self.ram = dict() #Variable to store data if Option "Data to RAM is enabled"
        self.ValidationSet = None
        self.Metrics = dict()
        #self.clip = QtGui.QApplication.clipboard() #This is how one defines a clipboard variable; one can put text on it via:#self.clip.setText("SomeText") 
        self.new_peaks = [] #list to store used defined peaks
        #######################################################################
        #######################################################################








        
        self.gridLayout_8.addWidget(self.table_dragdrop, 0, 0, 1, 1)
        self.splitter_2 = QtWidgets.QSplitter(self.splitter_3)
        self.splitter_2.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_2.setObjectName(_fromUtf8("splitter_2"))
        self.groupBox_DataOverview = QtWidgets.QGroupBox(self.splitter_2)
        self.groupBox_DataOverview.setObjectName(_fromUtf8("groupBox_DataOverview"))
        self.groupBox_DataOverview.setCheckable(True)
        self.groupBox_DataOverview.setChecked(True)
        self.groupBox_DataOverview.toggled.connect(self.dataOverviewOn)
        
        self.gridLayout_5 = QtWidgets.QGridLayout(self.groupBox_DataOverview)
        self.gridLayout_5.setObjectName(_fromUtf8("gridLayout_5"))
        self.tableWidget_Info = QtWidgets.QTableWidget(self.groupBox_DataOverview)
        self.tableWidget_Info.setMinimumSize(QtCore.QSize(0, 0))
        self.tableWidget_Info.setMaximumSize(QtCore.QSize(16777215, 16777215))
        #self.tableWidget_Info.setEditTriggers(QtWidgets.QAbstractItemView.AnyKeyPressed|QtWidgets.QAbstractItemView.DoubleClicked|QtWidgets.QAbstractItemView.EditKeyPressed|QtWidgets.QAbstractItemView.SelectedClicked)
        self.tableWidget_Info.setDragEnabled(False)
        #self.tableWidget_Info.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.tableWidget_Info.setAlternatingRowColors(True)
        self.tableWidget_Info.setObjectName(_fromUtf8("tableWidget_Info"))
        self.tableWidget_Info.setColumnCount(0)
        self.tableWidget_Info.setRowCount(0)
        self.gridLayout_5.addWidget(self.tableWidget_Info, 0, 0, 1, 1)


        self.tabWidget_DefineModel = QtWidgets.QTabWidget(self.splitter_2)
        self.tabWidget_DefineModel.setEnabled(True)
        self.tabWidget_DefineModel.setToolTip("")
        self.tabWidget_DefineModel.setObjectName("tabWidget_DefineModel")
        self.tab_DefineModel = QtWidgets.QWidget()
        self.tab_DefineModel.setObjectName("tab_DefineModel")
        self.gridLayout_11 = QtWidgets.QGridLayout(self.tab_DefineModel)
        self.gridLayout_11.setObjectName("gridLayout_11")
        self.scrollArea_defineModel = QtWidgets.QScrollArea(self.tab_DefineModel)
        self.scrollArea_defineModel.setWidgetResizable(True)
        self.scrollArea_defineModel.setObjectName("scrollArea_defineModel")
        self.scrollAreaWidgetContents_3 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_3.setGeometry(QtCore.QRect(0, 0, 598, 192))
        self.scrollAreaWidgetContents_3.setObjectName("scrollAreaWidgetContents_3")
        self.gridLayout_44 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents_3)
        self.gridLayout_44.setObjectName("gridLayout_44")
        self.gridLayout_newLoadModel = QtWidgets.QGridLayout()
        self.gridLayout_newLoadModel.setObjectName("gridLayout_newLoadModel")
        self.verticalLayout_newLoadModel = QtWidgets.QVBoxLayout()
        self.verticalLayout_newLoadModel.setObjectName("verticalLayout_newLoadModel")
        self.radioButton_NewModel = QtWidgets.QRadioButton(self.scrollAreaWidgetContents_3)
        self.radioButton_NewModel.setMinimumSize(QtCore.QSize(0, 20))
        self.radioButton_NewModel.setMaximumSize(QtCore.QSize(16777215, 20))
        self.radioButton_NewModel.setObjectName("radioButton_NewModel")
        self.verticalLayout_newLoadModel.addWidget(self.radioButton_NewModel)
        self.line_loadModel = QtWidgets.QFrame(self.scrollAreaWidgetContents_3)
        self.line_loadModel.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_loadModel.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_loadModel.setObjectName("line_loadModel")
        self.verticalLayout_newLoadModel.addWidget(self.line_loadModel)
        self.radioButton_LoadRestartModel = QtWidgets.QRadioButton(self.scrollAreaWidgetContents_3)
        self.radioButton_LoadRestartModel.setMinimumSize(QtCore.QSize(0, 20))
        self.radioButton_LoadRestartModel.setMaximumSize(QtCore.QSize(16777215, 20))
        self.radioButton_LoadRestartModel.setObjectName("radioButton_LoadRestartModel")
        self.radioButton_LoadRestartModel.clicked.connect(self.action_preview_model)
        self.verticalLayout_newLoadModel.addWidget(self.radioButton_LoadRestartModel)
        self.radioButton_LoadContinueModel = QtWidgets.QRadioButton(self.scrollAreaWidgetContents_3)
        self.radioButton_LoadContinueModel.setMinimumSize(QtCore.QSize(0, 20))
        self.radioButton_LoadContinueModel.setMaximumSize(QtCore.QSize(16777215, 20))
        self.radioButton_LoadContinueModel.setObjectName("radioButton_LoadContinueModel")
        self.radioButton_LoadContinueModel.clicked.connect(self.action_preview_model)
        self.verticalLayout_newLoadModel.addWidget(self.radioButton_LoadContinueModel)
        
        self.gridLayout_newLoadModel.addLayout(self.verticalLayout_newLoadModel, 0, 0, 1, 1)
        self.verticalLayout_newLoadModel_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_newLoadModel_2.setObjectName("verticalLayout_newLoadModel_2")
        self.comboBox_ModelSelection = QtWidgets.QComboBox(self.scrollAreaWidgetContents_3)
        self.comboBox_ModelSelection.setMinimumSize(QtCore.QSize(0, 20))
        self.comboBox_ModelSelection.setMaximumSize(QtCore.QSize(16777215, 20))
        self.comboBox_ModelSelection.setObjectName("comboBox_ModelSelection")
        self.predefined_models = ["None"] + model_zoo.get_predefined_models()
        self.comboBox_ModelSelection.addItems(self.predefined_models)        
        self.verticalLayout_newLoadModel_2.addWidget(self.comboBox_ModelSelection)
        self.line_loadModel_2 = QtWidgets.QFrame(self.scrollAreaWidgetContents_3)
        self.line_loadModel_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_loadModel_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_loadModel_2.setObjectName("line_loadModel_2")
        self.verticalLayout_newLoadModel_2.addWidget(self.line_loadModel_2)
        self.lineEdit_LoadModelPath = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.lineEdit_LoadModelPath.setMinimumSize(QtCore.QSize(0, 40))
        self.lineEdit_LoadModelPath.setMaximumSize(QtCore.QSize(16777215, 40))
        self.lineEdit_LoadModelPath.setObjectName("lineEdit_LoadModelPath")
        self.verticalLayout_newLoadModel_2.addWidget(self.lineEdit_LoadModelPath)
        self.gridLayout_newLoadModel.addLayout(self.verticalLayout_newLoadModel_2, 0, 1, 1, 1)
        self.gridLayout_44.addLayout(self.gridLayout_newLoadModel, 0, 0, 1, 1)
        self.gridLayout_cropNormEtc = QtWidgets.QGridLayout()
        self.gridLayout_cropNormEtc.setObjectName("gridLayout_cropNormEtc")
        self.horizontalLayout_colorMode = QtWidgets.QHBoxLayout()
        self.horizontalLayout_colorMode.setObjectName("horizontalLayout_colorMode")
        self.label_colorModeIcon = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_colorModeIcon.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_colorModeIcon.setObjectName("label_colorModeIcon")
        self.horizontalLayout_colorMode.addWidget(self.label_colorModeIcon)
        self.label_colorMode = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_colorMode.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_colorMode.setObjectName("label_colorMode")
        self.label_colorMode.setMinimumSize(QtCore.QSize(55,22))
        
        self.horizontalLayout_colorMode.addWidget(self.label_colorMode)
        self.gridLayout_cropNormEtc.addLayout(self.horizontalLayout_colorMode, 1, 2, 1, 1)
        self.horizontalLayout_nrEpochs = QtWidgets.QHBoxLayout()
        self.horizontalLayout_nrEpochs.setObjectName("horizontalLayout_nrEpochs")
        self.label_nrEpochsIcon = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_nrEpochsIcon.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_nrEpochsIcon.setObjectName("label_nrEpochsIcon")
        self.horizontalLayout_nrEpochs.addWidget(self.label_nrEpochsIcon)
        self.label_nrEpochs = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_nrEpochs.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_nrEpochs.setObjectName("label_nrEpochs")
        #self.label_nrEpochs.setMaximumSize(QtCore.QSize(50, 22))
        self.horizontalLayout_nrEpochs.addWidget(self.label_nrEpochs)
        self.gridLayout_cropNormEtc.addLayout(self.horizontalLayout_nrEpochs, 1, 0, 1, 1)
        self.horizontalLayout_normalization = QtWidgets.QHBoxLayout()
        self.horizontalLayout_normalization.setObjectName("horizontalLayout_normalization")
        self.label_NormalizationIcon = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_NormalizationIcon.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        #self.label_NormalizationIcon.setText("")
        self.label_NormalizationIcon.setObjectName("label_NormalizationIcon")
        self.horizontalLayout_normalization.addWidget(self.label_NormalizationIcon)
        self.label_Normalization = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_Normalization.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_Normalization.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_Normalization.setObjectName("label_Normalization")
        self.horizontalLayout_normalization.addWidget(self.label_Normalization)
        self.gridLayout_cropNormEtc.addLayout(self.horizontalLayout_normalization, 0, 2, 1, 1)
        self.horizontalLayout_crop = QtWidgets.QHBoxLayout()
        self.horizontalLayout_crop.setObjectName("horizontalLayout_crop")
        self.label_CropSpace = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_CropSpace.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_CropSpace.setObjectName("label_CropSpace")
        self.horizontalLayout_crop.addWidget(self.label_CropSpace)

        self.label_CropIcon = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_CropIcon.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_CropIcon.setObjectName("label_CropIcon")
        self.label_CropIcon.setMinimumSize(QtCore.QSize(20, 20))
        self.label_CropIcon.setMaximumSize(QtCore.QSize(20, 20))
        self.horizontalLayout_crop.addWidget(self.label_CropIcon)
        self.label_Crop = QtWidgets.QLabel(self.scrollAreaWidgetContents_3)
        self.label_Crop.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_Crop.setObjectName("label_Crop")
        self.horizontalLayout_crop.addWidget(self.label_Crop)
        self.gridLayout_cropNormEtc.addLayout(self.horizontalLayout_crop, 0, 0, 1, 1)
        self.comboBox_GrayOrRGB = QtWidgets.QComboBox(self.scrollAreaWidgetContents_3)
        self.comboBox_GrayOrRGB.setObjectName("comboBox_GrayOrRGB")
        self.gridLayout_cropNormEtc.addWidget(self.comboBox_GrayOrRGB, 1, 3, 1, 1)
        self.comboBox_Normalization = QtWidgets.QComboBox(self.scrollAreaWidgetContents_3)
        self.comboBox_Normalization.setMinimumSize(QtCore.QSize(200, 0))
        self.norm_methods = Default_dict["norm_methods"] #["None","Div. by 255", "StdScaling using mean and std of each image individually","StdScaling using mean and std of all training data"]
        self.comboBox_Normalization.addItems(self.norm_methods)
        self.comboBox_Normalization.setMinimumSize(QtCore.QSize(200,22))
        self.comboBox_Normalization.setMaximumSize(QtCore.QSize(200, 22))
        width=self.comboBox_Normalization.fontMetrics().boundingRect(max(self.norm_methods, key=len)).width()
        self.comboBox_Normalization.view().setFixedWidth(width+10)             
        self.comboBox_Normalization.setObjectName("comboBox_Normalization")
        self.gridLayout_cropNormEtc.addWidget(self.comboBox_Normalization, 0, 3, 1, 1)
        self.spinBox_imagecrop = QtWidgets.QSpinBox(self.scrollAreaWidgetContents_3)
        self.spinBox_imagecrop.setObjectName("spinBox_imagecrop")
        self.spinBox_imagecrop.setMinimum(1)
        self.spinBox_imagecrop.setMaximum(9E8)
        self.spinBox_imagecrop.setMaximumSize(QtCore.QSize(100, 22))
        self.gridLayout_cropNormEtc.addWidget(self.spinBox_imagecrop, 0, 1, 1, 1)
        self.spinBox_NrEpochs = QtWidgets.QSpinBox(self.scrollAreaWidgetContents_3)
        self.spinBox_NrEpochs.setObjectName("spinBox_NrEpochs")
        self.spinBox_NrEpochs.setMinimum(1)
        self.spinBox_NrEpochs.setMaximum(9E8)

        self.gridLayout_cropNormEtc.addWidget(self.spinBox_NrEpochs, 1, 1, 1, 1)
        self.gridLayout_44.addLayout(self.gridLayout_cropNormEtc, 1, 0, 1, 1)
        self.horizontalLayout_modelname = QtWidgets.QHBoxLayout()
        self.horizontalLayout_modelname.setObjectName("horizontalLayout_modelname")
        self.pushButton_modelname = QtWidgets.QPushButton(self.scrollAreaWidgetContents_3)
        self.pushButton_modelname.setObjectName("pushButton_modelname")
        self.pushButton_modelname.clicked.connect(self.action_set_modelpath_and_name)

        self.horizontalLayout_modelname.addWidget(self.pushButton_modelname)
        self.lineEdit_modelname = QtWidgets.QLineEdit(self.scrollAreaWidgetContents_3)
        self.lineEdit_modelname.setMinimumSize(QtCore.QSize(0, 22))
        self.lineEdit_modelname.setMaximumSize(QtCore.QSize(16777215, 22))
        self.lineEdit_modelname.setObjectName("lineEdit_modelname")
        self.horizontalLayout_modelname.addWidget(self.lineEdit_modelname)
        self.gridLayout_44.addLayout(self.horizontalLayout_modelname, 2, 0, 1, 1)
        self.scrollArea_defineModel.setWidget(self.scrollAreaWidgetContents_3)
        self.gridLayout_11.addWidget(self.scrollArea_defineModel, 0, 0, 1, 1)
        self.tabWidget_DefineModel.addTab(self.tab_DefineModel, "")

















        
        
        
        
        
        
        
        
        
        self.tab_kerasAug = QtWidgets.QWidget()
        self.tab_kerasAug.setObjectName(_fromUtf8("tab_kerasAug"))
        self.gridLayout_7 = QtWidgets.QGridLayout(self.tab_kerasAug)
        self.gridLayout_7.setObjectName(_fromUtf8("gridLayout_7"))
        self.verticalLayout_8 = QtWidgets.QVBoxLayout()
        self.verticalLayout_8.setObjectName(_fromUtf8("verticalLayout_8"))
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setObjectName(_fromUtf8("horizontalLayout_9"))
        self.label_RefreshAfterEpochs = QtWidgets.QLabel(self.tab_kerasAug)
        self.label_RefreshAfterEpochs.setObjectName(_fromUtf8("label_RefreshAfterEpochs"))
        self.horizontalLayout_9.addWidget(self.label_RefreshAfterEpochs)
        self.spinBox_RefreshAfterEpochs = QtWidgets.QSpinBox(self.tab_kerasAug)
        self.spinBox_RefreshAfterEpochs.setObjectName(_fromUtf8("spinBox_RefreshAfterEpochs"))
        self.spinBox_RefreshAfterEpochs.setMinimum(1)
        self.spinBox_RefreshAfterEpochs.setMaximum(9E8)
        self.horizontalLayout_9.addWidget(self.spinBox_RefreshAfterEpochs)
        self.verticalLayout_8.addLayout(self.horizontalLayout_9)
        self.verticalLayout_7 = QtWidgets.QVBoxLayout()
        self.verticalLayout_7.setObjectName(_fromUtf8("verticalLayout_7"))
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_8.setObjectName(_fromUtf8("horizontalLayout_8"))
        self.checkBox_HorizFlip = QtWidgets.QCheckBox(self.tab_kerasAug)
        self.checkBox_HorizFlip.setObjectName(_fromUtf8("checkBox_HorizFlip"))
        self.horizontalLayout_8.addWidget(self.checkBox_HorizFlip)
        self.checkBox_VertFlip = QtWidgets.QCheckBox(self.tab_kerasAug)
        self.checkBox_VertFlip.setObjectName(_fromUtf8("checkBox_VertFlip"))
        self.horizontalLayout_8.addWidget(self.checkBox_VertFlip)
        self.verticalLayout_7.addLayout(self.horizontalLayout_8)
        self.splitter = QtWidgets.QSplitter(self.tab_kerasAug)
        self.splitter.setOrientation(QtCore.Qt.Horizontal)
        self.splitter.setObjectName(_fromUtf8("splitter"))
        self.widget = QtWidgets.QWidget(self.splitter)
        self.widget.setObjectName(_fromUtf8("widget"))
        self.verticalLayout_6 = QtWidgets.QVBoxLayout(self.widget)
        self.verticalLayout_6.setContentsMargins(0, 0, 0, 0)
        
        self.onlyFloat = QtGui.QDoubleValidator()
        
        self.verticalLayout_6.setObjectName(_fromUtf8("verticalLayout_6"))
        self.label_Rotation = QtWidgets.QCheckBox(self.widget)
        self.label_Rotation.setObjectName(_fromUtf8("label_Rotation"))
        self.verticalLayout_6.addWidget(self.label_Rotation)
        self.label_width_shift = QtWidgets.QCheckBox(self.widget)
        self.label_width_shift.setObjectName(_fromUtf8("label_width_shift"))
        self.verticalLayout_6.addWidget(self.label_width_shift)
        self.label_height_shift = QtWidgets.QCheckBox(self.widget)
        self.label_height_shift.setObjectName(_fromUtf8("label_height_shift"))
        self.verticalLayout_6.addWidget(self.label_height_shift)
        self.label_zoom = QtWidgets.QCheckBox(self.widget)
        self.label_zoom.setObjectName(_fromUtf8("label_zoom"))
        self.verticalLayout_6.addWidget(self.label_zoom)
        self.label_shear = QtWidgets.QCheckBox(self.widget)
        self.label_shear.setObjectName(_fromUtf8("label_shear"))
        self.verticalLayout_6.addWidget(self.label_shear)
        self.widget1 = QtWidgets.QWidget(self.splitter)
        self.widget1.setObjectName(_fromUtf8("widget1"))
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.widget1)
        self.verticalLayout_5.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_5.setObjectName(_fromUtf8("verticalLayout_5"))
        self.lineEdit_Rotation = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit_Rotation.setObjectName(_fromUtf8("lineEdit_Rotation"))
        self.lineEdit_Rotation.setValidator(self.onlyFloat)
        self.verticalLayout_5.addWidget(self.lineEdit_Rotation)
        self.lineEdit_Rotation_2 = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit_Rotation_2.setObjectName(_fromUtf8("lineEdit_Rotation_2"))
        self.lineEdit_Rotation_2.setValidator(self.onlyFloat)

        self.verticalLayout_5.addWidget(self.lineEdit_Rotation_2)
        self.lineEdit_Rotation_3 = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit_Rotation_3.setObjectName(_fromUtf8("lineEdit_Rotation_3"))
        self.lineEdit_Rotation_3.setValidator(self.onlyFloat)

        self.verticalLayout_5.addWidget(self.lineEdit_Rotation_3)
        self.lineEdit_Rotation_4 = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit_Rotation_4.setObjectName(_fromUtf8("lineEdit_Rotation_4"))
        self.lineEdit_Rotation_4.setValidator(self.onlyFloat)

        self.verticalLayout_5.addWidget(self.lineEdit_Rotation_4)
        self.lineEdit_Rotation_5 = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit_Rotation_5.setObjectName(_fromUtf8("lineEdit_Rotation_5"))
        self.lineEdit_Rotation_5.setValidator(self.onlyFloat)

        self.verticalLayout_5.addWidget(self.lineEdit_Rotation_5)
        self.verticalLayout_7.addWidget(self.splitter)
        self.verticalLayout_8.addLayout(self.verticalLayout_7)
        self.gridLayout_7.addLayout(self.verticalLayout_8, 0, 0, 1, 1)
        self.tabWidget_DefineModel.addTab(self.tab_kerasAug, _fromUtf8(""))
        
        
        
        
        
        self.tab_BrightnessAug = QtWidgets.QWidget()
        self.tab_BrightnessAug.setObjectName("tab_BrightnessAug")
        self.gridLayout_42 = QtWidgets.QGridLayout(self.tab_BrightnessAug)
        self.gridLayout_42.setObjectName("gridLayout_42")
        self.scrollArea_2 = QtWidgets.QScrollArea(self.tab_BrightnessAug)
        self.scrollArea_2.setWidgetResizable(True)
        self.scrollArea_2.setObjectName("scrollArea_2")
        self.scrollAreaWidgetContents_2 = QtWidgets.QWidget()
        self.scrollAreaWidgetContents_2.setGeometry(QtCore.QRect(0, -2, 449, 269))
        self.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents_2")
        self.gridLayout_43 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents_2)
        self.gridLayout_43.setObjectName("gridLayout_43")
        self.groupBox_GaussianNoise = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.groupBox_GaussianNoise.setObjectName("groupBox_GaussianNoise")
        self.gridLayout_13 = QtWidgets.QGridLayout(self.groupBox_GaussianNoise)
        self.gridLayout_13.setObjectName("gridLayout_13")
        self.label_GaussianNoiseMean = QtWidgets.QCheckBox(self.groupBox_GaussianNoise)
        self.label_GaussianNoiseMean.setObjectName("label_GaussianNoiseMean")
        self.gridLayout_13.addWidget(self.label_GaussianNoiseMean, 0, 0, 1, 1)
        self.doubleSpinBox_GaussianNoiseMean = QtWidgets.QDoubleSpinBox(self.groupBox_GaussianNoise)
        self.doubleSpinBox_GaussianNoiseMean.setObjectName("spinBox_GaussianNoiseMean")
        self.gridLayout_13.addWidget(self.doubleSpinBox_GaussianNoiseMean, 0, 1, 1, 1)
        self.label_GaussianNoiseScale = QtWidgets.QCheckBox(self.groupBox_GaussianNoise)
        self.label_GaussianNoiseScale.setObjectName("label_GaussianNoiseScale")
        self.gridLayout_13.addWidget(self.label_GaussianNoiseScale, 1, 0, 1, 1)
        self.doubleSpinBox_GaussianNoiseScale = QtWidgets.QDoubleSpinBox(self.groupBox_GaussianNoise)
        self.doubleSpinBox_GaussianNoiseScale.setObjectName("spinBox_GaussianNoiseScale")
        self.gridLayout_13.addWidget(self.doubleSpinBox_GaussianNoiseScale, 1, 1, 1, 1)
        self.gridLayout_43.addWidget(self.groupBox_GaussianNoise, 2, 1, 1, 1)
        self.groupBox_colorAugmentation = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.groupBox_colorAugmentation.setCheckable(False)
        self.groupBox_colorAugmentation.setObjectName("groupBox_colorAugmentation")
        self.gridLayout_15 = QtWidgets.QGridLayout(self.groupBox_colorAugmentation)
        self.gridLayout_15.setObjectName("gridLayout_15")
        self.doubleSpinBox_contrastLower = QtWidgets.QDoubleSpinBox(self.groupBox_colorAugmentation)
        self.doubleSpinBox_contrastLower.setObjectName("doubleSpinBox_contrastLower")
        self.gridLayout_15.addWidget(self.doubleSpinBox_contrastLower, 0, 1, 1, 1)
        self.doubleSpinBox_saturationHigher = QtWidgets.QDoubleSpinBox(self.groupBox_colorAugmentation)
        self.doubleSpinBox_saturationHigher.setObjectName("doubleSpinBox_saturationHigher")
        self.gridLayout_15.addWidget(self.doubleSpinBox_saturationHigher, 1, 2, 1, 1)
        self.doubleSpinBox_contrastHigher = QtWidgets.QDoubleSpinBox(self.groupBox_colorAugmentation)
        self.doubleSpinBox_contrastHigher.setObjectName("doubleSpinBox_contrastHigher")
        self.gridLayout_15.addWidget(self.doubleSpinBox_contrastHigher, 0, 2, 1, 1)
        self.checkBox_contrast = QtWidgets.QCheckBox(self.groupBox_colorAugmentation)
        self.checkBox_contrast.setCheckable(True)
        self.checkBox_contrast.setObjectName("checkBox_contrast")
        self.gridLayout_15.addWidget(self.checkBox_contrast, 0, 0, 1, 1)
        self.doubleSpinBox_saturationLower = QtWidgets.QDoubleSpinBox(self.groupBox_colorAugmentation)
        self.doubleSpinBox_saturationLower.setObjectName("doubleSpinBox_saturationLower")
        self.gridLayout_15.addWidget(self.doubleSpinBox_saturationLower, 1, 1, 1, 1)
        self.doubleSpinBox_hueDelta = QtWidgets.QDoubleSpinBox(self.groupBox_colorAugmentation)
        self.doubleSpinBox_hueDelta.setObjectName("doubleSpinBox_hueDelta")
        self.gridLayout_15.addWidget(self.doubleSpinBox_hueDelta, 2, 1, 1, 1)
        self.checkBox_saturation = QtWidgets.QCheckBox(self.groupBox_colorAugmentation)
        self.checkBox_saturation.setObjectName("checkBox_saturation")
        self.gridLayout_15.addWidget(self.checkBox_saturation, 1, 0, 1, 1)
        self.checkBox_hue = QtWidgets.QCheckBox(self.groupBox_colorAugmentation)
        self.checkBox_hue.setObjectName("checkBox_hue")
        self.gridLayout_15.addWidget(self.checkBox_hue, 2, 0, 1, 1)
        self.gridLayout_43.addWidget(self.groupBox_colorAugmentation, 3, 0, 1, 1)
        
        self.groupBox_blurringAug = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.groupBox_blurringAug.setObjectName("groupBox_blurringAug")
        self.gridLayout_45 = QtWidgets.QGridLayout(self.groupBox_blurringAug)
        self.gridLayout_45.setObjectName("gridLayout_45")
        self.gridLayout_blur = QtWidgets.QGridLayout()
        self.gridLayout_blur.setObjectName("gridLayout_blur")
        self.label_avgBlurMin = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_avgBlurMin.setMaximumSize(QtCore.QSize(31, 16777215))
        self.label_avgBlurMin.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_avgBlurMin.setObjectName("label_avgBlurMin")
        self.gridLayout_blur.addWidget(self.label_avgBlurMin, 0, 1, 1, 1)
        self.spinBox_avgBlurMin = QtWidgets.QSpinBox(self.groupBox_blurringAug)
        self.spinBox_avgBlurMin.setObjectName("spinBox_avgBlurMin")
        self.gridLayout_blur.addWidget(self.spinBox_avgBlurMin, 0, 2, 1, 1)
        self.spinBox_avgBlurMax = QtWidgets.QSpinBox(self.groupBox_blurringAug)
        self.spinBox_avgBlurMax.setObjectName("spinBox_avgBlurMax")
        self.gridLayout_blur.addWidget(self.spinBox_avgBlurMax, 0, 4, 1, 1)
        self.checkBox_avgBlur = QtWidgets.QCheckBox(self.groupBox_blurringAug)
        self.checkBox_avgBlur.setObjectName("checkBox_avgBlur")
        self.gridLayout_blur.addWidget(self.checkBox_avgBlur, 0, 0, 1, 1)
        self.label_avgBlurMax = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_avgBlurMax.setMaximumSize(QtCore.QSize(31, 16777215))
        self.label_avgBlurMax.setObjectName("label_avgBlurMax")
        self.gridLayout_blur.addWidget(self.label_avgBlurMax, 0, 3, 1, 1)
        self.checkBox_avgBlur.setCheckable(True)
        self.spinBox_gaussBlurMax = QtWidgets.QSpinBox(self.groupBox_blurringAug)
        self.spinBox_gaussBlurMax.setObjectName("spinBox_gaussBlurMax")
        self.gridLayout_blur.addWidget(self.spinBox_gaussBlurMax, 1, 4, 1, 1)
        self.spinBox_gaussBlurMin = QtWidgets.QSpinBox(self.groupBox_blurringAug)
        self.spinBox_gaussBlurMin.setObjectName("spinBox_gaussBlurMin")
        self.gridLayout_blur.addWidget(self.spinBox_gaussBlurMin, 1, 2, 1, 1)
        self.label_gaussBlurMin = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_gaussBlurMin.setMaximumSize(QtCore.QSize(31, 16777215))
        self.label_gaussBlurMin.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_gaussBlurMin.setObjectName("label_gaussBlurMin")
        self.gridLayout_blur.addWidget(self.label_gaussBlurMin, 1, 1, 1, 1)
        self.checkBox_gaussBlur = QtWidgets.QCheckBox(self.groupBox_blurringAug)
        self.checkBox_gaussBlur.setObjectName("checkBox_gaussBlur")
        self.gridLayout_blur.addWidget(self.checkBox_gaussBlur, 1, 0, 1, 1)
        self.label_gaussBlurMax = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_gaussBlurMax.setMaximumSize(QtCore.QSize(31, 16777215))
        self.label_gaussBlurMax.setObjectName("label_gaussBlurMax")
        self.gridLayout_blur.addWidget(self.label_gaussBlurMax, 1, 3, 1, 1)
        self.checkBox_gaussBlur.setCheckable(True)
        self.label_motionBlurKernel = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_motionBlurKernel.setMaximumSize(QtCore.QSize(31, 16777215))
        self.label_motionBlurKernel.setObjectName("label_motionBlurKernel")
        self.gridLayout_blur.addWidget(self.label_motionBlurKernel, 2, 1, 1, 1)
        self.lineEdit_motionBlurAngle = QtWidgets.QLineEdit(self.groupBox_blurringAug)
        validator = QtGui.QRegExpValidator(QtCore.QRegExp("^[-+]?[0-9]\\d{0,3},(\\d{3})$"))
        self.lineEdit_motionBlurAngle.setValidator(validator)
        self.lineEdit_motionBlurAngle.setMaximumSize(QtCore.QSize(100, 16777215))
        self.lineEdit_motionBlurAngle.setInputMask("")
        self.lineEdit_motionBlurAngle.setObjectName("lineEdit_motionBlurAngle")
        self.gridLayout_blur.addWidget(self.lineEdit_motionBlurAngle, 2, 4, 1, 1)
        self.checkBox_motionBlur = QtWidgets.QCheckBox(self.groupBox_blurringAug)
        self.checkBox_motionBlur.setMaximumSize(QtCore.QSize(100, 16777215))
        self.checkBox_motionBlur.setObjectName("checkBox_motionBlur")
        self.gridLayout_blur.addWidget(self.checkBox_motionBlur, 2, 0, 1, 1)
        self.label_motionBlurAngle = QtWidgets.QLabel(self.groupBox_blurringAug)
        self.label_motionBlurAngle.setMaximumSize(QtCore.QSize(16777215, 16777215))
        self.label_motionBlurAngle.setObjectName("label_motionBlurAngle")
        self.gridLayout_blur.addWidget(self.label_motionBlurAngle, 2, 3, 1, 1)
        validator = QtGui.QRegExpValidator(QtCore.QRegExp("^\\d{1,3},(\\d{3})$"))
        self.lineEdit_motionBlurKernel = QtWidgets.QLineEdit(self.groupBox_blurringAug)
        self.lineEdit_motionBlurKernel.setValidator(validator)
        self.lineEdit_motionBlurKernel.setMaximumSize(QtCore.QSize(100, 16777215))
        self.lineEdit_motionBlurKernel.setInputMask("")
        self.lineEdit_motionBlurKernel.setMaxLength(32767)
        self.lineEdit_motionBlurKernel.setObjectName("lineEdit_motionBlurKernel")
        self.gridLayout_blur.addWidget(self.lineEdit_motionBlurKernel, 2, 2, 1, 1)
        self.gridLayout_45.addLayout(self.gridLayout_blur, 0, 0, 1, 1)
        self.gridLayout_43.addWidget(self.groupBox_blurringAug, 3, 1, 1, 1)
        self.checkBox_motionBlur.setCheckable(True)
        self.lineEdit_motionBlurKernel.setClearButtonEnabled(False)        
        


 
        self.groupBox_BrightnessAugmentation = QtWidgets.QGroupBox(self.scrollAreaWidgetContents_2)
        self.groupBox_BrightnessAugmentation.setObjectName("groupBox_BrightnessAugmentation")
        self.gridLayout_12 = QtWidgets.QGridLayout(self.groupBox_BrightnessAugmentation)
        self.gridLayout_12.setObjectName("gridLayout_12")
        self.label_Plus = QtWidgets.QCheckBox(self.groupBox_BrightnessAugmentation)
        self.label_Plus.setObjectName("label_Plus")
        self.gridLayout_12.addWidget(self.label_Plus, 1, 0, 1, 1)
        self.doubleSpinBox_MultLower = QtWidgets.QDoubleSpinBox(self.groupBox_BrightnessAugmentation)
        self.doubleSpinBox_MultLower.setObjectName("doubleSpinBox_MultLower")
        self.gridLayout_12.addWidget(self.doubleSpinBox_MultLower, 2, 1, 1, 1)
        self.spinBox_PlusUpper = QtWidgets.QSpinBox(self.groupBox_BrightnessAugmentation)
        self.spinBox_PlusUpper.setObjectName("spinBox_PlusUpper")
        self.gridLayout_12.addWidget(self.spinBox_PlusUpper, 1, 2, 1, 2)
        self.spinBox_PlusLower = QtWidgets.QSpinBox(self.groupBox_BrightnessAugmentation)
        self.spinBox_PlusLower.setObjectName("spinBox_PlusLower")
        self.gridLayout_12.addWidget(self.spinBox_PlusLower, 1, 1, 1, 1)
        self.label_Mult = QtWidgets.QCheckBox(self.groupBox_BrightnessAugmentation)
        self.label_Mult.setObjectName("label_Mult")
        self.gridLayout_12.addWidget(self.label_Mult, 2, 0, 1, 1)
        self.doubleSpinBox_MultUpper = QtWidgets.QDoubleSpinBox(self.groupBox_BrightnessAugmentation)
        self.doubleSpinBox_MultUpper.setObjectName("doubleSpinBox_MultUpper")
        self.gridLayout_12.addWidget(self.doubleSpinBox_MultUpper, 2, 2, 1, 2)
        self.gridLayout_43.addWidget(self.groupBox_BrightnessAugmentation, 2, 0, 1, 1)
        self.spinBox_RefreshAfterNrEpochs = QtWidgets.QSpinBox(self.scrollAreaWidgetContents_2)
        self.spinBox_RefreshAfterNrEpochs.setObjectName("spinBox_RefreshAfterNrEpochs")
        self.gridLayout_43.addWidget(self.spinBox_RefreshAfterNrEpochs, 1, 1, 1, 1)
        self.label_RefreshAfterNrEpochs = QtWidgets.QLabel(self.scrollAreaWidgetContents_2)
        self.label_RefreshAfterNrEpochs.setObjectName("label_RefreshAfterNrEpochs")
        self.gridLayout_43.addWidget(self.label_RefreshAfterNrEpochs, 1, 0, 1, 1)
        self.scrollArea_2.setWidget(self.scrollAreaWidgetContents_2)
        self.gridLayout_42.addWidget(self.scrollArea_2, 0, 0, 1, 1)
        self.tabWidget_DefineModel.addTab(self.tab_BrightnessAug, "")
        

        #################################ICONS#################################
        #use full ABSOLUTE path to the image, not relative
        self.radioButton_NewModel.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"model_new.png")))
        self.radioButton_LoadRestartModel.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"model_restart.png")))
        self.radioButton_LoadContinueModel.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"model_continue.png")))
        self.label_CropIcon.setPixmap(QtGui.QPixmap(os.path.join(dir_root,"art",Default_dict["Icon theme"],"cropping.png")))
        self.label_CropIcon.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.pushButton_modelname.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"model_path.png")))

        #self.label_CropIcon.setText("<html><img src='./art/cropping.png'> <p>Input image size</p> </html>")
        self.label_colorModeIcon.setPixmap(QtGui.QPixmap(os.path.join(dir_root,"art",Default_dict["Icon theme"],"color_mode.png")))
        self.label_NormalizationIcon.setPixmap(QtGui.QPixmap(os.path.join(dir_root,"art",Default_dict["Icon theme"],"normalization.png")))
        self.label_nrEpochsIcon.setPixmap(QtGui.QPixmap(os.path.join(dir_root,"art",Default_dict["Icon theme"],"nr_epochs.png")))

        self.checkBox_HorizFlip.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"horizontal_flip.png")))
        self.checkBox_VertFlip.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"vertical_flip.png")))
        self.label_Rotation.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"rotation.png")))
        self.label_Rotation.setChecked(True)
        self.label_Rotation.stateChanged.connect(self.keras_changed_rotation)
        self.label_width_shift.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"width_shift.png")))
        self.label_width_shift.setChecked(True)
        self.label_width_shift.stateChanged.connect(self.keras_changed_width_shift)
        self.label_height_shift.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"height_shift.png")))
        self.label_height_shift.setChecked(True)
        self.label_height_shift.stateChanged.connect(self.keras_changed_height_shift)
        self.label_zoom.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"zoom.png")))
        self.label_zoom.setChecked(True)
        self.label_zoom.stateChanged.connect(self.keras_changed_zoom)
        self.label_shear.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"shear.png")))
        self.label_shear.setChecked(True)
        self.label_shear.stateChanged.connect(self.keras_changed_shear)
        self.label_Plus.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"brightness_plus.png")))
        self.label_Plus.setChecked(True)
        self.label_Plus.stateChanged.connect(self.keras_changed_brightplus)
        self.label_Mult.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"brightness_mult.png")))
        self.label_Mult.setChecked(True)
        self.label_Mult.stateChanged.connect(self.keras_changed_brightmult)
        self.label_GaussianNoiseMean.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"gaussian_noise_mean.png")))
        self.label_GaussianNoiseMean.setChecked(True)
        self.label_GaussianNoiseMean.stateChanged.connect(self.keras_changed_noiseMean)
        self.label_GaussianNoiseScale.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"gaussian_noise_scale.png")))
        self.label_GaussianNoiseScale.setChecked(True)
        self.label_GaussianNoiseScale.stateChanged.connect(self.keras_changed_noiseScale)
        self.checkBox_contrast.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"contrast.png")))
        self.checkBox_contrast.stateChanged.connect(self.keras_changed_contrast)
        self.checkBox_saturation.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"saturation.png")))
        self.checkBox_saturation.stateChanged.connect(self.keras_changed_saturation)
        self.checkBox_hue.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"hue.png")))
        self.checkBox_hue.stateChanged.connect(self.keras_changed_hue)
        self.checkBox_avgBlur.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"average_blur.png")))
        #self.checkBox_avgBlur.stateChanged.connect(self.changed_averageBlur)
        self.checkBox_gaussBlur.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"gaussian_blur.png")))
        #self.checkBox_gaussBlur.stateChanged.connect(self.changed_gaussBlur)
        self.checkBox_motionBlur.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"motion_blur.png")))
        #self.checkBox_motionBlur.stateChanged.connect(self.changed_motionBlur)


        #There will be text on the label_colorMode (Color Mode), find out how "long" the text is and 
        #resize the label to that
        width=self.label_colorMode.fontMetrics().boundingRect(max(["Color Mode"], key=len)).width()
        height = self.label_colorMode.geometry().height()
        self.label_colorMode.setMaximumSize(QtCore.QSize(width, height))


        #Manual values on Build Model Tab
        self.spinBox_PlusUpper.setMinimum(-255)
        self.spinBox_PlusUpper.setMaximum(255)
        self.spinBox_PlusUpper.setSingleStep(1)
        self.spinBox_PlusLower.setMinimum(-255)
        self.spinBox_PlusLower.setMaximum(255)
        self.spinBox_PlusLower.setSingleStep(1)

        self.doubleSpinBox_MultLower.setMinimum(0)
        self.doubleSpinBox_MultLower.setMaximum(999999999)
        self.doubleSpinBox_MultLower.setSingleStep(0.1)
        self.doubleSpinBox_MultUpper.setMinimum(0)
        self.doubleSpinBox_MultUpper.setMaximum(999999999)
        self.doubleSpinBox_MultUpper.setSingleStep(0.1)

        self.doubleSpinBox_GaussianNoiseMean.setMinimum(-255)
        self.doubleSpinBox_GaussianNoiseMean.setMaximum(255)
        self.doubleSpinBox_GaussianNoiseMean.setSingleStep(0.1)

        self.doubleSpinBox_GaussianNoiseScale.setMinimum(0)
        self.doubleSpinBox_GaussianNoiseScale.setMaximum(999999999)
        self.doubleSpinBox_GaussianNoiseScale.setSingleStep(0.1)

        self.spinBox_RefreshAfterNrEpochs.setMinimum(1)
        self.spinBox_RefreshAfterNrEpochs.setMaximum(999999999)
        self.doubleSpinBox_hueDelta.setMaximum(0.5)
        self.doubleSpinBox_hueDelta.setSingleStep(0.01)
        self.doubleSpinBox_contrastHigher.setMaximum(100.0)
        self.doubleSpinBox_contrastHigher.setSingleStep(0.1)
        self.doubleSpinBox_contrastLower.setMaximum(100.0)
        self.doubleSpinBox_contrastLower.setSingleStep(0.1)

        self.doubleSpinBox_saturationLower.setMaximum(100.0)
        self.doubleSpinBox_saturationLower.setSingleStep(0.1)
        self.doubleSpinBox_saturationHigher.setMaximum(100.0)
        self.doubleSpinBox_saturationHigher.setSingleStep(0.1)
        
        self.spinBox_avgBlurMin.setMinimum(0)
        self.spinBox_avgBlurMin.setMaximum(255)
        #self.spinBox_avgBlurMin.setSingleStep(1)
        self.spinBox_avgBlurMax.setMinimum(0)
        self.spinBox_avgBlurMax.setMaximum(255)
        #self.spinBox_avgBlurMax.setSingleStep(0.1)
        
        self.spinBox_gaussBlurMin.setMinimum(0)
        self.spinBox_gaussBlurMin.setMaximum(255)
        #self.spinBox_gaussBlurMin.setSingleStep(0.1)
        self.spinBox_gaussBlurMax.setMinimum(0)
        self.spinBox_gaussBlurMax.setMaximum(255)
        #self.spinBox_gaussBlurMax.setSingleStep(0.1)
               
        
        
        

        
        
        
        
        
        
        
        self.tab_ExampleImgs = QtWidgets.QWidget()
        self.tab_ExampleImgs.setObjectName(_fromUtf8("tab_ExampleImgs"))
        self.gridLayout_9 = QtWidgets.QGridLayout(self.tab_ExampleImgs)
        self.gridLayout_9.setObjectName(_fromUtf8("gridLayout_9"))
        self.splitter_4 = QtWidgets.QSplitter(self.tab_ExampleImgs)
        self.splitter_4.setOrientation(QtCore.Qt.Vertical)
        self.splitter_4.setObjectName(_fromUtf8("splitter_4"))
        self.widget2 = QtWidgets.QWidget(self.splitter_4)
        self.widget2.setObjectName(_fromUtf8("widget2"))
        self.horizontalLayout_ExampleImgs = QtWidgets.QHBoxLayout(self.widget2)
        self.horizontalLayout_ExampleImgs.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_ExampleImgs.setObjectName(_fromUtf8("horizontalLayout_ExampleImgs"))
        self.comboBox_ShowTrainOrValid = QtWidgets.QComboBox(self.widget2)
        #Insert option for training or valid
        self.comboBox_ShowTrainOrValid.addItems(["Training","Validation"])        
        self.comboBox_ShowTrainOrValid.setObjectName(_fromUtf8("comboBox_ShowTrainOrValid"))
        self.horizontalLayout_ExampleImgs.addWidget(self.comboBox_ShowTrainOrValid)
        self.comboBox_ShowWOrWoAug = QtWidgets.QComboBox(self.widget2)
        self.comboBox_ShowWOrWoAug.addItems(["With Augmentation","Original image"])        
        self.comboBox_ShowWOrWoAug.setObjectName(_fromUtf8("comboBox_ShowWOrWoAug"))
        self.horizontalLayout_ExampleImgs.addWidget(self.comboBox_ShowWOrWoAug)
        self.label_ShowIndex = QtWidgets.QLabel(self.widget2)
        self.label_ShowIndex.setObjectName(_fromUtf8("label_ShowIndex"))
        self.label_ShowIndex.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.horizontalLayout_ExampleImgs.addWidget(self.label_ShowIndex)
        self.spinBox_ShowIndex = QtWidgets.QSpinBox(self.widget2)
        self.spinBox_ShowIndex.setMinimum(0)
        self.spinBox_ShowIndex.setMaximum(9E8)

        self.spinBox_ShowIndex.setObjectName(_fromUtf8("spinBox_ShowIndex"))
        self.horizontalLayout_ExampleImgs.addWidget(self.spinBox_ShowIndex)
        self.pushButton_ShowExamleImgs = QtWidgets.QPushButton(self.widget2)
        self.pushButton_ShowExamleImgs.setObjectName(_fromUtf8("pushButton_ShowExamleImgs"))
        self.pushButton_ShowExamleImgs.clicked.connect(self.action_show_example_imgs)

        self.horizontalLayout_ExampleImgs.addWidget(self.pushButton_ShowExamleImgs)
        self.widget_ViewImages = QtWidgets.QWidget(self.splitter_4)
        self.widget_ViewImages.setObjectName(_fromUtf8("widget_ViewImages"))
        self.gridLayout_9.addWidget(self.splitter_4, 0, 0, 1, 1)
        self.tabWidget_DefineModel.addTab(self.tab_ExampleImgs, _fromUtf8(""))


        

        self.tab_expert = QtWidgets.QWidget()
        self.tab_expert.setObjectName("tab_expert")
        self.gridLayout_34 = QtWidgets.QGridLayout(self.tab_expert)
        self.gridLayout_34.setObjectName("gridLayout_34")
        self.groupBox_expertMode = QtWidgets.QGroupBox(self.tab_expert)
        self.groupBox_expertMode.setEnabled(True)
        self.groupBox_expertMode.setCheckable(True)
        self.groupBox_expertMode.setChecked(False)
        self.groupBox_expertMode.setObjectName("groupBox_expertMode")
        self.groupBox_expertMode.toggled.connect(self.expert_mode_off)

        self.gridLayout_35 = QtWidgets.QGridLayout(self.groupBox_expertMode)
        self.gridLayout_35.setObjectName("gridLayout_35")
        self.scrollArea = QtWidgets.QScrollArea(self.groupBox_expertMode)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, -25, 425, 218))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_37 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_37.setObjectName("gridLayout_37")

        self.groupBox_modelKerasFit = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.groupBox_modelKerasFit.setObjectName("groupBox_modelKerasFit")
        self.gridLayout_10 = QtWidgets.QGridLayout(self.groupBox_modelKerasFit)
        self.gridLayout_10.setObjectName("gridLayout_10")
        
        self.verticalLayout_19 = QtWidgets.QVBoxLayout()
        self.verticalLayout_19.setObjectName("verticalLayout_19")

        self.horizontalLayout_32 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_32.setObjectName("horizontalLayout_32")
        self.label_batchSize = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_batchSize.setObjectName("label_batchSize")
        self.horizontalLayout_32.addWidget(self.label_batchSize)
        self.spinBox_batchSize = QtWidgets.QSpinBox(self.scrollAreaWidgetContents)
        self.spinBox_batchSize.setObjectName("spinBox_batchSize")
        self.horizontalLayout_32.addWidget(self.spinBox_batchSize)
        self.verticalLayout_19.addLayout(self.horizontalLayout_32)
        self.horizontalLayout_33 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_33.setObjectName("horizontalLayout_33")
        self.label_epochs = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        self.label_epochs.setObjectName("label_epochs")
        self.horizontalLayout_33.addWidget(self.label_epochs)
        self.spinBox_epochs = QtWidgets.QSpinBox(self.scrollAreaWidgetContents)
        self.spinBox_epochs.setObjectName("spinBox_epochs")
        self.horizontalLayout_33.addWidget(self.spinBox_epochs)
        
        self.verticalLayout_19.addLayout(self.horizontalLayout_33)
        self.gridLayout_10.addLayout(self.verticalLayout_19, 0, 0, 1, 1)
        self.gridLayout_37.addWidget(self.groupBox_modelKerasFit, 1, 0, 1, 1)

        self.groupBox_regularization = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.groupBox_regularization.setObjectName("groupBox_regularization")

        self.gridLayout_46 = QtWidgets.QGridLayout(self.groupBox_regularization)
        self.gridLayout_46.setObjectName("gridLayout_46")        

        self.horizontalLayout_expt_loss = QtWidgets.QHBoxLayout()
        self.horizontalLayout_expt_loss.setObjectName("horizontalLayout_expt_loss")
        self.checkBox_expt_loss = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_expt_loss.setObjectName("checkBox_expt_loss")
        self.checkBox_expt_loss.stateChanged.connect(self.expert_loss_off)

        self.horizontalLayout_expt_loss.addWidget(self.checkBox_expt_loss)
        self.comboBox_expt_loss = QtWidgets.QComboBox(self.groupBox_regularization)
        self.comboBox_expt_loss.setEnabled(False)
        self.comboBox_expt_loss.setObjectName("comboBox_expt_loss")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.comboBox_expt_loss.addItem("")
        self.horizontalLayout_expt_loss.addWidget(self.comboBox_expt_loss)
        self.gridLayout_46.addLayout(self.horizontalLayout_expt_loss, 0, 0, 1, 1)




        self.horizontalLayout_34 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_34.setObjectName("horizontalLayout_34")
        self.checkBox_learningRate = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.checkBox_learningRate.setObjectName("checkBox_learningRate")
        self.checkBox_learningRate.stateChanged.connect(self.expert_learningrate_off)

        self.horizontalLayout_34.addWidget(self.checkBox_learningRate)
        self.doubleSpinBox_learningRate = QtWidgets.QDoubleSpinBox(self.scrollAreaWidgetContents)
        self.doubleSpinBox_learningRate.setDecimals(6)
        self.doubleSpinBox_learningRate.setMaximum(999.0)
        self.doubleSpinBox_learningRate.setSingleStep(0.0001)
        self.doubleSpinBox_learningRate.setValue(Default_dict["doubleSpinBox_learningRate_Adam"])
        self.doubleSpinBox_learningRate.setObjectName("doubleSpinBox_learningRate")
        self.horizontalLayout_34.addWidget(self.doubleSpinBox_learningRate)

        self.checkBox_optimizer = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_optimizer.setObjectName("checkBox_optimizer")
        self.checkBox_optimizer.stateChanged.connect(self.expert_optimizer_off)

        self.horizontalLayout_34.addWidget(self.checkBox_optimizer)
        self.comboBox_optimizer = QtWidgets.QComboBox(self.groupBox_regularization)
        self.comboBox_optimizer.setEnabled(False)
        self.comboBox_optimizer.setObjectName("comboBox_optimizer")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.addItem("")
        self.comboBox_optimizer.currentTextChanged.connect(self.expert_optimizer_changed)
        self.horizontalLayout_34.addWidget(self.comboBox_optimizer)


        self.gridLayout_46.addLayout(self.horizontalLayout_34, 2, 0, 1, 1)
        self.horizontalLayout_35 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_35.setObjectName("horizontalLayout_35")
        self.checkBox_trainLastNOnly = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_trainLastNOnly.setObjectName("checkBox_trainLastNOnly")
        self.horizontalLayout_35.addWidget(self.checkBox_trainLastNOnly)
        self.spinBox_trainLastNOnly = QtWidgets.QSpinBox(self.groupBox_regularization)
        self.spinBox_trainLastNOnly.setObjectName("spinBox_trainLastNOnly")
        self.horizontalLayout_35.addWidget(self.spinBox_trainLastNOnly)
        self.checkBox_trainDenseOnly = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_trainDenseOnly.setObjectName("checkBox_trainDenseOnly")
        self.horizontalLayout_35.addWidget(self.checkBox_trainDenseOnly)
        self.gridLayout_46.addLayout(self.horizontalLayout_35, 3, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.checkBox_dropout = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_dropout.setObjectName("checkBox_dropout")
        self.horizontalLayout_2.addWidget(self.checkBox_dropout)
        self.lineEdit_dropout = QtWidgets.QLineEdit(self.groupBox_regularization)
        self.lineEdit_dropout.setObjectName("lineEdit_dropout")
        validator = QtGui.QRegExpValidator(QtCore.QRegExp("^[0-9 . ,]+$")) #validator allows numbers, dots and commas
        #aternatively, I could use "^[0-9 . , \[ \] ]+$" - this would also allow the user to put the brackets. But why? I just do it in the program
        self.lineEdit_dropout.setValidator(validator)        
        
        self.horizontalLayout_2.addWidget(self.lineEdit_dropout)
        self.gridLayout_46.addLayout(self.horizontalLayout_2, 4, 0, 1, 1)
        
        self.horizontalLayout_partialTrainability = QtWidgets.QHBoxLayout()
        self.horizontalLayout_partialTrainability.setObjectName("horizontalLayout_partialTrainability")
        self.checkBox_partialTrainability = QtWidgets.QCheckBox(self.groupBox_regularization)
        self.checkBox_partialTrainability.setObjectName("checkBox_partialTrainability")
        self.horizontalLayout_partialTrainability.addWidget(self.checkBox_partialTrainability)
        self.lineEdit_partialTrainability = QtWidgets.QLineEdit(self.groupBox_regularization)
        self.lineEdit_partialTrainability.setEnabled(False)
        self.lineEdit_partialTrainability.setObjectName("lineEdit_partialTrainability")
        self.horizontalLayout_partialTrainability.addWidget(self.lineEdit_partialTrainability)
        self.pushButton_partialTrainability = QtWidgets.QPushButton(self.groupBox_regularization)
        self.pushButton_partialTrainability.setObjectName("pushButton_partialTrainability")
        self.horizontalLayout_partialTrainability.addWidget(self.pushButton_partialTrainability)
        self.gridLayout_46.addLayout(self.horizontalLayout_partialTrainability, 5, 0, 1, 1)
        self.gridLayout_37.addWidget(self.groupBox_regularization, 2, 0, 1, 1)


        self.pushButton_partialTrainability.setEnabled(False)
        self.pushButton_partialTrainability.setMinimumSize(QtCore.QSize(0, 0))
        self.pushButton_partialTrainability.setMaximumSize(QtCore.QSize(40, 16777215))
        self.pushButton_partialTrainability.clicked.connect(self.partialTrainability)
        
        self.horizontalLayout_lossW = QtWidgets.QHBoxLayout()
        self.horizontalLayout_lossW.setObjectName("horizontalLayout_lossW")
        self.checkBox_lossW = QtWidgets.QCheckBox(self.scrollAreaWidgetContents)
        self.checkBox_lossW.setObjectName("checkBox_lossW")
        self.horizontalLayout_lossW.addWidget(self.checkBox_lossW)
        self.lineEdit_lossW = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        self.lineEdit_lossW.setEnabled(False)
        self.lineEdit_lossW.setObjectName("lineEdit_lossW")
        self.horizontalLayout_lossW.addWidget(self.lineEdit_lossW)
        self.pushButton_lossW = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.pushButton_lossW.setObjectName("pushButton_lossW")
        self.pushButton_lossW.setEnabled(False)
        self.horizontalLayout_lossW.addWidget(self.pushButton_lossW)
        self.gridLayout_46.addLayout(self.horizontalLayout_lossW, 6, 0, 1, 1)
        self.pushButton_lossW.setMinimumSize(QtCore.QSize(0, 0))
        self.pushButton_lossW.setMaximumSize(QtCore.QSize(40, 16777215))

        self.groupBox_expt_imgProc = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.groupBox_expt_imgProc.setObjectName("groupBox_expt_imgProc")
        self.gridLayout_48 = QtWidgets.QGridLayout(self.groupBox_expt_imgProc)
        self.gridLayout_48.setObjectName("gridLayout_48")
        self.checkBox_expt_paddingMode = QtWidgets.QCheckBox(self.groupBox_expt_imgProc)
        self.checkBox_expt_paddingMode.setObjectName("checkBox_expt_paddingMode")
        self.gridLayout_48.addWidget(self.checkBox_expt_paddingMode, 0, 0, 1, 1)
        self.comboBox_expt_paddingMode = QtWidgets.QComboBox(self.groupBox_expt_imgProc)
        self.comboBox_expt_paddingMode.setEnabled(False)
        self.comboBox_expt_paddingMode.setObjectName("comboBox_expt_paddingMode")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.comboBox_expt_paddingMode.addItem("")
        self.gridLayout_48.addWidget(self.comboBox_expt_paddingMode, 0, 1, 1, 1)
        self.gridLayout_37.addWidget(self.groupBox_expt_imgProc, 4, 0, 1, 1)


        self.groupBox_expertMetrics = QtWidgets.QGroupBox(self.scrollAreaWidgetContents)
        self.groupBox_expertMetrics.setObjectName("groupBox_expertMetrics")
        self.gridLayout = QtWidgets.QGridLayout(self.groupBox_expertMetrics)
        self.gridLayout.setObjectName("gridLayout")
        self.checkBox_expertAccuracy = QtWidgets.QCheckBox(self.groupBox_expertMetrics)
        self.checkBox_expertAccuracy.setChecked(True)
        self.checkBox_expertAccuracy.setEnabled(False) #Accuracy is ALWAYS tracked!
        
        self.checkBox_expertAccuracy.setObjectName("checkBox_expertAccuracy")
        self.gridLayout.addWidget(self.checkBox_expertAccuracy, 0, 0, 1, 1)
        self.checkBox_expertF1 = QtWidgets.QCheckBox(self.groupBox_expertMetrics)
        self.checkBox_expertF1.setChecked(False)
        self.checkBox_expertF1.setObjectName("checkBox_expertF1")
        self.gridLayout.addWidget(self.checkBox_expertF1, 0, 1, 1, 1)
        self.checkBox_expertPrecision = QtWidgets.QCheckBox(self.groupBox_expertMetrics)
        self.checkBox_expertPrecision.setObjectName("checkBox_expertPrecision")
        self.gridLayout.addWidget(self.checkBox_expertPrecision, 0, 2, 1, 1)
        self.checkBox_expertRecall = QtWidgets.QCheckBox(self.groupBox_expertMetrics)
        self.checkBox_expertRecall.setObjectName("checkBox_expertRecall")
        self.gridLayout.addWidget(self.checkBox_expertRecall, 0, 3, 1, 1)
        self.gridLayout_37.addWidget(self.groupBox_expertMetrics, 7, 0, 1, 1)



        
        
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_35.addWidget(self.scrollArea, 0, 0, 1, 1)
        self.gridLayout_34.addWidget(self.groupBox_expertMode, 0, 0, 1, 1)

        self.tabWidget_DefineModel.addTab(self.tab_expert, "")
        
        self.groupBox_Finalize = QtWidgets.QGroupBox(self.splitter_5)
        self.groupBox_Finalize.setObjectName(_fromUtf8("groupBox_Finalize"))
        self.gridLayout_6 = QtWidgets.QGridLayout(self.groupBox_Finalize)
        self.gridLayout_6.setObjectName(_fromUtf8("gridLayout_6"))
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName(_fromUtf8("horizontalLayout_3"))
        self.textBrowser_Info = QtWidgets.QTextBrowser(self.groupBox_Finalize)
        self.textBrowser_Info.setMinimumSize(QtCore.QSize(0, 60))
        self.textBrowser_Info.setMaximumSize(QtCore.QSize(16777215, 500))
        self.textBrowser_Info.setObjectName(_fromUtf8("textBrowser_Info"))
        self.horizontalLayout_3.addWidget(self.textBrowser_Info)
        self.pushButton_FitModel = QtWidgets.QPushButton(self.groupBox_Finalize)
        self.pushButton_FitModel.setMinimumSize(QtCore.QSize(111, 60))
        self.pushButton_FitModel.setMaximumSize(QtCore.QSize(111, 60))
        self.pushButton_FitModel.setObjectName(_fromUtf8("pushButton_FitModel"))
        self.pushButton_FitModel.clicked.connect(self.action_initialize_model)

        self.horizontalLayout_3.addWidget(self.pushButton_FitModel)
        self.gridLayout_6.addLayout(self.horizontalLayout_3, 0, 0, 1, 1)
        self.gridLayout_17.addWidget(self.splitter_5, 0, 0, 1, 1)
        self.tabWidget_Modelbuilder.addTab(self.tab_Build, _fromUtf8(""))
        self.tab_History = QtWidgets.QWidget()
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tab_History.sizePolicy().hasHeightForWidth())
        self.tab_History.setSizePolicy(sizePolicy)
        self.tab_History.setObjectName(_fromUtf8("tab_History"))
        self.gridLayout_3 = QtWidgets.QGridLayout(self.tab_History)
        self.gridLayout_3.setObjectName(_fromUtf8("gridLayout_3"))
        self.verticalLayout_9 = QtWidgets.QVBoxLayout()
        self.verticalLayout_9.setObjectName(_fromUtf8("verticalLayout_9"))
        self.verticalLayout_HistoryLoad = QtWidgets.QVBoxLayout()
        self.verticalLayout_HistoryLoad.setSizeConstraint(QtWidgets.QLayout.SetFixedSize)
        self.verticalLayout_HistoryLoad.setObjectName(_fromUtf8("verticalLayout_HistoryLoad"))
        self.horizontalLayout_HistoryLoad = QtWidgets.QHBoxLayout()
        self.horizontalLayout_HistoryLoad.setSizeConstraint(QtWidgets.QLayout.SetMaximumSize)
        self.horizontalLayout_HistoryLoad.setObjectName(_fromUtf8("horizontalLayout_HistoryLoad"))
        self.pushButton_Live = QtWidgets.QPushButton(self.tab_History)
        self.pushButton_Live.clicked.connect(self.action_load_history_current)
        self.pushButton_Live.setMinimumSize(QtCore.QSize(93, 28))
        self.pushButton_Live.setMaximumSize(QtCore.QSize(93, 28))
        self.pushButton_Live.setObjectName(_fromUtf8("pushButton_Live"))
        self.horizontalLayout_HistoryLoad.addWidget(self.pushButton_Live)
        self.pushButton_LoadHistory = QtWidgets.QPushButton(self.tab_History)
        self.pushButton_LoadHistory.setMinimumSize(QtCore.QSize(93, 28))
        self.pushButton_LoadHistory.setMaximumSize(QtCore.QSize(93, 28))
        self.pushButton_LoadHistory.clicked.connect(self.action_load_history)
        self.pushButton_LoadHistory.setObjectName(_fromUtf8("pushButton_LoadHistory"))
        self.horizontalLayout_HistoryLoad.addWidget(self.pushButton_LoadHistory)
        self.lineEdit_LoadHistory = QtWidgets.QLineEdit(self.tab_History)
        self.lineEdit_LoadHistory.setDisabled(True)
        self.lineEdit_LoadHistory.setObjectName(_fromUtf8("lineEdit_LoadHistory"))
        self.horizontalLayout_HistoryLoad.addWidget(self.lineEdit_LoadHistory)
        self.verticalLayout_HistoryLoad.addLayout(self.horizontalLayout_HistoryLoad)
        self.horizontalLayout_HistoryLoadInfo = QtWidgets.QHBoxLayout()
        self.horizontalLayout_HistoryLoadInfo.setSizeConstraint(QtWidgets.QLayout.SetMaximumSize)
        self.horizontalLayout_HistoryLoadInfo.setObjectName(_fromUtf8("horizontalLayout_HistoryLoadInfo"))
        self.tableWidget_HistoryItems = QtWidgets.QTableWidget(self.tab_History)
        self.tableWidget_HistoryItems.setMinimumSize(QtCore.QSize(0, 100))
        self.tableWidget_HistoryItems.setMaximumSize(QtCore.QSize(16777215, 140))
        self.tableWidget_HistoryItems.setObjectName(_fromUtf8("tableWidget_HistoryItems"))
        self.tableWidget_HistoryItems.setColumnCount(7)
        self.tableWidget_HistoryItems.setRowCount(0)
        self.horizontalLayout_HistoryLoadInfo.addWidget(self.tableWidget_HistoryItems)

        self.verticalLayout_UpdatePlot = QtWidgets.QVBoxLayout()
        self.pushButton_UpdateHistoryPlot = QtWidgets.QPushButton(self.tab_History)
        self.pushButton_UpdateHistoryPlot.clicked.connect(self.update_historyplot)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_UpdateHistoryPlot.sizePolicy().hasHeightForWidth())
        self.pushButton_UpdateHistoryPlot.setSizePolicy(sizePolicy)
        self.pushButton_UpdateHistoryPlot.setObjectName(_fromUtf8("pushButton_UpdateHistoryPlot"))
        self.verticalLayout_UpdatePlot.addWidget(self.pushButton_UpdateHistoryPlot)

        self.horizontalLayout_rollmedi = QtWidgets.QHBoxLayout()
        self.checkBox_rollingMedian = QtWidgets.QCheckBox(self.tab_History)
        self.checkBox_rollingMedian.setMinimumSize(QtCore.QSize(100, 19))
        self.checkBox_rollingMedian.setMaximumSize(QtCore.QSize(125, 25))
        self.checkBox_rollingMedian.toggled.connect(self.checkBox_rollingMedian_statechange)
        
        self.checkBox_rollingMedian.setObjectName(_fromUtf8("checkBox_rollingMedian"))
        self.horizontalLayout_rollmedi.addWidget(self.checkBox_rollingMedian)        
        self.horizontalSlider_rollmedi = QtWidgets.QSlider(self.tab_History)
        self.horizontalSlider_rollmedi.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider_rollmedi.setMinimumSize(QtCore.QSize(50, 19))
        self.horizontalSlider_rollmedi.setMaximumSize(QtCore.QSize(50, 25))
        #Adjust the horizontalSlider_rollmedi
        self.horizontalSlider_rollmedi.setSingleStep(1)
        self.horizontalSlider_rollmedi.setMinimum(1)
        self.horizontalSlider_rollmedi.setMaximum(50)
        self.horizontalSlider_rollmedi.setValue(10)
        self.horizontalSlider_rollmedi.setEnabled(False)
        
        self.horizontalSlider_rollmedi.setObjectName(_fromUtf8("horizontalSlider_rollmedi"))
        self.horizontalLayout_rollmedi.addWidget(self.horizontalSlider_rollmedi)
        self.verticalLayout_UpdatePlot.addLayout(self.horizontalLayout_rollmedi)

        self.checkBox_linearFit = QtWidgets.QCheckBox(self.tab_History)
        self.checkBox_linearFit.setObjectName(_fromUtf8("checkBox_linearFit"))
        self.verticalLayout_UpdatePlot.addWidget(self.checkBox_linearFit)
        self.horizontalLayout_HistoryLoadInfo.addLayout(self.verticalLayout_UpdatePlot)
        
        self.verticalLayout_HistoryLoad.addLayout(self.horizontalLayout_HistoryLoadInfo)
        self.verticalLayout_9.addLayout(self.verticalLayout_HistoryLoad)
        self.widget_Scatterplot = pg.GraphicsLayoutWidget(self.tab_History)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.widget_Scatterplot.sizePolicy().hasHeightForWidth())
        self.widget_Scatterplot.setSizePolicy(sizePolicy)
        self.widget_Scatterplot.setMinimumSize(QtCore.QSize(491, 350))
        self.widget_Scatterplot.setObjectName(_fromUtf8("widget_Scatterplot"))
        self.verticalLayout_9.addWidget(self.widget_Scatterplot)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setSizeConstraint(QtWidgets.QLayout.SetMaximumSize)
        self.horizontalLayout_7.setObjectName(_fromUtf8("horizontalLayout_7"))
        self.verticalLayout_convert = QtWidgets.QVBoxLayout()
        self.verticalLayout_convert.setSizeConstraint(QtWidgets.QLayout.SetMaximumSize)
        self.verticalLayout_convert.setObjectName(_fromUtf8("verticalLayout_convert"))

        self.verticalLayout_4 = QtWidgets.QVBoxLayout()
        self.verticalLayout_4.setObjectName(_fromUtf8("verticalLayout_4"))
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName(_fromUtf8("horizontalLayout_2"))
        self.combobox_initial_format = QtWidgets.QComboBox(self.tab_History)
        self.combobox_initial_format.setObjectName(_fromUtf8("combobox_initial_format"))
        self.horizontalLayout_2.addWidget(self.combobox_initial_format)
        self.verticalLayout_4.addLayout(self.horizontalLayout_2)
        self.pushButton_LoadModel = QtWidgets.QPushButton(self.tab_History)
        self.pushButton_LoadModel.setMinimumSize(QtCore.QSize(123, 61))
        self.pushButton_LoadModel.setMaximumSize(QtCore.QSize(150, 61))
        self.pushButton_LoadModel.setObjectName(_fromUtf8("pushButton_LoadModel"))
        self.pushButton_LoadModel.clicked.connect(self.history_tab_get_model_path)
        self.verticalLayout_4.addWidget(self.pushButton_LoadModel)
        self.horizontalLayout_7.addLayout(self.verticalLayout_4)
        self.textBrowser_SelectedModelInfo = QtWidgets.QTextBrowser(self.tab_History)
        self.textBrowser_SelectedModelInfo.setMinimumSize(QtCore.QSize(0, 120))
        self.textBrowser_SelectedModelInfo.setMaximumSize(QtCore.QSize(16777215, 120))
        self.textBrowser_SelectedModelInfo.setObjectName(_fromUtf8("textBrowser_SelectedModelInfo"))
        self.horizontalLayout_7.addWidget(self.textBrowser_SelectedModelInfo)
        
        self.comboBox_convertTo = QtWidgets.QComboBox(self.tab_History)
        self.comboBox_convertTo.setObjectName(_fromUtf8("comboBox_convertTo"))
        self.verticalLayout_convert.addWidget(self.comboBox_convertTo)

        self.pushButton_convertModel = QtWidgets.QPushButton(self.tab_History)
        self.pushButton_convertModel.setMinimumSize(QtCore.QSize(0, 61))
        self.pushButton_convertModel.setMaximumSize(QtCore.QSize(16777215, 61))
        self.pushButton_convertModel.setObjectName(_fromUtf8("pushButton_convertModel"))
        self.pushButton_convertModel.setEnabled(True)
        self.pushButton_convertModel.clicked.connect(self.history_tab_convertModel)
        self.verticalLayout_convert.addWidget(self.pushButton_convertModel)
        
        self.horizontalLayout_7.addLayout(self.verticalLayout_convert)
                
        self.verticalLayout_9.addLayout(self.horizontalLayout_7)
        self.gridLayout_3.addLayout(self.verticalLayout_9, 0, 0, 1, 1)
        self.tabWidget_Modelbuilder.addTab(self.tab_History, _fromUtf8(""))




        #########################Assess Model tab##############################


        self.tab_AssessModel = QtWidgets.QWidget()
        self.tab_AssessModel.setObjectName("tab_AssessModel")
        self.gridLayout_23 = QtWidgets.QGridLayout(self.tab_AssessModel)
        self.gridLayout_23.setObjectName("gridLayout_23")
        self.splitter_7 = QtWidgets.QSplitter(self.tab_AssessModel)
        self.splitter_7.setOrientation(QtCore.Qt.Vertical)
        self.splitter_7.setObjectName("splitter_7")
        self.widget = QtWidgets.QWidget(self.splitter_7)
        self.widget.setObjectName("widget")
        self.verticalLayout_10 = QtWidgets.QVBoxLayout(self.widget)
        self.verticalLayout_10.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.groupBox_loadModel = QtWidgets.QGroupBox(self.widget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.groupBox_loadModel.sizePolicy().hasHeightForWidth())
        self.groupBox_loadModel.setSizePolicy(sizePolicy)
        self.groupBox_loadModel.setMinimumSize(QtCore.QSize(0, 101))
        self.groupBox_loadModel.setMaximumSize(QtCore.QSize(16777215, 101))
        self.groupBox_loadModel.setObjectName("groupBox_loadModel")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.groupBox_loadModel)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.horizontalLayout_10 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_10.setObjectName("horizontalLayout_10")
        self.pushButton_LoadModel_2 = QtWidgets.QPushButton(self.groupBox_loadModel)
        self.pushButton_LoadModel_2.setMinimumSize(QtCore.QSize(123, 24))
        self.pushButton_LoadModel_2.setMaximumSize(QtCore.QSize(123, 24))
        self.pushButton_LoadModel_2.setObjectName("pushButton_LoadModel_2")
        self.horizontalLayout_10.addWidget(self.pushButton_LoadModel_2)
        self.lineEdit_LoadModel_2 = QtWidgets.QLineEdit(self.groupBox_loadModel)
        self.lineEdit_LoadModel_2.setEnabled(False)
        self.lineEdit_LoadModel_2.setObjectName("lineEdit_LoadModel_2")
        self.horizontalLayout_10.addWidget(self.lineEdit_LoadModel_2)
        self.comboBox_loadedRGBorGray = QtWidgets.QComboBox(self.groupBox_loadModel)
        self.comboBox_loadedRGBorGray.setEnabled(False)
        self.comboBox_loadedRGBorGray.setObjectName("comboBox_loadedRGBorGray")
        self.horizontalLayout_10.addWidget(self.comboBox_loadedRGBorGray)
        self.gridLayout_4.addLayout(self.horizontalLayout_10, 0, 0, 1, 1)
        self.horizontalLayout_11 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_11.setObjectName("horizontalLayout_11")
        self.label_ModelIndex_2 = QtWidgets.QLabel(self.groupBox_loadModel)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_ModelIndex_2.sizePolicy().hasHeightForWidth())
        self.label_ModelIndex_2.setSizePolicy(sizePolicy)
        self.label_ModelIndex_2.setMinimumSize(QtCore.QSize(68, 25))
        self.label_ModelIndex_2.setMaximumSize(QtCore.QSize(68, 25))
        self.label_ModelIndex_2.setObjectName("label_ModelIndex_2")
        self.horizontalLayout_11.addWidget(self.label_ModelIndex_2)
        self.spinBox_ModelIndex_2 = QtWidgets.QSpinBox(self.groupBox_loadModel)
        self.spinBox_ModelIndex_2.setEnabled(False)
        self.spinBox_ModelIndex_2.setMinimumSize(QtCore.QSize(46, 22))
        self.spinBox_ModelIndex_2.setMaximumSize(QtCore.QSize(46, 22))
        self.spinBox_ModelIndex_2.setObjectName("spinBox_ModelIndex_2")
        self.horizontalLayout_11.addWidget(self.spinBox_ModelIndex_2)
        self.lineEdit_ModelSelection_2 = QtWidgets.QLineEdit(self.groupBox_loadModel)
        self.lineEdit_ModelSelection_2.setEnabled(False)
        self.lineEdit_ModelSelection_2.setObjectName("lineEdit_ModelSelection_2")
        self.horizontalLayout_11.addWidget(self.lineEdit_ModelSelection_2)
        self.label_Normalization_2 = QtWidgets.QLabel(self.groupBox_loadModel)
        self.label_Normalization_2.setObjectName("label_Normalization_2")
        self.horizontalLayout_11.addWidget(self.label_Normalization_2)
        self.comboBox_Normalization_2 = QtWidgets.QComboBox(self.groupBox_loadModel)
        self.comboBox_Normalization_2.setEnabled(False)
        self.comboBox_Normalization_2.setObjectName("comboBox_Normalization_2")
        self.horizontalLayout_11.addWidget(self.comboBox_Normalization_2)
        self.label_Crop_2 = QtWidgets.QLabel(self.groupBox_loadModel)
        self.label_Crop_2.setObjectName("label_Crop_2")
        self.horizontalLayout_11.addWidget(self.label_Crop_2)
        self.spinBox_Crop_2 = QtWidgets.QSpinBox(self.groupBox_loadModel)
        self.spinBox_Crop_2.setEnabled(False)
        self.spinBox_Crop_2.setObjectName("spinBox_Crop_2")
        self.horizontalLayout_11.addWidget(self.spinBox_Crop_2)
        self.label_OutClasses_2 = QtWidgets.QLabel(self.groupBox_loadModel)
        self.label_OutClasses_2.setObjectName("label_OutClasses_2")
        self.horizontalLayout_11.addWidget(self.label_OutClasses_2)
        self.spinBox_OutClasses_2 = QtWidgets.QSpinBox(self.groupBox_loadModel)
        self.spinBox_OutClasses_2.setEnabled(False)
        self.spinBox_OutClasses_2.setObjectName("spinBox_OutClasses_2")
        self.horizontalLayout_11.addWidget(self.spinBox_OutClasses_2)
        self.gridLayout_4.addLayout(self.horizontalLayout_11, 1, 0, 1, 1)
        self.verticalLayout_10.addWidget(self.groupBox_loadModel)
        self.horizontalLayout_16 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_16.setObjectName("horizontalLayout_16")
        self.groupBox_validData = QtWidgets.QGroupBox(self.widget)
        self.groupBox_validData.setMaximumSize(QtCore.QSize(150, 250))
        self.groupBox_validData.setObjectName("groupBox_validData")
        self.gridLayout_14 = QtWidgets.QGridLayout(self.groupBox_validData)
        self.gridLayout_14.setObjectName("gridLayout_14")
        self.horizontalLayout_17 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_17.setObjectName("horizontalLayout_17")
        self.pushButton_ImportValidFromNpy = QtWidgets.QPushButton(self.groupBox_validData)
        self.pushButton_ImportValidFromNpy.setMinimumSize(QtCore.QSize(65, 28))
        self.pushButton_ImportValidFromNpy.setObjectName("pushButton_ImportValidFromNpy")
        self.horizontalLayout_17.addWidget(self.pushButton_ImportValidFromNpy)
        self.pushButton_ExportValidToNpy = QtWidgets.QPushButton(self.groupBox_validData)
        self.pushButton_ExportValidToNpy.setMinimumSize(QtCore.QSize(55, 0))
        self.pushButton_ExportValidToNpy.setObjectName("pushButton_ExportValidToNpy")
        self.horizontalLayout_17.addWidget(self.pushButton_ExportValidToNpy)
        self.gridLayout_14.addLayout(self.horizontalLayout_17, 1, 0, 1, 1)
        self.tableWidget_Info_2 = QtWidgets.QTableWidget(self.groupBox_validData)
        self.tableWidget_Info_2.setObjectName("tableWidget_Info_2")
        self.tableWidget_Info_2.setColumnCount(0)
        self.tableWidget_Info_2.setRowCount(0)
        self.gridLayout_14.addWidget(self.tableWidget_Info_2, 0, 0, 1, 1)
        self.horizontalLayout_16.addWidget(self.groupBox_validData)
        self.verticalLayout_9 = QtWidgets.QVBoxLayout()
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.groupBox_InferenceTime = QtWidgets.QGroupBox(self.widget)
        self.groupBox_InferenceTime.setMinimumSize(QtCore.QSize(0, 71))
        self.groupBox_InferenceTime.setMaximumSize(QtCore.QSize(16777215, 71))
        self.groupBox_InferenceTime.setObjectName("groupBox_InferenceTime")
        self.gridLayout_20 = QtWidgets.QGridLayout(self.groupBox_InferenceTime)
        self.gridLayout_20.setObjectName("gridLayout_20")
        self.lineEdit_InferenceTime = QtWidgets.QLineEdit(self.groupBox_InferenceTime)
        self.lineEdit_InferenceTime.setObjectName("lineEdit_InferenceTime")
        self.gridLayout_20.addWidget(self.lineEdit_InferenceTime, 0, 2, 1, 1)
        self.pushButton_CompInfTime = QtWidgets.QPushButton(self.groupBox_InferenceTime)
        self.pushButton_CompInfTime.setMinimumSize(QtCore.QSize(121, 31))
        self.pushButton_CompInfTime.setMaximumSize(QtCore.QSize(121, 31))
        self.pushButton_CompInfTime.setObjectName("pushButton_CompInfTime")
        self.gridLayout_20.addWidget(self.pushButton_CompInfTime, 0, 0, 1, 1)
        self.spinBox_inftime_nr_images = QtWidgets.QSpinBox(self.groupBox_InferenceTime)
        self.spinBox_inftime_nr_images.setObjectName("spinBox_inftime_nr_images")
        self.gridLayout_20.addWidget(self.spinBox_inftime_nr_images, 0, 1, 1, 1)
        self.verticalLayout_9.addWidget(self.groupBox_InferenceTime)
        
        
        self.groupBox_classify = QtWidgets.QGroupBox(self.widget)
        self.groupBox_classify.setObjectName("groupBox_classify")
        self.gridLayout_36 = QtWidgets.QGridLayout(self.groupBox_classify)
        self.gridLayout_36.setObjectName("gridLayout_36")
        self.comboBox_scoresOrPrediction = QtWidgets.QComboBox(self.groupBox_classify)
        self.comboBox_scoresOrPrediction.setObjectName("comboBox_scoresOrPrediction")
        self.gridLayout_36.addWidget(self.comboBox_scoresOrPrediction, 0, 1, 1, 1)
        self.pushButton_classify = QtWidgets.QPushButton(self.groupBox_classify)
        self.pushButton_classify.setObjectName("pushButton_classify")
        self.gridLayout_36.addWidget(self.pushButton_classify, 0, 2, 1, 1)
        self.horizontalLayout_36 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_36.setObjectName("horizontalLayout_36")
        self.radioButton_selectAll = QtWidgets.QRadioButton(self.groupBox_classify)
        self.radioButton_selectAll.setObjectName("radioButton_selectAll")
        self.horizontalLayout_36.addWidget(self.radioButton_selectAll)
        self.radioButton_selectDataSet = QtWidgets.QRadioButton(self.groupBox_classify)
        self.radioButton_selectDataSet.setMinimumSize(QtCore.QSize(10, 22))
        self.radioButton_selectDataSet.setMaximumSize(QtCore.QSize(22, 16))
        self.radioButton_selectDataSet.setText("")
        self.radioButton_selectDataSet.setObjectName("radioButton_selectDataSet")
        self.horizontalLayout_36.addWidget(self.radioButton_selectDataSet)
        self.comboBox_selectData = QtWidgets.QComboBox(self.groupBox_classify)
        self.comboBox_selectData.setObjectName("comboBox_selectData")
        self.horizontalLayout_36.addWidget(self.comboBox_selectData)
        self.gridLayout_36.addLayout(self.horizontalLayout_36, 0, 0, 1, 1)
        self.verticalLayout_9.addWidget(self.groupBox_classify)
        
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.groupBox_settings = QtWidgets.QGroupBox(self.widget)
        self.groupBox_settings.setMinimumSize(QtCore.QSize(0, 99))
        self.groupBox_settings.setMaximumSize(QtCore.QSize(16777215, 99))
        self.groupBox_settings.setObjectName("groupBox_settings")
        self.gridLayout_22 = QtWidgets.QGridLayout(self.groupBox_settings)
        self.gridLayout_22.setObjectName("gridLayout_22")
        self.horizontalLayout_AssessModelSettings = QtWidgets.QHBoxLayout()
        self.horizontalLayout_AssessModelSettings.setObjectName("horizontalLayout_AssessModelSettings")
        self.verticalLayout_AssessModelSettings = QtWidgets.QVBoxLayout()
        self.verticalLayout_AssessModelSettings.setObjectName("verticalLayout_AssessModelSettings")
        self.horizontalLayout_AssessModelSettings_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_AssessModelSettings_2.setObjectName("horizontalLayout_AssessModelSettings_2")
        self.label_SortingIndex = QtWidgets.QLabel(self.groupBox_settings)
        self.label_SortingIndex.setObjectName("label_SortingIndex")
        self.horizontalLayout_AssessModelSettings_2.addWidget(self.label_SortingIndex)
        self.spinBox_indexOfInterest = QtWidgets.QSpinBox(self.groupBox_settings)
        self.spinBox_indexOfInterest.setObjectName("spinBox_indexOfInterest")
        self.horizontalLayout_AssessModelSettings_2.addWidget(self.spinBox_indexOfInterest)
        self.verticalLayout_AssessModelSettings.addLayout(self.horizontalLayout_AssessModelSettings_2)
        self.horizontalLayout_AssessModelSettings_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_AssessModelSettings_3.setObjectName("horizontalLayout_AssessModelSettings_3")
        self.checkBox_SortingThresh = QtWidgets.QCheckBox(self.groupBox_settings)
        self.checkBox_SortingThresh.setObjectName("checkBox_SortingThresh")
        self.horizontalLayout_AssessModelSettings_3.addWidget(self.checkBox_SortingThresh)
        self.doubleSpinBox_sortingThresh = QtWidgets.QDoubleSpinBox(self.groupBox_settings)
        self.doubleSpinBox_sortingThresh.setObjectName("doubleSpinBox_sortingThresh")
        self.horizontalLayout_AssessModelSettings_3.addWidget(self.doubleSpinBox_sortingThresh)
        self.verticalLayout_AssessModelSettings.addLayout(self.horizontalLayout_AssessModelSettings_3)
        self.horizontalLayout_AssessModelSettings.addLayout(self.verticalLayout_AssessModelSettings)
        self.verticalLayout_13 = QtWidgets.QVBoxLayout()
        self.verticalLayout_13.setObjectName("verticalLayout_13")
        self.pushButton_AssessModel = QtWidgets.QPushButton(self.groupBox_settings)
        self.pushButton_AssessModel.setMinimumSize(QtCore.QSize(0, 36))
        self.pushButton_AssessModel.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_AssessModel.setFont(font)
        self.pushButton_AssessModel.setAutoFillBackground(False)
        self.pushButton_AssessModel.setShortcut("")
        self.pushButton_AssessModel.setCheckable(False)
        self.pushButton_AssessModel.setChecked(False)
        self.pushButton_AssessModel.setAutoDefault(False)
        self.pushButton_AssessModel.setDefault(False)
        self.pushButton_AssessModel.setFlat(False)
        self.pushButton_AssessModel.setObjectName("pushButton_AssessModel")
        self.verticalLayout_13.addWidget(self.pushButton_AssessModel)
        self.comboBox_probability_histogram = QtWidgets.QComboBox(self.groupBox_settings)
        self.comboBox_probability_histogram.setMaximumSize(QtCore.QSize(16777215, 18))
        self.comboBox_probability_histogram.setObjectName("comboBox_probability_histogram")
        self.verticalLayout_13.addWidget(self.comboBox_probability_histogram)
        self.horizontalLayout_AssessModelSettings.addLayout(self.verticalLayout_13)
        self.gridLayout_22.addLayout(self.horizontalLayout_AssessModelSettings, 0, 0, 1, 1)
        self.horizontalLayout.addWidget(self.groupBox_settings)
        self.groupBox_3rdPlotSettings = QtWidgets.QGroupBox(self.widget)
        self.groupBox_3rdPlotSettings.setMinimumSize(QtCore.QSize(0, 91))
        self.groupBox_3rdPlotSettings.setMaximumSize(QtCore.QSize(16777215, 91))
        self.groupBox_3rdPlotSettings.setObjectName("groupBox_3rdPlotSettings")
        self.gridLayout_21 = QtWidgets.QGridLayout(self.groupBox_3rdPlotSettings)
        self.gridLayout_21.setObjectName("gridLayout_21")
        self.horizontalLayout_20 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_20.setObjectName("horizontalLayout_20")
        self.verticalLayout_3rdPlotSettings = QtWidgets.QVBoxLayout()
        self.verticalLayout_3rdPlotSettings.setObjectName("verticalLayout_3rdPlotSettings")
        self.label_3rdPlot = QtWidgets.QLabel(self.groupBox_3rdPlotSettings)
        self.label_3rdPlot.setObjectName("label_3rdPlot")
        self.verticalLayout_3rdPlotSettings.addWidget(self.label_3rdPlot)
        self.comboBox_3rdPlot = QtWidgets.QComboBox(self.groupBox_3rdPlotSettings)
        self.comboBox_3rdPlot.setObjectName("comboBox_3rdPlot")
        self.verticalLayout_3rdPlotSettings.addWidget(self.comboBox_3rdPlot)
        self.horizontalLayout_20.addLayout(self.verticalLayout_3rdPlotSettings)
        self.horizontalLayout_12 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_12.setObjectName("horizontalLayout_12")
        self.verticalLayout_3rdPlotSettings_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3rdPlotSettings_2.setObjectName("verticalLayout_3rdPlotSettings_2")
        self.label_Indx1 = QtWidgets.QLabel(self.groupBox_3rdPlotSettings)
        self.label_Indx1.setObjectName("label_Indx1")
        self.verticalLayout_3rdPlotSettings_2.addWidget(self.label_Indx1)
        self.spinBox_Indx1 = QtWidgets.QSpinBox(self.groupBox_3rdPlotSettings)
        self.spinBox_Indx1.setEnabled(False)
        self.spinBox_Indx1.setObjectName("spinBox_Indx1")
        self.verticalLayout_3rdPlotSettings_2.addWidget(self.spinBox_Indx1)
        self.horizontalLayout_12.addLayout(self.verticalLayout_3rdPlotSettings_2)
        self.verticalLayout_3rdPlotSettings_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3rdPlotSettings_3.setObjectName("verticalLayout_3rdPlotSettings_3")
        self.label_Indx2 = QtWidgets.QLabel(self.groupBox_3rdPlotSettings)
        self.label_Indx2.setObjectName("label_Indx2")
        self.verticalLayout_3rdPlotSettings_3.addWidget(self.label_Indx2)
        self.spinBox_Indx2 = QtWidgets.QSpinBox(self.groupBox_3rdPlotSettings)
        self.spinBox_Indx2.setEnabled(False)
        self.spinBox_Indx2.setObjectName("spinBox_Indx2")
        self.verticalLayout_3rdPlotSettings_3.addWidget(self.spinBox_Indx2)
        self.horizontalLayout_12.addLayout(self.verticalLayout_3rdPlotSettings_3)
        self.horizontalLayout_20.addLayout(self.horizontalLayout_12)
        self.gridLayout_21.addLayout(self.horizontalLayout_20, 0, 0, 1, 1)
        self.horizontalLayout.addWidget(self.groupBox_3rdPlotSettings)
        self.verticalLayout_9.addLayout(self.horizontalLayout)
        self.horizontalLayout_16.addLayout(self.verticalLayout_9)
        self.verticalLayout_10.addLayout(self.horizontalLayout_16)
        self.splitter_8 = QtWidgets.QSplitter(self.splitter_7)
        self.splitter_8.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_8.setObjectName("splitter_8")
        self.groupBox_confusionMatrixPlot = QtWidgets.QGroupBox(self.splitter_8)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.groupBox_confusionMatrixPlot.sizePolicy().hasHeightForWidth())
        self.groupBox_confusionMatrixPlot.setSizePolicy(sizePolicy)
        self.groupBox_confusionMatrixPlot.setBaseSize(QtCore.QSize(250, 500))
        self.groupBox_confusionMatrixPlot.setObjectName("groupBox_confusionMatrixPlot")
        self.gridLayout_16 = QtWidgets.QGridLayout(self.groupBox_confusionMatrixPlot)
        self.gridLayout_16.setObjectName("gridLayout_16")
        self.horizontalLayout_24 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_24.setObjectName("horizontalLayout_24")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout()
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.horizontalLayout_21 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_21.setObjectName("horizontalLayout_21")
        self.label_True_CM1 = QtWidgets.QLabel(self.groupBox_confusionMatrixPlot)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setStrikeOut(False)
        font.setKerning(True)
        font.setStyleStrategy(QtGui.QFont.PreferDefault)
        self.label_True_CM1.setFont(font)
        self.label_True_CM1.setObjectName("label_True_CM1")
        self.horizontalLayout_21.addWidget(self.label_True_CM1)
        self.tableWidget_CM1 = QtWidgets.QTableWidget(self.groupBox_confusionMatrixPlot)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget_CM1.sizePolicy().hasHeightForWidth())
        self.tableWidget_CM1.setSizePolicy(sizePolicy)
        self.tableWidget_CM1.setBaseSize(QtCore.QSize(250, 250))
        self.tableWidget_CM1.setObjectName("tableWidget_CM1")
        self.tableWidget_CM1.setColumnCount(0)
        self.tableWidget_CM1.setRowCount(0)
        self.horizontalLayout_21.addWidget(self.tableWidget_CM1)
        self.verticalLayout_11.addLayout(self.horizontalLayout_21)
        self.horizontalLayout_18 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_18.setObjectName("horizontalLayout_18")
        self.pushButton_CM1_to_Clipboard = QtWidgets.QPushButton(self.groupBox_confusionMatrixPlot)
        self.pushButton_CM1_to_Clipboard.setObjectName("pushButton_CM1_to_Clipboard")
        self.horizontalLayout_18.addWidget(self.pushButton_CM1_to_Clipboard)
        self.label_Pred_CM1 = QtWidgets.QLabel(self.groupBox_confusionMatrixPlot)
        font = QtGui.QFont()
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.label_Pred_CM1.setFont(font)
        self.label_Pred_CM1.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_Pred_CM1.setAutoFillBackground(False)
        self.label_Pred_CM1.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.label_Pred_CM1.setObjectName("label_Pred_CM1")
        self.horizontalLayout_18.addWidget(self.label_Pred_CM1)
        self.verticalLayout_11.addLayout(self.horizontalLayout_18)
        self.horizontalLayout_24.addLayout(self.verticalLayout_11)
        self.verticalLayout_12 = QtWidgets.QVBoxLayout()
        self.verticalLayout_12.setObjectName("verticalLayout_12")
        self.horizontalLayout_23 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_23.setObjectName("horizontalLayout_23")
        self.label_True_CM2 = QtWidgets.QLabel(self.groupBox_confusionMatrixPlot)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setStrikeOut(False)
        font.setKerning(True)
        font.setStyleStrategy(QtGui.QFont.PreferDefault)
        self.label_True_CM2.setFont(font)
        self.label_True_CM2.setObjectName("label_True_CM2")
        self.horizontalLayout_23.addWidget(self.label_True_CM2)
        self.tableWidget_CM2 = QtWidgets.QTableWidget(self.groupBox_confusionMatrixPlot)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget_CM2.sizePolicy().hasHeightForWidth())
        self.tableWidget_CM2.setSizePolicy(sizePolicy)
        self.tableWidget_CM2.setBaseSize(QtCore.QSize(250, 250))
        self.tableWidget_CM2.setObjectName("tableWidget_CM2")
        self.tableWidget_CM2.setColumnCount(0)
        self.tableWidget_CM2.setRowCount(0)
        self.horizontalLayout_23.addWidget(self.tableWidget_CM2)
        self.verticalLayout_12.addLayout(self.horizontalLayout_23)
        self.horizontalLayout_22 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_22.setObjectName("horizontalLayout_22")
        self.pushButton_CM2_to_Clipboard = QtWidgets.QPushButton(self.groupBox_confusionMatrixPlot)
        self.pushButton_CM2_to_Clipboard.setObjectName("pushButton_CM2_to_Clipboard")
        self.horizontalLayout_22.addWidget(self.pushButton_CM2_to_Clipboard)
        self.label_Pred_CM2 = QtWidgets.QLabel(self.groupBox_confusionMatrixPlot)
        font = QtGui.QFont()
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.label_Pred_CM2.setFont(font)
        self.label_Pred_CM2.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_Pred_CM2.setLineWidth(1)
        self.label_Pred_CM2.setMidLineWidth(0)
        self.label_Pred_CM2.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.label_Pred_CM2.setIndent(-1)
        self.label_Pred_CM2.setObjectName("label_Pred_CM2")
        self.horizontalLayout_22.addWidget(self.label_Pred_CM2)
        self.verticalLayout_12.addLayout(self.horizontalLayout_22)
        self.horizontalLayout_24.addLayout(self.verticalLayout_12)
        self.gridLayout_16.addLayout(self.horizontalLayout_24, 0, 0, 1, 1)
        self.tableWidget_AccPrecSpec = QtWidgets.QTableWidget(self.groupBox_confusionMatrixPlot)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget_AccPrecSpec.sizePolicy().hasHeightForWidth())
        self.tableWidget_AccPrecSpec.setSizePolicy(sizePolicy)
        self.tableWidget_AccPrecSpec.setObjectName("tableWidget_AccPrecSpec")
        self.gridLayout_16.addWidget(self.tableWidget_AccPrecSpec, 1, 0, 1, 1)
        self.groupBox_probHistPlot = QtWidgets.QGroupBox(self.splitter_8)
        self.groupBox_probHistPlot.setObjectName("groupBox_probHistPlot")
        self.gridLayout_19 = QtWidgets.QGridLayout(self.groupBox_probHistPlot)
        self.gridLayout_19.setObjectName("gridLayout_19")
        self.widget_probHistPlot = pg.GraphicsLayoutWidget(self.groupBox_probHistPlot)#QtWidgets.QWidget(self.groupBox_probHistPlot)
        self.widget_probHistPlot.setObjectName("widget_probHistPlot")
        self.gridLayout_19.addWidget(self.widget_probHistPlot, 0, 0, 1, 1)
        self.groupBox_3rdPlot = QtWidgets.QGroupBox(self.splitter_8)
        self.groupBox_3rdPlot.setObjectName("groupBox_3rdPlot")
        self.gridLayout_18 = QtWidgets.QGridLayout(self.groupBox_3rdPlot)
        self.gridLayout_18.setObjectName("gridLayout_18")
        self.widget_3rdPlot = pg.GraphicsLayoutWidget(self.groupBox_3rdPlot)
        self.widget_3rdPlot.setObjectName("widget_3rdPlot")
        self.gridLayout_18.addWidget(self.widget_3rdPlot, 1, 0, 1, 1)
        self.gridLayout_23.addWidget(self.splitter_7, 0, 0, 1, 1)
        self.tabWidget_Modelbuilder.addTab(self.tab_AssessModel, "")

        ##################Tab Plotting and Peakdetermination##################

        self.tab_Plotting = QtWidgets.QWidget()
        self.tab_Plotting.setObjectName("tab_Plotting")
        self.gridLayout_25 = QtWidgets.QGridLayout(self.tab_Plotting)
        self.gridLayout_25.setObjectName("gridLayout_25")
        self.comboBox_chooseRtdcFile = QtWidgets.QComboBox(self.tab_Plotting)
        self.comboBox_chooseRtdcFile.setObjectName("comboBox_chooseRtdcFile")
        self.gridLayout_25.addWidget(self.comboBox_chooseRtdcFile, 0, 0, 1, 1)
        self.splitter_15 = QtWidgets.QSplitter(self.tab_Plotting)
        self.splitter_15.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_15.setObjectName("splitter_15")
        self.splitter_14 = QtWidgets.QSplitter(self.splitter_15)
        self.splitter_14.setOrientation(QtCore.Qt.Vertical)
        self.splitter_14.setObjectName("splitter_14")
        self.groupBox_plottingregion = QtWidgets.QGroupBox(self.splitter_14)
        self.groupBox_plottingregion.setObjectName("groupBox_plottingregion")
        self.gridLayout_27 = QtWidgets.QGridLayout(self.groupBox_plottingregion)
        self.gridLayout_27.setObjectName("gridLayout_27")
        self.horizontalLayout_27 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_27.setObjectName("horizontalLayout_27")
        self.comboBox_featurey = QtWidgets.QComboBox(self.groupBox_plottingregion)
        self.comboBox_featurey.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.comboBox_featurey.setCurrentText("")
        self.comboBox_featurey.setObjectName("comboBox_featurey")
        self.horizontalLayout_27.addWidget(self.comboBox_featurey)
        self.verticalLayout_15 = QtWidgets.QVBoxLayout()
        self.verticalLayout_15.setObjectName("verticalLayout_15")
        self.splitter_13 = QtWidgets.QSplitter(self.groupBox_plottingregion)
        self.splitter_13.setOrientation(QtCore.Qt.Vertical)
        self.splitter_13.setObjectName("splitter_13")
        self.splitter_12 = QtWidgets.QSplitter(self.splitter_13)
        self.splitter_12.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_12.setObjectName("splitter_12")
        self.widget_histx = pg.GraphicsLayoutWidget(self.splitter_12)
        self.widget_histx.resize(QtCore.QSize(251, 75))
        self.widget_histx.setObjectName("widget_histx")
        self.widget_infoBox = QtWidgets.QWidget(self.splitter_12)
        self.widget_infoBox.setObjectName("widget_infoBox")
        self.gridLayout_30 = QtWidgets.QGridLayout(self.widget_infoBox)
        self.gridLayout_30.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_30.setObjectName("gridLayout_30")
        self.spinBox_cellInd = QtWidgets.QSpinBox(self.widget_infoBox)
        self.spinBox_cellInd.setMaximumSize(QtCore.QSize(16777215, 22))
        self.spinBox_cellInd.setMaximum(999999999)
        self.spinBox_cellInd.setObjectName("spinBox_cellInd")
        self.gridLayout_30.addWidget(self.spinBox_cellInd, 1, 0, 1, 1)
        self.horizontalSlider_cellInd = QtWidgets.QSlider(self.widget_infoBox)
        self.horizontalSlider_cellInd.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider_cellInd.setObjectName("horizontalSlider_cellInd")
        self.gridLayout_30.addWidget(self.horizontalSlider_cellInd, 0, 0, 1, 1)
        self.splitter_9 = QtWidgets.QSplitter(self.splitter_13)
        self.splitter_9.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_9.setObjectName("splitter_9")
        self.widget_scatter = pg.GraphicsLayoutWidget(self.splitter_9)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(1)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.widget_scatter.sizePolicy().hasHeightForWidth())
        self.widget_scatter.setSizePolicy(sizePolicy)
        #self.widget_scatter.setMinimumSize(QtCore.QSize(251, 251))
        self.widget_scatter.setSizeIncrement(QtCore.QSize(1, 1))
        self.widget_scatter.setBaseSize(QtCore.QSize(1, 1))
        self.widget_scatter.setObjectName("widget_scatter")
        self.widget_histy = pg.GraphicsLayoutWidget(self.splitter_9)
        #self.widget_histy.setMinimumSize(QtCore.QSize(75, 251))
        self.widget_histy.setObjectName("widget_histy")
        self.verticalLayout_15.addWidget(self.splitter_13)
        self.comboBox_featurex = QtWidgets.QComboBox(self.groupBox_plottingregion)
        self.comboBox_featurex.setMinimumSize(QtCore.QSize(326, 30))
        self.comboBox_featurex.setObjectName("comboBox_featurex")
        self.verticalLayout_15.addWidget(self.comboBox_featurex)
        self.horizontalLayout_27.addLayout(self.verticalLayout_15)
        self.gridLayout_27.addLayout(self.horizontalLayout_27, 0, 0, 1, 1)
        self.groupBox_plottingOptions = QtWidgets.QGroupBox(self.splitter_14)
        self.groupBox_plottingOptions.setObjectName("groupBox_plottingOptions")
        self.gridLayout_26 = QtWidgets.QGridLayout(self.groupBox_plottingOptions)
        self.gridLayout_26.setObjectName("gridLayout_26")
        self.verticalLayout_14 = QtWidgets.QVBoxLayout()
        self.verticalLayout_14.setObjectName("verticalLayout_14")
        self.horizontalLayout_29 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_29.setObjectName("horizontalLayout_29")
        self.checkBox_fl1 = QtWidgets.QCheckBox(self.groupBox_plottingOptions)
        self.checkBox_fl1.setObjectName("checkBox_fl1")
        self.horizontalLayout_29.addWidget(self.checkBox_fl1)
        self.checkBox_fl2 = QtWidgets.QCheckBox(self.groupBox_plottingOptions)
        self.checkBox_fl2.setObjectName("checkBox_fl2")
        self.horizontalLayout_29.addWidget(self.checkBox_fl2)
        self.checkBox_fl3 = QtWidgets.QCheckBox(self.groupBox_plottingOptions)
        self.checkBox_fl3.setObjectName("checkBox_fl3")
        self.horizontalLayout_29.addWidget(self.checkBox_fl3)
        self.checkBox_centroid = QtWidgets.QCheckBox(self.groupBox_plottingOptions)
        self.checkBox_centroid.setObjectName("checkBox_centroid")
        self.horizontalLayout_29.addWidget(self.checkBox_centroid)

        self.verticalLayout_14.addLayout(self.horizontalLayout_29)
        self.horizontalLayout_26 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_26.setObjectName("horizontalLayout_26")
        self.label_coloring = QtWidgets.QLabel(self.groupBox_plottingOptions)
        self.label_coloring.setObjectName("label_coloring")
        self.horizontalLayout_26.addWidget(self.label_coloring)
        self.comboBox_coloring = QtWidgets.QComboBox(self.groupBox_plottingOptions)
        self.comboBox_coloring.setObjectName("comboBox_coloring")
        self.horizontalLayout_26.addWidget(self.comboBox_coloring)
        self.checkBox_colorLog = QtWidgets.QCheckBox(self.groupBox_plottingOptions)
        self.checkBox_colorLog.setObjectName("checkBox_colorLog")
        self.horizontalLayout_26.addWidget(self.checkBox_colorLog)
        self.pushButton_updateScatterPlot = QtWidgets.QPushButton(self.groupBox_plottingOptions)
        self.pushButton_updateScatterPlot.setObjectName("pushButton_updateScatterPlot")
        self.horizontalLayout_26.addWidget(self.pushButton_updateScatterPlot)
        self.verticalLayout_14.addLayout(self.horizontalLayout_26)
        self.gridLayout_26.addLayout(self.verticalLayout_14, 0, 0, 1, 1)
        self.groupBox = QtWidgets.QGroupBox(self.splitter_14)
        self.groupBox.setObjectName("groupBox")
        self.gridLayout_33 = QtWidgets.QGridLayout(self.groupBox)
        self.gridLayout_33.setObjectName("gridLayout_33")
        self.textBrowser_fileInfo = QtWidgets.QTextBrowser(self.groupBox)
        self.textBrowser_fileInfo.setObjectName("textBrowser_fileInfo")
        self.gridLayout_33.addWidget(self.textBrowser_fileInfo, 0, 0, 1, 1)
        self.tabWidget_filter_peakdet = QtWidgets.QTabWidget(self.splitter_15)
        self.tabWidget_filter_peakdet.setObjectName("tabWidget_filter_peakdet")
        self.tab_filter = QtWidgets.QWidget()
        self.tab_filter.setObjectName("tab_filter")
        self.gridLayout_24 = QtWidgets.QGridLayout(self.tab_filter)
        self.gridLayout_24.setObjectName("gridLayout_24")
        self.tableWidget_filterOptions = QtWidgets.QTableWidget(self.tab_filter)
        self.tableWidget_filterOptions.setObjectName("tableWidget_filterOptions")
        self.tableWidget_filterOptions.setColumnCount(0)
        self.tableWidget_filterOptions.setRowCount(0)
        self.gridLayout_24.addWidget(self.tableWidget_filterOptions, 0, 0, 1, 1)
        self.tabWidget_filter_peakdet.addTab(self.tab_filter, "")
        self.tab_peakdet = QtWidgets.QWidget()
        self.tab_peakdet.setObjectName("tab_peakdet")
        self.gridLayout_29 = QtWidgets.QGridLayout(self.tab_peakdet)
        self.gridLayout_29.setObjectName("gridLayout_29")
        self.splitter_10 = QtWidgets.QSplitter(self.tab_peakdet)
        self.splitter_10.setOrientation(QtCore.Qt.Vertical)
        self.splitter_10.setObjectName("splitter_10")
        self.groupBox_showCell = QtWidgets.QGroupBox(self.splitter_10)
        self.groupBox_showCell.setObjectName("groupBox_showCell")
        self.gridLayout_32 = QtWidgets.QGridLayout(self.groupBox_showCell)
        self.gridLayout_32.setObjectName("gridLayout_32")
        self.widget_showFltrace = pg.GraphicsLayoutWidget(self.groupBox_showCell)
        self.widget_showFltrace.setMinimumSize(QtCore.QSize(0, 81))
        #self.widget_showFltrace.setMaximumSize(QtCore.QSize(16777215, 81))
        self.widget_showFltrace.setObjectName("widget_showFltrace")
        self.gridLayout_32.addWidget(self.widget_showFltrace, 1, 0, 1, 1)
        self.widget_showCell = pg.ImageView(self.groupBox_showCell)
        self.widget_showCell.setMinimumSize(QtCore.QSize(0, 91))
        #self.widget_showCell.setMaximumSize(QtCore.QSize(16777215, 91))
        self.widget_showCell.ui.histogram.hide()
        self.widget_showCell.ui.roiBtn.hide()
        self.widget_showCell.ui.menuBtn.hide()
        self.widget_showCell.setObjectName("widget_showCell")
        self.gridLayout_32.addWidget(self.widget_showCell, 0, 0, 1, 1)
        
        self.groupBox_showSelectedPeaks = QtWidgets.QGroupBox(self.splitter_10)
        self.groupBox_showSelectedPeaks.setObjectName("groupBox_showSelectedPeaks")
        self.gridLayout_28 = QtWidgets.QGridLayout(self.groupBox_showSelectedPeaks)
        self.gridLayout_28.setObjectName("gridLayout_28")
        self.splitter_11 = QtWidgets.QSplitter(self.groupBox_showSelectedPeaks)
        self.splitter_11.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_11.setObjectName("splitter_11")
        self.widget = QtWidgets.QWidget(self.splitter_11)
        self.widget.setObjectName("widget")
        self.verticalLayout_18 = QtWidgets.QVBoxLayout(self.widget)
        self.verticalLayout_18.setSizeConstraint(QtWidgets.QLayout.SetMinimumSize)
        self.verticalLayout_18.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_18.setObjectName("verticalLayout_18")
        self.horizontalLayout_30 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_30.setObjectName("horizontalLayout_30")
        self.pushButton_selectPeakPos = QtWidgets.QPushButton(self.widget)
        self.pushButton_selectPeakPos.setMinimumSize(QtCore.QSize(50, 28))
        self.pushButton_selectPeakPos.setMaximumSize(QtCore.QSize(50, 28))
        self.pushButton_selectPeakPos.setObjectName("pushButton_selectPeakPos")
        self.horizontalLayout_30.addWidget(self.pushButton_selectPeakPos)
        self.pushButton_selectPeakRange = QtWidgets.QPushButton(self.widget)
        self.pushButton_selectPeakRange.setMinimumSize(QtCore.QSize(50, 28))
        self.pushButton_selectPeakRange.setMaximumSize(QtCore.QSize(50, 28))
        self.pushButton_selectPeakRange.setObjectName("pushButton_selectPeakRange")
        self.horizontalLayout_30.addWidget(self.pushButton_selectPeakRange)
        self.verticalLayout_18.addLayout(self.horizontalLayout_30)
        self.label_automatic = QtWidgets.QLabel(self.widget)
        self.label_automatic.setObjectName("label_automatic")
        self.verticalLayout_18.addWidget(self.label_automatic)
        self.horizontalLayout_28 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_28.setObjectName("horizontalLayout_28")
        self.pushButton_highestXPercent = QtWidgets.QPushButton(self.widget)
        self.pushButton_highestXPercent.setMinimumSize(QtCore.QSize(82, 28))
        self.pushButton_highestXPercent.setMaximumSize(QtCore.QSize(82, 28))
        self.pushButton_highestXPercent.setObjectName("pushButton_highestXPercent")
        self.horizontalLayout_28.addWidget(self.pushButton_highestXPercent)
        self.doubleSpinBox_highestXPercent = QtWidgets.QDoubleSpinBox(self.widget)
        self.doubleSpinBox_highestXPercent.setMinimumSize(QtCore.QSize(51, 22))
        self.doubleSpinBox_highestXPercent.setMaximumSize(QtCore.QSize(51, 22))
        self.doubleSpinBox_highestXPercent.setObjectName("doubleSpinBox_highestXPercent")
        self.horizontalLayout_28.addWidget(self.doubleSpinBox_highestXPercent)
        self.verticalLayout_18.addLayout(self.horizontalLayout_28)
        spacerItem = QtWidgets.QSpacerItem(157, 13, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.verticalLayout_18.addItem(spacerItem)
        self.label_remove = QtWidgets.QLabel(self.widget)
        self.label_remove.setObjectName("label_remove")
        self.verticalLayout_18.addWidget(self.label_remove)
        self.horizontalLayout_31 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_31.setObjectName("horizontalLayout_31")
        self.pushButton_removeSelectedPeaks = QtWidgets.QPushButton(self.widget)
        self.pushButton_removeSelectedPeaks.setMinimumSize(QtCore.QSize(60, 28))
        self.pushButton_removeSelectedPeaks.setMaximumSize(QtCore.QSize(60, 28))
        self.pushButton_removeSelectedPeaks.setObjectName("pushButton_removeSelectedPeaks")
        self.horizontalLayout_31.addWidget(self.pushButton_removeSelectedPeaks)
        self.pushButton_removeAllPeaks = QtWidgets.QPushButton(self.widget)
        self.pushButton_removeAllPeaks.setMinimumSize(QtCore.QSize(40, 28))
        self.pushButton_removeAllPeaks.setMaximumSize(QtCore.QSize(40, 28))
        self.pushButton_removeAllPeaks.setObjectName("pushButton_removeAllPeaks")
        self.horizontalLayout_31.addWidget(self.pushButton_removeAllPeaks)
        self.verticalLayout_18.addLayout(self.horizontalLayout_31)
        self.widget_showSelectedPeaks = pg.GraphicsLayoutWidget(self.splitter_11)
        self.widget_showSelectedPeaks.setObjectName("widget_showSelectedPeaks")
        self.tableWidget_showSelectedPeaks = QtWidgets.QTableWidget(self.splitter_11)
        self.tableWidget_showSelectedPeaks.setObjectName("tableWidget_showSelectedPeaks")
        self.gridLayout_28.addWidget(self.splitter_11, 0, 0, 1, 1)


        self.groupBox_peakDetModel = QtWidgets.QGroupBox(self.splitter_10)
        self.groupBox_peakDetModel.setObjectName("groupBox_peakDetModel")
        self.gridLayout_31 = QtWidgets.QGridLayout(self.groupBox_peakDetModel)
        self.gridLayout_31.setObjectName("gridLayout_31")
        self.horizontalLayout_25 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_25.setObjectName("horizontalLayout_25")
        self.verticalLayout_16 = QtWidgets.QVBoxLayout()
        self.verticalLayout_16.setObjectName("verticalLayout_16")
        self.comboBox_peakDetModel = QtWidgets.QComboBox(self.groupBox_peakDetModel)
        self.comboBox_peakDetModel.setObjectName("comboBox_peakDetModel")
        self.verticalLayout_16.addWidget(self.comboBox_peakDetModel)
        self.pushButton_fitPeakDetModel = QtWidgets.QPushButton(self.groupBox_peakDetModel)
        self.pushButton_fitPeakDetModel.setObjectName("pushButton_fitPeakDetModel")
        self.verticalLayout_16.addWidget(self.pushButton_fitPeakDetModel)
        self.pushButton_SavePeakDetModel = QtWidgets.QPushButton(self.groupBox_peakDetModel)
        self.pushButton_SavePeakDetModel.setObjectName("pushButton_SavePeakDetModel")
        self.verticalLayout_16.addWidget(self.pushButton_SavePeakDetModel)
        self.pushButton_loadPeakDetModel = QtWidgets.QPushButton(self.groupBox_peakDetModel)
        self.pushButton_loadPeakDetModel.setObjectName("pushButton_loadPeakDetModel")
        self.verticalLayout_16.addWidget(self.pushButton_loadPeakDetModel)
        self.horizontalLayout_25.addLayout(self.verticalLayout_16)
        self.tableWidget_peakModelParameters = QtWidgets.QTableWidget(self.groupBox_peakDetModel)
        self.tableWidget_peakModelParameters.setObjectName("tableWidget_peakModelParameters")
        self.tableWidget_peakModelParameters.setColumnCount(0)
        self.tableWidget_peakModelParameters.setRowCount(0)
        self.horizontalLayout_25.addWidget(self.tableWidget_peakModelParameters)
        self.gridLayout_31.addLayout(self.horizontalLayout_25, 0, 0, 1, 1)
        self.verticalLayout_17 = QtWidgets.QVBoxLayout()
        self.verticalLayout_17.setObjectName("verticalLayout_17")
        self.radioButton_exportSelected = QtWidgets.QRadioButton(self.groupBox_peakDetModel)
        self.radioButton_exportSelected.setChecked(True)
        self.radioButton_exportSelected.setObjectName("radioButton_exportSelected")
        self.verticalLayout_17.addWidget(self.radioButton_exportSelected)
        self.radioButton_exportAll = QtWidgets.QRadioButton(self.groupBox_peakDetModel)
        self.radioButton_exportAll.setObjectName("radioButton_exportAll")
        self.verticalLayout_17.addWidget(self.radioButton_exportAll)
        self.comboBox_toFlOrUserdef = QtWidgets.QComboBox(self.groupBox_peakDetModel)
        self.comboBox_toFlOrUserdef.setObjectName("comboBox_toFlOrUserdef")
        self.verticalLayout_17.addWidget(self.comboBox_toFlOrUserdef)
        self.pushButton_export = QtWidgets.QPushButton(self.groupBox_peakDetModel)
        self.pushButton_export.setObjectName("pushButton_export")
        self.verticalLayout_17.addWidget(self.pushButton_export)
        self.gridLayout_31.addLayout(self.verticalLayout_17, 0, 1, 1, 1)
        self.gridLayout_29.addWidget(self.splitter_10, 0, 0, 1, 1)
        self.tabWidget_filter_peakdet.addTab(self.tab_peakdet, "")
        self.tab_defineModel = QtWidgets.QWidget()
        self.tab_defineModel.setObjectName("tab_defineModel")
        self.tabWidget_filter_peakdet.addTab(self.tab_defineModel, "")
        self.gridLayout_25.addWidget(self.splitter_15, 1, 0, 1, 1)
        self.tabWidget_Modelbuilder.addTab(self.tab_Plotting, "")






        ####################Tab Python Editor/Console##########################
        self.tab_python = QtWidgets.QWidget()
        self.tab_python.setObjectName("tab_python")
        self.gridLayout_41 = QtWidgets.QGridLayout(self.tab_python)
        self.gridLayout_41.setObjectName("gridLayout_41")
        self.verticalLayout_python_1 = QtWidgets.QVBoxLayout()
        self.verticalLayout_python_1.setObjectName("verticalLayout_python_1")
        self.groupBox_pythonMenu = QtWidgets.QGroupBox(self.tab_python)
        self.groupBox_pythonMenu.setMaximumSize(QtCore.QSize(16777215, 71))
        self.groupBox_pythonMenu.setObjectName("groupBox_pythonMenu")
        self.gridLayout_40 = QtWidgets.QGridLayout(self.groupBox_pythonMenu)
        self.gridLayout_40.setObjectName("gridLayout_40")
        self.horizontalLayout_pythonMenu = QtWidgets.QHBoxLayout()
        self.horizontalLayout_pythonMenu.setObjectName("horizontalLayout_pythonMenu")
        self.label_pythonCurrentFile = QtWidgets.QLabel(self.groupBox_pythonMenu)
        self.label_pythonCurrentFile.setObjectName("label_pythonCurrentFile")
        self.horizontalLayout_pythonMenu.addWidget(self.label_pythonCurrentFile)
        self.lineEdit_pythonCurrentFile = QtWidgets.QLineEdit(self.groupBox_pythonMenu)
        self.lineEdit_pythonCurrentFile.setEnabled(False)
        self.lineEdit_pythonCurrentFile.setObjectName("lineEdit_pythonCurrentFile")
        self.horizontalLayout_pythonMenu.addWidget(self.lineEdit_pythonCurrentFile)
        self.gridLayout_40.addLayout(self.horizontalLayout_pythonMenu, 0, 0, 1, 1)
        self.verticalLayout_python_1.addWidget(self.groupBox_pythonMenu)
        self.splitter_python_1 = QtWidgets.QSplitter(self.tab_python)
        self.splitter_python_1.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_python_1.setObjectName("splitter_python_1")
        self.groupBox_pythonEditor = QtWidgets.QGroupBox(self.splitter_python_1)
        self.groupBox_pythonEditor.setObjectName("groupBox_pythonEditor")
        self.gridLayout_38 = QtWidgets.QGridLayout(self.groupBox_pythonEditor)
        self.gridLayout_38.setObjectName("gridLayout_38")
        self.verticalLayout_editor_1 = QtWidgets.QVBoxLayout()
        self.verticalLayout_editor_1.setObjectName("verticalLayout_editor_1")
        self.horizontalLayout_editor_1 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_editor_1.setObjectName("horizontalLayout_editor_1")
        self.pushButton_pythonInOpen = QtWidgets.QPushButton(self.groupBox_pythonEditor)
        self.pushButton_pythonInOpen.setObjectName("pushButton_pythonInOpen")
        self.horizontalLayout_editor_1.addWidget(self.pushButton_pythonInOpen)
        self.pushButton_pythonSaveAs = QtWidgets.QPushButton(self.groupBox_pythonEditor)
        self.pushButton_pythonSaveAs.setObjectName("pushButton_pythonSaveAs")
        self.horizontalLayout_editor_1.addWidget(self.pushButton_pythonSaveAs)
        self.pushButton_pythonInClear = QtWidgets.QPushButton(self.groupBox_pythonEditor)
        self.pushButton_pythonInClear.setObjectName("pushButton_pythonInClear")
        self.horizontalLayout_editor_1.addWidget(self.pushButton_pythonInClear)
        self.pushButton_pythonInRun = QtWidgets.QPushButton(self.groupBox_pythonEditor)
        self.pushButton_pythonInRun.setObjectName("pushButton_pythonInRun")
        self.horizontalLayout_editor_1.addWidget(self.pushButton_pythonInRun)
        self.verticalLayout_editor_1.addLayout(self.horizontalLayout_editor_1)
        self.plainTextEdit_pythonIn = QtWidgets.QPlainTextEdit(self.groupBox_pythonEditor)
        self.plainTextEdit_pythonIn.setObjectName("plainTextEdit_pythonIn")
        self.verticalLayout_editor_1.addWidget(self.plainTextEdit_pythonIn)
        self.gridLayout_38.addLayout(self.verticalLayout_editor_1, 0, 0, 1, 1)
        self.groupBox_pythonConsole = QtWidgets.QGroupBox(self.splitter_python_1)
        self.groupBox_pythonConsole.setObjectName("groupBox_pythonConsole")
        self.gridLayout_39 = QtWidgets.QGridLayout(self.groupBox_pythonConsole)
        self.gridLayout_39.setObjectName("gridLayout_39")
        self.verticalLayout_console_1 = QtWidgets.QVBoxLayout()
        self.verticalLayout_console_1.setObjectName("verticalLayout_console_1")
        self.horizontalLayout_console_1 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_console_1.setObjectName("horizontalLayout_console_1")
        self.pushButton_pythonOutClear = QtWidgets.QPushButton(self.groupBox_pythonConsole)
        self.pushButton_pythonOutClear.setObjectName("pushButton_pythonOutClear")
        self.horizontalLayout_console_1.addWidget(self.pushButton_pythonOutClear)
        self.pushButton_pythonOutRun = QtWidgets.QPushButton(self.groupBox_pythonConsole)
        self.pushButton_pythonOutRun.setObjectName("pushButton_pythonOutRun")
        self.horizontalLayout_console_1.addWidget(self.pushButton_pythonOutRun)
        self.verticalLayout_console_1.addLayout(self.horizontalLayout_console_1)
        self.textBrowser_pythonOut = QtWidgets.QTextBrowser(self.groupBox_pythonConsole)
        self.textBrowser_pythonOut.setEnabled(True)
        self.textBrowser_pythonOut.setObjectName("textBrowser_pythonOut")
        self.verticalLayout_console_1.addWidget(self.textBrowser_pythonOut)
        self.gridLayout_39.addLayout(self.verticalLayout_console_1, 0, 0, 1, 1)
        self.verticalLayout_python_1.addWidget(self.splitter_python_1)
        self.gridLayout_41.addLayout(self.verticalLayout_python_1, 0, 0, 1, 1)
        self.tabWidget_Modelbuilder.addTab(self.tab_python, "")
        

        
        
        
        
        
        
        
        
        




        ######################Connections######################################
        self.doubleSpinBox_learningRate.setEnabled(False)
        self.spinBox_trainLastNOnly.setEnabled(False)
        self.lineEdit_dropout.setEnabled(False)
        self.checkBox_learningRate.toggled['bool'].connect(self.doubleSpinBox_learningRate.setEnabled)        
        self.checkBox_trainLastNOnly.toggled['bool'].connect(self.spinBox_trainLastNOnly.setEnabled)
        self.checkBox_dropout.toggled['bool'].connect(self.lineEdit_dropout.setEnabled)













        #######################################################################
        ##########################Manual manipulation##########################
        ##########################Manual manipulation##########################
        #######################################################################
        ###########this makes it easier if the Ui should be changed############



        ####################Define Model####################################
        #self.label_Crop.setChecked(True)
        #self.label_Crop.stateChanged.connect(self.activate_deactivate_spinbox)
        
        self.spinBox_imagecrop.valueChanged.connect(self.delete_ram)
        self.colorModes = ["Grayscale","RGB"]
        self.comboBox_GrayOrRGB.addItems(self.colorModes)
        self.comboBox_GrayOrRGB.setCurrentIndex(0)             
        self.comboBox_GrayOrRGB.currentIndexChanged.connect(self.gray_or_rgb_augmentation)            
        #By default, Color Mode is Grayscale. Therefore switch off saturation and hue
        self.checkBox_contrast.setEnabled(True)
        self.checkBox_contrast.setChecked(True)
        self.doubleSpinBox_contrastLower.setEnabled(True)
        self.doubleSpinBox_contrastHigher.setEnabled(True)
        self.checkBox_saturation.setEnabled(False)
        self.checkBox_saturation.setChecked(False)
        self.doubleSpinBox_saturationLower.setEnabled(False)
        self.doubleSpinBox_saturationHigher.setEnabled(False)
        self.checkBox_hue.setEnabled(False)
        self.checkBox_hue.setChecked(False)
        self.doubleSpinBox_hueDelta.setEnabled(False)


        ###########################Expert mode Tab################################
        self.spinBox_batchSize.setMinimum(1)       
        self.spinBox_batchSize.setMaximum(1E6)       
        self.spinBox_batchSize.setValue(128)       
        self.spinBox_epochs.setMinimum(1)       
        self.spinBox_epochs.setMaximum(1E6)       
        self.spinBox_epochs.setValue(1)       
        self.doubleSpinBox_learningRate.setDecimals(9)
        self.doubleSpinBox_learningRate.setMinimum(0.0)       
        self.doubleSpinBox_learningRate.setMaximum(1E6)       
        self.doubleSpinBox_learningRate.setValue(0.001)       
        self.doubleSpinBox_learningRate.setSingleStep(0.0001)
        self.spinBox_trainLastNOnly.setMinimum(0)       
        self.spinBox_trainLastNOnly.setMaximum(1E6)       
        self.spinBox_trainLastNOnly.setValue(0)    
        self.checkBox_trainDenseOnly.setChecked(False)
        self.checkBox_partialTrainability.toggled.connect(self.partialtrainability_activated)
        self.checkBox_lossW.clicked.connect(lambda on_or_off: self.lossWeights_activated(on_or_off,-1))
        self.pushButton_lossW.clicked.connect(lambda: self.lossWeights_popup(-1))

        self.checkBox_expt_paddingMode.stateChanged.connect(self.expert_paddingMode_off)

        ###########################History Tab################################
        self.tableWidget_HistoryItems.doubleClicked.connect(self.tableWidget_HistoryItems_dclick)
        conversion_methods_source = ["Keras TensorFlow", "Frozen TensorFlow .pb"]
        conversion_methods_target = [".nnet","Frozen TensorFlow .pb", "Optimized TensorFlow .pb", "ONNX (via keras2onnx)", "ONNX (via MMdnn)","CoreML", "PyTorch Script","Caffe Script","CNTK Script","MXNet Script","ONNX Script","TensorFlow Script","Keras Script"]
        self.comboBox_convertTo.addItems(conversion_methods_target)
        self.comboBox_convertTo.setMinimumSize(QtCore.QSize(200,22))
        self.comboBox_convertTo.setMaximumSize(QtCore.QSize(200, 22))
        width=self.comboBox_convertTo.fontMetrics().boundingRect(max(conversion_methods_target, key=len)).width()
        self.comboBox_convertTo.view().setFixedWidth(width+10)
        self.combobox_initial_format.setCurrentIndex(0)             
        #self.comboBox_convertTo.setEnabled(False)
        
        self.combobox_initial_format.addItems(conversion_methods_source)
        self.combobox_initial_format.setMinimumSize(QtCore.QSize(200,22))
        self.combobox_initial_format.setMaximumSize(QtCore.QSize(200, 22))
        width=self.combobox_initial_format.fontMetrics().boundingRect(max(conversion_methods_source, key=len)).width()
        self.combobox_initial_format.view().setFixedWidth(width+10)             
        self.combobox_initial_format.setCurrentIndex(0)             
        #self.combobox_initial_format.setEnabled(False)


        ###########################Assess Model################################
        self.comboBox_loadedRGBorGray.addItems(["Grayscale","RGB"])
        self.groupBox_loadModel.setMaximumSize(QtCore.QSize(16777215, 120))
        self.label_ModelIndex_2.setMinimumSize(QtCore.QSize(68, 25))
        self.label_ModelIndex_2.setMaximumSize(QtCore.QSize(68, 25))
        self.spinBox_ModelIndex_2.setEnabled(False) 
        self.spinBox_ModelIndex_2.setMinimum(0)
        self.spinBox_ModelIndex_2.setMaximum(9E8)
        self.spinBox_ModelIndex_2.setMinimumSize(QtCore.QSize(40, 22))
        self.spinBox_ModelIndex_2.setMaximumSize(QtCore.QSize(75, 22))
        self.spinBox_Crop_2.setMinimumSize(QtCore.QSize(40, 22))
        self.spinBox_Crop_2.setMaximumSize(QtCore.QSize(50, 22))
        self.spinBox_Crop_2.setMinimum(1)
        self.spinBox_Crop_2.setMaximum(9E8)
        self.spinBox_OutClasses_2.setMinimumSize(QtCore.QSize(40, 22))
        self.spinBox_OutClasses_2.setMaximumSize(QtCore.QSize(50, 22))
        self.spinBox_OutClasses_2.setMinimum(1)
        self.spinBox_OutClasses_2.setMaximum(9E8)
        
        self.lineEdit_ModelSelection_2.setMinimumSize(QtCore.QSize(0, 20))
        self.lineEdit_ModelSelection_2.setMaximumSize(QtCore.QSize(16777215, 20))

        self.pushButton_LoadModel_2.setMinimumSize(QtCore.QSize(123, 24))
        self.pushButton_LoadModel_2.setMaximumSize(QtCore.QSize(123, 24))
        self.pushButton_LoadModel_2.clicked.connect(self.assessmodel_tab_load_model)

        self.comboBox_Normalization_2.addItems(self.norm_methods)
        self.comboBox_Normalization_2.setMinimumSize(QtCore.QSize(200,22))
        self.comboBox_Normalization_2.setMaximumSize(QtCore.QSize(200, 22))
        width=self.comboBox_Normalization_2.fontMetrics().boundingRect(max(self.norm_methods, key=len)).width()
        self.comboBox_Normalization_2.view().setFixedWidth(width+10)             

        self.pushButton_ImportValidFromNpy.setMinimumSize(QtCore.QSize(65, 28))
        self.pushButton_ImportValidFromNpy.clicked.connect(self.import_valid_from_rtdc)
        self.pushButton_ExportValidToNpy.setMinimumSize(QtCore.QSize(55, 0))
        self.pushButton_ExportValidToNpy.clicked.connect(self.export_valid_to_rtdc)

        self.lineEdit_InferenceTime.setMinimumSize(QtCore.QSize(0, 30))
        self.lineEdit_InferenceTime.setMaximumSize(QtCore.QSize(16777215, 30))
        self.spinBox_inftime_nr_images.setMinimum(10)
        self.spinBox_inftime_nr_images.setMaximum(9E8)
        self.spinBox_inftime_nr_images.setValue(1000)
        
        self.groupBox_validData.setMaximumSize(QtCore.QSize(400, 250))
        self.tableWidget_Info_2.setColumnCount(0)
        self.tableWidget_Info_2.setRowCount(0)
        self.tableWidget_Info_2.clicked.connect(self.tableWidget_Info_2_click)
        self.groupBox_settings.setMaximumSize(QtCore.QSize(16777215, 250))
        self.spinBox_indexOfInterest.setMinimum(0)
        self.spinBox_indexOfInterest.setMaximum(9E8)

        self.doubleSpinBox_sortingThresh.setMinimum(0)
        self.doubleSpinBox_sortingThresh.setMaximum(1)
        self.doubleSpinBox_sortingThresh.setValue(0.5)
        self.doubleSpinBox_sortingThresh.setSingleStep(0.1)
        self.doubleSpinBox_sortingThresh.setDecimals(5)

        self.comboBox_selectData.setMinimumSize(QtCore.QSize(200,22))
        self.comboBox_selectData.setMaximumSize(QtCore.QSize(200, 32))

        
        #3rd Plot
        items_3rd_plot = ["None","Conc. vs. Threshold","Enrichment vs. Threshold","Yield vs. Threshold","ROC-AUC","Precision-Recall"]
        self.comboBox_3rdPlot.addItems(items_3rd_plot)
        self.comboBox_3rdPlot.setMinimumSize(QtCore.QSize(100,22))
        self.comboBox_3rdPlot.setMaximumSize(QtCore.QSize(100,22))
        width=self.comboBox_3rdPlot.fontMetrics().boundingRect(max(items_3rd_plot, key=len)).width()
        self.comboBox_3rdPlot.view().setFixedWidth(width+10)             
        self.comboBox_3rdPlot.currentTextChanged.connect(self.thirdplot)
        
        self.spinBox_Indx1.setMinimum(0)
        self.spinBox_Indx1.setMaximum(9E8)
        self.spinBox_Indx1.setEnabled(False)
        self.spinBox_Indx2.setMinimum(0)
        self.spinBox_Indx2.setMaximum(9E8)
        self.spinBox_Indx2.setEnabled(False)
#        self.groupBox_confusionMatrixPlot.setMinimumSize(QtCore.QSize(100, 16777215))
#        self.groupBox_confusionMatrixPlot.setMaximumSize(QtCore.QSize(600, 16777215))
        self.tableWidget_CM1.setColumnCount(0)
        self.tableWidget_CM1.setRowCount(0)
        self.tableWidget_CM1.doubleClicked.connect(self.export_to_rtdc)
        self.tableWidget_CM2.setColumnCount(0)
        self.tableWidget_CM2.setRowCount(0)
        self.tableWidget_CM2.doubleClicked.connect(self.export_to_rtdc)
        self.pushButton_CM1_to_Clipboard.clicked.connect(lambda: self.copy_cm_to_clipboard(1)) #1 tells the function to connect to CM1
        self.pushButton_CM2_to_Clipboard.clicked.connect(lambda: self.copy_cm_to_clipboard(2)) #2 tells the function to connect to CM2
        self.pushButton_CompInfTime.clicked.connect(self.inference_time)
        self.pushButton_AssessModel.clicked.connect(self.assess_model_plotting)
        #self.comboBox_probability_histogram.clicked.connect(self.probability_histogram)
        self.comboBox_probability_histogram.addItems(["Style1","Style2","Style3","Style4","Style5"])

        ##################Plot/Peak Tab########################################
        self.tabWidget_filter_peakdet.setCurrentIndex(1)
        self.comboBox_chooseRtdcFile.currentIndexChanged.connect(self.update_comboBox_feature_xy)
        self.pushButton_updateScatterPlot.clicked.connect(self.updateScatterPlot)        
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)  
        self.pushButton_updateScatterPlot.setFont(font)        
        #set infobox to minimal size
        #self.widget_infoBox.resize(QtCore.QSize(100, 100))
        self.widget_infoBox.setMaximumSize(QtCore.QSize(100, 100))
        self.comboBox_featurey.setMinimumSize(QtCore.QSize(30, 326))
        self.comboBox_featurey.setMaximumSize(QtCore.QSize(75, 326))
        self.comboBox_featurey.view().setFixedWidth(150)#Wider box for the dropdown text

        #Add plot        
        self.scatter_xy = self.widget_scatter.addPlot()
        self.scatter_xy.showGrid(x=True,y=True)
       
        #Fill histogram for x-axis; widget_histx
        self.hist_x = self.widget_histx.addPlot()
        #hide the x-axis
        self.hist_x.hideAxis('bottom')
        self.hist_x.setXLink(self.scatter_xy) ## test linking by name
        self.hist_x.showGrid(x=True,y=True)

        #Initiate histogram for y-axis; widget_histy
        self.hist_y = self.widget_histy.addPlot()
        #hide the axes
        self.hist_y.hideAxis('left')
        self.hist_y.setYLink(self.scatter_xy) ## test linking by name
        self.hist_y.showGrid(x=True,y=True)

        #PlotItem for the Fl-trace:
        self.plot_fl_trace = self.widget_showFltrace.addPlot()
        self.plot_fl_trace.showGrid(x=True,y=True)
        
        #When horizontalSlider_cellInd is changed, initiate an onClick event and the points have to be accordigly
        self.horizontalSlider_cellInd.valueChanged.connect(self.onIndexChange)
        self.spinBox_cellInd.valueChanged.connect(self.onIndexChange)

        #Select a peak
        self.pushButton_selectPeakPos.clicked.connect(self.selectPeakPos)
        #Select range
        self.pushButton_selectPeakRange.clicked.connect(self.selectPeakRange)
        #Initiate the peak-table
        #self.tableWidget_showSelectedPeaks.append("FL_max\tFl_pos\tpos_x")
        self.tableWidget_showSelectedPeaks.setColumnCount(0)
        self.tableWidget_showSelectedPeaks.setColumnCount(5)
        self.tableWidget_showSelectedPeaks.setRowCount(0)
        header_labels = ["FL_max","Fl_pos_[us]","pos_x_[um]","Fl_pos_[]","pos_x_[]"]
        self.tableWidget_showSelectedPeaks.setHorizontalHeaderLabels(header_labels) 
        header = self.tableWidget_showSelectedPeaks.horizontalHeader()
        for i in range(3):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        
        
        self.pushButton_removeSelectedPeaks.clicked.connect(self.actionRemoveSelectedPeaks_function)
        self.pushButton_removeAllPeaks.clicked.connect(self.actionRemoveAllPeaks_function)

        self.selectedPeaksPlot = self.widget_showSelectedPeaks.addPlot()
        self.selectedPeaksPlot.showGrid(x=True,y=True)
        self.selectedPeaksPlot.setLabel('bottom', 'pos_x', units='um')
        self.selectedPeaksPlot.setLabel('left', 'flx_pos', units='us')

        self.comboBox_peakDetModel.addItems(["Linear dependency and max in range"])
        self.comboBox_peakDetModel.setMinimumSize(QtCore.QSize(200,22))
        self.comboBox_peakDetModel.setMaximumSize(QtCore.QSize(200, 22))
        width=self.comboBox_peakDetModel.fontMetrics().boundingRect(max(self.norm_methods, key=len)).width()
        self.comboBox_peakDetModel.view().setFixedWidth(width+10)             
        self.pushButton_fitPeakDetModel.clicked.connect(self.update_peak_plot)

        self.pushButton_highestXPercent.clicked.connect(self.addHighestXPctPeaks)
        self.pushButton_SavePeakDetModel.clicked.connect(self.savePeakDetModel)
        self.checkBox_fl1.setChecked(True)
        self.checkBox_fl2.setChecked(True)
        self.checkBox_fl3.setChecked(True)
        self.checkBox_centroid.setChecked(True)
        self.doubleSpinBox_highestXPercent.setMinimum(0)
        self.doubleSpinBox_highestXPercent.setMaximum(100)
        self.doubleSpinBox_highestXPercent.setValue(5)

        self.tableWidget_AccPrecSpec.cellClicked.connect(lambda: self.copy_cm_to_clipboard(3))
        self.pushButton_loadPeakDetModel.clicked.connect(self.loadPeakDetModel)
        overwrite_methods = ["Overwrite Fl_max and Fl_pos","Save to userdef"]
        self.comboBox_toFlOrUserdef.addItems(overwrite_methods)
        width=self.comboBox_toFlOrUserdef.fontMetrics().boundingRect(max(overwrite_methods, key=len)).width()
        self.comboBox_toFlOrUserdef.view().setFixedWidth(width+10)             
        self.pushButton_export.clicked.connect(self.applyPeakModel_and_export)
        ex_opt = ["Scores and predictions to Excel sheet","Add predictions to .rtdc file (userdef0)"]
        ex_opt.append("Add pred&scores to .rtdc file (userdef0 to 9)")
        self.comboBox_scoresOrPrediction.addItems(ex_opt)
        self.pushButton_classify.clicked.connect(self.classify)



        #########################Python Editor/Console#########################
        self.pushButton_pythonInRun.clicked.connect(self.pythonInRun)
        self.pushButton_pythonInClear.clicked.connect(self.pythonInClear)
        self.pushButton_pythonSaveAs.clicked.connect(self.pythonInSaveAs)
        self.pushButton_pythonInOpen.clicked.connect(self.pythonInOpen)
        self.pushButton_pythonOutClear.clicked.connect(self.pythonOutClear)
        self.pushButton_pythonOutRun.clicked.connect(self.pythonInRun)


        #############################MenuBar###################################
        self.gridLayout_2.addWidget(self.tabWidget_Modelbuilder, 0, 1, 1, 1)
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 912, 26))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName(_fromUtf8("menuFile"))
        self.menuEdit = QtWidgets.QMenu(self.menubar)
        self.menuEdit.setObjectName(_fromUtf8("menuEdit"))
        
        self.menu_Options = QtWidgets.QMenu(self.menubar)
        self.menu_Options.setObjectName("menu_Options")
        self.menu_Options.setToolTipsVisible(True)

        self.menuLayout = QtWidgets.QMenu(self.menu_Options)
        self.menuLayout.setObjectName("menuLayout")
        self.menuExport = QtWidgets.QMenu(self.menu_Options)
        self.menuExport.setObjectName("menuExport")
        self.menuZoomOrder = QtWidgets.QMenu(self.menu_Options)
        self.menuZoomOrder.setObjectName("menuZoomOrder")
#        self.menuColorMode = QtWidgets.QMenu(self.menu_Options)
#        self.menuColorMode.setObjectName("menuColorMode")
#        self.menuColorMode.setToolTipsVisible(True)

        self.menu_Help = QtWidgets.QMenu(self.menubar)
        self.menu_Help.setObjectName("menu_Help")
        self.menu_Help.setToolTipsVisible(True)

        self.actionDocumentation = QtWidgets.QAction(self)
        self.actionDocumentation.triggered.connect(self.actionDocumentation_function)
        self.actionDocumentation.setObjectName(_fromUtf8("actionDocumentation"))
        self.menu_Help.addAction(self.actionDocumentation)
        self.actionSoftware = QtWidgets.QAction(self)
        self.actionSoftware.triggered.connect(self.actionSoftware_function)
        self.actionSoftware.setObjectName(_fromUtf8("actionSoftware"))
        self.menu_Help.addAction(self.actionSoftware)
        self.actionAbout = QtWidgets.QAction(self)
        self.actionAbout.triggered.connect(self.actionAbout_function)
        self.actionAbout.setObjectName(_fromUtf8("actionAbout"))
        self.menu_Help.addAction(self.actionAbout)
        self.actionUpdate = QtWidgets.QAction(self)
        self.actionUpdate.triggered.connect(self.actionUpdate_check_function)
        self.actionUpdate.setObjectName(_fromUtf8("actionUpdate"))
        self.menu_Help.addAction(self.actionUpdate)



        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        self.setStatusBar(self.statusbar)
        self.statusbar_cpuRam = QtWidgets.QLabel("CPU: xx%  RAM: xx%   ")
        self.statusbar.addPermanentWidget(self.statusbar_cpuRam)        
        
        self.actionLoadSession = QtWidgets.QAction(self)
        self.actionLoadSession.triggered.connect(self.actionLoadSession_function)
        self.actionLoadSession.setObjectName(_fromUtf8("actionLoadSession"))
        self.actionSaveSession = QtWidgets.QAction(self)
        self.actionSaveSession.triggered.connect(self.actionSaveSession_function)
        self.actionSaveSession.setObjectName(_fromUtf8("actionSaveSession"))
        
        self.actionQuit = QtWidgets.QAction(self)
        self.actionQuit.setObjectName(_fromUtf8("actionQuit"))
        self.actionQuit.triggered.connect(self.quit_app)
        
        self.actionDataToRamNow = QtWidgets.QAction(self)
        self.actionDataToRamNow.triggered.connect(self.actionDataToRamNow_function)
        self.actionDataToRamNow.setObjectName(_fromUtf8("actionDataToRamNow"))
        self.actionDataToRam = QtWidgets.QAction(self,checkable=True)
        self.actionDataToRam.setChecked(True)
        self.actionDataToRam.setObjectName(_fromUtf8("actionDataToRam"))
        self.actionVerbose = QtWidgets.QAction(self,checkable=True)
        self.actionVerbose.setObjectName(_fromUtf8("actionVerbose"))
        self.actionClearList = QtWidgets.QAction(self)
        self.actionClearList.triggered.connect(self.actionClearList_function)
        self.actionClearList.setObjectName(_fromUtf8("actionClearList"))
        self.actionRemoveSelected = QtWidgets.QAction(self)
        self.actionRemoveSelected.setObjectName(_fromUtf8("actionRemoveSelected"))
        self.actionRemoveSelected.triggered.connect(self.actionRemoveSelected_function)
        self.actionSaveToPng = QtWidgets.QAction(self)
        self.actionSaveToPng.setObjectName(_fromUtf8("actionSaveToPng"))
        self.actionSaveToPng.triggered.connect(self.actionSaveToPng_function)
        
        self.actionGroup_Export = QtWidgets.QActionGroup(self,exclusive=True)
        self.actionExport_Off = QtWidgets.QAction(self)
        self.actionExport_Off.setCheckable(True)
        self.actionExport_Off.setObjectName("actionExport_Off")
        self.actionExport_Original = QtWidgets.QAction(self)
        self.actionExport_Original.setCheckable(True)
        self.actionExport_Original.setChecked(True)
        self.actionExport_Original.setObjectName("actionExport_Original")
        self.actionExport_Cropped = QtWidgets.QAction(self)
        self.actionExport_Cropped.setCheckable(True)
        self.actionExport_Cropped.setChecked(False)
        self.actionExport_Cropped.setObjectName("actionExport_Cropped")
        a = self.actionGroup_Export.addAction(self.actionExport_Off)
        self.menuExport.addAction(a)
        a = self.actionGroup_Export.addAction(self.actionExport_Original)
        self.menuExport.addAction(a)
        a = self.actionGroup_Export.addAction(self.actionExport_Cropped)
        self.menuExport.addAction(a)


        self.actionGroup_ZoomOrder = QtWidgets.QActionGroup(self,exclusive=True)
        self.actionOrder0 = QtWidgets.QAction(self)
        self.actionOrder0.setCheckable(True)
        self.actionOrder0.setChecked(True)
        self.actionOrder0.setObjectName("actionOrder0")
        self.actionOrder1 = QtWidgets.QAction(self)
        self.actionOrder1.setCheckable(True)
        self.actionOrder1.setChecked(False)
        self.actionOrder1.setObjectName("actionOrder1")
        self.actionOrder2 = QtWidgets.QAction(self)
        self.actionOrder2.setCheckable(True)
        self.actionOrder2.setChecked(False)
        self.actionOrder2.setObjectName("actionOrder2")
        self.actionOrder3 = QtWidgets.QAction(self)
        self.actionOrder3.setCheckable(True)
        self.actionOrder3.setChecked(False)
        self.actionOrder3.setObjectName("actionOrder3")
        self.actionOrder4 = QtWidgets.QAction(self)
        self.actionOrder4.setCheckable(True)
        self.actionOrder4.setChecked(False)
        self.actionOrder4.setObjectName("actionOrder4")
        self.actionOrder5 = QtWidgets.QAction(self)
        self.actionOrder5.setCheckable(True)
        self.actionOrder5.setChecked(False)
        self.actionOrder5.setObjectName("actionOrder5")

        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder0)
        self.menuZoomOrder.addAction(a)
        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder1)
        self.menuZoomOrder.addAction(a)
        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder2)
        self.menuZoomOrder.addAction(a)
        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder3)
        self.menuZoomOrder.addAction(a)
        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder4)
        self.menuZoomOrder.addAction(a)
        a = self.actionGroup_ZoomOrder.addAction(self.actionOrder5)
        self.menuZoomOrder.addAction(a)
        
        self.actionGroup_Layout = QtWidgets.QActionGroup(self,exclusive=True)
        self.actionLayout_Normal = QtWidgets.QAction(self)
        self.actionLayout_Normal.setCheckable(True)
        self.actionLayout_Normal.setObjectName("actionLayout_Normal")
        self.actionLayout_Dark = QtWidgets.QAction(self)
        self.actionLayout_Dark.setCheckable(True)
        self.actionLayout_Dark.setObjectName("actionLayout_Dark")
        self.actionLayout_DarkOrange = QtWidgets.QAction(self)
        self.actionLayout_DarkOrange.setCheckable(True)
        self.actionLayout_DarkOrange.setObjectName("actionLayout_DarkOrange")

        if Default_dict["Layout"] == "Normal":
            self.actionLayout_Normal.setChecked(True)
        elif Default_dict["Layout"] == "Dark":
            self.actionLayout_Dark.setChecked(True)
        elif Default_dict["Layout"] == "DarkOrange":
            self.actionLayout_DarkOrange.setChecked(True)

        a = self.actionGroup_Layout.addAction(self.actionLayout_Normal)
        self.menuLayout.addAction(a)
        a = self.actionGroup_Layout.addAction(self.actionLayout_Dark)
        self.menuLayout.addAction(a)
        a = self.actionGroup_Layout.addAction(self.actionLayout_DarkOrange)
        self.menuLayout.addAction(a)
        self.menuLayout.addSeparator()
        self.actionTooltipOnOff = QtWidgets.QAction(self,checkable=True)
        self.actionTooltipOnOff.setChecked(True)
        self.actionTooltipOnOff.setObjectName(_fromUtf8("actionDataToRam"))
        self.menuLayout.addAction(self.actionTooltipOnOff)        

        self.menuLayout.addSeparator()

        self.actionGroup_IconTheme = QtWidgets.QActionGroup(self,exclusive=True)
        self.actionIconTheme_1 = QtWidgets.QAction(self)
        self.actionIconTheme_1.setCheckable(True)
        self.actionIconTheme_1.setObjectName("actionIconTheme_1")
        self.actionIconTheme_2 = QtWidgets.QAction(self)
        self.actionIconTheme_2.setCheckable(True)
        self.actionIconTheme_2.setObjectName("actionIconTheme_2")

        if Default_dict["Icon theme"] == "Icon theme 1":
            self.actionIconTheme_1.setChecked(True)
        elif Default_dict["Icon theme"] == "Icon theme 2":
            self.actionIconTheme_2.setChecked(True)

        a = self.actionGroup_IconTheme.addAction(self.actionIconTheme_1)
        self.menuLayout.addAction(a)
        a = self.actionGroup_IconTheme.addAction(self.actionIconTheme_2)
        self.menuLayout.addAction(a)

        self.menu_Options.addAction(self.menuLayout.menuAction())
        self.menu_Options.addAction(self.menuExport.menuAction())
        self.menu_Options.addAction(self.menuZoomOrder.menuAction())

        self.menuFile.addAction(self.actionLoadSession)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionSaveSession)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionQuit)

        self.menuEdit.addAction(self.actionDataToRamNow)        
        self.menuEdit.addAction(self.actionDataToRam)
        self.menuEdit.addSeparator()
        self.menuEdit.addAction(self.actionClearList)
        self.menuEdit.addAction(self.actionRemoveSelected)

        self.menuEdit.addSeparator()
        self.menuEdit.addAction(self.actionSaveToPng)

        self.menuEdit.addSeparator()
        self.menuEdit.addAction(self.actionVerbose)

        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuEdit.menuAction())
        self.menubar.addAction(self.menu_Options.menuAction())
        self.menubar.addAction(self.menu_Help.menuAction())

        #Add Default values:
        self.spinBox_imagecrop.setValue(Default_dict ["Input image size"])
        default_norm = Default_dict["Normalization"]
        index = self.comboBox_Normalization.findText(default_norm, QtCore.Qt.MatchFixedString)
        if index >= 0:
            self.comboBox_Normalization.setCurrentIndex(index)
        self.spinBox_NrEpochs.setValue(Default_dict ["Nr. epochs"])
        self.spinBox_RefreshAfterEpochs.setValue(Default_dict ["Keras refresh after nr. epochs"])
        self.checkBox_HorizFlip.setChecked(Default_dict ["Horz. flip"])
        self.checkBox_VertFlip.setChecked(Default_dict ["Vert. flip"])

        self.lineEdit_Rotation.setText(str(Default_dict ["rotation"]))
 
        self.lineEdit_Rotation.setText(str(Default_dict ["rotation"]))
        self.lineEdit_Rotation_2.setText(str(Default_dict ["width_shift"]))
        self.lineEdit_Rotation_3.setText(str(Default_dict ["height_shift"]))
        self.lineEdit_Rotation_4.setText(str(Default_dict ["zoom"]))
        self.lineEdit_Rotation_5.setText(str(Default_dict ["shear"]))
        self.spinBox_RefreshAfterNrEpochs.setValue(Default_dict ["Brightness refresh after nr. epochs"])
        self.spinBox_PlusLower.setValue(Default_dict ["Brightness add. lower"])
        self.spinBox_PlusUpper.setValue(Default_dict ["Brightness add. upper"])
        self.doubleSpinBox_MultLower.setValue(Default_dict ["Brightness mult. lower"])        
        self.doubleSpinBox_MultUpper.setValue(Default_dict ["Brightness mult. upper"])
        self.doubleSpinBox_GaussianNoiseMean.setValue(Default_dict ["Gaussnoise Mean"])
        self.doubleSpinBox_GaussianNoiseScale.setValue(Default_dict ["Gaussnoise Scale"])


        self.checkBox_contrast.setChecked(Default_dict["Contrast On"])
        self.doubleSpinBox_contrastLower.setEnabled(Default_dict["Contrast On"])
        self.doubleSpinBox_contrastHigher.setEnabled(Default_dict["Contrast On"])
        self.doubleSpinBox_contrastLower.setValue(Default_dict["Contrast min"])
        self.doubleSpinBox_contrastHigher.setValue(Default_dict["Contrast max"])
        
        self.checkBox_saturation.setChecked(Default_dict["Saturation On"])
        self.doubleSpinBox_saturationLower.setEnabled(Default_dict["Saturation On"])
        self.doubleSpinBox_saturationHigher.setEnabled(Default_dict["Saturation On"])
        self.doubleSpinBox_saturationLower.setValue(Default_dict["Saturation min"])
        self.doubleSpinBox_saturationHigher.setValue(Default_dict["Saturation max"])

        self.checkBox_hue.setChecked(Default_dict["Hue On"])
        self.doubleSpinBox_hueDelta.setEnabled(Default_dict["Hue On"])
        self.doubleSpinBox_hueDelta.setValue(Default_dict["Hue range"])

        self.checkBox_avgBlur.setChecked(Default_dict["AvgBlur On"])
        self.label_avgBlurMin.setEnabled(Default_dict["AvgBlur On"])
        self.spinBox_avgBlurMin.setEnabled(Default_dict["AvgBlur On"])
        self.label_avgBlurMax.setEnabled(Default_dict["AvgBlur On"])
        self.spinBox_avgBlurMax.setEnabled(Default_dict["AvgBlur On"])
        self.spinBox_avgBlurMin.setValue(Default_dict["AvgBlur min"])
        self.spinBox_avgBlurMax.setValue(Default_dict["AvgBlur max"])

        self.checkBox_gaussBlur.setChecked(Default_dict["GaussBlur On"])
        self.label_gaussBlurMin.setEnabled(Default_dict["GaussBlur On"])
        self.spinBox_gaussBlurMin.setEnabled(Default_dict["GaussBlur On"])
        self.label_gaussBlurMax.setEnabled(Default_dict["GaussBlur On"])
        self.spinBox_gaussBlurMax.setEnabled(Default_dict["GaussBlur On"])
        self.spinBox_gaussBlurMin.setValue(Default_dict["GaussBlur min"])
        self.spinBox_gaussBlurMax.setValue(Default_dict["GaussBlur max"])

        self.checkBox_motionBlur.setChecked(Default_dict["MotionBlur On"])
        self.label_motionBlurKernel.setEnabled(Default_dict["MotionBlur On"])
        self.lineEdit_motionBlurKernel.setEnabled(Default_dict["MotionBlur On"])
        self.label_motionBlurAngle.setEnabled(Default_dict["MotionBlur On"])
        self.lineEdit_motionBlurAngle.setEnabled(Default_dict["MotionBlur On"])
        self.lineEdit_motionBlurKernel.setText(str(Default_dict["MotionBlur Kernel"]))
        self.lineEdit_motionBlurAngle.setText(str(Default_dict["MotionBlur Angle"]))



        self.actionLayout_Normal.triggered.connect(self.onLayoutChange)
        self.actionLayout_Dark.triggered.connect(self.onLayoutChange)
        self.actionLayout_DarkOrange.triggered.connect(self.onLayoutChange)

        self.actionTooltipOnOff.triggered.connect(self.onTooltipOnOff)

        self.actionIconTheme_1.triggered.connect(self.onIconThemeChange)
        self.actionIconTheme_2.triggered.connect(self.onIconThemeChange)

        self.retranslateUi()
        
#        self.checkBox_contrast.clicked['bool'].connect(self.doubleSpinBox_contrastLower.setEnabled)
#        self.checkBox_contrast.clicked['bool'].connect(self.doubleSpinBox_contrastHigher.setEnabled)
        self.checkBox_avgBlur.clicked['bool'].connect(self.spinBox_avgBlurMin.setEnabled)
        self.checkBox_avgBlur.clicked['bool'].connect(self.spinBox_avgBlurMax.setEnabled)
        self.checkBox_gaussBlur.clicked['bool'].connect(self.spinBox_gaussBlurMin.setEnabled)
        self.checkBox_gaussBlur.clicked['bool'].connect(self.spinBox_gaussBlurMax.setEnabled)
        self.checkBox_motionBlur.clicked['bool'].connect(self.label_motionBlurKernel.setEnabled)
        self.checkBox_motionBlur.clicked['bool'].connect(self.lineEdit_motionBlurKernel.setEnabled)
        self.checkBox_motionBlur.clicked['bool'].connect(self.label_motionBlurAngle.setEnabled)
        self.checkBox_motionBlur.clicked['bool'].connect(self.lineEdit_motionBlurAngle.setEnabled)
        self.checkBox_gaussBlur.clicked['bool'].connect(self.label_gaussBlurMin.setEnabled)
        self.checkBox_gaussBlur.clicked['bool'].connect(self.label_gaussBlurMax.setEnabled)
        self.checkBox_avgBlur.clicked['bool'].connect(self.label_avgBlurMin.setEnabled)
        self.checkBox_avgBlur.clicked['bool'].connect(self.label_avgBlurMax.setEnabled)
        self.checkBox_optimizer.toggled['bool'].connect(self.comboBox_optimizer.setEnabled)
        self.checkBox_expt_loss.toggled['bool'].connect(self.comboBox_expt_loss.setEnabled)
        self.checkBox_expt_paddingMode.toggled['bool'].connect(self.comboBox_expt_paddingMode.setEnabled)

        #Start running show_cpu_ram function and run it all the time
        worker_cpu_ram = Worker(self.cpu_ram_worker)
        self.threadpool.start(worker_cpu_ram)

        self.tabWidget_Modelbuilder.setCurrentIndex(0)
        self.tabWidget_DefineModel.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(self)
    
    def retranslateUi(self):
        self.setWindowTitle(_translate("MainWindow", "AIDeveloper v."+VERSION, None))
        self.groupBox_dragdrop.setTitle(_translate("MainWindow", "Drag and drop data (.rtdc) here", None))
        self.groupBox_dragdrop.setToolTip(_translate("MainWindow", "<html><head/><body><p>Drag and drop files (.rtdc) or folders with images here. Valid .rtdc files have to contain at least 'images', 'pos_x' and 'pos_y'. If folders with images are dropped, the contents will be converted to a single .rtdc file (speeds up loading in the future).<br>After dropping data, you can specify the ‘Class’ of the images and if it should be used for training (T) or validation (V).<br>Double-click on the filename to show a random image of the data set. The original and cropped (using ‘Input image size’) image is shown<br>Click on the button 'Plot' to open a popup where you can get a histogram or scatterplot of enclosed data.<br>'Cells/Epoch' defines the nr. of images that are used in each epoch for training. Random images are drawn in each epoch during training. For the validation set, images are drawn once at the beginning and kept constant.<br>Deactivate ‘Shuffle’ if you don’t want to use random data. Then all images of this file are used.<br>Zoom allows you to increase or decrease resolution. Zoom=1 does nothing; Zoom=2 zooms in; Zoom=0.5 zooms out. This is useful if you have data acquired at 40x but you want to use it to train a model for 20x data. Use 'Options'->'Zooming order' to define the method used for zooming.<br>Hint for RT-DC users: Use ShapeOut to gate for particular subpopulations and save the filtered data as .rtdc. Make sure to export at least 'images', 'pos_x' and 'pos_y'.</p></body></html>", None))
        self.groupBox_DataOverview.setTitle(_translate("MainWindow", "Data Overview", None))
        self.groupBox_DataOverview.setToolTip(_translate("MainWindow", "<html><head/><body><p>The numbers of events of each class are added up. To do so the properties of each file are read. This happens each time you click into the table above. Unfortunately, reading is quite slow, so maybe disable this box, while you are assembling your data-set (especially when working with many large files). Use the column 'Name' to specify custom class-names, which help you later to remember, the meaning of each class. This table is saved to meta.xlsx when training is started.</p></body></html>", None))

        self.tab_ExampleImgs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Show random example images of the training data</p></body></html>", None))
        self.comboBox_ModelSelection.setToolTip(_translate("MainWindow", "<html><head/><body><p>Select the model architecture. MLP means Multilayer perceptron. Currently only MLPs are fast enough for AI based sorting. The definition of the architectures can be found in the model_zoo.py. If you want to implement custom Neural net architectures, you can edit model_zoo.py accordingly. Restart AIDeveloper in order to re-load model_zoo and import the new definitions </p></body></html>", None))
        self.radioButton_NewModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Select a model architecture in the dropdown menu</p></body></html>", None))
        self.radioButton_NewModel.setText(_translate("MainWindow", "New", None))
        self.radioButton_LoadRestartModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Load an existing architecture (from .model or .arch), and start training using random initial weights</p></body></html>", None))
        self.radioButton_LoadRestartModel.setText(_translate("MainWindow", "Load and restart", None))
        self.radioButton_LoadContinueModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Load an existing model (.model) and continue fitting. This option could be used to load a trained model to optimize it (transfer learning)</p></body></html>", None))
        self.radioButton_LoadContinueModel.setText(_translate("MainWindow", "Load and continue", None))
        self.lineEdit_LoadModelPath.setToolTip(_translate("MainWindow", "<html><head/><body><p>When you use 'Load and restart' or 'Load and continue', the filename of the chosen model will apprear here</p></body></html>", None))
        self.label_Crop.setToolTip(_translate("MainWindow", "<html><head/><body><p>Models need a defined input size image. Choose wisely since cutting off from large objects might result in information loss.</p></body></html>", None))
        self.label_Crop.setText(_translate("MainWindow", "<html><head/><body><p>Input image size</p></body></html>", None))
        self.label_Normalization.setToolTip(_translate("Form", "<html><head/><body><p>Image normalization method.<br><br>None: No normalization is applied.<br><br>'Div. by 255': Each input image is divided by 255 (useful since pixelvalues go from 0 to 255, so the result will be in range 0-1).<br><br>StdScaling using mean and std of each image individually: The mean and standard deviation of each input image itself is used to scale it by first subtracting the mean and then dividing by the standard deviation.<br><br>StdScaling using mean and std of all training data: First, all pixels of the entire training dataset are used to calc. a mean and a std. deviation. This mean and standard deviation is used to scale images during training by first subtracting the mean and then dividing by the standard deviation.<br><br>Only 'None' and 'Div. by 255' are currently supported in the Sorting-Software</p></body></html>", None))
        self.label_Normalization.setText(_translate("MainWindow", "Normalization", None))
        self.comboBox_Normalization.setToolTip(_translate("Form", "<html><head/><body><p>Image normalization method.<br><br>None: No normalization is applied.<br><br>'Div. by 255': Each input image is divided by 255 (useful since pixelvalues go from 0 to 255, so the result will be in range 0-1).<br><br>StdScaling using mean and std of each image individually: The mean and standard deviation of each input image itself is used to scale it by first subtracting the mean and then dividing by the standard deviation.<br><br>StdScaling using mean and std of all training data: First, all pixels of the entire training dataset are used to calc. a mean and a std. deviation. This mean and standard deviation is used to scale images during training by first subtracting the mean and then dividing by the standard deviation.<br><br>Only 'None' and 'Div. by 255' are currently supported in the Sorting-Software</p></body></html>", None))
        self.label_colorMode.setText(_translate("MainWindow", "Color Mode", None))
        self.label_colorMode.setToolTip(_translate("MainWindow", "<html><head/><body><p>The Color Mode defines the input image depth. Color images have three channels -RGB. Grayscale images only have a single channel. Models are automatically built accordingly. Models trained for Grayscale cannot be applied to RGB images (unless images are converted). Same is true the other way around.</p></body></html>", None))
    
        self.comboBox_GrayOrRGB.setToolTip(_translate("MainWindow", "<html><head/><body><p>The Color Mode defines the input image depth. Color images have three channels -RGB. Grayscale images only have a single channel. Models are automatically built accordingly. Models trained for Grayscale cannot be applied to RGB images (unless images are converted). Same is true the other way around.</p></body></html>", None))
        self.label_nrEpochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Number of epochs (iterations over the training data). In each epoch, a subset of the training data is used to update the weights if 'Shuffle' is on.</p></body></html>", None))
        self.label_nrEpochs.setText(_translate("MainWindow", "Nr. epochs", None))
        self.pushButton_modelname.setToolTip(_translate("MainWindow", "Define path and filename for the model you want to fit", None))
        self.pushButton_modelname.setText(_translate("MainWindow", "Model path:", None))
        self.tabWidget_DefineModel.setTabText(self.tabWidget_DefineModel.indexOf(self.tab_DefineModel), _translate("MainWindow", "Define Model", None))
        self.label_RefreshAfterEpochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Affine image augmentation takes quite long; so maybe use the same images for this nr. of epochs</p></body></html>", None))
        self.label_RefreshAfterEpochs.setText(_translate("MainWindow", "Refresh after nr. epochs:", None))
        self.tab_kerasAug.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define settings for  affine image augmentations</p></body></html>", None))
        self.checkBox_HorizFlip.setToolTip(_translate("MainWindow", "<html><head/><body><p>Flip some training images randomly along horiz. axis (left becomes right; right becomes left)</p></body></html>", None))
        self.checkBox_VertFlip.setToolTip(_translate("MainWindow", "<html><head/><body><p>Flip some training images randomly along vert. axis (bottom up; top down)</p></body></html>", None))
        self.label_Rotation.setToolTip(_translate("MainWindow", "<html><head/><body><p>Degree range for random rotations</p></body></html>", None))
        self.label_width_shift.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define random shift of width<br>Fraction of total width, if &lt; 1. Otherwise pixels if>=1.<br>Value defines an interval (-width_shift_range, +width_shift_range) from which random numbers are created</p></body></html>", None))
        self.label_height_shift.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define random shift of height<br>Fraction of total height if &lt; 1. Otherwise pixels if>=1.<br>Value defines an interval (-height_shift_range, +height_shift_range) from which random numbers are created   </p></body></html>", None))
        self.label_zoom.setToolTip(_translate("MainWindow", "<html><head/><body><p>Range for random zoom</p></body></html>", None))
        self.label_shear.setToolTip(_translate("MainWindow", "<html><head/><body><p>Shear Intensity (Shear angle in counter-clockwise direction in degrees)</p></body></html>", None))        
        self.lineEdit_Rotation.setToolTip(_translate("MainWindow", "<html><head/><body><p>Degree range for random rotations</p></body></html>", None))
        self.lineEdit_Rotation_2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define random shift of width<br>Fraction of total width, if &lt; 1. Otherwise pixels if&gt;=1.<br>Value defines an interval (-width_shift_range, +width_shift_range) from which random numbers are created</p></body></html>", None))
        self.lineEdit_Rotation_3.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define random shift of height<br>Fraction of total height if &lt; 1. Otherwise pixels if&gt;=1.<br>Value defines an interval (-height_shift_range, +height_shift_range) from which random numbers are created   </p></body></html>", None))
        self.lineEdit_Rotation_4.setToolTip(_translate("MainWindow", "<html><head/><body><p>Range for random zoom</p></body></html>", None))
        self.lineEdit_Rotation_5.setToolTip(_translate("MainWindow", "<html><head/><body><p>Shear Intensity (Shear angle in counter-clockwise direction in degrees)</p></body></html>", None))
        self.spinBox_RefreshAfterEpochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Affine image augmentation takes quite long; so maybe use the same images for this nr. of epochs</p></body></html>", None))
        self.checkBox_HorizFlip.setText(_translate("MainWindow", "Horiz. flip", None))
        self.checkBox_VertFlip.setText(_translate("MainWindow", "Vert. flip", None))
        self.label_Rotation.setText(_translate("MainWindow", "Rotation", None))
        self.label_width_shift.setText(_translate("MainWindow", "Width shift", None))
        self.label_height_shift.setText(_translate("MainWindow", "Height shift", None))
        self.label_zoom.setText(_translate("MainWindow", "Zoom", None))
        self.label_shear.setText(_translate("MainWindow", "Shear", None))
        self.tabWidget_DefineModel.setTabText(self.tabWidget_DefineModel.indexOf(self.tab_kerasAug), _translate("MainWindow", "Affine img. augm.", None))
        self.label_RefreshAfterNrEpochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Brightness augmentation is really fast, so best you refresh images for each epoch (set to 1)</p></body></html>", None))
        self.label_RefreshAfterNrEpochs.setText(_translate("MainWindow", "Refresh after nr. epochs:", None))
        self.spinBox_RefreshAfterNrEpochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Brightness augmentation is really fast, so best you refresh images for each epoch (set to 1)</p></body></html>", None))
        self.groupBox_BrightnessAugmentation.setTitle(_translate("MainWindow", "Brightness augmentation", None))

        self.groupBox_BrightnessAugmentation.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define add/multiply offset to make image randomly slightly brighter or darker. Additive offset (A) is one number that is added to all pixels values; Multipl. offset (M) is a value to multiply each pixel value with: NewImage = A + M*Image</p></body></html>", None))

        self.label_Plus.setText(_translate("MainWindow", "Add.", None))
        self.label_Plus.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define lower threshold for additive offset</p></body></html>", None))
        self.spinBox_PlusLower.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define lower threshold for additive offset</p></body></html>", None))
        #self.label_PlusTo.setText(_translate("MainWindow", "...", None))
        #self.label_Plus.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define upper threshold for additive offset</p></body></html>", None))
        self.spinBox_PlusUpper.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define upper threshold for additive offset</p></body></html>", None))

        self.label_Mult.setText(_translate("MainWindow", "Mult.", None))
        self.doubleSpinBox_MultLower.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define lower threshold for multiplicative offset</p></body></html>", None))
        self.doubleSpinBox_MultUpper.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define upper threshold for multiplicative offset</p></body></html>", None))

        #self.label_Rotation_MultTo.setText(_translate("MainWindow", "...", None))
        self.groupBox_GaussianNoise.setTitle(_translate("MainWindow", "Gaussian noise", None))
        self.groupBox_GaussianNoise.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define Gaussian Noise, which is added to the image</p></body></html>", None))
        self.label_GaussianNoiseMean.setText(_translate("MainWindow", "Mean", None))
        self.label_GaussianNoiseMean.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the mean of the Gaussian noise. Typically this should be zero. If you use a positive number it would mean that your noise tends to be positive, i.e. bright.</p></body></html>", None))

        self.label_GaussianNoiseScale.setText(_translate("MainWindow", "Scale", None))
        self.label_GaussianNoiseScale.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the standard deviation of the Gaussian noise. A larger number means a wider distribution of the noise, which results in an image that looks more noisy. Prefer to change this parameter over chainging the mean.</p></body></html>", None))

        self.groupBox_colorAugmentation.setTitle(_translate("Form", "Color augm.", None))
        tooltip_colorAugmentation = "Define methods to randomly alter the contrast (applicable for grayscale and RGB), saturation (RGB only) or hue (RGB only) of your images"
        self.groupBox_colorAugmentation.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+tooltip_colorAugmentation+"</p></body></html>", None))
        self.checkBox_contrast.setText(_translate("Form", "Contrast", None))
        self.checkBox_contrast.setToolTip(_translate("Form", "<html><head/><body><p>Augment contrast using a random factor. Applicable for Grayscale and RGB. Left spinbox (lower factor) has to be >0. '0.70' to '1.3' means plus/minus 30% contrast (at random)</p></body></html>", None))
        self.checkBox_saturation.setText(_translate("Form", "Saturation", None))
        self.checkBox_saturation.setToolTip(_translate("Form", "<html><head/><body><p>Augment saturation using a random factor. Applicable for RGB only. Left spinbox (lower factor) has to be >0. '0.70' to '1.3' means plus/minus 30% saturation (at random)</p></body></html>", None))
        self.checkBox_hue.setText(_translate("Form", "Hue", None))
        self.checkBox_hue.setToolTip(_translate("Form", "<html><head/><body><p>Augment hue using a random factor. Applicable for RGB only. Left spinbox (lower factor) has to be >0. '0.70' to '1.3' means plus/minus 30% hue (at random)</p></body></html>", None))

        self.groupBox_blurringAug.setTitle(_translate("MainWindow", "Blurring", None))
        tooltip_blurringAug = "Define methods to randomly blur images."
        self.groupBox_blurringAug.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+tooltip_blurringAug+"</p></body></html>", None))
        self.label_motionBlurKernel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define kernels by giving a range [min,max]. Values in this range are then randomly picked for each image</p></body></html>", None))
        self.label_motionBlurKernel.setText(_translate("MainWindow", "Kernel", None))
        self.lineEdit_motionBlurAngle.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define angle for the motion blur by defining a range \"min degree,max degree\". Values in this range are then randomly picked for each image</p></body></html>", None))
        #self.lineEdit_motionBlurAngle.setText(_translate("MainWindow", "-10,10", None))
        self.label_avgBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for average blur</p></body></html>", None))
        self.label_avgBlurMin.setText(_translate("MainWindow", "Min", None))
        self.spinBox_gaussBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for gaussian blur</p></body></html>", None))
        tooltip_motionBlur = "Apply random motion blurring. Motion blur is defined by an intensity ('kernel') and a direction ('angle'). Please define a range for 'kernel' and 'angle'. AID will pick a random value (within each range) for each image."
        self.checkBox_motionBlur.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+tooltip_motionBlur+"</p></body></html>", None))
        self.checkBox_motionBlur.setText(_translate("MainWindow", "Motion", None))
        self.spinBox_avgBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for average blur</p></body></html>", None))
        self.spinBox_gaussBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for gaussian blur</p></body></html>", None))
        self.label_motionBlurAngle.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define a range of angles for the motion blur: 'min degree,max degree'. Values from this range are then randomly picked for each image</p></body></html>", None))
        self.label_motionBlurAngle.setText(_translate("MainWindow", "Angle", None))
        self.label_gaussBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for gaussian blur</p></body></html>", None))
        self.label_gaussBlurMin.setText(_translate("MainWindow", "Min", None))
        self.label_gaussBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for gaussian blur</p></body></html>", None))
        tooltip_gaussianBlur = "Apply random gaussian blurring. For gaussian blurring, a gaussian kernel of defined size is used. Please define a min. and max. kernel size. For each image a random value is picked from this range to generate a gaussian kernel."
        self.checkBox_gaussBlur.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+tooltip_gaussianBlur+"</p></body></html>", None))
        self.checkBox_gaussBlur.setText(_translate("MainWindow", "Gauss", None))
        self.spinBox_avgBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for average blur</p></body></html>", None))
        self.label_gaussBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for gaussian blur</p></body></html>", None))
        self.label_gaussBlurMax.setText(_translate("MainWindow", "Max", None))
        self.label_gaussBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for gaussian blur</p></body></html>", None))
        tooltip_avgBlur = "Apply random average blurring. Define a range of kernel sizes for the average blur (min. and max. kernel size). Values from this range are then randomly picked for each image. To blur an image, all pixels within the kernel area used to compute an average pixel value. The central element of the kernel area in the image is then set to this value. This operation is carried out over the entire image"
        self.checkBox_avgBlur.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+tooltip_avgBlur+"</p></body></html>", None))        
        self.checkBox_avgBlur.setText(_translate("MainWindow", "Average", None))
        
        self.label_avgBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for average blur</p></body></html>", None))
        self.label_avgBlurMax.setText(_translate("MainWindow", "Max", None))
        self.spinBox_avgBlurMin.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the minimum kernel size for average blur</p></body></html>", None))
        self.spinBox_avgBlurMax.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define the maximum kernel size for average blur</p></body></html>", None))
        self.lineEdit_motionBlurKernel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Define kernels by giving a range \"min,max\". Values from this range are then randomly picked for each image</p></body></html>", None))
        #self.lineEdit_motionBlurKernel.setText(_translate("MainWindow", "0,5", None))


        self.tabWidget_DefineModel.setTabText(self.tabWidget_DefineModel.indexOf(self.tab_BrightnessAug), _translate("MainWindow", "Brightn/Color augm.", None))
        self.label_ShowIndex.setText(_translate("MainWindow", "Class", None))
        self.pushButton_ShowExamleImgs.setText(_translate("MainWindow", "Show", None))
        self.tabWidget_DefineModel.setTabText(self.tabWidget_DefineModel.indexOf(self.tab_ExampleImgs), _translate("MainWindow", "Example imgs.", None))
        
        self.groupBox_expertMode.setTitle(_translate("MainWindow", "Expert Mode", None))
        self.groupBox_modelKerasFit.setTitle(_translate("MainWindow", "In model_keras.fit()", None))
        self.groupBox_regularization.setTitle(_translate("MainWindow", "Loss / Regularization", None))

        self.label_batchSize.setText(_translate("MainWindow", "Batch size", None))
        self.label_epochs.setText(_translate("MainWindow", "Epochs", None))
        #self.label_others.setText(_translate("MainWindow", "Others", None))
        self.checkBox_learningRate.setText(_translate("MainWindow", "Learning Rate", None))
        self.checkBox_trainLastNOnly.setText(_translate("MainWindow", "Train only last N layers", None))
        self.checkBox_trainDenseOnly.setText(_translate("MainWindow", "Train only Dense layers", None))
        self.checkBox_dropout.setText(_translate("MainWindow", "Dropout", None))
        self.tabWidget_DefineModel.setTabText(self.tabWidget_DefineModel.indexOf(self.tab_expert), _translate("MainWindow", "Expert", None))


        self.groupBox_expertMode.setToolTip(_translate("MainWindow", "<html><head/><body><p>Expert mode allows changing the learning rate and you can even set parts of the neural net to \'not trainable\' in order to perform transfer learning and fine tune models. Also dropout rates can be adjusted. When expert mode is turned off again, the initial values are applied again.</p></body></html>", None))
        self.checkBox_learningRate.setToolTip(_translate("MainWindow", "<html><head/><body><p>Change the learning rate. The default optimizer is \'adam\' with a learning rate of 0.001</p></body></html>", None))
        self.doubleSpinBox_learningRate.setToolTip(_translate("MainWindow", "<html><head/><body><p>Change the learning rate. Typically, the optimizer \'adam\' is used and the default value is 0.001</p></body></html>", None))
        self.checkBox_optimizer.setText(_translate("MainWindow", "Optimizer", None))
        self.comboBox_optimizer.setItemText(0, _translate("MainWindow", "Adam", None))
        self.comboBox_optimizer.setItemText(1, _translate("MainWindow", "SGD", None))
        self.comboBox_optimizer.setItemText(2, _translate("MainWindow", "RMSprop", None))
        self.comboBox_optimizer.setItemText(3, _translate("MainWindow", "Adagrad", None))
        self.comboBox_optimizer.setItemText(4, _translate("MainWindow", "Adadelta", None))
        self.comboBox_optimizer.setItemText(5, _translate("MainWindow", "Adamax", None))
        self.comboBox_optimizer.setItemText(6, _translate("MainWindow", "Nadam", None))
        self.checkBox_trainLastNOnly.setToolTip(_translate("MainWindow", "<html><head/><body><p>When checked, only the last n layers of the model, which have >0 parameters will stay trainable. Layers before are set to trainable=False. Please specify n using the spinbox. After this change, the model has to be recompiled, which means the optimizer values are deleted.</p></body></html>", None))
        self.spinBox_trainLastNOnly.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify the number of last layer in your model that should be kept trainable. Only layers with >0 parameters are counted. Use the checkbox to apply this option. After this change, the model has to be recompiled, which means the optimizer values are deleted. </p></body></html>", None))
        self.checkBox_trainDenseOnly.setToolTip(_translate("MainWindow", "<html><head/><body><p>When checked, only the dense layers are kept trainable (if they have >0 parameters). Other layers are set to trainable=False. After this change, the model has to be recompiled, which means the optimizer values are deleted.</p></body></html>", None))
        self.label_batchSize.setToolTip(_translate("MainWindow", "<html><head/><body><p>Number of samples per gradient update. If unspecified, batch_size will default to 128. (Source: Keras documentation)</p></body></html>", None))
        self.spinBox_batchSize.setToolTip(_translate("MainWindow", "<html><head/><body><p>Number of samples per gradient update. If unspecified, batch_size will default to 128. (Source: Keras documentation)</p></body></html>", None))
        self.label_epochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Number of epochs to train the model on an identical batch.</p></body></html>", None))
        self.spinBox_epochs.setToolTip(_translate("MainWindow", "<html><head/><body><p>Number of epochs to train the model on identical batch.</p></body></html>", None))
        
        self.checkBox_expt_loss.setText(_translate("MainWindow", "Loss", None))
        self.comboBox_expt_loss.setItemText(0, _translate("MainWindow", "categorical_crossentropy", None))
        #self.comboBox_expt_loss.setItemText(1, _translate("MainWindow", "sparse_categorical_crossentropy", None))
        self.comboBox_expt_loss.setItemText(1, _translate("MainWindow", "mean_squared_error", None))
        self.comboBox_expt_loss.setItemText(2, _translate("MainWindow", "mean_absolute_error", None))
        self.comboBox_expt_loss.setItemText(3, _translate("MainWindow", "mean_absolute_percentage_error", None))
        self.comboBox_expt_loss.setItemText(4, _translate("MainWindow", "mean_squared_logarithmic_error", None))
        self.comboBox_expt_loss.setItemText(5, _translate("MainWindow", "squared_hinge", None))
        self.comboBox_expt_loss.setItemText(6, _translate("MainWindow", "hinge", None))
        self.comboBox_expt_loss.setItemText(7, _translate("MainWindow", "categorical_hinge", None))
        self.comboBox_expt_loss.setItemText(8, _translate("MainWindow", "logcosh", None))
        #self.comboBox_expt_loss.setItemText(9, _translate("MainWindow", "huber_loss", None))
        self.comboBox_expt_loss.setItemText(9, _translate("MainWindow", "binary_crossentropy", None))
        self.comboBox_expt_loss.setItemText(10, _translate("MainWindow", "kullback_leibler_divergence", None))
        self.comboBox_expt_loss.setItemText(11, _translate("MainWindow", "poisson", None))
        self.comboBox_expt_loss.setItemText(12, _translate("MainWindow", "cosine_proximity", None))
        #self.comboBox_expt_loss.setItemText(13, _translate("MainWindow", "is_categorical_crossentropy", None))
        
        self.checkBox_dropout.setToolTip(_translate("MainWindow", "<html><head/><body><p>If your model has one or more dropout layers, you can change the dropout rates here. Insert into the lineEdit one value (e.g. 0.5) to apply this one value to all dropout layers, or insert a list of values to specify the dropout rates for each dropout layer individually (e.g. for three dropout layers: [ 0.2 , 0.5 , 0.25 ]. The model will be recompiled, but the optimizer weights are not deleted.</p></body></html>", None))
        self.lineEdit_dropout.setToolTip(_translate("MainWindow", "<html><head/><body><p>If your model has one or more dropout layers, you can change the dropout rates here. Insert into the lineEdit one value (e.g. 0.5) to apply this one value to all dropout layers, or insert a list of values to specify the dropout rates for each dropout layer individually (e.g. for three dropout layers: [ 0.2 , 0.5 , 0.25 ]. The model will be recompiled, but the optimizer weights are not deleted.</p></body></html>", None))
        self.checkBox_partialTrainability.setText(_translate("MainWindow", "Partial trainablility", None))
        self.checkBox_partialTrainability.setToolTip(_translate("MainWindow", "<html><head/><body><p>Partial trainability allows you to make parts of a layer non-trainable. Hence, this option makes most sense in combination with 'Load and continue' training a model. After checking this box, the model you chose on 'Define model'-tab is initialized. The line on the right shows the trainability of each layer in the model. Use the button on the right ('...') to open a popup menu, which allows you to specify individual trainabilities for each layer.</p></body></html>", None))
        self.lineEdit_partialTrainability.setToolTip(_translate("MainWindow", "<html><head/><body><p>Partial trainability allows you to make parts of a layer non-trainable. Hence, this option makes most sense in combination with 'Load and continue' training a model. After checking this box, the model you chose on 'Define model'-tab is initialized. The line on the right shows the trainability of each layer in the model. Use the button on the right ('...') to open a popup menu, which allows you to specify individual trainabilities for each layer.</p></body></html>", None))
        self.pushButton_partialTrainability.setText(_translate("MainWindow", "...", None))
        self.pushButton_partialTrainability.setToolTip(_translate("MainWindow", "<html><head/><body><p>Partial trainability allows you to make parts of a layer non-trainable. Hence, this option makes most sense in combination with 'Load and continue' training a model. After checking this box, the model you chose on 'Define model'-tab is initialized. The line on the right shows the trainability of each layer in the model. Use the button on the right ('...') to open a popup menu, which allows you to specify individual trainabilities for each layer.</p></body></html>", None))

        self.checkBox_lossW.setText(_translate("MainWindow", "Loss weights", None))
        self.checkBox_lossW.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify scalar coefficients to weight the loss contributions of different classes.</p></body></html>", None))
        self.lineEdit_lossW.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify scalar coefficients to weight the loss contributions of different classes.</p></body></html>", None))
        self.pushButton_lossW.setText(_translate("MainWindow", "...", None))

        self.groupBox_expt_imgProc.setTitle(_translate("MainWindow", "Image processing", None))
        self.checkBox_expt_paddingMode.setText(_translate("MainWindow", "Padding mode", None))
        self.comboBox_expt_paddingMode.setToolTip(_translate("MainWindow", "<html><head/><body><p>By default, the padding mode is \"constant\", which means that zeros are padded.\n"
"\"edge\": Pads with the edge values of array.\n"
"\"linear_ramp\": Pads with the linear ramp between end_value and the array edge value.\n"
"\"maximum\": Pads with the maximum value of all or part of the vector along each axis.\n"
"\"mean\": Pads with the mean value of all or part of the vector along each axis.\n"
"\"median\": Pads with the median value of all or part of the vector along each axis.\n"
"\"minimum\": Pads with the minimum value of all or part of the vector along each axis.\n"
"\"reflect\": Pads with the reflection of the vector mirrored on the first and last values of the vector along each axis.\n"
"\"symmetric\": Pads with the reflection of the vector mirrored along the edge of the array.\n"
"\"wrap\": Pads with the wrap of the vector along the axis. The first values are used to pad the end and the end values are used to pad the beginning.\n"
"Text copied from https://docs.scipy.org/doc/numpy/reference/generated/numpy.pad.html</p></body></html>", None))
        self.comboBox_expt_paddingMode.setItemText(0, _translate("MainWindow", "constant", None))
        self.comboBox_expt_paddingMode.setItemText(1, _translate("MainWindow", "edge", None))
        self.comboBox_expt_paddingMode.setItemText(2, _translate("MainWindow", "linear_ramp", None))
        self.comboBox_expt_paddingMode.setItemText(3, _translate("MainWindow", "maximum", None))
        self.comboBox_expt_paddingMode.setItemText(4, _translate("MainWindow", "mean", None))
        self.comboBox_expt_paddingMode.setItemText(5, _translate("MainWindow", "median", None))
        self.comboBox_expt_paddingMode.setItemText(6, _translate("MainWindow", "minimum", None))
        self.comboBox_expt_paddingMode.setItemText(7, _translate("MainWindow", "reflect", None))
        self.comboBox_expt_paddingMode.setItemText(8, _translate("MainWindow", "symmetric", None))
        self.comboBox_expt_paddingMode.setItemText(9, _translate("MainWindow", "wrap", None))

        self.groupBox_expertMetrics.setTitle(_translate("MainWindow", "Metrics", None))
        text_metrics_tt = "Define metrics, that are computed after each training iteration ('epoch'). Those metrics are can then also be displayed in real-time during training and are tracked/saved in the meta.xlsx file. Each model where any metric for the validation set breaks a new record is saved (minimum val. loss achived -> model is saved. maximum val. accuracy achieved-> model is saved)."
        self.groupBox_expertMetrics.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+text_metrics_tt+"</p></body></html>", None))

        self.checkBox_expertAccuracy.setText(_translate("MainWindow", "Accuracy", None))
        text_acc_tt = "Compute accuracy and validation accuracy after each epoch. Each model, where the corresponding metric for the validatio-set achieves a new record will be saved."
        self.checkBox_expertAccuracy.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+text_acc_tt+"</p></body></html>", None))
        
        self.checkBox_expertF1.setText(_translate("MainWindow", "F1 score", None))
        text_f1_tt = "Compute F1 and validation F1 score after each epoch. Each model, where the corresponding metric for the validatio-set achieves a new record will be saved."
        self.checkBox_expertF1.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+text_f1_tt+"</p></body></html>", None))
        
        self.checkBox_expertPrecision.setText(_translate("MainWindow", "Precision", None))
        text_precision_tt = "Compute precision and validation precision after each epoch for each class. Each model, where the corresponding metric for the validatio-set achieves a new record will be saved."
        self.checkBox_expertPrecision.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+text_precision_tt+"</p></body></html>", None))
        
        self.checkBox_expertRecall.setText(_translate("MainWindow", "Recall", None))
        text_recall_tt = "Compute recall and validation recall after each epoch for each class. Each model, where the corresponding metric for the validatio-set achieves a new record will be saved."
        self.checkBox_expertRecall.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+text_recall_tt+"</p></body></html>", None))


        self.groupBox_Finalize.setTitle(_translate("MainWindow", "Model summary and Fit", None))
        self.pushButton_FitModel.setText(_translate("MainWindow", "Initialize/Fit\nModel", None))
        self.pushButton_FitModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Afer defining all model parameters, hit this button to build/initialize the model, load the data to RAM (if 'Load data to RAM' is chosen in 'Edit') and start the fitting process. You can also do only the initialization to check the model summary (appears in textbox on the left).</p></body></html>", None))

        self.tabWidget_Modelbuilder.setTabText(self.tabWidget_Modelbuilder.indexOf(self.tab_Build), _translate("MainWindow", "Build", None))
        self.pushButton_Live.setToolTip(_translate("MainWindow", "<html><head/><body><p>Load and display the history of the model which is currently fitted</p></body></html>", None))
        self.pushButton_Live.setText(_translate("MainWindow", "Live!", None))
        self.pushButton_LoadHistory.setToolTip(_translate("MainWindow", "<html><head/><body><p>Select a history file to be plotted</p></body></html>", None))
        self.pushButton_LoadHistory.setText(_translate("MainWindow", "Load History", None))
        self.lineEdit_LoadHistory.setToolTip(_translate("MainWindow", "Enter path/filename of a meta-file (meta.xlsx). The history is contained in this file.", None))
        self.tableWidget_HistoryItems.setToolTip(_translate("MainWindow", "<html><head/><body><p>Information of the history file is shown here\nDouble-click to enter menu for changing color</p></body></html>", None))
        self.pushButton_UpdateHistoryPlot.setText(_translate("MainWindow", "Update plot", None))
        self.checkBox_rollingMedian.setText(_translate("MainWindow", "Rolling Median", None))
        self.checkBox_rollingMedian.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check to add a rolling median. Use the slider to change the window size for which the median is computed.</p></body></html>", None))
        self.horizontalSlider_rollmedi.setToolTip(_translate("MainWindow", "<html><head/><body><p>Use this slider to change the window size for the rolling median between 1 and 50.</p></body></html>", None))
        
        self.checkBox_linearFit.setText(_translate("MainWindow", "Linear Fit", None))
        self.checkBox_linearFit.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check if you want to add a liner fit. A movable region will appear. Only data inside this region will be used for the fit.</p></body></html>", None))

        #self.label_ModelIndex.setText(_translate("MainWindow", "Load a model and convert", None))
        self.pushButton_LoadModel.setText(_translate("MainWindow", "Load model", None))
        
        
        self.pushButton_LoadModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Load a model from disk for conversion to other formats. Please specify the format of the model using the dropbox above. Next, define the target format using the dropdown menu on the right-> and finally, hit 'Convert'.</p></body></html>", None))

        self.pushButton_convertModel.setText(_translate("MainWindow", "Convert", None))
        self.pushButton_convertModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Convert chosen model to the format indicated by the dropbox above. This might be useful to deply models to other platforms. AIDeveloper is only compatible with Keras TensorFlow models to perform inference. For usage of the model with soRT-DC, please convert it to .nnet (currently only MLPs are supported by soRT-DC software!).</p></body></html>", None))
        self.tabWidget_Modelbuilder.setTabText(self.tabWidget_Modelbuilder.indexOf(self.tab_History), _translate("MainWindow", "History", None))

        self.groupBox_loadModel.setTitle(_translate("MainWindow", "Load Model", None))
        self.label_ModelIndex_2.setText(_translate("MainWindow", "Model index", None))
        self.lineEdit_ModelSelection_2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Model architecture name, read from meta.xlsx ('Chosen Model') is displayed here</p></body></html>", None))
        self.label_Normalization_2.setText(_translate("MainWindow", "Normalization", None))
        self.label_Crop_2.setText(_translate("MainWindow", "Input size", None))
        self.label_OutClasses_2.setText(_translate("MainWindow", "Output Nr. of classes", None))
        self.pushButton_LoadModel_2.setText(_translate("MainWindow", "Load model", None))
        self.lineEdit_LoadModel_2.setToolTip(_translate("MainWindow", "Enter path/filename of a model (.model)", None))
        self.tableWidget_Info_2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify validation data via 'Build'-tab or load .rtdc file (->From .rtdc).<br>Use column 'Name' to specify proper cell names (for presentation purposes).<br>Use column 'clr' to specify plotting color of that cell-type</p></body></html>", None))

        self.pushButton_ExportValidToNpy.setToolTip(_translate("MainWindow", "<html><head/><body><p>Export the validation data (images and labels). Optionally the cropped images can exported->use 'Options'->'Export' to change. Normalization method of the chosen model is not yet applied. Please use the \'Build\' tab to define data</p></body></html>", None))
        self.groupBox_validData.setTitle(_translate("MainWindow", "Data", None))
        self.pushButton_ExportValidToNpy.setText(_translate("MainWindow", "To .rtdc", None))
        self.pushButton_ImportValidFromNpy.setText(_translate("MainWindow", "From .rtdc",None))
        self.pushButton_ImportValidFromNpy.setToolTip(_translate("MainWindow", "<html><head/><body><p>Import validation data (images from .rtdc and labels from .txt) file. Cropped and non-cropped images can be imported. If necessary they will be cropped to the correct format. If the loaded images are smaller than the required size, there will be zero-padding.</p></body></html>", None))

        self.groupBox_settings.setTitle(_translate("MainWindow", "Settings", None))
        self.groupBox_InferenceTime.setTitle(_translate("MainWindow", "Inference time", None))
        self.groupBox_InferenceTime.setToolTip(_translate("MainWindow", "<html><head/><body><p>Inference time is the time required to predict a single image. To get a meaningful value, several (define how many using spinbox->) images are predicted one by one. The given Nr. (spinbox->) is divided by 10. The resulting nr. of images is predicted one by one and an average computing time is obtained. This process is repqeated 10 times</p></body></html>", None))
  
        self.pushButton_CompInfTime.setText(_translate("MainWindow", "Compute for N imgs", None))
        self.groupBox_classify.setTitle(_translate("MainWindow", "Classify unlabeled data", None))
        self.pushButton_classify.setText(_translate("MainWindow", "Classify", None))
        self.radioButton_selectAll.setText(_translate("MainWindow", "all", None))
        self.groupBox_settings.setTitle(_translate("MainWindow", "Settings", None))
        
        
        self.groupBox_settings.setTitle(_translate("MainWindow", "Settings", None))
        self.label_SortingIndex.setText(_translate("MainWindow", "Sorting class", None))
        self.label_SortingIndex.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify the class you intend to sort for ('Target cell type'). This is important if you want to know the final concentration in the target region</p></body></html>", None))
        self.checkBox_SortingThresh.setText(_translate("MainWindow", "Sorting threshold", None))
        self.checkBox_SortingThresh.setToolTip(_translate("MainWindow", "<html><head/><body><p>Specify a probability threshold above which a cell is classified as a target cell (specified using 'Sorting class'). Typically cells with a probability above 0.5 are sorted, but you could also increase the threshold in order to only have cells in the target region that are more confidently classified</p></body></html>", None))

        self.pushButton_AssessModel.setText(_translate("MainWindow", "Update Plots", None))
        
        self.comboBox_probability_histogram.setToolTip(_translate("MainWindow", "<html><head/><body><p>Select a plotting style for the probability plot<br>Style1: Only outline, width 5<br>Style2: Only outline, width 10<br>Style4: Filled hist, alpha 0.6<br>Style3: Filled hist, alpha 0.7<br>Style5: Filled hist, alpha 0.8</p></body></html>",None))
        
        #self.comboBox_probability_histogram.setText(_translate("MainWindow", "Prob. hist.",None))

        self.groupBox_3rdPlotSettings.setTitle(_translate("MainWindow", "3rd plot settings", None))
        
        thirdrdplot_text = "ROC (Receiver Operating Characteristic) curves summarize the trade-off between the true positive rate and false positive rate for a predictive model using different probability thresholds.<br>Precision-Recall curves summarize the trade-off between the true positive rate and the positive predictive value for a predictive model using different probability thresholds.<br>ROC curves are appropriate when the observations are balanced between each class, whereas precision-recall curves are appropriate for imbalanced datasets."
        self.groupBox_3rdPlot.setToolTip(_translate("MainWindow", "<html><head/><body><p>"+thirdrdplot_text+"</p></body></html>", None))

        self.label_3rdPlot.setText(_translate("MainWindow", "What to plot", None))
        self.comboBox_3rdPlot.setToolTip(_translate("MainWindow", "<html><head/><body><p>Use the combobox to define what is shown in the third plot. Some options might only be valid for binary problems. For such cases please use the spinboxes Indx1 and Indx2 to define two cell types</p></body></html>", None))
        self.label_Indx1.setToolTip(_translate("MainWindow", "<html><head/><body><p>Some options of the combobox (\'Third plot\') are only valid for binary problems. For such cases please use the spinboxes Indx1 and Indx2 to define two cell types</p></body></html>", None))
        self.label_Indx1.setText(_translate("MainWindow", "Indx1", None))
        self.spinBox_Indx1.setToolTip(_translate("MainWindow", "<html><head/><body><p>Some options of the combobox (\'Third plot\') are only valid for binary problems. For such cases please use the spinboxes Indx1 and Indx2 to define two cell types</p></body></html>", None))
        self.label_Indx2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Some options of the combobox (\'Third plot\') are only valid for binary problems. For such cases please use the spinboxes Indx1 and Indx2 to define two cell types</p></body></html>", None))
        self.label_Indx2.setText(_translate("MainWindow", "Indx2", None))
        self.spinBox_Indx2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Some options of the combobox (\'Third plot\') are only valid for binary problems. For such cases please use the spinboxes Indx1 and Indx2 to define two cell types</p></body></html>", None))
        self.groupBox_confusionMatrixPlot.setTitle(_translate("MainWindow", "Classification Metrics", None))
        self.tableWidget_CM1.setToolTip(_translate("MainWindow", "<html><head/><body><p>Confusion matrix shows the total Nrs. of cells. Doubleclick a field in the matrix to save the corresponding images to .rtdc</p></body></html>", None))
        self.label_True_CM1.setText(_translate("MainWindow", "T\n"
"R\n"
"U\n"
"E",None))
        self.pushButton_CM1_to_Clipboard.setText(_translate("MainWindow", "To Clipboard",None))
        self.label_Pred_CM1.setText(_translate("MainWindow", "PREDICTED",None))
        self.tableWidget_CM2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Normalized Confusion matrix shows the relative amount of cells. Doubleclick a field in the matrix to save the corresponding images to .rtdc </p></body></html>", None))

        self.label_True_CM2.setText(_translate("MainWindow", "T\n"
"R\n"
"U\n"
"E",None))
        self.pushButton_CM2_to_Clipboard.setText(_translate("MainWindow", "To Clipboard",None))
        self.label_Pred_CM2.setText(_translate("MainWindow", "PREDICTED",None))
        text_scikit_learn = "Compute precision, recall, F-measure and support for each class<br>\
        The precision is the ratio tp / (tp + fp) where tp is the number of \
        true positives and fp the number of false positives. The precision is \
        intuitively the ability of the classifier not to label as positive a \
        sample that is negative.<br>The recall is the ratio tp / (tp + fn) where \
        tp is the number of true positives and fn the number of false negatives. \
        The recall is intuitively the ability of the classifier to find all the \
        positive samples.<br>The F-beta score can be interpreted as a weighted \
        harmonic mean of the precision and recall, where an F-beta score reaches \
        its best value at 1 and worst score at 0. The F-beta score weights recall \
        more than precision by a factor of beta. beta == 1.0 means recall and \
        precision are equally important.<br>The support is the number of occurrences \
        of each class in y_true.<br>micro average - averaging the total true \
        positives false negatives and false positives<br>macro average - averaging \
        the unweighted mean per label<br>weighted average - averaging the \
        support-weighted mean per label<br>(Source:scikit-learn.org)"
        
        self.tableWidget_AccPrecSpec.setToolTip(_translate("MainWindow", "<html><head/><body><p>Classification Metrics appear after hitting 'Update Plots'<br>Copy to clipboad by clicking somewhere on table<br><br>"+text_scikit_learn+"</p></body></html>", None))
        self.groupBox_probHistPlot.setTitle(_translate("MainWindow", "Probability histogram", None))
        self.groupBox_probHistPlot.setToolTip(_translate("MainWindow", "<html><head/><body><p>Probability histogram appears after hitting 'Prob. hist'. It shows the predicted probability of each cell to remain to class 'Sorting clas'. Colors indicate the true label. A good model returns high probabilities for cells of 'Sorting class', while cells of other indices (different cell types) have very low probabilities. This plot also allows to guess a reasonable 'Sorting threshold'</p></body></html>", None))

        self.groupBox_3rdPlot.setTitle(_translate("MainWindow", "3rd plot", None))
        self.tabWidget_Modelbuilder.setTabText(self.tabWidget_Modelbuilder.indexOf(self.tab_AssessModel), _translate("MainWindow", "Assess Model", None))


        #Plotting Peakdet-Tab
        self.groupBox_plottingregion.setTitle(_translate("MainWindow", "Plotting region", None))
        self.groupBox_plottingOptions.setTitle(_translate("MainWindow", "Plotting options", None))
        self.checkBox_fl1.setText(_translate("MainWindow", "FL1", None))
        self.checkBox_fl2.setText(_translate("MainWindow", "FL2", None))
        self.checkBox_fl3.setText(_translate("MainWindow", "FL3", None))
        self.checkBox_centroid.setText(_translate("MainWindow", "Centroid", None))
        self.label_coloring.setText(_translate("MainWindow", "Coloring", None))
        self.checkBox_colorLog.setText(_translate("MainWindow", "Logscaled", None))
        self.pushButton_updateScatterPlot.setText(_translate("MainWindow", "Update", None))
        self.groupBox.setTitle(_translate("MainWindow", "Info", None))
        self.tabWidget_filter_peakdet.setTabText(self.tabWidget_filter_peakdet.indexOf(self.tab_filter), _translate("MainWindow", "Placeholder", None))
        self.groupBox_showCell.setTitle(_translate("MainWindow", "Show cell", None))
        self.groupBox_showSelectedPeaks.setTitle(_translate("MainWindow", "Select peaks manually", None))
        self.label_automatic.setText(_translate("MainWindow", "Automatic", None))
        self.pushButton_highestXPercent.setText(_translate("MainWindow", "Highest x %", None))
        self.label_remove.setText(_translate("MainWindow", "Remove", None))
        self.pushButton_selectPeakPos.setText(_translate("MainWindow", "Peak", None))
        self.pushButton_selectPeakRange.setText(_translate("MainWindow", "Range", None))
        self.pushButton_removeSelectedPeaks.setText(_translate("MainWindow", "Selected", None))
        self.pushButton_removeAllPeaks.setText(_translate("MainWindow", "All", None))
        self.groupBox_peakDetModel.setTitle(_translate("MainWindow", "Peak detection Model", None))
        self.pushButton_fitPeakDetModel.setText(_translate("MainWindow", "Fit model to peaks", None))
        self.pushButton_SavePeakDetModel.setText(_translate("MainWindow", "Save model", None))
        self.pushButton_loadPeakDetModel.setText(_translate("MainWindow", "Load model", None))
        self.radioButton_exportSelected.setText(_translate("MainWindow", "only selected", None))
        self.radioButton_exportAll.setText(_translate("MainWindow", "all", None))
        self.pushButton_export.setText(_translate("MainWindow", "Export to...", None))
        self.tabWidget_filter_peakdet.setTabText(self.tabWidget_filter_peakdet.indexOf(self.tab_peakdet), _translate("MainWindow", "Peakdetection", None))
        self.tabWidget_filter_peakdet.setTabText(self.tabWidget_filter_peakdet.indexOf(self.tab_defineModel), _translate("MainWindow", "Placeholder", None))
        self.tabWidget_Modelbuilder.setTabText(self.tabWidget_Modelbuilder.indexOf(self.tab_Plotting), _translate("MainWindow", "Plot/Peak", None))

        ##############################Python Tab###############################
        self.groupBox_pythonMenu.setTitle(_translate("MainWindow", "File", None))
        self.label_pythonCurrentFile.setText(_translate("MainWindow", "Current file:", None))
        self.groupBox_pythonEditor.setTitle(_translate("MainWindow", "Editor", None))
        self.pushButton_pythonInOpen.setText(_translate("MainWindow", "Open file..", None))
        self.pushButton_pythonSaveAs.setText(_translate("MainWindow", "Save as...", None))
        self.pushButton_pythonInClear.setText(_translate("MainWindow", "Clear", None))
        self.pushButton_pythonInRun.setText(_translate("MainWindow", "Run", None))
        self.groupBox_pythonConsole.setTitle(_translate("MainWindow", "Console", None))
        self.pushButton_pythonOutClear.setText(_translate("MainWindow", "Clear", None))
        self.pushButton_pythonOutRun.setText(_translate("MainWindow", "Run", None))
        self.tabWidget_Modelbuilder.setTabText(self.tabWidget_Modelbuilder.indexOf(self.tab_python), _translate("MainWindow", "Python", None))

        #Tooltips setToolTip(_translate("MainWindow", "<html><head/><body><p> </p></body></html>", None))
        self.comboBox_chooseRtdcFile.setToolTip(_translate("MainWindow", "<html><head/><body><p>Choose a file</p></body></html>", None))
        self.comboBox_featurey.setToolTip(_translate("MainWindow", "<html><head/><body><p>Dropdown menu shows all availble features of the chosen .rtdc-file. The chosen feature will be plotted on the y-axis</p></body></html>", None))
        self.comboBox_featurex.setToolTip(_translate("MainWindow", "<html><head/><body><p>Dropdown menu shows all availble features of the chosen .rtdc-file. The chosen feature will be plotted on the x-axis</p></body></html>", None))
        self.widget_histx.setToolTip(_translate("MainWindow", "<html><head/><body><p>Histogram projection of the x-dimension</p></body></html>", None))
        self.widget_histy.setToolTip(_translate("MainWindow", "<html><head/><body><p>Histogram projection of the y-dimension</p></body></html>", None))
        self.horizontalSlider_cellInd.setToolTip(_translate("MainWindow", "<html><head/><body><p>Use this slider to choose a cell in the data-set. The respective cell and trace will be shown in tab 'Peakdetection'</p></body></html>", None))
        self.spinBox_cellInd.setToolTip(_translate("MainWindow", "<html><head/><body><p>Use this to index and choose an event in the data-set. The respective cell and trace will be shown in tab 'Peakdetection'</p></body></html>", None))
        self.widget_scatter.setToolTip(_translate("MainWindow", "<html><head/><body><p>Click into this scatterplot to choose an event. The respective cell and trace will be shown in tab 'Peakdetection'</p></body></html>", None))
        self.checkBox_fl1.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check this if you want to plot the trace for Fl1. Automatic peak finding (Highest x% in 'Peakdetection'-tab) will then also search in those traces.</p></body></html>", None))
        self.checkBox_fl2.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check this if you want to plot the trace for Fl2. Automatic peak finding (Highest x% in 'Peakdetection'-tab) will then also search in those traces.</p></body></html>", None))
        self.checkBox_fl3.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check this if you want to plot the trace for Fl3. Automatic peak finding (Highest x% in 'Peakdetection'-tab) will then also search in those traces.</p></body></html>", None))
        self.checkBox_centroid.setToolTip(_translate("MainWindow", "<html><head/><body><p>Check this if you want to plot the centroid of a chosen event in 'Show cell' on 'Peakdetection'-tab).</p></body></html>", None))
        self.pushButton_selectPeakPos.setToolTip(_translate("MainWindow", "<html><head/><body><p>After you moved the range on the trace-plot you can select a peak using this button. This data of the peak will be shown in the table in the right and will be used to fit the peak-detection model.</p></body></html>", None))
        self.pushButton_selectPeakRange.setToolTip(_translate("MainWindow", "<html><head/><body><p>After you changed the range on the trace-plot you can select the range-width using this button. This range will be shown in the table below and will be used in the peak-detection model.</p></body></html>", None))
        self.pushButton_highestXPercent.setToolTip(_translate("MainWindow", "<html><head/><body><p>The highest x% of FLx_max peaks are looked up for each x (=1,2,3 if Checkboxes are activated) and inserted into the table on the right.</p></body></html>", None))
        self.doubleSpinBox_highestXPercent.setToolTip(_translate("MainWindow", "<html><head/><body><p>The highest x% of FLx_max peaks are looked up for each x (=1,2,3 if Checkboxes are activated) and inserted into the table on the right.</p></body></html>", None))
        self.pushButton_removeSelectedPeaks.setToolTip(_translate("MainWindow", "<html><head/><body><p>Select a peak in the table or in the plot on the right and remove it using this button.</p></body></html>", None))
        self.pushButton_removeAllPeaks.setToolTip(_translate("MainWindow", "<html><head/><body><p>Remove all peaks from the table the right.</p></body></html>", None))
        self.widget_showSelectedPeaks.setToolTip(_translate("MainWindow", "<html><head/><body><p>Scatterplot shows all selcted peaks. Click on a peak to highlight the respective position in the table. After that you can also use the button 'Selected' to remove this point</p></body></html>", None))
        self.tableWidget_showSelectedPeaks.setToolTip(_translate("MainWindow", "<html><head/><body><p>Table shows all seelcted peaks. After clicking on a row you can use the button 'Selected' to remove this point</p></body></html>", None))
        self.groupBox_showCell.setToolTip(_translate("MainWindow", "<html><head/><body><p>Plot shows image of the recorded cell and the respective fluorescence traces (optional- use checkboxes to show FL1/2/3). Optionally the centroid is shown.</p></body></html>", None))
        self.pushButton_updateScatterPlot.setToolTip(_translate("MainWindow", "<html><head/><body><p>Hit this button to read the chosen features and plot the scatterplot above.</p></body></html>", None))
        self.tableWidget_peakModelParameters.setToolTip(_translate("MainWindow", "<html><head/><body><p>After fitting a peak-detection model, the model parameters are shown here. Each parameter can also be manipulated right here.</p></body></html>", None))

        self.comboBox_peakDetModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Choose a peak detection model.</p></body></html>", None))
        self.pushButton_fitPeakDetModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Start fitting a model usig the selected peaks shown above</p></body></html>", None))
        self.pushButton_SavePeakDetModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Save model to an excel file</p></body></html>", None))
        self.pushButton_loadPeakDetModel.setToolTip(_translate("MainWindow", "<html><head/><body><p>Load peak detection model from an excel file</p></body></html>", None))
        self.radioButton_exportSelected.setToolTip(_translate("MainWindow", "<html><head/><body><p>Apply the peak detection model only on the single chosen file</p></body></html>", None))
        self.radioButton_exportAll.setToolTip(_translate("MainWindow", "<html><head/><body><p>Apply the peak detection model on all files on the 'Build'-Tab</p></body></html>", None))


        self.menuFile.setTitle(_translate("MainWindow", "File", None))
        self.menuEdit.setTitle(_translate("MainWindow", "Edit", None))
        self.menu_Options.setTitle(_translate("MainWindow", "Options",None))
        self.menuLayout.setTitle(_translate("MainWindow", "Layout",None))
        self.menuExport.setTitle(_translate("MainWindow", "Export",None))
        self.menuZoomOrder.setTitle(_translate("MainWindow", "Zoom Order",None))
        self.menu_Options.setToolTip(_translate("MainWindow", "<html><head/><body><p>menu_Options tooltip</p></body></html>", None))


        self.menu_Help.setTitle(_translate("MainWindow", "Help",None))
        self.actionDocumentation.setText(_translate("MainWindow", "Documentation",None))
        self.actionSoftware.setText(_translate("MainWindow", "Software",None))
        self.actionAbout.setText(_translate("MainWindow", "About",None))
        self.actionUpdate.setText(_translate("MainWindow", "Check for updates...",None))


        self.actionLoadSession.setText(_translate("MainWindow", "Load Session", None))
        self.actionSaveSession.setText(_translate("MainWindow", "Save Session", None))
        self.actionQuit.setText(_translate("MainWindow", "Quit", None))
        self.actionDataToRam.setText(_translate("MainWindow", "Data to RAM upon Initialization of Model", None))
        self.actionVerbose.setText(_translate("MainWindow", "Verbose", None))
#        self.actionShowDataOverview.setText(_translate("MainWindow", "Show Data Overview", None))
        self.actionClearList.setText(_translate("MainWindow", "Clear List", None))
        self.actionDataToRamNow.setText(_translate("MainWindow", "Data to RAM now", None))
        self.actionRemoveSelected.setText(_translate("MainWindow", "Remove selected", None))
        self.actionSaveToPng.setText(_translate("MainWindow", "Export selected to .png/.jpg", None))
        self.actionExport_Off.setText(_translate("MainWindow", "No exporting",None))
        self.actionExport_Original.setText(_translate("MainWindow", "Export Original Images",None))
        self.actionExport_Cropped.setText(_translate("MainWindow", "Export Cropped Images",None))
        self.actionLayout_Normal.setText(_translate("MainWindow", "Normal layout",None))
        self.actionLayout_Dark.setText(_translate("MainWindow", "Dark layout",None))
        self.actionLayout_DarkOrange.setText(_translate("MainWindow", "DarkOrange layout",None))
        self.actionIconTheme_1.setText(_translate("MainWindow", "Icon theme 1",None))
        self.actionIconTheme_2.setText(_translate("MainWindow", "Icon theme 2",None))
        self.actionOrder0.setText(_translate("MainWindow", "0 (nearest neighbor)",None))
        self.actionOrder1.setText(_translate("MainWindow", "1 (linear interp.)",None))
        self.actionOrder2.setText(_translate("MainWindow", "2 (quadratic interp.)",None))
        self.actionOrder3.setText(_translate("MainWindow", "3 (cubic interp.)",None))
        self.actionOrder4.setText(_translate("MainWindow", "4",None))
        self.actionOrder5.setText(_translate("MainWindow", "5",None))

        self.actionTooltipOnOff.setText(_translate("MainWindow", "Show tooltips",None))

    def dataDropped(self, l):
        #If there is data stored on ram tell user that RAM needs to be refreshed!
        if len(self.ram)>0:
            self.statusbar.showMessage("Newly added data is not yet in RAM. Only RAM data will be used. Use ->Edit->Data to RAM now to update RAM",5000)
        #l is a list of some filenames (.rtdc) or folders (containing .jpg, jpeg, .png)
        
        #Iterate over l and check if it is a folder or a file (directory)
        isfile = [os.path.isfile(url) for url in l]
        isfolder = [os.path.isdir(url) for url in l]


        #####################For folders with images:##########################            
        #where are folders?
        ind_true = np.where(np.array(isfolder)==True)[0]
        foldernames = list(np.array(l)[ind_true]) #select the indices that are valid
        #On mac, there is a trailing / in case of folders; remove them
        foldernames = [os.path.normpath(url) for url in foldernames]

        basename = [os.path.basename(f) for f in foldernames]
        #Look quickly inside the folders and ask the user if he wants to convert
        #to .rtdc (might take a while!)
        if len(foldernames)>0: #User dropped (also) folders (which may contain images)
#            filecounts = []
#            for i in range(len(foldernames)):
#                url = foldernames[i]
#                files = os.listdir(url)
#                files_full = [os.path.join(url,files[i]) for i in range(len(files))]
#                filecounts.append(len([f for f in files_full if os.path.isfile(f)]))
#            Text = []
#            for b,n in zip(basename,filecounts):
#                Text.append(b+": "+str(n)+" images")
#            Text = "\n".join(Text)

            Text = "Images from single folders are read and saved to individual .rtdc files with the same name like the corresponding folder.<b>If you have RGB images you can either save the full RGB information, or do a conversion to Grayscale (saves some diskspace but information about color is lost). RGB is recommended since AID will automatically do the conversion to grayscale later if required.<b>If you have Grayscale images, a conversion to RGB will just copy the info to all channels, which allows you to use RGB-mode and Grayscale-mode lateron. Conversion to Grayscale is will "

            Text = Text+"\nImages from following folders will be converted:\n"+"\n".join(basename)
            #Show the user a summary with all the found folders and how many files are
            #contained. Ask if he want to convert
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Question)
            text = "<html><head/><body><p>Should the images of the chosen folder(s) be converted to .rtdc using <b>RGB</b> or <b>Grayscale</b> format? <b>(RGB is recommended!)</b>  Either option might take some time. You can reuse the .rtdc file next time.</p></body></html>"
            msg.setText(text)
            msg.setDetailedText(Text)
            msg.setWindowTitle("Format for conversion to .rtdc (RGB/Grayscale)")
            msg.addButton(QtGui.QPushButton('Convert to Grayscale'), QtGui.QMessageBox.YesRole)
            msg.addButton(QtGui.QPushButton('Convert to RGB'), QtGui.QMessageBox.NoRole)
            msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
            retval = msg.exec_()

            #Conversion of images in folders is (almost) independent from what 
            #is going to be fitted (So I leave the option menu still!)
            #In options: Color Mode one can still use RGB mode and export here as
            #Grayscale (but this would actually not work since RGB information is lost).
            #The other way around works. Therefore it is recommended to export RGB!
            if retval==0: 
                color_mode = "Grayscale"
                channels = 1
            elif retval==1:
                color_mode = "RGB"
                channels = 3
            else:
                return
            self.statusbar.showMessage("Color mode' "+color_mode+"' is used",5000)
            url_converted = []
            for i in range(len(foldernames)):
                url = foldernames[i]
                print("Start converting images in\n"+url)
                #try:
                #get a list of files inside this directory:
                images,pos_x,pos_y = [],[],[]
                for root, dirs, files in os.walk(url):
                    for file in files:
                        try:
                            path = os.path.join(root, file)
                            img = load_img(path,color_mode=color_mode.lower()) #This uses PIL and supports many many formats!
                            images.append(np.array(img)) #append nice numpy array to list
                            #create pos_x and pos_y
                            pos_x.append( int(np.round(img.width/2.0,0)) ) 
                            pos_y.append( int(np.round(img.height/2.0,0)) )  
                        except:
                            pass
                
                #Thanks to andko76 for pointing that unequal image sizes cause an error:
                #https://github.com/maikherbig/AIDeveloper/issues/1
                #Check that all images have the same size
#                img_shape_errors = 0
#                text_error = "Images have unequal dimensions:"
#                img_h = [a.shape[0] for a in images]
#                img_h_uni = len(np.unique(img_h))
#                if img_h_uni!=1:
#                    text_error += "\n- found unequal heights"
#                    img_shape_errors=1
#                img_w = [a.shape[1] for a in images]
#                img_w_uni = len(np.unique(img_w))
#                if img_w_uni!=1:
#                    text_error += "\n- found unequal widths"
#                    img_shape_errors=1
#                img_c = [len(a.shape) for a in images]
#                img_c_uni = len(np.unique(img_c))
#                if img_c_uni!=1:
#                    text_error += "\n- found unequal numbers of channels"
#                    img_shape_errors=1
#                #If there were issues detected, show error message
#                if img_shape_errors==1:
#                    msg = QtWidgets.QMessageBox()
#                    msg.setIcon(QtWidgets.QMessageBox.Warning)       
#                    msg.setText(str(text_error))
#                    msg.setWindowTitle("Error: Unequal image shapes")
#                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                    msg.exec_()
#                    return

                #Get a list of occuring image dimensions (width and height)
                img_shape = [a.shape[0] for a in images] + [a.shape[1] for a in images]
                dims = np.unique(img_shape)
                #Get a list of occurences of image shapes
                img_shape = [str(a.shape[0])+" x "+str(a.shape[1]) for a in images]
                occurences = np.unique(img_shape,return_counts=True)
                #inform user if there is more than one img shape 
                if len(occurences[0])>1 or len(dims)>1:
                    text_detail = "Path: "+url
                    text_detail += "\nFollowing image shapes are present"
                    for i in range(len(occurences[0])):
                        text_detail+="\n- "+str(occurences[1][i])+" times: "+str(occurences[0][i])
                    
                    self.popup_imgRes = QtGui.QDialog()
                    self.popup_imgRes_ui = aid_frontend.popup_imageLoadResize()
                    self.popup_imgRes_ui.setupUi(self.popup_imgRes) #open a popup to show options for image resizing (make image equally sized)
                    #self.popup_imgRes.setWindowModality(QtCore.Qt.WindowModal)
                    self.popup_imgRes.setWindowModality(QtCore.Qt.ApplicationModal)
                    #Insert information into textBrowser
                    self.popup_imgRes_ui.textBrowser_imgResize_occurences.setText(text_detail)
                    Image_import_dimension = Default_dict["Image_import_dimension"]
                    self.popup_imgRes_ui.spinBox_ingResize_h_1.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_h_2.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_w_1.setValue(Image_import_dimension)
                    self.popup_imgRes_ui.spinBox_ingResize_w_2.setValue(Image_import_dimension)
                    Image_import_interpol_method = Default_dict["Image_import_interpol_method"]
                    index = self.popup_imgRes_ui.comboBox_resizeMethod.findText(Image_import_interpol_method, QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.popup_imgRes_ui.comboBox_resizeMethod.setCurrentIndex(index)
                    #Define function for the OK button:
                    def popup_imgRes_ok(images,channels,pos_x,pos_y):
                        print("Start resizing operation")
                        #Get info from GUI
                        final_h = int(self.popup_imgRes_ui.spinBox_ingResize_h_1.value())
                        print("Height:"+str(final_h))
                        final_w = int(self.popup_imgRes_ui.spinBox_ingResize_w_1.value())
                        print("Width:"+str(final_w))
                        Default_dict["Image_import_dimension"] = final_h

                        if self.popup_imgRes_ui.radioButton_imgResize_cropPad.isChecked():#cropping and padding method
                            images = aid_img.image_resize_crop_pad(images,pos_x,pos_y,final_h,final_w,channels,verbose=False,padding_mode='constant')
                        elif self.popup_imgRes_ui.radioButton_imgResize_interpolate.isChecked():
                            interpolation_method = str(self.popup_imgRes_ui.comboBox_resizeMethod.currentText())
                            Default_dict["Image_import_interpol_method"] = interpolation_method
                            images = aid_img.image_resize_scale(images,pos_x,pos_y,final_h,final_w,channels,interpolation_method,verbose=False)
                        else:
                            print("Invalid image resize method!")
                        #Save the Default_dict
                        aid_bin.save_aid_settings(Default_dict) 
                        self.popup_imgRes.accept()
                        return images
                    
                    #Define function for the Cancel button:                    
                    def popup_imgRes_cancel():
                        self.popup_imgRes.close()
                        return

                    self.popup_imgRes_ui.pushButton_imgResize_ok.clicked.connect(lambda: popup_imgRes_ok(images,channels,pos_x,pos_y))
                    self.popup_imgRes_ui.pushButton_imgResize_cancel.clicked.connect(popup_imgRes_cancel)
                    
                    retval = self.popup_imgRes.exec_()
                    #retval is 0 if the user clicked cancel or just closed the window; in this case just exist the function
                    if retval==0:
                        return

                #get new pos_x, pos_y (after cropping, the pixel value for the middle of the image is different!)
                pos_x = [int(np.round(img.shape[1]/2.0,0)) for img in images]
                pos_y = [int(np.round(img.shape[0]/2.0,0)) for img in images]
                
                #Now, all images are of identical shape and can be converted to a numpy array
                images = np.array((images), dtype="uint8")
                pos_x = np.array((pos_x), dtype="uint8")
                pos_y = np.array((pos_y), dtype="uint8")
                
                #Save as foldername.rtdc
                fname = url+".rtdc"
                if os.path.isfile(fname):
                    #ask user if file can be overwritten
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Question)
                    text = "<html><head/><body><p>File:"+fname+" already exists. Should it be overwritten?</p></body></html>"
                    msg.setText(text)
                    msg.setWindowTitle("Overwrite file?")
                    msg.addButton(QtGui.QPushButton('Yes'), QtGui.QMessageBox.YesRole)
                    msg.addButton(QtGui.QPushButton('No'), QtGui.QMessageBox.NoRole)
                    msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
                    retval = msg.exec_()
        
                    if retval==0:
                        try:
                            os.remove(fname)
                            aid_img.imgs_2_rtdc(fname,images,pos_x,pos_y)
                            url_converted.append(fname)
                        except Exception as e:
                            msg = QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Information)
                            msg.setText(str(e))
                            msg.setWindowTitle("Error")
                            retval = msg.exec_()
                    elif retval==1:
                        pass
                    else:
                        pass
                else:#file does not yet exist. Create it
                    aid_img.imgs_2_rtdc(fname,images,pos_x,pos_y)
                    url_converted.append(fname)

            print("Finished converting! Final dimension of image tensor is:"+str(images.shape))
            #Now load the created files directly to drag/drop-region!
            self.dataDropped(url_converted)

        #####################For .rtdc files:##################################            
        #where are files?
        ind_true = np.where(np.array(isfile)==True)[0]
        filenames = list(np.array(l)[ind_true]) #select the indices that are valid
        #check if the file can be opened and get some information
        fileinfo = []
        for i in range(len(filenames)):
            rtdc_path = filenames[i]
            try:
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                
                features = rtdc_ds.features
                #Make sure that there is "images", "pos_x" and "pos_y" available
                if "image" in features and "pos_x" in features and "pos_y" in features:
                    nr_images = rtdc_ds["image"].len()
                    pix = rtdc_ds.config["imaging"]["pixel size"]
                    fileinfo.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"pix":pix})
                else:
                    missing = []
                    for feat in ["image","pos_x","pos_y"]:
                        if feat not in features:
                            missing.append(feat)    
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Essential feature(s) are missing in data-set")
                    msg.setDetailedText("Data-set: "+rtdc_path+"\nis missing "+str(missing))
                    msg.setWindowTitle("Missing essential features")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()                      
            except:
                pass
        
        #Add the stuff to the combobox on Plot/Peak Tab
        url_list = [fileinfo[iterator]["rtdc_path"] for iterator in range(len(fileinfo))]
        self.comboBox_chooseRtdcFile.addItems(url_list)
        self.comboBox_selectData.addItems(url_list)
        if len(url_list)==0: #This fixes the issue that the prog. crashes if accidentially a tableitem is dragged and "dropped" on the table
            return
        width=self.comboBox_selectData.fontMetrics().boundingRect(max(url_list, key=len)).width()
        self.comboBox_selectData.view().setFixedWidth(width+10)             
        
        for rowNumber in range(len(fileinfo)):#for url in l:
            url = fileinfo[rowNumber]["rtdc_path"]
            #add to table
            rowPosition = self.table_dragdrop.rowCount()
            self.table_dragdrop.insertRow(rowPosition)
            
            columnPosition = 0
            line = QtWidgets.QLabel(self.table_dragdrop)
            line.setText(url)
            line.setDisabled(True)
            line.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, line)            
            
#            item = QtWidgets.QTableWidgetItem(url) 
#            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
#            print(item.textAlignment())
#            item.setTextAlignment(QtCore.Qt.AlignRight) # change the alignment
#            #item.setTextAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AnchorRight) # change the alignment
#            self.table_dragdrop.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 1
            spinb = QtWidgets.QSpinBox(self.table_dragdrop)
            spinb.valueChanged.connect(self.dataOverviewOn)
            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, spinb)            

            for columnPosition in range(2,4):
                #for each item, also create 2 checkboxes (train/valid)
                item = QtWidgets.QTableWidgetItem()#("item {0} {1}".format(rowNumber, columnNumber))
                item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
                item.setCheckState(QtCore.Qt.Unchecked)
                self.table_dragdrop.setItem(rowPosition, columnPosition, item)
            
            columnPosition = 4
            #Place a button which allows to show a plot (scatter, histo...lets see)
            btn = QtWidgets.QPushButton(self.table_dragdrop)
            btn.setMinimumSize(QtCore.QSize(50, 30))
            btn.setMaximumSize(QtCore.QSize(50, 30))
            btn.clicked.connect(self.button_hist)
            btn.setText('Plot')
            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, btn)            
            self.table_dragdrop.resizeRowsToContents()

#            columnPosition = 5
#            #Place a combobox with the available features
#            cb = QtWidgets.QComboBox(self.table_dragdrop)
#            cb.addItems(fileinfo[rowNumber]["features"])
#            cb.setMinimumSize(QtCore.QSize(70, 30))
#            cb.setMaximumSize(QtCore.QSize(70, 30))            
#            width=cb.fontMetrics().boundingRect(max(fileinfo[rowNumber]["features"], key=len)).width()
#            cb.view().setFixedWidth(width+30)             
#            self.table_dragdrop.setCellWidget(rowPosition, columnPosition, cb)            
          

            columnPosition = 5
            #Place a combobox with the available features
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, fileinfo[rowNumber]["nr_images"])
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 6
            #Field to user-define nr. of cells/epoch
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole,100)
            #item.cellChanged.connect(self.dataOverviewOn)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 7
            #Pixel size
            item = QtWidgets.QTableWidgetItem()
            pix = fileinfo[rowNumber]["pix"]
            item.setData(QtCore.Qt.EditRole,pix)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 8           
            #Should data be shuffled (random?)
            item = QtWidgets.QTableWidgetItem()#("item {0} {1}".format(rowNumber, columnNumber))
            item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
            item.setCheckState(QtCore.Qt.Checked)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)

            columnPosition = 9
            #Zooming factor
            item = QtWidgets.QTableWidgetItem()
            zoom = 1.0
            item.setData(QtCore.Qt.EditRole,zoom)
            self.table_dragdrop.setItem(rowPosition, columnPosition, item)



    #Functions for Keras augmentation checkboxes
    def keras_changed_rotation(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation.setText(str(0))
            self.lineEdit_Rotation.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation.setText(str(Default_dict ["rotation"]))
            self.lineEdit_Rotation.setEnabled(True)
        else:
            return
    def keras_changed_width_shift(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation_2.setText(str(0))
            self.lineEdit_Rotation_2.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation_2.setText(str(Default_dict ["width_shift"]))
            self.lineEdit_Rotation_2.setEnabled(True)
        else:
            return
    def keras_changed_height_shift(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation_3.setText(str(0))
            self.lineEdit_Rotation_3.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation_3.setText(str(Default_dict ["height_shift"]))
            self.lineEdit_Rotation_3.setEnabled(True)
        else:
            return
    def keras_changed_zoom(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation_4.setText(str(0))
            self.lineEdit_Rotation_4.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation_4.setText(str(Default_dict ["zoom"]))
            self.lineEdit_Rotation_4.setEnabled(True)
        else:
            return
    def keras_changed_shear(self,on_or_off):
        if on_or_off==0:
            self.lineEdit_Rotation_5.setText(str(0))
            self.lineEdit_Rotation_5.setEnabled(False)
        elif on_or_off==2:
            self.lineEdit_Rotation_5.setText(str(Default_dict ["shear"]))
            self.lineEdit_Rotation_5.setEnabled(True)
        else:
            return
    def keras_changed_brightplus(self,on_or_off):
        if on_or_off==0:
            self.spinBox_PlusLower.setValue(0)
            self.spinBox_PlusLower.setEnabled(False)
            self.spinBox_PlusUpper.setValue(0)
            self.spinBox_PlusUpper.setEnabled(False)
        elif on_or_off==2:
            self.spinBox_PlusLower.setValue(Default_dict ["Brightness add. lower"])
            self.spinBox_PlusLower.setEnabled(True)
            self.spinBox_PlusUpper.setValue(Default_dict ["Brightness add. upper"])
            self.spinBox_PlusUpper.setEnabled(True)
        else:
            return
    def keras_changed_brightmult(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_MultLower.setValue(1.0)
            self.doubleSpinBox_MultLower.setEnabled(False)
            self.doubleSpinBox_MultUpper.setValue(1.0)
            self.doubleSpinBox_MultUpper.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_MultLower.setValue(Default_dict ["Brightness mult. lower"])
            self.doubleSpinBox_MultLower.setEnabled(True)
            self.doubleSpinBox_MultUpper.setValue(Default_dict ["Brightness mult. upper"])
            self.doubleSpinBox_MultUpper.setEnabled(True)
        else:
            return
    def keras_changed_noiseMean(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_GaussianNoiseMean.setValue(0.0)
            self.doubleSpinBox_GaussianNoiseMean.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_GaussianNoiseMean.setValue(Default_dict ["Gaussnoise Mean"])
            self.doubleSpinBox_GaussianNoiseMean.setEnabled(True)
        else:
            return
    def keras_changed_noiseScale(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_GaussianNoiseScale.setValue(0.0)
            self.doubleSpinBox_GaussianNoiseScale.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_GaussianNoiseScale.setValue(Default_dict ["Gaussnoise Scale"])
            self.doubleSpinBox_GaussianNoiseScale.setEnabled(True)
        else:
            return
    def keras_changed_contrast(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_contrastLower.setEnabled(False)
            self.doubleSpinBox_contrastHigher.setEnabled(False)

        elif on_or_off==2:
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)
        else:
            return
    def keras_changed_saturation(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_saturationLower.setEnabled(False)
            self.doubleSpinBox_saturationHigher.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_saturationLower.setEnabled(True)
            self.doubleSpinBox_saturationHigher.setEnabled(True)
        else:
            return
    def keras_changed_hue(self,on_or_off):
        if on_or_off==0:
            self.doubleSpinBox_hueDelta.setEnabled(False)
        elif on_or_off==2:
            self.doubleSpinBox_hueDelta.setEnabled(True)
        else:
            return

    def expert_mode_off(self,on_or_off):
        """
        Reset all values on the expert tab to the default values, excluding the metrics
        metrics are defined only once when starting fitting and should not be changed
        """
        if on_or_off==0: #switch off
            self.spinBox_batchSize.setValue(Default_dict["spinBox_batchSize"])
            self.spinBox_epochs.setValue(1)
            self.checkBox_expt_loss.setChecked(False)
            self.expert_loss_off(0)
            self.checkBox_learningRate.setChecked(False)        
            self.expert_learningrate_off(0)
            self.checkBox_optimizer.setChecked(False)
            self.expert_optimizer_off(0)
            self.checkBox_expt_paddingMode.setChecked(False)            
            self.expert_paddingMode_off(0)

    def expert_loss_off(self,on_or_off):
        if on_or_off==0: #switch off
            #switch back to categorical_crossentropy 
            index = self.comboBox_expt_loss.findText("categorical_crossentropy", QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_expt_loss.setCurrentIndex(index)
        
    def expert_learningrate_off(self,on_or_off):
        if on_or_off==0: #switch off
            #which optimizer is used? (there are different default learning-rates
            #for each optimizer!)
            optimizer = str(self.comboBox_optimizer.currentText())
            self.doubleSpinBox_learningRate.setValue(Default_dict["doubleSpinBox_learningRate_"+optimizer])

    def expert_optimizer_off(self,on_or_off):
        if on_or_off==0: #switch off, set back to categorical_crossentropy
            optimizer = "Adam"
            index = self.comboBox_optimizer.findText(optimizer, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_optimizer.setCurrentIndex(index)
                #also reset the learning rate to the default
                self.doubleSpinBox_learningRate.setValue(Default_dict["doubleSpinBox_learningRate_"+optimizer])

    def expert_optimizer_changed(self,value):
        #set the learning rate to the default for this optimizer
        optimizer = value
        value_current = float(self.doubleSpinBox_learningRate.value())
        value_wanted = Default_dict["doubleSpinBox_learningRate_"+optimizer]
        if value_current!=value_wanted:
            print("Update learning rate spinbox")
            self.doubleSpinBox_learningRate.setValue(value_wanted)
            #Inform user
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setWindowTitle("Learning rate to default")
            msg.setText("Learning rate was set to the default for "+optimizer)
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

    def expert_paddingMode_off(self,on_or_off):
        if on_or_off==0: #switch off
            #switch back to "constant" padding 
            index = self.comboBox_expt_paddingMode.findText("constant", QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_expt_paddingMode.setCurrentIndex(index)
        
    def update_hist1(self):
        feature = str(self.comboBox_feat1.currentText())
        feature_values = self.rtdc_ds[feature]
        #if len(feature_values)==len(self.rtdc_ds['area_cvx']):
#        self.histogram = pg.GraphicsWindow()        
        #plt1 = self.histogram.addPlot()
        y,x = np.histogram(feature_values, bins='auto')
        self.plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
#        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#        self.w.show()
    def update_hist2(self):
        feature = str(self.comboBox_feat2.currentText())
        feature_values = self.rtdc_ds[feature]
        #if len(feature_values)==len(self.rtdc_ds['area_cvx']):
        #self.histogram = pg.GraphicsWindow()        
        #plt1 = self.histogram.addPlot()
        y,x = np.histogram(feature_values, bins='auto')
        self.plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
#        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#        self.w.show()
    def update_scatter(self):
        feature_x = str(self.comboBox_feat1.currentText())
        feature_x_values = self.rtdc_ds[feature_x]
        feature_y = str(self.comboBox_feat2.currentText())
        feature_y_values = self.rtdc_ds[feature_y]
        if len(feature_x_values)==len(feature_y_values):
            #self.histogram = pg.GraphicsWindow()        
            #plt1 = self.histogram.addPlot()
            #y,x = np.histogram(feature_values, bins='auto')
            self.plt1.plot(feature_x_values, feature_y_values,pen=None,symbol='o',clear=True)
#            self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
#            self.w.show()

    def button_hist(self,item):
        buttonClicked = self.sender()
        index = self.table_dragdrop.indexAt(buttonClicked.pos())
        rowPosition = index.row()
        rtdc_path = self.table_dragdrop.cellWidget(rowPosition, 0).text()
        rtdc_path = str(rtdc_path)

        failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
                    
        self.rtdc_ds = rtdc_ds
#        feature_values = rtdc_ds[feature]
        #Init a popup window
        self.w = MyPopup()
        self.w.setWindowTitle(rtdc_path)
        self.w.setObjectName(_fromUtf8("w"))
        self.gridLayout_w2 = QtWidgets.QGridLayout(self.w)
        self.gridLayout_w2.setContentsMargins(0, 0, 0, 0)

        self.gridLayout_w2.setObjectName(_fromUtf8("gridLayout_w2"))
        self.widget = QtWidgets.QWidget(self.w)
        self.widget.setMinimumSize(QtCore.QSize(0, 65))
        self.widget.setMaximumSize(QtCore.QSize(16777215, 65))
        self.widget.setObjectName(_fromUtf8("widget"))
        self.horizontalLayout_w3 = QtWidgets.QHBoxLayout(self.widget)        
        self.horizontalLayout_w3.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_w3.setObjectName(_fromUtf8("horizontalLayout_w3"))
        self.verticalLayout_w = QtWidgets.QVBoxLayout()
        self.verticalLayout_w.setObjectName(_fromUtf8("verticalLayout_w"))
        self.horizontalLayout_w = QtWidgets.QHBoxLayout()
        self.horizontalLayout_w.setObjectName(_fromUtf8("horizontalLayout_w"))
        self.comboBox_feat1 = QtWidgets.QComboBox(self.widget)
        self.comboBox_feat1.setObjectName(_fromUtf8("comboBox_feat1"))
        self.comboBox_feat1.addItems(self.rtdc_ds.features)
        self.horizontalLayout_w.addWidget(self.comboBox_feat1)
        self.comboBox_feat2 = QtWidgets.QComboBox(self.widget)
        self.comboBox_feat2.setObjectName(_fromUtf8("comboBox_feat2"))
        self.comboBox_feat2.addItems(self.rtdc_ds.features)
        self.horizontalLayout_w.addWidget(self.comboBox_feat2)
        self.verticalLayout_w.addLayout(self.horizontalLayout_w)
        self.horizontalLayout_w2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_w2.setObjectName(_fromUtf8("horizontalLayout_w2"))
        self.pushButton_Hist1 = QtWidgets.QPushButton(self.widget)
        self.pushButton_Hist1.setObjectName(_fromUtf8("pushButton_Hist1"))
        self.horizontalLayout_w2.addWidget(self.pushButton_Hist1)
        self.pushButton_Hist2 = QtWidgets.QPushButton(self.widget)
        self.pushButton_Hist2.setObjectName(_fromUtf8("pushButton_Hist2"))
        self.horizontalLayout_w2.addWidget(self.pushButton_Hist2)
        self.verticalLayout_w.addLayout(self.horizontalLayout_w2)
        self.horizontalLayout_w3.addLayout(self.verticalLayout_w)
        self.verticalLayout_w2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_w2.setObjectName(_fromUtf8("verticalLayout_w2"))
        self.pushButton_Scatter = QtWidgets.QPushButton(self.widget)
        self.pushButton_Scatter.setObjectName(_fromUtf8("pushButton_Scatter"))
        self.verticalLayout_w2.addWidget(self.pushButton_Scatter)
        self.checkBox_ScalePix = QtWidgets.QCheckBox(self.widget)
        self.checkBox_ScalePix.setObjectName(_fromUtf8("checkBox_ScalePix"))
        self.verticalLayout_w2.addWidget(self.checkBox_ScalePix)
        self.horizontalLayout_w3.addLayout(self.verticalLayout_w2)
        self.gridLayout_w2.addWidget(self.widget, 0, 0, 1, 1)
      
        self.pushButton_Hist1.setText("Hist")
        self.pushButton_Hist1.clicked.connect(self.update_hist1)
        self.pushButton_Hist2.setText("Hist")
        self.pushButton_Hist2.clicked.connect(self.update_hist2)
        self.pushButton_Scatter.setText("Scatter")
        self.pushButton_Scatter.clicked.connect(self.update_scatter)

        self.checkBox_ScalePix.setText("Scale by pix")

        self.histogram = pg.GraphicsWindow()        
        self.plt1 = self.histogram.addPlot()
#        y,x = np.histogram(feature_values, bins='auto')
#        plt1.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150))
        self.gridLayout_w2.addWidget(self.histogram,1, 0, 1, 1)        
        self.w.show()

    def update_historyplot_pop(self,listindex):
        #listindex = self.popupcounter-1 #len(self.fittingpopups_ui)-1
        #After the first epoch there are checkboxes available. Check, if user checked some:
        colcount = int(self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.columnCount())
        #Collect items that are checked
        selected_items,Colors = [],[]
        for colposition in range(colcount):  
            #is it checked for train?
            cb = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(0, colposition)
            if not cb==None:
                if cb.checkState() == QtCore.Qt.Checked:
                    selected_items.append(str(cb.text()))
                    Colors.append(cb.background())
        self.Colors = Colors
        Histories = self.fittingpopups_ui[listindex].Histories
        DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
        if len(DF1)>0:
            DF1 = pd.concat(DF1)
        else:
            return
        self.fittingpopups_ui[listindex].widget_pop.clear()
        
        #Create fresh plot
        plt1 = self.fittingpopups_ui[listindex].widget_pop.addPlot()
        plt1.showGrid(x=True,y=True)
        plt1.addLegend()
        plt1.setLabel('bottom', 'Epoch', units='')
        #Create a dict that stores plots for each metric (for real time plotting)
        self.fittingpopups_ui[listindex].historyscatters = dict()
        for i in range(len(selected_items)):
            key = selected_items[i]
            df = DF1[key]
            color = self.Colors[i]
            pen_rollmedi = list(color.color().getRgb())
            pen_rollmedi = pg.mkColor(pen_rollmedi)
            pen_rollmedi = pg.mkPen(color=pen_rollmedi,width=6)
            color = list(color.color().getRgb())
            color[-1] = int(0.6*color[-1])
            color = tuple(color)                
            pencolor = pg.mkColor(color)
            brush = pg.mkBrush(color=pencolor)
            historyscatter = plt1.plot(range(len(df)), df,pen=None,symbol='o',symbolPen=None,symbolBrush=brush,name=key,clear=False)
            #self.fittingpopups_ui[listindex].historyscatters.append(historyscatter)
            self.fittingpopups_ui[listindex].historyscatters[key]=historyscatter


    def stop_fitting_pop(self,listindex):
        #listindex = len(self.fittingpopups_ui)-1
        epochs = self.fittingpopups_ui[listindex].epoch_counter                            
        #Stop button on the fititng popup
        #Should stop the fitting process and save the metafile
        #1. Change the nr. requested epochs to a smaller number
        self.fittingpopups_ui[listindex].spinBox_NrEpochs_pop.setValue(epochs-1)
        #2. Check the box which will cause that the new parameters are applied at next epoch
        self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.setChecked(True)


    def pause_fitting_pop(self,listindex):
        #Just change the text on the button
        if str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())==" ":
            #If the the text on the button was Pause, change it to Continue
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setText("")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setStyleSheet("background-color: green")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"continue.png")))

        elif str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())=="":
            #If the the text on the button was Continue, change it to Pause
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setText(" ")
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"pause.png")))
            self.fittingpopups_ui[listindex].pushButton_Pause_pop.setStyleSheet("")



    def saveTextWindow_pop(self,listindex):
        #Get the entire content of textBrowser_FittingInfo_pop
        text = str(self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.toPlainText())
        #Ask the user where to save the stuff
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Fitting info', Default_dict["Path of last model"]," (*.txt)")
        filename = filename[0]
        #Save to this filename
        if len(filename)>0:
            f = open(filename,'w')
            f.write(text)
            f.close()                

    def clearTextWindow_pop(self,listindex):
        self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.clear()
        
    def showModelSumm_pop(self,listindex):
        text5 = "Model summary:\n"
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        text = text5+summary
        self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)

    def saveModelSumm_pop(self,listindex):
        text5 = "Model summary:\n"
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        text = text5+summary
        #Ask the user where to save the stuff
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Model summary', Default_dict["Path of last model"]," (*.txt)")
        filename = filename[0]
        #Save to this filename
        f = open(filename,'w')
        f.write(text)
        f.close()                

    #class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert) #
    def get_class_weight(self,SelectedFiles,lossW_expert,custom_check_classes=False):
        t1 = time.time()
        print("Getting dictionary for class_weight")
        if lossW_expert=="None":
            return None
        elif lossW_expert=="":
            return None
        elif lossW_expert=="Balanced": #Definition of SelectedFiles:SelectedFiles.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"hash":hash_}) 
            #Which are training files
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_train = list(np.array(SelectedFiles)[ind])
            classes = [int(selectedfile["class"]) for selectedfile in SelectedFiles_train]
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train]
            classes_uni = np.unique(classes)
            counter = {}
            for class_ in classes_uni:
                ind = np.where(np.array(classes)==class_)[0]
                nr_events_epoch_class = np.array(nr_events_epoch)[ind]
                counter[class_] = np.sum(nr_events_epoch_class)
            max_val = float(max(counter.values()))
            return {class_id : max_val/num_images for class_id, num_images in counter.items()}
            
        elif lossW_expert.startswith("{"):#Custom loss weights
            class_weights = eval(lossW_expert)
            if custom_check_classes:#Check that each element in classes_uni is contained in class_weights.keys()
                ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
                ind = np.where(np.array(ind)==True)[0]
                SelectedFiles_train = list(np.array(SelectedFiles)[ind])
                classes = [int(selectedfile["class"]) for selectedfile in SelectedFiles_train]
                classes_uni = np.unique(classes)
                classes_uni = np.sort(classes_uni)
                class_weights_keys = np.sort([int(a) for a in class_weights.keys()])
                #each element in classes_uni has to be equal to class_weights_keys
                equal = np.array_equal(classes_uni,class_weights_keys)
                if equal == True:
                    return class_weights
                else:    
                    #If the equal is false I'm really in trouble...
                    #run the function again, but request 'Balanced' weights. I'm not sure if this should be the default...
                    class_weights = self.get_class_weight(SelectedFiles,"Balanced")
                    return ["Balanced",class_weights]
            else:
                return class_weights
        t2 = time.time()
        dt = np.round(t2-t1,2)
        print("Comp. time = "+str(dt))
            
            
            
            
    def partialtrainability_activated_pop(self,listindex):#same function like partialTrainability but on fitting popup
        print("Not implemented yet")
        print("Placeholder")
        print("Building site")

    def lossWeights_activated(self,on_or_off,listindex):
        if listindex==-1:
            item_ui = self
        else:
            item_ui = self.fittingpopups_ui[listindex]
        
        if on_or_off==False:#0 means switched OFF
            item_ui.lineEdit_lossW.setText("")
            item_ui.pushButton_lossW.setEnabled(False)

        #this happens when the user activated the expert option "loss weights"
        elif on_or_off==True:#2 means switched ON
            #Activate button
            item_ui.pushButton_lossW.setEnabled(True)
            self.lossWeights_popup(listindex)

        
    def lossWeights_popup(self,listindex):
        if listindex==-1:
            item_ui = self
            SelectedFiles = self.items_clicked()
        else:
            item_ui = self.fittingpopups_ui[listindex]
            SelectedFiles = item_ui.SelectedFiles
            
        item_ui.popup_lossW = MyPopup()
        item_ui.popup_lossW_ui = aid_frontend.popup_lossweights()
        item_ui.popup_lossW_ui.setupUi(item_ui.popup_lossW) #open a popup to show the numbers of events in each class in a table

        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]
        #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
        item_ui.popup_lossW_ui.tableWidget_lossW.setColumnCount(5)
        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = nr_ind
        item_ui.popup_lossW_ui.tableWidget_lossW.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class", "Events tot." ,"Events/Epoch", "Events/Epoch[%]", "Loss weight"]
        item_ui.popup_lossW_ui.tableWidget_lossW.setHorizontalHeaderLabels(header_labels) 
        header = item_ui.popup_lossW_ui.tableWidget_lossW.horizontalHeader()
        for i in range(len(header_labels)):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        

        #Fill the table 
        rowPosition = 0      
        #Training info
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]
        nr_events_train_total = np.sum([int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train])

        #Total nr of cells for each index
        for index in np.unique(indices_train):
            colPos = 0 #"Class" #put the index (class!) in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)
            
            #Get the training files of that index
            ind = np.where(indices_train==index)[0]
            SelectedFiles_train_index = np.array(SelectedFiles_train)[ind]
    
            colPos = 1 #"Events tot."
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)

            colPos = 2 #"Events/Epoch"
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)
            
            colPos = 3 #"Events/Epoch[%]"
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole, str(np.round(np.sum(nr_events_epoch)/float(nr_events_train_total),2)))
            item_ui.popup_lossW_ui.tableWidget_lossW.setItem(rowPosition, colPos, item)

            colPos = 4 #"Loss weights"
            #for each item create a spinbopx (trainability)
            spinb = QtWidgets.QDoubleSpinBox(item_ui.popup_lossW_ui.tableWidget_lossW)
            spinb.setEnabled(False)
            spinb.setMinimum(-99999)
            spinb.setMaximum(99999)
            spinb.setSingleStep(0.1)
            spinb.setValue(1.0) #Default in Keras is "None", which means class_weight=1.0
            item_ui.popup_lossW_ui.tableWidget_lossW.setCellWidget(rowPosition, colPos, spinb)            
            
            rowPosition += 1

        item_ui.popup_lossW_ui.tableWidget_lossW.resizeColumnsToContents()            
        item_ui.popup_lossW_ui.tableWidget_lossW.resizeRowsToContents()

        item_ui.popup_lossW.show()
                
        item_ui.popup_lossW_ui.pushButton_pop_lossW_cancel.clicked.connect(lambda: self.lossW_cancel(listindex))
        item_ui.popup_lossW_ui.pushButton_pop_lossW_ok.clicked.connect(lambda: self.lossW_ok(listindex))
        item_ui.popup_lossW_ui.comboBox_lossW.currentIndexChanged.connect(lambda on_or_off: self.lossW_comboB(on_or_off,listindex))



    def onLayoutChange(self):
        #Get the text of the triggered layout
        layout_trig = (self.sender().text()).split(" layout")[0]
        layout_current = Default_dict["Layout"]

        if layout_trig == layout_current:
            self.statusbar.showMessage(layout_current+" layout is already in use",2000)
            return
        
        elif layout_trig == "Normal":
            #Change Layout in Defaultdict to "Normal", such that next start will use Normal layout
            Default_dict["Layout"] = "Normal"
            app.setStyleSheet("")
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)

        elif layout_trig == "Dark":
            #Change Layout in Defaultdict to "Dark", such that next start will use Dark layout
            Default_dict["Layout"] = "Dark"
            dir_layout = os.path.join(dir_root,"layout_dark.txt")#dir to settings
            f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
            f = f.read()
            app.setStyleSheet(f)
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)
        
        elif layout_trig == "DarkOrange":
            #Change Layout in Defaultdict to "Dark", such that next start will use Dark layout
            Default_dict["Layout"] = "DarkOrange"
            dir_layout = os.path.join(dir_root,"layout_darkorange.txt")#dir to settings
            f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/nphase/qt-ping-grapher/blob/master/resources/darkorange.stylesheet
            f = f.read()
            app.setStyleSheet(f)
            #Standard is with tooltip
            self.actionTooltipOnOff.setChecked(True)

        #Save the layout to Default_dict
        with open(dir_settings, 'w') as f:
            json.dump(Default_dict,f)
        
    def onTooltipOnOff(self):
        #what is the current layout?
        if bool(self.actionLayout_Normal.isChecked())==True: #use normal layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                app.setStyleSheet("")
            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                app.setStyleSheet("""QToolTip {
                                         opacity: 0
                                           }""")

        elif bool(self.actionLayout_Dark.isChecked())==True: #use dark layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                dir_layout = os.path.join(dir_root,"layout_dark.txt")#dir to settings
                f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
                f = f.read()
                app.setStyleSheet(f)

            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                dir_layout = os.path.join(dir_root,"layout_dark_notooltip.txt")#dir to settings
                f = open(dir_layout, "r")#I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
                f = f.read()
                app.setStyleSheet(f)

        elif bool(self.actionLayout_DarkOrange.isChecked())==True: #use darkorange layout
            if bool(self.actionTooltipOnOff.isChecked())==True: #with tooltips
                dir_layout = os.path.join(dir_root,"layout_darkorange.txt")#dir to settings
                f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/nphase/qt-ping-grapher/blob/master/resources/darkorange.stylesheet
                f = f.read()
                app.setStyleSheet(f)

            elif bool(self.actionTooltipOnOff.isChecked())==False: #no tooltips
                dir_layout = os.path.join(dir_root,"layout_darkorange_notooltip.txt")#dir to settings
                f = open(dir_layout, "r")
                f = f.read()
                app.setStyleSheet(f)

    def onIconThemeChange(self):
        #Get the text of the triggered icon theme
        icontheme_trig = self.sender().text()
        icontheme_currenent = Default_dict["Icon theme"]

        if icontheme_trig == icontheme_currenent:
            self.statusbar.showMessage(icontheme_currenent+" is already in use",2000)
            return
        
        elif icontheme_trig == "Icon theme 1":
            Default_dict["Icon theme"] = "Icon theme 1"
            self.statusbar.showMessage("Icon theme 1 will be used after restart",2000)

        elif icontheme_trig == "Icon theme 2":
            Default_dict["Icon theme"] = "Icon theme 2"
            self.statusbar.showMessage("Icon theme 2 will be used after restart",2000)
        
        #Save the layout to Default_dict
        with open(dir_settings, 'w') as f:
            json.dump(Default_dict,f)


    def items_clicked(self):
        #This function checks, which data has been checked on table_dragdrop and returns the necessary data
        rowCount = self.table_dragdrop.rowCount()
        #Collect urls to files that are checked
        SelectedFiles = []
        for rowPosition in range(rowCount):  
            #get the filename/path
            rtdc_path = str(self.table_dragdrop.cellWidget(rowPosition, 0).text())
            #get the index (celltype) of it
            index = int(self.table_dragdrop.cellWidget(rowPosition, 1).value())
            #is it checked for train?
            cb_t = self.table_dragdrop.item(rowPosition, 2)
            #How many Events contains dataset in total?
            nr_events = int(self.table_dragdrop.item(rowPosition, 5).text())
            #how many cells/epoch during training or validation?
            nr_events_epoch = int(self.table_dragdrop.item(rowPosition, 6).text())            
            #should the dataset be randomized (shuffled?)            
            shuffle = bool(self.table_dragdrop.item(rowPosition, 8).checkState())           
            #should the images be zoomed in/out by a factor?
            zoom_factor = float(self.table_dragdrop.item(rowPosition, 9).text())            
            
            if cb_t.checkState() == QtCore.Qt.Checked and nr_events_epoch>0: #add to training files if the user wants more than 0 images per epoch
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    
                hash_ = rtdc_ds.hash
                features = rtdc_ds.features
                nr_images = rtdc_ds["image"].len()
                SelectedFiles.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"hash":hash_})
            
            cb_v = self.table_dragdrop.item(rowPosition, 3)
            if cb_v.checkState() == QtCore.Qt.Checked and nr_events_epoch>0:
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                hash_ = rtdc_ds.hash
                features = rtdc_ds.features
                nr_images = rtdc_ds["image"].len()
                SelectedFiles.append({"rtdc_ds":rtdc_ds,"rtdc_path":rtdc_path,"features":features,"nr_images":nr_images,"class":index,"TrainOrValid":"Valid","nr_events":nr_events,"nr_events_epoch":nr_events_epoch,"shuffle":shuffle,"zoom_factor":zoom_factor,"hash":hash_})
        return SelectedFiles


    def items_clicked_no_rtdc_ds(self):
        #This function checks, which data has been checked on table_dragdrop and returns the necessary data
        rowCount = self.table_dragdrop.rowCount()
        #Collect urls to files that are checked
        SelectedFiles = []
        for rowPosition in range(rowCount):  
            #get the index (celltype) of it
            index = int(self.table_dragdrop.cellWidget(rowPosition, 1).value())
            #is it checked for train?
            cb_t = self.table_dragdrop.item(rowPosition, 2)
            #How many Events contains dataset in total?
            nr_events = int(self.table_dragdrop.item(rowPosition, 5).text())
            #how many cells/epoch during training or validation?
            nr_events_epoch = int(self.table_dragdrop.item(rowPosition, 6).text())            

            if cb_t.checkState() == QtCore.Qt.Checked and nr_events_epoch>0: #add to training files if the user wants more than 0 images per epoch
                SelectedFiles.append({"nr_images":nr_events,"class":index,"TrainOrValid":"Train","nr_events":nr_events,"nr_events_epoch":nr_events_epoch})
            
            cb_v = self.table_dragdrop.item(rowPosition, 3)
            if cb_v.checkState() == QtCore.Qt.Checked and nr_events_epoch>0:
                SelectedFiles.append({"nr_images":nr_events,"class":index,"TrainOrValid":"Valid","nr_events":nr_events,"nr_events_epoch":nr_events_epoch})
        return SelectedFiles




    def uncheck_if_zero(self,item):
        #If the Nr. of epochs is changed to zero:
        #uncheck the dataset for train/valid
        row = item.row()
        col = item.column()
        #if the user changed Nr. of cells per epoch to zero
        if col==6 and int(item.text())==0:
            #get the checkstate of the coresponding T/V
            cb_t = self.table_dragdrop.item(row, 2)
            if cb_t.checkState() == QtCore.Qt.Checked:
                cb_t.setCheckState(False)
            cb_v = self.table_dragdrop.item(row, 3)
            if cb_v.checkState() == QtCore.Qt.Checked:
                cb_v.setCheckState(False)
            
    def item_click(self,item): 
        colPosition = item.column()
        rowPosition = item.row()
        #if Shuffle was clicked (col=8), check if this checkbox is not deactivated
        if colPosition==8:
            if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==False:
                rtdc_path = self.table_dragdrop.cellWidget(rowPosition, 0).text()
                rtdc_path = str(rtdc_path)

                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                nr_images = rtdc_ds["image"].len()
        
                columnPosition = 6
                item = QtWidgets.QTableWidgetItem()
                item.setData(QtCore.Qt.DisplayRole, nr_images)
                item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                self.table_dragdrop.setItem(rowPosition, columnPosition, item)
            if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==True:
                #Inspect this table item. If shuffle was checked before, it will be grayed out. Invert normal cell then
                item = self.table_dragdrop.item(rowPosition, 6)
                item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )

        if len(self.ram)>0:
            self.statusbar.showMessage("Make sure to update RAM (->Edit->Data to RAM now) after changing Data-set",2000)
            self.ram = dict() #clear the ram, since the data was changed
            
        self.dataOverviewOn()

        #When data is clicked, always reset the validation set (only important for 'Assess Model'-tab)
        self.ValidationSet = None
        self.Metrics = dict() #Also reset the metrics

    def dataOverviewOn(self):
        if self.groupBox_DataOverview.isChecked()==True:
            if self.threadpool_single_queue == 0:
                SelectedFiles = self.items_clicked_no_rtdc_ds()
                self.update_data_overview(SelectedFiles)
                self.update_data_overview_2(SelectedFiles)

    def dataOverviewOn_OnChange(self,item):
        #When a value is entered in Events/Epoch and enter is hit
        #there is no update of the table called
        if self.groupBox_DataOverview.isChecked()==True:
            if self.threadpool_single_queue == 0:
                rowPosition = item.row()
                colPosition = item.column()
                if colPosition==6:#one when using the spinbox (Class),or when entering a new number in "Events/Epoch", the table is not updated. 
                    #get the new value
                    nr_cells = self.table_dragdrop.cellWidget(rowPosition, colPosition)
                    if nr_cells==None:
                        return
                    else:
                        SelectedFiles = self.items_clicked_no_rtdc_ds()
                        self.update_data_overview(SelectedFiles)
                        self.update_data_overview_2(SelectedFiles)
                            
    def update_data_overview(self,SelectedFiles):
        #Check if there are custom class names (determined by user)
        rows = self.tableWidget_Info.rowCount()
        self.classes_custom = [] #by default assume there are no custom classes
        classes_custom_bool = False
        if rows>0:#if >0, then there is already a table existing
            classes,self.classes_custom = [],[]
            for row in range(rows):
                try:
                    class_ = self.tableWidget_Info.item(row,0).text()
                    if class_.isdigit():    
                        classes.append(class_)#get the classes
                except:
                    pass
                try:
                    self.classes_custom.append(self.tableWidget_Info.item(row,3).text())#get the classes
                except:
                    pass
            classes = np.unique(classes)
            if len(classes)==len(self.classes_custom):#equal in length
                same = [i for i, j in zip(classes, self.classes_custom) if i == j] #which items are identical?
                if len(same)==0:
                    #apparently there are custom classes! Save them
                    classes_custom_bool = True                            
            
        if len(SelectedFiles)==0:#reset the table
            #Table1
            #Prepare a table in tableWidget_Info
            self.tableWidget_Info.setColumnCount(0)
            self.tableWidget_Info.setRowCount(0)
            self.tableWidget_Info.setColumnCount(4)
            header = self.tableWidget_Info.horizontalHeader()
            header_labels = ["Class","Events tot.","Events/Epoch","Name"]
            self.tableWidget_Info.setHorizontalHeaderLabels(header_labels) 
            header = self.tableWidget_Info.horizontalHeader()
            for i in range(4):
                header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
            return
        #Prepare a table in tableWidget_Info
        self.tableWidget_Info.setColumnCount(0)
        self.tableWidget_Info.setRowCount(0)

        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]
        self.tableWidget_Info.setColumnCount(4)
        header = self.tableWidget_Info.horizontalHeader()

        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = 2*nr_ind+2 #add two rows for intermediate headers (Train/Valid)
        self.tableWidget_Info.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Events tot.","Events/Epoch","Name"]
        self.tableWidget_Info.setHorizontalHeaderLabels(header_labels) 
        #self.tableWidget_Info.resizeColumnsToContents()            
        header = self.tableWidget_Info.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        #Training info
        rowPosition = 0
        self.tableWidget_Info.setSpan(rowPosition, 0, 1, 2) 
        item = QtWidgets.QTableWidgetItem("Train. data") 
        item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
        self.tableWidget_Info.setItem(rowPosition, 0, item)            
        rowPosition += 1
        ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_train = np.array(SelectedFiles)[ind]
        SelectedFiles_train = list(SelectedFiles_train)
        indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]

        classes = np.unique(indices_train)
        if len(classes)==len(self.classes_custom):
            classes_custom_bool = True
        else:
            classes_custom_bool = False
 
       #display information for each individual class            
        for index_ in range(len(classes)):
        #for index in np.unique(indices_train):
            index = classes[index_]
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info.setItem(rowPosition, 0, item)
            #Get the training files of that index
            ind = np.where(indices_train==index)[0]
            SelectedFiles_train_index = np.array(SelectedFiles_train)[ind]
            #Total nr of cells for each class
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            self.tableWidget_Info.setItem(rowPosition, 1, item)
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_train_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info.setItem(rowPosition, 2, item)    

            item = QtWidgets.QTableWidgetItem()
            if classes_custom_bool==False:
                item.setData(QtCore.Qt.EditRole,str(index))
            else:
                item.setData(QtCore.Qt.EditRole,self.classes_custom[index_])
            self.tableWidget_Info.setItem(rowPosition, 3, item)                

            rowPosition += 1
        



        #Validation info
        self.tableWidget_Info.setSpan(rowPosition, 0, 1, 2) 
        item = QtWidgets.QTableWidgetItem("Val. data")  
        item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
        self.tableWidget_Info.setItem(rowPosition, 0, item)            
        rowPosition += 1
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        #Total nr of cells for each index
        for index in np.unique(indices_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(indices_valid==index)[0]
            SelectedFiles_valid_index = np.array(SelectedFiles_valid)[ind]
            nr_events = [int(selectedfile["nr_events"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events)))
            self.tableWidget_Info.setItem(rowPosition, 1, item)
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info.setItem(rowPosition, 2, item)
            rowPosition += 1
        self.tableWidget_Info.resizeColumnsToContents()            
        self.tableWidget_Info.resizeRowsToContents()

    def update_data_overview_2(self,SelectedFiles):
        if len(SelectedFiles)==0:
            #Table2
            self.tableWidget_Info_2.setColumnCount(0)
            self.tableWidget_Info_2.setRowCount(0)               
            #In case user specified X_valid and y_valid before, delete it again:
            self.ValidationSet = None
            self.Metrics = dict() #Also reset the metrics
            #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
            self.tableWidget_Info_2.setColumnCount(4)
            header_labels = ["Class","Nr of cells","Clr","Name"]
            self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
            header = self.tableWidget_Info_2.horizontalHeader()
            for i in range(4):
                header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)  
            return

        #Prepare a table in tableWidget_Info
        self.tableWidget_Info_2.setColumnCount(0)
        self.tableWidget_Info_2.setRowCount(0)
        
        #In case user specified X_valid and y_valid before, delete it again:
        self.ValidationSet = None
        self.Metrics = dict() #Also reset the metrics

        indices = [SelectedFiles[i]["class"] for i in range(len(SelectedFiles))]
        #Initiate the table with 4 columns : this will be ["Index","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setColumnCount(4)
        nr_ind = len(set(indices)) #each index could occur for train and valid
        nr_rows = nr_ind
        self.tableWidget_Info_2.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
        header = self.tableWidget_Info_2.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        rowPosition = 0      
        #Validation info
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        #Total nr of cells for each index
        for index in np.unique(indices_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(indices_valid==index)[0]
            SelectedFiles_valid_index = np.array(SelectedFiles_valid)[ind]
            nr_events_epoch = [int(selectedfile["nr_events_epoch"]) for selectedfile in SelectedFiles_valid_index]
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info_2.setItem(rowPosition, 1, item)
            
            #Column for color
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, "")
            item.setBackground(QtGui.QColor(self.colorsQt[index]))            
            self.tableWidget_Info_2.setItem(rowPosition, 2, item)

            #Column for User specified name
            item = QtWidgets.QTableWidgetItem()
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 3, item)
            
            rowPosition += 1
        self.tableWidget_Info_2.resizeColumnsToContents()            
        self.tableWidget_Info_2.resizeRowsToContents()


    def tableWidget_Info_2_click(self,item):
        if item is not None:
            if item.column()==2:
                tableitem = self.tableWidget_Info_2.item(item.row(), item.column())
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)

    def tableWidget_HistoryItems_dclick(self,item):
        if item is not None:
            tableitem = self.tableWidget_HistoryItems.item(item.row(), item.column())
            if str(tableitem.text())!="Show saved only":
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)
                    self.update_historyplot()
 
    def select_all(self,col):
        apply_at_col = [2,3,8]
        if col not in apply_at_col:
            return
        #otherwiese continue
        rows = range(self.table_dragdrop.rowCount()) #Number of rows of the table
        
        tableitems = [self.table_dragdrop.item(row, col) for row in rows]
        checkStates = [tableitem.checkState() for tableitem in tableitems]
        #Checked?
        checked = [state==QtCore.Qt.Checked for state in checkStates]
        if set(checked)=={True}:#all are checked!
            #Uncheck all!
            for tableitem in tableitems:
                tableitem.setCheckState(QtCore.Qt.Unchecked)
        else:#otherwise check all   
            for tableitem in tableitems:
                tableitem.setCheckState(QtCore.Qt.Checked)
                
        #If shuffle column was clicked do some extra
        if col==8:
            for rowPosition in rows:
                if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==False:
                    rtdc_path = self.table_dragdrop.cellWidget(rowPosition, 0).text()
                    rtdc_path = str(rtdc_path)
    
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    if failed:
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Critical)       
                        msg.setText(str(rtdc_ds))
                        msg.setWindowTitle("Error occurred during loading file")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        return
                    nr_images = rtdc_ds["image"].len()
            
                    columnPosition = 6
                    item = QtWidgets.QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, nr_images)
                    item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                    self.table_dragdrop.setItem(rowPosition, columnPosition, item)
                if bool(self.table_dragdrop.item(rowPosition, 8).checkState())==True:
                    #Inspect this table item. If shuffle was checked before, it will be grayed out. Invert normal cell then
                    item = self.table_dragdrop.item(rowPosition, 6)
                    item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )

        #Finally, update the Data-Overview-Box
        self.dataOverviewOn()#update the overview box

    def item_dclick(self, item):
        #Check/Uncheck if item is from column 2 or 3
        tableitem = self.table_dragdrop.item(item.row(), item.column())
        if item.column() in [2,3]:
            #If the item is unchecked ->check it!
            if tableitem.checkState() == QtCore.Qt.Unchecked:
                tableitem.setCheckState(QtCore.Qt.Checked)
            #else, the other way around
            elif tableitem.checkState() == QtCore.Qt.Checked:
                tableitem.setCheckState(QtCore.Qt.Unchecked)  
        
        #Show example image if item on column 0 was dclicked
        if item.column() == 0: 
            #rtdc_path = str(item.text())
            #rtdc_path = tableitem.text()
            rtdc_path = self.table_dragdrop.cellWidget(item.row(), item.column()).text()

            failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
            if failed:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)       
                msg.setText(str(rtdc_ds))
                msg.setWindowTitle("Error occurred during loading file")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                    
            nr_images = rtdc_ds["image"].len()
            ind = np.random.randint(0,nr_images)
            img = rtdc_ds["image"][ind]
            if len(img.shape)==2:
                height, width = img.shape
                channels = 1
            elif len(img.shape)==3:
                height, width, channels = img.shape
            else:
                print("Invalid image format: "+str(img.shape))
                return
            self.w = MyPopup()
            self.gridLayout_w = QtWidgets.QGridLayout(self.w)
            self.label_image = QtWidgets.QLabel(self.w)
            self.label_cropimage = QtWidgets.QLabel(self.w)

            #zoom image such that longest side is 512
            factor = np.round(float(512.0/np.max(img.shape)),0)
            #Get the order, specified in Options->Zoom Order
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            order = np.where(np.array(zoom_methods)==True)[0]
            if channels==1:
                img_zoom = ndimage.zoom(img, zoom=factor,order=int(order)) #Order 0 means nearest neighbor interplation
            if channels==3:
                img_zoom = ndimage.zoom(img, zoom=(factor,factor,1),order=int(order)) #Order 0 means nearest neighbor interplation
            img_zoom = np.ascontiguousarray(img_zoom)
            
            if channels==1:
                height, width = img_zoom.shape
            if channels==3:
                height, width, _ = img_zoom.shape
            
            if channels==1:
                qi=QtGui.QImage(img_zoom.data, width, height,width, QtGui.QImage.Format_Indexed8)
            if channels==3:
                qi = QtGui.QImage(img_zoom.data,img_zoom.shape[1], img_zoom.shape[0], QtGui.QImage.Format_RGB888)
                
            self.label_image.setPixmap(QtGui.QPixmap.fromImage(qi))
            self.gridLayout_w.addWidget(self.label_image, 1,1)
              
            #get the location of the cell
            rowPosition = item.row()
            pix = float(self.table_dragdrop.item(rowPosition, 7).text())
            #pix = rtdc_ds.config["imaging"]["pixel size"]
            PIX = pix
            
            pos_x,pos_y = rtdc_ds["pos_x"][ind]/PIX,rtdc_ds["pos_y"][ind]/PIX
            cropsize = self.spinBox_imagecrop.value()
            y1 = int(round(pos_y))-cropsize/2                
            x1 = int(round(pos_x))-cropsize/2 
            y2 = y1+cropsize                
            x2 = x1+cropsize
            img_crop = img[int(y1):int(y2),int(x1):int(x2)]
            #zoom image such that the height gets the same as for non-cropped img
            factor = float(float(height)/np.max(img_crop.shape[0]))
            if factor == np.inf:
                factor = 1
                if self.actionVerbose.isChecked()==True:
                    print("Set resize factor to 1. Before, it was: "+str(factor))     
            #img_crop = zoom(img_crop,factor)
            #Get the order, specified in Options->Zoom Order
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            order = np.where(np.array(zoom_methods)==True)[0]
            if channels==1:
                img_crop = ndimage.zoom(img_crop, zoom=factor,order=int(order))
            if channels==3:
                img_crop = ndimage.zoom(img_crop, zoom=(factor,factor,1),order=int(order))                
            img_crop = np.ascontiguousarray(img_crop)
            if channels==1:
                height, width = img_crop.shape
            if channels==3:
                height, width, _ = img_crop.shape
            if channels==1:
                qi=QtGui.QImage(img_crop.data, width, height,width, QtGui.QImage.Format_Indexed8)
            if channels==3:
                qi = QtGui.QImage(img_crop.data,img_crop.shape[1], img_crop.shape[0], QtGui.QImage.Format_RGB888)
            
            self.label_cropimage.setPixmap(QtGui.QPixmap.fromImage(qi))
            self.gridLayout_w.addWidget(self.label_cropimage, 1,2)
            self.w.show()

    def get_norm_from_modelparafile(self):
        #Get the normalization method from a modelparafile
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta file (*meta.xlsx)")
        filename = filename[0]
        if len(str(filename))==0:
            return
        norm = pd.read_excel(filename,sheetname='Parameters')["Normalization"]
        norm = str(norm[0])
        index = self.comboBox_Normalization.findText(norm, QtCore.Qt.MatchFixedString)
        if index >= 0:
            self.comboBox_Normalization.setCurrentIndex(index)
            self.w.close()
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Invalid normalization method was specified.\
            Likely this version of AIDeveloper does not support that normalization method\
            Please define a valid normalization method")
            msg.setDetailedText("Supported normalization methods are: "+"\n".join(self.norm_methods))
            msg.setWindowTitle("Invalid Normalization method")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            #raise ValueError("Invalid Normalization method")

    def update_comboBox_feature_xy(self):           
        #Get current text of combobox (url to data set)
        url = str(self.comboBox_chooseRtdcFile.currentText())
        if len(url)>0:

            failed,rtdc_ds = aid_bin.load_rtdc(url)
            if failed:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)       
                msg.setText(str(rtdc_ds))
                msg.setWindowTitle("Error occurred during loading file")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            features = rtdc_ds.features
            #Fill those feautues in the comboboxes at the scatterplot
            self.comboBox_featurex.addItems(features)
            self.comboBox_featurey.addItems(features)

    def activate_deactivate_spinbox(self,newstate):
        #get the checkstate of the Input model crop 
        if newstate==2:
            #activate the spinbox
            self.spinBox_imagecrop.setEnabled(True)
        elif newstate==0:
            self.spinBox_imagecrop.setEnabled(False)


    def gray_or_rgb_augmentation(self,index):
        #When Color-Mode is changed:
        #Get the new colormode:
        new_colormode = self.colorModes[index]
        #when the new Color Mode is Grayscale, disable saturation and hue augmentation
        if new_colormode=="Grayscale":

            self.checkBox_contrast.setEnabled(True)
            self.checkBox_contrast.setChecked(True)
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)

            self.checkBox_saturation.setEnabled(False)
            self.checkBox_saturation.setChecked(False)
            self.doubleSpinBox_saturationLower.setEnabled(False)
            self.doubleSpinBox_saturationHigher.setEnabled(False)

            self.checkBox_hue.setEnabled(False)
            self.checkBox_hue.setChecked(False)
            self.doubleSpinBox_hueDelta.setEnabled(False)

        elif new_colormode=="RGB":
            self.checkBox_contrast.setEnabled(True)
            self.checkBox_contrast.setChecked(True)
            self.doubleSpinBox_contrastLower.setEnabled(True)
            self.doubleSpinBox_contrastHigher.setEnabled(True)

            self.checkBox_saturation.setEnabled(True)
            self.checkBox_saturation.setChecked(True)
            self.doubleSpinBox_saturationLower.setEnabled(True)
            self.doubleSpinBox_saturationHigher.setEnabled(True)

            self.checkBox_hue.setEnabled(True)
            self.checkBox_hue.setChecked(True)
            self.doubleSpinBox_hueDelta.setEnabled(True)
        else:
            print("Invalid Color Mode")

           
    def onClick(self,points,pointermethod):
        #delete the last item if the user selected already one:
        if self.point_was_selected_before:
            self.scatter_xy.removeItem(self.scatter_xy.listDataItems()[-1])
            try:
                self.widget_showCell.removeItem(self.dot)
            except:
                pass
        if pointermethod=="point":
            points = points[0]
            p = points.pos()
            clicked_x, clicked_y = p.x(), p.y()
            a1 = (clicked_x)/float(np.max(self.feature_x))            
            a2 = (clicked_y)/float(np.max(self.feature_y))
            #Which is the closest scatter point?
            dist = np.sqrt(( a1-self.scatter_x_norm )**2 + ( a2-self.scatter_y_norm )**2)
            index =  np.argmin(dist)

        elif pointermethod=="index":
            index = points
            
        clicked_x = self.feature_x[index]
        clicked_y = self.feature_y[index]
#            print("index,clicked_x,clicked_y")
#            print(index,clicked_x,clicked_y)
        
        self.scatter_xy.plot([clicked_x], [clicked_y],pen=None,symbol='o',symbolPen='w',clear=False)
        self.point_was_selected_before = True

        #I dont care if the user click or used the slider->always adjust spinboc and slider without running the onChange functions 
        self.changedbyuser = False
        self.spinBox_cellInd.setValue(index)
        self.horizontalSlider_cellInd.setValue(index)
        self.changedbyuser = True

        rtdc_ds = self.rtdc_ds
        #Get the corresponding image:
        if "image" in rtdc_ds.features:
            img = rtdc_ds["image"][index]

            if len(img.shape)==2:
                channels = 1
            elif len(img.shape)==3:
                height, width, channels = img.shape
            else:
                print("Invalid image format: "+str(img.shape))
                return
            
            #zoom image such that longest side is 64              
            factor = 1#float(64.0/np.max(img.shape))
            #Get the order, specified in Options->Zoom Order
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            order = np.where(np.array(zoom_methods)==True)[0]

            if channels==1:
                img_zoom = ndimage.zoom(img, zoom=factor,order=int(order)) #Order 0 means nearest neighbor interplation
            if channels==3:
                img_zoom = ndimage.zoom(img, zoom=(factor,factor,1),order=int(order)) #Order 0 means nearest neighbor interplation
            img_zoom = np.ascontiguousarray(img_zoom)

            #from 0 to 255
#            img_zoom = img_zoom-np.min(img_zoom)
#            fac = np.max(img_zoom)
#            img_zoom = (img_zoom/fac)*255.0
            img_zoom = img_zoom.astype(np.uint8)

            if channels==1:
                self.widget_showCell.setImage(img_zoom.T,autoRange=False)
            elif channels==3:                
                self.widget_showCell.setImage(img_zoom.T,autoRange=False)
                #self.widget_showCell.setImage(np.swapaxes(img_zoom,0,1),autoRange=False)
            self.widget_showCell.ui.histogram.hide()
            self.widget_showCell.ui.roiBtn.hide()
            self.widget_showCell.ui.menuBtn.hide()
            
            #Indicate the centroid of the cell
            pix = rtdc_ds.config["imaging"]["pixel size"]
            pos_x = rtdc_ds["pos_x"][index]/pix
            pos_y = rtdc_ds["pos_y"][index]/pix
            if self.checkBox_centroid.isChecked():
                self.dot = pg.CircleROI(pos=(pos_x-2, pos_y-2), size=4, pen=QtGui.QPen(QtCore.Qt.red, 0.1), movable=False)
                self.widget_showCell.getView().addItem(self.dot)
                self.widget_showCell.show()
        
        #Fluorescence traces: clear first
        try:
            self.plot_fl_trace_.clear() #clear the plot
            self.plot_fl_trace.clear() #clear the plot
        except:
            pass
        if "trace" in rtdc_ds.features:
            #Show the flourescence traces
            trace = rtdc_ds["trace"]
            fl_keys = list(trace.keys())
            feature_keys = rtdc_ds.features
            fl1_max,fl1_pos,fl2_max,fl2_pos,fl3_max,fl3_pos = 0,0,0,0,0,0
            Traces_flx = []
            for i in range(len(fl_keys)):
                if "fl1_median" in fl_keys[i] and self.checkBox_fl1.isChecked():
                    trace_flx = trace[fl_keys[i]][index]
                    pencolor = "g"
                    Traces_flx.append(trace_flx)
                    self.plot_fl_trace_ = self.plot_fl_trace.plot(range(len(trace_flx)),trace_flx,width=6,pen=pencolor,clear=False)
                    if "fl1_max" in feature_keys and "fl1_pos" in feature_keys: #if also the maxima and position of the max are available: use it to put the region accordingly
                        fl1_max,fl1_pos = rtdc_ds["fl1_max"][index],rtdc_ds["fl1_pos"][index]
                elif "fl2_median" in fl_keys[i] and self.checkBox_fl2.isChecked():
                    trace_flx = trace[fl_keys[i]][index]
                    Traces_flx.append(trace_flx)
                    pencolor = (255,128,0) #orange
                    self.plot_fl_trace_ = self.plot_fl_trace.plot(range(len(trace_flx)),trace_flx,width=6,pen=pencolor,clear=False)
                    if "fl2_max" in feature_keys and "fl2_pos" in feature_keys: #if also the maxima and position of the max are available: use it to put the region accordingly
                        fl2_max,fl2_pos = rtdc_ds["fl2_max"][index],rtdc_ds["fl2_pos"][index]
                elif "fl3_median" in fl_keys[i] and self.checkBox_fl3.isChecked():
                    trace_flx = trace[fl_keys[i]][index]
                    Traces_flx.append(trace_flx)
                    pencolor = "r"
                    self.plot_fl_trace_ = self.plot_fl_trace.plot(range(len(trace_flx)),trace_flx,width=6,pen=pencolor,clear=False)
                    if "fl3_max" in feature_keys and "fl3_pos" in feature_keys: #if also the maxima and position of the max are available: use it to put the region accordingly
                        fl3_max,fl3_pos = rtdc_ds["fl3_max"][index],rtdc_ds["fl3_pos"][index]
    
            #get the maximum of [fl1_max,fl2_max,fl3_max] and put the region to the corresponding fl-position
            ind = np.argmax(np.array([fl1_max,fl2_max,fl3_max]))
            region_pos = np.array([fl1_pos,fl2_pos,fl3_pos])[ind] #this region is already given in us. translate this back to range
            peak_height = np.array([fl1_max,fl2_max,fl3_max])[ind]
            sample_rate = rtdc_ds.config["fluorescence"]["sample rate"]
            fl_pos_ind = float((sample_rate*region_pos))/1E6 #
            #Indicate the used flx_max and flx_pos by a scatter dot
            self.peak_dot = self.plot_fl_trace.plot([float(fl_pos_ind)], [float(peak_height)],pen=None,symbol='o',symbolPen='w',clear=False)
    
            #Place a LinearRegionItem on the plot. FL_max is searched in this region
            #Search, if a region_width was defined already
            if not hasattr(self, 'region_width'): #if there was no region_width defined yet...
    #            self.region_width = 50 #width of the region in us
    #            self.region_width = float((sample_rate*self.region_width))/1E6 #
                #to get a reasonable initial range, use 20% of the nr. of availeble samples
                samples_per_event = self.rtdc_ds.config["fluorescence"]["samples per event"]
                self.region_width = 0.2*samples_per_event #width of the region in samples
                #Convert to SI unit:
                sample_rate = self.rtdc_ds.config["fluorescence"]["sample rate"]
                self.region_width = (float(self.region_width)/float(sample_rate))*1E6 #range[samples]*(1/sample_rate[1/s]) = range[s]; div by 1E6 to convert to us
    
            region_width_samples = (self.region_width*float(sample_rate))/1E6
            self.region = pg.LinearRegionItem([fl_pos_ind-region_width_samples/2.0, fl_pos_ind+region_width_samples/2.0], bounds=[-20,33000], movable=True)
            self.plot_fl_trace.addItem(self.region)
            def region_changed():
                #delete the current maximum indicator
                self.plot_fl_trace.removeItem(self.peak_dot)
                #where did the user drag the region to?
                new_region = self.region.getRegion()
                #for each fl-trace, search for the maximum in that region
                Fl_max_pos_,Fl_max_ = [],[]
                for i in range(len(Traces_flx)):
                    trace_flx = Traces_flx[i]
                    trace_flx_ = trace_flx[int(new_region[0]):int(new_region[1])]
                    trace_flx_pos_ = range(len(trace_flx))[int(new_region[0]):int(new_region[1])]
                    ind = np.argmax(trace_flx_)
                    fl_max_ = trace_flx_[ind]
                    fl_max_pos_ = trace_flx_pos_[ind]
                    Fl_max_pos_.append(fl_max_pos_)
                    Fl_max_.append(fl_max_)
                #Get the highest maximum across all the traces
                ind = np.argmax(np.array(Fl_max_))
                fl_max__ = Fl_max_[ind]
                fl_max_pos__ = Fl_max_pos_[ind]
                self.new_peak = {"fl_pos":fl_max_pos__,"fl_max":fl_max__,"pos_x":pos_x}
                self.new_peaks.append(self.new_peak)
                self.peak_dot = self.plot_fl_trace.plot([float(fl_max_pos__)], [float(fl_max__)],pen=None,symbol='o',symbolPen='w',clear=False)
    
            self.region.sigRegionChangeFinished.connect(region_changed)

    def onScatterClick(self,event, points):
        pointermethod = 'point'
        if self.changedbyuser:
            self.onClick(points,pointermethod)
    def onIndexChange(self,index):
        pointermethod = 'index'
        if self.changedbyuser:
            self.onClick(index,pointermethod)
        #Set self.changedbyuser to False and change the spinbox and slider. changedbyuser=False prevents onClick function
        self.changedbyuser = False
        self.spinBox_cellInd.setValue(index)
        self.horizontalSlider_cellInd.setValue(index)
        self.changedbyuser = True

    def updateScatterPlot(self):
        #If the Plot is updated, delete the dot in the cell-image
        try:
            self.widget_showCell.removeItem(self.dot)
        except:
            pass
        self.point_was_selected_before = False
        #read url from current comboBox_chooseRtdcFile
        url = str(self.comboBox_chooseRtdcFile.currentText())
        if len(url)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please use the 'Build' tab to load files first")
            msg.setWindowTitle("No file selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        failed,rtdc_ds = aid_bin.load_rtdc(url)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        self.rtdc_ds = rtdc_ds

        feature_x_name = str(self.comboBox_featurex.currentText())
        feature_y_name = str(self.comboBox_featurey.currentText())
        
        features = self.rtdc_ds.features
        if feature_x_name in features:
            self.feature_x = self.rtdc_ds[feature_x_name]
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Feature on x axis is not contained in data set")
            msg.setWindowTitle("Invalid x feature")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        if feature_y_name in features:    
            self.feature_y = self.rtdc_ds[feature_y_name]
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Feature on y axis is not contained in data set")
            msg.setWindowTitle("Invalid y feature")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return            

        self.changedbyuser = True #variable used to prevent plotting if spinbox or slider is changed programmatically
        
        #Add plot  
        self.scatter = self.scatter_xy.plot(np.array(self.feature_x), np.array(self.feature_y),pen=None,symbol='o',symbolPen=None,symbolBrush='b',clear=True)        
        self.scatter.sigPointsClicked.connect(self.onScatterClick) #When scatterplot is clicked, show the desired cell

        #Fill histogram for x-axis; widget_histx
        y,x = np.histogram(self.feature_x, bins='auto')
        self.hist_x.plot(x, y, stepMode=True, fillLevel=0, brush=(0,0,255,150),clear=True)
        #Manually clear y hist first. Only clear=True did not do the job
        self.hist_y.clear()
        #Fill histogram for y-axis; widget_histy
        y,x = np.histogram(self.feature_y, bins='auto')
        curve = pg.PlotCurveItem(-1.*x, y, stepMode=True, fillLevel=0, brush=(0, 0, 255, 150),clear=True)
        curve.rotate(-90)
        self.hist_y.addItem(curve)

        self.scatter_x_norm = (self.feature_x.astype(float))/float(np.max(self.feature_x))
        self.scatter_y_norm = (self.feature_y.astype(float))/float(np.max(self.feature_y))

        #Adjust the horizontalSlider_cellInd and spinBox_cellInd
        self.horizontalSlider_cellInd.setSingleStep(1)
        self.horizontalSlider_cellInd.setMinimum(0)
        self.horizontalSlider_cellInd.setMaximum(len(self.feature_x)-1)
        self.spinBox_cellInd.setMinimum(0)
        self.spinBox_cellInd.setMaximum(len(self.feature_x)-1)                  

    def selectPeakPos(self):
        #Check if self.region exists
        #If not, show a message and return:
        if not hasattr(self, 'region'):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no region defined yet")
            msg.setWindowTitle("No region defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return   
        #Try to get the user defined peak position
        if not hasattr(self, 'new_peak'):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no peak defined yet")
            msg.setWindowTitle("No peak defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
                
        #how much rows are already in table?
        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        self.tableWidget_showSelectedPeaks.setRowCount(rowcount+1)
        rowPosition = rowcount        
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["fl_max"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 0, item)
        item = QtWidgets.QTableWidgetItem()
        fl_pos_us = float(float(self.new_peak["fl_pos"])*float(1E6))/float(self.rtdc_ds.config["fluorescence"]["sample rate"])
        item.setData(QtCore.Qt.EditRole,fl_pos_us)
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 1, item)
        item = QtWidgets.QTableWidgetItem()
        pos_x_um = float(self.new_peak["pos_x"])*float(self.rtdc_ds.config["imaging"]["pixel size"])
        item.setData(QtCore.Qt.EditRole,pos_x_um)
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["fl_pos"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.new_peak["pos_x"]))
        self.tableWidget_showSelectedPeaks.setItem(rowPosition, 4, item)
        self.tableWidget_showSelectedPeaks.resizeColumnsToContents()            
        self.tableWidget_showSelectedPeaks.resizeRowsToContents()

        #Update the widget_showSelectedPeaks
        self.update_peak_plot()

    def selectPeakRange(self):
        new_region = self.region.getRegion()
        region_width = np.max(new_region) - np.min(new_region) #in [samples]
        sample_rate = self.rtdc_ds.config["fluorescence"]["sample rate"]
        region_width = (float(region_width)/float(sample_rate))*1E6 #range[samples]*(1/sample_rate[1/s]) = range[s]; div by 1E6 to conver to us
        self.region_width = region_width
        #put this in the table
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, "Range [us]")
        self.tableWidget_peakModelParameters.setItem(1, 0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(self.region_width))
        self.tableWidget_peakModelParameters.setItem(1, 1, item)
        item = QtWidgets.QTableWidgetItem()
        
    def onPeaksPlotClick(self,event, points):
        points = points[0]
        p = points.pos()
        clicked_x, clicked_y = p.x(), p.y()
        a1 = (clicked_x)/float(np.max(self.Pos_x))            
        a2 = (clicked_y)/float(np.max(self.Fl_pos))
        
        #Which is the closest scatter point?
        pos_x_norm = self.Pos_x/np.max(self.Pos_x)#normalized pos_x
        fl_pos_norm = self.Fl_pos/np.max(self.Fl_pos)#normalized fl_pos
        dist = np.sqrt(( a1-pos_x_norm )**2 + ( a2-fl_pos_norm )**2)
        index =  np.argmin(dist)
        #Highlight this row
        self.tableWidget_showSelectedPeaks.selectRow(index)
        #Delete the highlighted rows
#        try:
#            self.actionRemoveSelectedPeaks_function()
#        except:
#            pass
        
    def update_peak_plot(self):
        #This function reads tableWidget_showSelectedPeaks and 
        #fits a function and 
        #puts fitting parameters on tableWidget_peakModelParameters
        
        #read the data on tableWidget_showSelectedPeaks        
        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        Fl_pos,Pos_x = [],[]
        for row in range(rowcount):
            line = [float(self.tableWidget_showSelectedPeaks.item(row, col).text()) for col in [1,2]] #use the values for [us] and [um]
            Fl_pos.append(line[0])
            Pos_x.append(line[1])
            
        self.Fl_pos = np.array(Fl_pos)
        self.Pos_x = np.array(Pos_x)
        
        self.selectedPeaksPlotPlot = self.selectedPeaksPlot.plot(self.Pos_x, self.Fl_pos,pen=None,symbol='o',symbolPen=None,symbolBrush='b',clear=True)
        #if user clicks in the plot, show him the corresponding row in the table
        self.selectedPeaksPlotPlot.sigPointsClicked.connect(self.onPeaksPlotClick)

        if not hasattr(self, 'region_width'): #if there was no region_width defined yet...
            #to get a reasonable initial range, use 20% of the nr. of availeble samples
            samples_per_event = self.rtdc_ds.config["fluorescence"]["samples per event"]
            self.region_width = 0.2*samples_per_event #width of the region in samples
            #Convert to SI unit:
            sample_rate = self.rtdc_ds.config["fluorescence"]["sample rate"]
            self.region_width = (float(self.region_width)/float(sample_rate))*1E6 #range[samples]*(1/sample_rate[1/s]) = range[s]; div by 1E6 to convert to us

        #which model should be used?
        if str(self.comboBox_peakDetModel.currentText()) == "Linear dependency and max in range" and len(Pos_x)>1:
            slope,intercept = np.polyfit(Pos_x, Fl_pos,deg=1) #Linear FIT, y=mx+n; y=FL_pos[us] x=Pos_x[um]
            xlin = np.round(np.linspace(np.min(Pos_x),np.max(Pos_x),25),1)
            ylin = intercept + slope*xlin
            self.selectedPeaksPlot.plot(xlin, ylin,width=6,pen='b',clear=False)
            
            #Put info to tableWidget_peakModelParameters
            self.tableWidget_peakModelParameters.setColumnCount(2)
            self.tableWidget_peakModelParameters.setRowCount(5)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Model")
            self.tableWidget_peakModelParameters.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Linear dependency and max in range")
            self.tableWidget_peakModelParameters.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Range [us]")
            self.tableWidget_peakModelParameters.setItem(1, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(self.region_width))
            self.tableWidget_peakModelParameters.setItem(1, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Intercept [us]")
            self.tableWidget_peakModelParameters.setItem(2, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(intercept))
            self.tableWidget_peakModelParameters.setItem(2, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Slope [us/um]")
            self.tableWidget_peakModelParameters.setItem(3, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(slope))
            self.tableWidget_peakModelParameters.setItem(3, 1, item)
            #Calculate velocity
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Velocity[m/s]")
            self.tableWidget_peakModelParameters.setItem(4, 0, item)
            velocity = float(1.0/float(slope))
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(velocity))
            self.tableWidget_peakModelParameters.setItem(4, 1, item)

            
    def addHighestXPctPeaks(self):
        #how many x%?
        x_pct = float(self.doubleSpinBox_highestXPercent.value())
        #Get the flourescence traces and maxima/positions of maxima
        #->it could be that the user did not yet load the dataset:
        if not hasattr(self,"rtdc_ds"):
            #run the function updateScatterPlot()
            self.updateScatterPlot()
            
        trace = self.rtdc_ds["trace"]
        fl_keys = list(trace.keys())
        fl1_max,fl1_pos,fl2_max,fl2_pos,fl3_max,fl3_pos,pos_x = [],[],[],[],[],[],[]
        for i in range(len(fl_keys)):
            if "fl1_median" in fl_keys[i] and self.checkBox_fl1.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl1_max.append(trace_flx[ind])
                    fl1_pos.append(ind)
                #Get the x% maxima
                fl1_max = np.array(fl1_max)
                fl1_pos = np.array(fl1_pos)
                sorter = np.argsort(fl1_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl1_max))]
                fl1_max = fl1_max[sorter]
                fl1_pos = fl1_pos[sorter]
                pos_x.append(self.rtdc_ds["pos_x"][sorter])
                
            elif "fl2_median" in fl_keys[i] and self.checkBox_fl2.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl2_max.append(trace_flx[ind])
                    fl2_pos.append(ind)
                #Get the x% maxima
                fl2_max = np.array(fl2_max)
                fl2_pos = np.array(fl2_pos)
                sorter = np.argsort(fl2_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl2_max))]
                fl2_max = fl2_max[sorter]
                fl2_pos = fl2_pos[sorter]
                pos_x.append(self.rtdc_ds["pos_x"][sorter])

            elif "fl3_median" in fl_keys[i] and self.checkBox_fl3.isChecked():
                for index in range(len(trace[fl_keys[i]])):
                    trace_flx = trace[fl_keys[i]][index]
                    ind = np.argmax(trace_flx)
                    fl3_max.append(trace_flx[ind])
                    fl3_pos.append(ind)
                #Get the x% maxima
                fl3_max = np.array(fl3_max)
                fl3_pos = np.array(fl3_pos)
                sorter = np.argsort(fl3_max)[::-1]
                sorter = sorter[0:int(x_pct/100.0*len(fl3_max))]
                fl3_max = fl3_max[sorter]
                fl3_pos = fl3_pos[sorter]
                pos_x.append(self.rtdc_ds["pos_x"][sorter])

        #Add fl1 fl2 and fl3 information
        flx_max = np.array(list(fl1_max)+list(fl2_max)+list(fl3_max))
        flx_pos = np.array(list(fl1_pos)+list(fl2_pos)+list(fl3_pos))
        pos_x_um = np.concatenate(np.atleast_2d(np.array(pos_x)))
        pix = self.rtdc_ds.config["imaging"]["pixel size"]
        pos_x = pos_x_um/pix #convert from um to pix

        rowcount = self.tableWidget_showSelectedPeaks.rowCount()
        self.tableWidget_showSelectedPeaks.setRowCount(rowcount+len(flx_max))
        
        for i in range(len(flx_max)):
            rowPosition = rowcount+i
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(flx_max[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 0, item)
            item = QtWidgets.QTableWidgetItem()
            fl_pos_us = float(float(flx_pos[i])*float(1E6))/float(self.rtdc_ds.config["fluorescence"]["sample rate"])
            item.setData(QtCore.Qt.EditRole,fl_pos_us)
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 1, item)
            item = QtWidgets.QTableWidgetItem()
            #pos_x_um = float(pos_x[i])*float(self.rtdc_ds.config["imaging"]["pixel size"])
            item.setData(QtCore.Qt.EditRole,float(pos_x_um[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(flx_pos[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 3, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(pos_x[i]))
            self.tableWidget_showSelectedPeaks.setItem(rowPosition, 4, item)
        #Update the widget_showSelectedPeaks
        self.update_peak_plot()



    def savePeakDetModel(self):
        #Get tableWidget_peakModelParameters and write it to excel file
        #Get filename from user:
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save peak fitting model', Default_dict["Path of last model"],"Excel file (*.xlsx)")
        filename = filename[0]
        if len(filename)==0:
            return
        #add the suffix .csv
        if not filename.endswith(".xlsx"):
            filename = filename +".xlsx"               

        table = self.tableWidget_peakModelParameters
        cols = table.columnCount()
        header = range(cols)
        rows = table.rowCount()
        model_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    model_df.iloc[i, j] = table.item(i, j).text()
                except:
                    model_df.iloc[i, j] = np.nan

        table = self.tableWidget_showSelectedPeaks
        cols = table.columnCount()
        header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        rows = table.rowCount()
        peaks_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    peaks_df.iloc[i, j] = table.item(i, j).text()
                except:
                    peaks_df.iloc[i, j] = np.nan

        writer = pd.ExcelWriter(filename, engine='openpyxl')
        #Used files go to a separate sheet on the MetaFile.xlsx
        pd.DataFrame().to_excel(writer,sheet_name='Model') #initialize empty Sheet
        model_df.to_excel(writer,sheet_name='Model') #initialize empty Sheet
        pd.DataFrame().to_excel(writer,sheet_name='Peaks') #initialize empty Sheet
        peaks_df.to_excel(writer,sheet_name='Peaks')
        writer.save()
        writer.close()

    def loadPeakDetModel(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open peak fitting model', Default_dict["Path of last model"],"Excel file (*.xlsx)")
        filename = filename[0]
        if len(str(filename))==0:
            return
        peak_model_df = pd.read_excel(filename,sheetname='Model')
        model = peak_model_df.iloc[0,1]
        if model=="Linear dependency and max in range":
            #set the combobox accordingly
            index = self.comboBox_peakDetModel.findText(model, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.comboBox_peakDetModel.setCurrentIndex(index)
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Could not find a valid model in the chosen file. Did you accidentially load a session or history file?!")
                msg.setWindowTitle("No valid model found")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return                

            range_ = float(peak_model_df.iloc[1,1])
            intercept = float(peak_model_df.iloc[2,1])
            slope = float(peak_model_df.iloc[3,1])
            velocity = float(peak_model_df.iloc[4,1])

            #put the information in the table
            xlin = np.round(np.linspace(np.min(0),np.max(100),25),1)
            ylin = intercept + slope*xlin
            self.selectedPeaksPlot.plot(xlin, ylin,width=6,pen='b',clear=False)
            
            #Put info to tableWidget_peakModelParameters
            self.tableWidget_peakModelParameters.setColumnCount(2)
            self.tableWidget_peakModelParameters.setRowCount(5)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Model")
            self.tableWidget_peakModelParameters.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Linear dependency and max in range")
            self.tableWidget_peakModelParameters.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Range [us]")
            self.tableWidget_peakModelParameters.setItem(1, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(range_))
            self.tableWidget_peakModelParameters.setItem(1, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Intercept [us]")
            self.tableWidget_peakModelParameters.setItem(2, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(intercept))
            self.tableWidget_peakModelParameters.setItem(2, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Slope [us/um]")
            self.tableWidget_peakModelParameters.setItem(3, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(slope))
            self.tableWidget_peakModelParameters.setItem(3, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, "Velocity[m/s]")
            self.tableWidget_peakModelParameters.setItem(4, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, float(velocity))
            self.tableWidget_peakModelParameters.setItem(4, 1, item)


    def applyPeakModel_and_export(self):
        #On which files should the action be performed?
        Files = []
        if self.radioButton_exportAll.isChecked():
            #Grab all items of comboBox_chooseRtdcFile
            Files = [self.comboBox_chooseRtdcFile.itemText(i) for i in range(self.comboBox_chooseRtdcFile.count())]
        else:
            file = self.comboBox_chooseRtdcFile.currentText()
            Files.append(str(file))
        
        #Get the model from tableWidget_peakModelParameters
        table = self.tableWidget_peakModelParameters
        cols = table.columnCount()
        header = range(cols)
        rows = table.rowCount()
        model_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    model_df.iloc[i, j] = table.item(i, j).text()
                except:
                    model_df.iloc[i, j] = np.nan
        model = model_df.iloc[0,1]
        if model == "Linear dependency and max in range":
            range_us = float(model_df.iloc[1,1]) #[us]
            intercept_us = float(model_df.iloc[2,1]) 
            slope_us_um = float(model_df.iloc[3,1])
            #velocity_m_s = float(model_df.iloc[4,1])

        #Get a directory from the user!
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, 'Select directory', Default_dict["Path of last model"])
        if len(folder)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Invalid directory")
            msg.setWindowTitle("Invalid directory")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        for rtdc_path in Files:
            path, rtdc_file = os.path.split(rtdc_path)
            savename = os.path.join(folder,rtdc_file)
            
            #Avoid to save to an existing file:
            addon = 1
            while os.path.isfile(savename):
                savename = savename.split(".rtdc")[0]
                if addon>1:
                    savename = savename.split("_"+str(addon-1))[0]
                savename = savename+"_"+str(addon)+".rtdc"
                addon += 1
            print(savename)

            failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
            if failed:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)       
                msg.setText(str(rtdc_ds))
                msg.setWindowTitle("Error occurred during loading file")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            #Convert quantities to [index]
            sample_rate = rtdc_ds.config["fluorescence"]["sample rate"]
            range_ = (range_us*float(sample_rate))/1E6 #range was given in us->Divide by 1E6 to get to s and then multiply by the sample rate
            
#            #check if an rtdc_ds is already chosen:
#            if not hasattr(self,'rtdc_ds'):
#                msg = QtWidgets.QMessageBox()
#                msg.setIcon(QtWidgets.QMessageBox.Information)       
#                msg.setText("No measurement chosen yet. Use 'Update' button")
#                msg.setWindowTitle("No measurement")
#                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                msg.exec_()
#                return
                                
            trace = rtdc_ds["trace"]
            fl_keys = list(trace.keys()) #Which traces are available
            fl1_max,fl1_pos,fl2_max,fl2_pos,fl3_max,fl3_pos,pos_x = [],[],[],[],[],[],[]
            
            #Iterate over the available cells
            pos_x = rtdc_ds["pos_x"] #is already given in [um]
            indices = range(len(pos_x))
    
            if model == "Linear dependency and max in range":
                #Use the linear model to get the estimated location of the fluorescence peaks
                fl_peak_position_us = intercept_us+slope_us_um*pos_x
                #Convert to index
                fl_peak_position_ = (fl_peak_position_us*float(sample_rate))/1E6
                #Now we have the estimated peak position of each cell. Look at the traces on these spots            

            def ind_to_us(x):
                return x*1E6/sample_rate
    
            #iterate over the cells:
            for cellindex in range(len(pos_x)):
                #Iterate over the availble traces
                for i in range(len(fl_keys)):
                    if "_median" in fl_keys[i]:
                        trace_flx = trace[fl_keys[i]][cellindex]
                        trace_pos = np.array(range(len(trace_flx)))
                        left = int(fl_peak_position_[cellindex]-range_/2.0)
                        right = int(fl_peak_position_[cellindex]+range_/2.0)
                        trace_flx_range = trace_flx[left:right] 
                        trace_pos_range = trace_pos[left:right] 
                        ind = np.argmax(trace_flx_range)
                        if "fl1_median" in fl_keys[i]:
                            fl1_max.append(trace_flx_range[ind])
                            fl1_pos.append(ind_to_us(trace_pos_range[ind]))
                        if "fl2_median" in fl_keys[i]:
                            fl2_max.append(trace_flx_range[ind])
                            fl2_pos.append(ind_to_us(trace_pos_range[ind]))
                        if "fl3_median" in fl_keys[i]:
                            fl3_max.append(trace_flx_range[ind])
                            fl3_pos.append(ind_to_us(trace_pos_range[ind]))
              
            #Save those new fluorescence features into free spots in .rtdc file
            #Those names can be found via dclab.dfn.feature_names called (userdef0...userdef9)
            
            #get metadata of the dataset
            meta = {}
            # only export configuration meta data (no user-defined config)
            for sec in dclab.definitions.CFG_METADATA:
                if sec in ["fmt_tdms"]:
                    # ignored sections
                    continue
                if sec in rtdc_ds.config:
                    meta[sec] = rtdc_ds.config[sec].copy()
            
            #features = rtdc_ds._events.keys() #Get the names of the online features
            compression = 'gzip'    
            nev = len(rtdc_ds)
            
            #["Overwrite Fl_max and Fl_pos","Save to userdef"]
            features = rtdc_ds._events.keys()
            if str(self.comboBox_toFlOrUserdef.currentText())=='Save to userdef':
                features = features+["userdef"+str(i) for i in range(10)]
                
            with dclab.rtdc_dataset.write_hdf5.write(path_or_h5file=savename,meta=meta, mode="append") as h5obj:
                # write each feature individually
                for feat in features:
                    print(feat)
                    # event-wise, because
                    # - tdms-based datasets don't allow indexing with numpy
                    # - there might be memory issues
                    if feat == "contour":
                        cont_list = [rtdc_ds["contour"][ii] for ii in indices]
                        dclab.rtdc_dataset.write_hdf5.write(h5obj,
                              data={"contour": cont_list},
                              mode="append",
                              compression=compression)
                    elif feat == "userdef0":
                        if "fl1_median" in fl_keys: 
                            print("writing fl1_max to userdef0")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef0": np.array(fl1_max)},
                                  mode="append",
                                  compression=compression)
                    elif feat == "userdef1":
                        if "fl2_median" in fl_keys: 
                            print("writing fl2_max to userdef1")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef1": np.array(fl2_max)},
                                  mode="append",
                                  compression=compression)
                    elif feat == "userdef2":
                        if "fl3_median" in fl_keys:     
                            print("writing fl3_max to userdef2")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef2": np.array(fl3_max)},
                                  mode="append",
                                  compression=compression)
                            
                    elif feat == "userdef3":
                        if "fl1_pos" in features: 
                            print("writing fl1_pos to userdef3")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef3": np.array(fl1_pos)},
                                  mode="append",
                                  compression=compression)
                    elif feat == "userdef4":
                        if "fl2_pos" in features: 
                            print("writing fl2_pos to userdef4")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef4": np.array(fl2_pos)},
                                  mode="append",
                                  compression=compression)
                    elif feat == "userdef5":
                        if "fl3_pos" in features:     
                            print("writing fl3_pos to userdef5")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"userdef5": np.array(fl3_pos)},
                                  mode="append",
                                  compression=compression)
                               
                    elif feat in ["userdef"+str(i) for i in range(5,10)]:
                        pass

                    elif feat == "fl1_max":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl1_max")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl1_max)},mode="append",compression=compression)
                    elif feat == "fl2_max":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl2_max")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl2_max)},mode="append",compression=compression)
                    elif feat == "fl3_max":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl3_max")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl3_max)},mode="append",compression=compression)

                    elif feat == "fl1_pos":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl1_pos")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl1_pos)},mode="append",compression=compression)
                    elif feat == "fl2_pos":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl2_pos")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl2_pos)},mode="append",compression=compression)
                    elif feat == "fl3_pos":                    
                        if str(self.comboBox_toFlOrUserdef.currentText())=='Overwrite Fl_max and Fl_pos':
                            print("overwriting fl3_pos")
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: np.array(fl3_pos)},mode="append",compression=compression)
                            
                    elif feat == "index":
                        dclab.rtdc_dataset.write_hdf5.write(h5obj,
                              data={"index": np.array(indices)+1}, #ShapeOut likes to start with index=1
                              mode="append",
                              compression=compression)
            
                    elif feat in ["mask", "image"]:
                        # store image stacks (reduced file size and save time)
                        m = 64
                        if feat=='mask':
                            im0 = rtdc_ds[feat][0]
                        if feat=="image":
                            im0 = rtdc_ds[feat][0]
                        imstack = np.zeros((m, im0.shape[0], im0.shape[1]),
                                           dtype=im0.dtype)
                        jj = 0
                        if feat=='mask':
                            image_list = [rtdc_ds[feat][ii] for ii in indices]
                        elif feat=='image':
                            image_list = [rtdc_ds[feat][ii] for ii in indices]
                        for ii in range(len(image_list)):
                            dat = image_list[ii]
                            #dat = rtdc_ds[feat][ii]
                            imstack[jj] = dat
                            if (jj + 1) % m == 0:
                                jj = 0
                                dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                      data={feat: imstack},
                                      mode="append",
                                      compression=compression)
                            else:
                                jj += 1
                        # write rest
                        if jj:
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={feat: imstack[:jj, :, :]},
                                  mode="append",
                                  compression=compression)
                    elif feat == "trace":
                        for tr in rtdc_ds["trace"].keys():
                            tr0 = rtdc_ds["trace"][tr][0]
                            trdat = np.zeros((nev, tr0.size), dtype=tr0.dtype)
                            jj = 0
                            trace_list = [rtdc_ds["trace"][tr][ii] for ii in indices]
                            for ii in range(len(trace_list)):
                                trdat[jj] = trace_list[ii]
                                jj += 1
                            dclab.rtdc_dataset.write_hdf5.write(h5obj,
                                  data={"trace": {tr: trdat}},
                                  mode="append",
                                  compression=compression)
                    else:
                        dclab.rtdc_dataset.write_hdf5.write(h5obj,
                              data={feat: rtdc_ds[feat][indices]},mode="append")
                    
                h5obj.close()











    def partialtrainability_activated(self,on_or_off):
        if on_or_off==False:#0 means switched OFF 
            
            self.lineEdit_partialTrainability.setText("")
            #self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(False)
            #Also, remove the model from self!
            self.model_keras = None

            self.radioButton_NewModel.setChecked(False)
            self.radioButton_LoadRestartModel.setChecked(False)
            self.radioButton_LoadContinueModel.setChecked(False)
            self.lineEdit_LoadModelPath.setText("")#put the filename in the lineedit
            
        #this happens when the user activated the expert option "partial trainability"
        elif on_or_off==True:#2 means switched ON
            #Has the user already chosen a model?
            if self.model_keras == None: #if there is no model yet chosen
                self.action_initialize_model(only_initialize=True)
            #If there is still no model...
            if self.model_keras == None:# or self.model_keras_path==None: #if there is no model yet chosen
                #Tell the user to initiate a model first!
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("<html><head/><body><p>To use this option please first select and load a model. To do that choose/load a model in 'Define Model'-Tab and hit the button 'Initialize/Fit Model'. Choose to only initialize the model.</p></body></html>")
                msg.setWindowTitle("Please load a model first")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                
                #Switch off
                self.lineEdit_partialTrainability.setText("")
                self.radioButton_NewModel.setChecked(False)
                self.radioButton_LoadRestartModel.setChecked(False)
                self.radioButton_LoadContinueModel.setChecked(False)
                self.lineEdit_LoadModelPath.setText("")
                #self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
                self.pushButton_partialTrainability.setEnabled(False)
                self.checkBox_partialTrainability.setChecked(False)
                return

            #Otherwise, there is a model on self and we can continue :)
            
            #Collections are not supported
            if type(self.model_keras)==tuple:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("<html><head/><body><p>Partial trainability is not available for collections of models. Please specify a single model.</p></body></html>")
                msg.setWindowTitle("Collections of models not supported for collections of models")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
            
            #Switch on lineedit and the button
            #self.lineEdit_partialTrainability.setEnabled(True)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(True)#enables the lineEdit which shows the trainability status of each layer.

            #Load trainability states of the model           
            Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
            #Count Dense and Conv layers
            is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
            index = np.where(np.array(is_dense_or_conv)==True)[0]
            Layer_train_status = [self.model_keras.layers[layerindex].trainable for layerindex in index]
            self.lineEdit_partialTrainability.setText(str(Layer_train_status))#enables the lineEdit which shows the trainability status of each layer.



    def partialTrainability(self):
        self.popup_trainability = MyPopup()
        self.popup_trainability_ui = aid_frontend.popup_trainability()
        self.popup_trainability_ui.setupUi(self.popup_trainability) #open a popup to show the layers in a table

        #One can only activate this function when there was a model loaded already!
        #self.model_keras has to exist!!!

        if self.model_keras == None: #if there is no model yet chosen
            self.action_initialize_model()

        if self.model_keras == None: #if there is still no model...            
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("<html><head/><body><p>To use this option please first select and load a model. To do that choose/load a model in 'Define Model'-Tab and hit the button 'Initialize/Fit Model'. Choose to only initialize the model.</p></body></html>")
            msg.setWindowTitle("Please load a model first")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            #Switch this On in the final version
            self.lineEdit_partialTrainability.setText("")
            self.lineEdit_partialTrainability.setEnabled(False)#enables the lineEdit which shows the trainability status of each layer.
            self.pushButton_partialTrainability.setEnabled(False)
            return

        #Fill information about the model 
        if self.radioButton_NewModel.isChecked():#a new model is loaded          
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("New model")
        elif self.radioButton_LoadRestartModel.isChecked():#a new model is loaded          
            load_model_path =  str(self.lineEdit_LoadModelPath.text())
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("Restart model: "+load_model_path)
        elif self.radioButton_LoadContinueModel.isChecked():#a new model is loaded          
            load_model_path =  str(self.lineEdit_LoadModelPath.text())
            self.popup_trainability_ui.lineEdit_pop_pTr_modelPath.setText("Continue model: "+load_model_path)

        in_dim = self.model_keras.input_shape
        #Retrieve the color_mode from the model (nr. of channels in last in_dim)
        channels = in_dim[-1] #TensorFlow: channels in last dimension
        out_dim = self.model_keras.output_shape[-1]
        
        self.popup_trainability_ui.spinBox_pop_pTr_inpSize.setValue(int(in_dim[1]))
        self.popup_trainability_ui.spinBox_pop_pTr_outpSize.setValue(int(out_dim))
        if channels==1:
            self.popup_trainability_ui.comboBox_pop_pTr_colorMode.addItem("Grayscale")
        elif channels==3:
            self.popup_trainability_ui.comboBox_pop_pTr_colorMode.addItem("RGB")
            
        #Model summary to textBrowser_pop_pTr_modelSummary
        summary = []
        self.model_keras.summary(print_fn=summary.append)
        summary = "\n".join(summary)
        self.popup_trainability_ui.textBrowser_pop_pTr_modelSummary.setText(summary)
                
        #Work on the tableWidget_pop_pTr_layersTable
        Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
        #Count Dense and Conv layers
        is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
        index = np.where(np.array(is_dense_or_conv)==True)[0]
        nr_layers = len(index) #total nr. of dense and conv layers with parameters

        for rowNumber in range(nr_layers):
            layerindex = index[rowNumber]
            columnPosition = 0
            layer = self.model_keras.layers[layerindex]
            rowPosition = self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.rowCount()
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.insertRow(rowPosition)
            Name = layer.name
            item = QtWidgets.QTableWidgetItem(Name) 
            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            item.setTextAlignment(QtCore.Qt.AlignCenter) # change the alignment
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 1
            layer_type = layer.__class__.__name__
            item = QtWidgets.QTableWidgetItem(layer_type)             
            item.setFlags( QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            item.setTextAlignment(QtCore.Qt.AlignCenter) # change the alignment
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition , columnPosition, item ) #

            columnPosition = 2
            Params = layer.count_params()
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, Params)
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition, columnPosition, item)

            columnPosition = 3
            if layer_type == "Dense":
                split_property = "units" #'units' are the number of nodes in dense layers
            elif layer_type == "Conv2D":
                split_property = "filters"
            else:
                print("other splitprop!")
                return
            layer_config = layer.get_config()
            nr_units = layer_config[split_property] #units are either nodes or filters for dense and convolutional layer, respectively
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, int(nr_units))
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setItem(rowPosition, columnPosition, item)

            columnPosition = 4
            #for each item create a spinbopx (trainability)
            spinb = QtWidgets.QDoubleSpinBox(self.popup_trainability_ui.tableWidget_pop_pTr_layersTable)
            spinb.setMinimum(0)
            spinb.setMaximum(1)
            spinb.setSingleStep(0.1)
            trainability = int(layer.trainable) #.trainable actually returns True or False. Make it integer
            spinb.setValue(trainability) #this should be always 1
            self.popup_trainability_ui.tableWidget_pop_pTr_layersTable.setCellWidget(rowPosition, columnPosition, spinb)            

        self.popup_trainability.show()
        
        #self.popup_trainability_ui.pushButton_pop_pTr_reset.clicked.connect(self.pop_pTr_reset)
        self.popup_trainability_ui.pushButton_pop_pTr_update.clicked.connect(self.pop_pTr_update_2)
        self.popup_trainability_ui.pushButton_pop_pTr_ok.clicked.connect(self.pop_pTr_ok)






    ###############Functions for the partial trainability popup################

    def pop_pTr_reset(self):
        #Reset the model to initial state, with partial trainability
        print("Not implemented yet")
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("<html><head/><body><p>Not implemented yet.</p></body></html>")
        msg.setWindowTitle("Not implemented")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def pop_pTr_update_1(self):#main worker function
        #Apply the requested changes and display updated model in table
        pTr_table = self.popup_trainability_ui.tableWidget_pop_pTr_layersTable

        #Read the table: 
        Layer_names,Layer_trainabilities = [],[]
        rowCount = pTr_table.rowCount()
        for row in range(rowCount):
            #Layer_indices.append(str(pTr_table.item(row, 0).text()))
            Layer_names.append(str(pTr_table.item(row, 0).text()))
            Layer_trainabilities.append(float(pTr_table.cellWidget(row, 4).value()))
        Layer_trainabilities = np.array(Layer_trainabilities)
        
        #What are the current trainability statuses of the model
        Layer_trainabilities_orig = np.array([self.model_keras.get_layer(l_name).trainable for l_name in Layer_names])
        diff = abs( Layer_trainabilities - Layer_trainabilities_orig )
        ind = np.where( diff>0 )[0]
        
        #Where do we have a trainability between 0 and 1
        #ind = np.where( (Layer_trainabilities>0) & (Layer_trainabilities<1) )[0]
        if len(ind)>0:
            Layer_trainabilities = list(Layer_trainabilities[ind])
            Layer_names = list(np.array(Layer_names)[ind])
            #Update the model using user-specified trainabilities
            self.model_keras = partial_trainability(self.model_keras,Layer_names,Layer_trainabilities)

            #Update lineEdit_partialTrainability
            Layer_types = [self.model_keras.layers[i].__class__.__name__ for i in range(len(self.model_keras.layers))]
            #Count Dense and Conv layers
            is_dense_or_conv = [layer_type in ["Dense","Conv2D"] for layer_type in Layer_types]  
            index = np.where(np.array(is_dense_or_conv)==True)[0]
            Layer_train_status = [self.model_keras.layers[layerindex].trainable for layerindex in index]
            self.lineEdit_partialTrainability.setText(str(Layer_train_status))#enables the lineEdit which shows the trainability status of each layer.
        else:
            print("Nothing to do. All trainabilities are either 0 or 1")

    def pop_pTr_update_2(self):#call pop_pTr_update_1 to the the work and then update the window
        try:
            self.pop_pTr_update_1()#Change the model on self.model_keras according to the table
            self.partialTrainability()#Update the popup window by calling the partialTrainability function
        except Exception as e: 
            #There is an issue building the model!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(e))
            msg.setWindowTitle("Error occured when building model:")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return       
        
    def pop_pTr_ok(self):
        self.pop_pTr_update_1()#Change the model on self.model_keras according to the table; If 'Update' was used before, there will not be done work again, but the model is used as it is
        #To make the model accessible, it has to be saved to a new .model file
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save model', Default_dict["Path of last model"],"AIDeveloper model file (*.model)")
        filename = filename[0]
        path, fname = os.path.split(filename)
        if len(fname)==0:
            return
        #add the suffix _session.xlsx
        if not fname.endswith(".model"):
            fname = fname +".model"
        filename = os.path.join(path,fname)
        
        self.model_keras.save(filename)
        #Activate 'load and restart' and put this file
        #Avoid the automatic popup
        self.radioButton_NewModel.setChecked(False)
        self.radioButton_LoadRestartModel.setChecked(False)
        self.radioButton_LoadContinueModel.setChecked(True)

        self.lineEdit_LoadModelPath.setText(filename)#put the filename in the lineedit
        #Destroy the window
        self.popup_trainability = None

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("<html><head/><body><p>The model was successfully saved. 'Load and continue' was automatically selected  in 'Define Model'-Tab, so your model with partial trainability will be loaded when you start fitting. The model architecture is documented in each .model file and the .arch file. Change of partial trainability during training is not supported yet (but it is theoretically no problem to implement it).</p></body></html>")
        msg.setWindowTitle("Sucessfully created and selected model")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        


    def lossW_comboB(self,state_nr,listindex):
        if listindex==-1:
            ui_item = self.popup_lossW_ui
        else:
            ui_item = self.fittingpopups_ui[listindex].popup_lossW_ui

        state_str = ui_item.comboBox_lossW.itemText(int(state_nr))
        rows_nr = int(ui_item.tableWidget_lossW.rowCount())
        if rows_nr==0:
            state_str = "None"
            
        if state_str=="None":
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(False)
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setValue(1.0)
            
        elif state_str=="Custom":
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(True)

        elif state_str=="Balanced":
            #How many cells in total per epoch
            events_epoch = [int(ui_item.tableWidget_lossW.item(rowPos,2).text()) for rowPos in range(rows_nr)]
            classes = [int(ui_item.tableWidget_lossW.item(rowPos,0).text()) for rowPos in range(rows_nr)]
            counter = {}
            for i in range(len(classes)):
                counter[classes[i]]=events_epoch[i]
                
            max_val = float(max(counter.values()))       
            class_weights = {class_id : max_val/num_images for class_id, num_images in counter.items()}                     
            class_weights = list(class_weights.values())
            for rowPos in range(rows_nr):
                colPos = 4 #"Loss weights"
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setEnabled(False)
                ui_item.tableWidget_lossW.cellWidget(rowPos,colPos).setValue(class_weights[rowPos])

    def lossW_ok(self,listindex):
        #This happens when the user presses the OK button on the popup for 
        #custom loss weights
        if listindex==-1:
            ui_item = self
        else:
            ui_item = self.fittingpopups_ui[listindex]
        
        #Which option was used on comboBox_lossW?
        state_str = ui_item.popup_lossW_ui.comboBox_lossW.currentText()
        if state_str=="None":#User left None. This actually means its off
            ui_item.lineEdit_lossW.setText("")
            ui_item.pushButton_lossW.setEnabled(False)
            ui_item.checkBox_lossW.setChecked(False)

        elif state_str=="Custom":#User left None. This actually means its off
            #There are custom values         
            #Read the loss values on the table
            rows_nr = int(ui_item.popup_lossW_ui.tableWidget_lossW.rowCount())
            classes = [int(ui_item.popup_lossW_ui.tableWidget_lossW.item(rowPos,0).text()) for rowPos in range(rows_nr)]
            loss_weights = [float(ui_item.popup_lossW_ui.tableWidget_lossW.cellWidget(rowPos,4).value()) for rowPos in range(rows_nr)]
            counter = {}
            for i in range(len(classes)):
                counter[classes[i]]=loss_weights[i]
            #Put counter (its a dictionary) to lineedit
            ui_item.lineEdit_lossW.setText(str(counter))
        
        elif state_str=="Balanced":#Balanced, the values are computed later fresh, even when user changes the cell-numbers again
            ui_item.lineEdit_lossW.setText("Balanced")

        #Destroy the window        
        ui_item.popup_lossW = None
        
    def lossW_cancel(self,listindex):
        #This happens when the user presses the Cancel button on the popup for 
        #custom loss weights       
        if listindex==-1:
            ui_item = self
        else:
            ui_item = self.fittingpopups_ui[listindex]

        if ui_item.lineEdit_lossW.text()=="":
        #if state_str=="None":#User left None. This actually means its off
            ui_item.lineEdit_lossW.setText("")
            ui_item.pushButton_lossW.setEnabled(False)
            ui_item.checkBox_lossW.setChecked(False)
            ui_item.popup_lossW = None
            return
        #Destroy the window
        ui_item.popup_lossW = None



    def get_norm_from_manualselection(self):
        norm = self.comboBox_w.currentText()
        index = self.comboBox_Normalization.findText(norm, QtCore.Qt.MatchFixedString)
        if index >= 0:
            self.comboBox_Normalization.setCurrentIndex(index)
            self.w.close()
    
    def popup_normalization(self):
            self.w = MyPopup()
            self.gridLayout_w = QtWidgets.QGridLayout(self.w)
            self.gridLayout_w.setObjectName(_fromUtf8("gridLayout"))
            self.verticalLayout_w = QtWidgets.QVBoxLayout()
            self.verticalLayout_w.setObjectName(_fromUtf8("verticalLayout"))
            self.label_w = QtWidgets.QLabel(self.w)
            self.label_w.setAlignment(QtCore.Qt.AlignCenter)
            self.label_w.setObjectName(_fromUtf8("label_w"))
            self.verticalLayout_w.addWidget(self.label_w)
            self.horizontalLayout_2_w = QtWidgets.QHBoxLayout()
            self.horizontalLayout_2_w.setObjectName(_fromUtf8("horizontalLayout_2"))
            self.pushButton_w = QtWidgets.QPushButton(self.w)
            self.pushButton_w.setObjectName(_fromUtf8("pushButton"))
            self.horizontalLayout_2_w.addWidget(self.pushButton_w)
            self.horizontalLayout_w = QtWidgets.QHBoxLayout()
            self.horizontalLayout_w.setObjectName(_fromUtf8("horizontalLayout"))
            self.label_2_w = QtWidgets.QLabel(self.w)
            self.label_2_w.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
            self.label_2_w.setObjectName(_fromUtf8("label_2_w"))
            self.horizontalLayout_w.addWidget(self.label_2_w)
            self.comboBox_w = QtWidgets.QComboBox(self.w)
            self.comboBox_w.setObjectName(_fromUtf8("comboBox"))
            self.comboBox_w.addItems(["Select"]+self.norm_methods)
            self.comboBox_w.setMinimumSize(QtCore.QSize(200,22))
            self.comboBox_w.setMaximumSize(QtCore.QSize(200, 22))
            width=self.comboBox_w.fontMetrics().boundingRect(max(self.norm_methods, key=len)).width()
            self.comboBox_w.view().setFixedWidth(width+10)             
            self.comboBox_w.currentIndexChanged.connect(self.get_norm_from_manualselection)
            self.horizontalLayout_w.addWidget(self.comboBox_w)
            self.horizontalLayout_2_w.addLayout(self.horizontalLayout_w)
            self.verticalLayout_w.addLayout(self.horizontalLayout_2_w)
            self.gridLayout_w.addLayout(self.verticalLayout_w, 0, 0, 1, 1)

            self.w.setWindowTitle("Select normalization method")
            self.label_w.setText("You are about to continue training a pretrained model\n"
    "Please select the meta file of that model to load the normalization method\n"
    "or choose the normalization method manually")
            self.pushButton_w.setText("Load meta file")
            self.label_2_w.setText("Manual \n"
    "selection")

            #one button that allows to load a meta file containing the norm-method
            self.pushButton_w.clicked.connect(self.get_norm_from_modelparafile)
            self.w.show()

    def action_preview_model(self,enabled):
        if enabled: 
            #if the "Load and restart" radiobutton was clicked:
            if self.radioButton_LoadRestartModel.isChecked():
                modelname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open model architecture', Default_dict["Path of last model"],"Architecture or model (*.arch *.model)")
                modelname = modelname[0]
                #modelname_for_dict = modelname
            #if the "Load and continue" radiobutton was clicked:
            elif self.radioButton_LoadContinueModel.isChecked():
                modelname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open model with all parameters', Default_dict["Path of last model"],"Keras model (*.model)")
                modelname = modelname[0]
                #modelname_for_dict = modelname
            self.lineEdit_LoadModelPath.setText(modelname) #Put the filename to the line edit

            #Remember the location for next time
            if len(str(modelname))>0:
                Default_dict["Path of last model"] = os.path.split(modelname)[0]
                aid_bin.save_aid_settings(Default_dict)
            #If user wants to load and restart a model
            if self.radioButton_LoadRestartModel.isChecked():
                #load the model and print summary
                if modelname.endswith(".arch"):
                    json_file = open(modelname, 'r')
                    model_config = json_file.read()
                    json_file.close()
                    model_config = json.loads(model_config)
                    #cut the .json off
                    modelname = modelname.split(".arch")[0]
                                    
                #Or a .model (FULL model with trained weights) , but for display only load the architecture        
                elif modelname.endswith(".model"):
                    #Load the model config (this is the architecture)                  
                    model_full_h5 = h5py.File(modelname, 'r')
                    model_config = model_full_h5.attrs['model_config']
                    model_full_h5.close() #close the hdf5
                    model_config = json.loads(str(model_config)[2:-1])
                    #model = model_from_config(model_config)
                    modelname = modelname.split(".model")[0]
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    #raise ValueError("No valid file was chosen")
                    
                text1 = "Architecture: loaded from .arch\nWeights: will be randomly initialized'\n"                    
                    
                #Try to find the corresponding .meta
                #All models have a number:
                metaname = modelname.rsplit('_',1)[0]+"_meta.xlsx"
                if os.path.isfile(metaname):
                    #open the metafile
                    meta = pd.read_excel(metaname,sheetname="Parameters")
                    if "Chosen Model" in list(meta.keys()):
                        chosen_model = meta["Chosen Model"].iloc[-1]
                    else:
                        #Try to get the model architecture and adjust the combobox
                        try:
                            ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                        except:#No model could be identified
                            chosen_model = "None"
                else:
                    #Try to get the model architecture and adjust the combobox
                    try:
                        ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                    except:#No model could be identified
                        chosen_model = "None"

                if chosen_model is not None:
                    #chosen_model is a string that should be contained in comboBox_ModelSelection
                    index = self.comboBox_ModelSelection.findText(chosen_model, QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_ModelSelection.setCurrentIndex(index)
                else:
                    index = self.comboBox_ModelSelection.findText('None', QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_ModelSelection.setCurrentIndex(index)
                    

            #Otherwise, user wants to load and continue training a model
            elif self.radioButton_LoadContinueModel.isChecked():
                #User can only choose a .model (FULL model with trained weights) , but for display only load the architecture
                if modelname.endswith(".model"):
                    #Load the model config (this is the architecture)                  
                    model_full_h5 = h5py.File(modelname, 'r')
                    model_config = model_full_h5.attrs['model_config']
                    model_full_h5.close() #close the hdf5
                    model_config = json.loads(str(model_config)[2:-1])
                    #model = model_from_config(model_config)
                    modelname = modelname.split(".model")[0]

                    #Try to find the corresponding .meta
                    #All models have a number:
                    metaname = modelname.rsplit('_',1)[0]+"_meta.xlsx"
                    if os.path.isfile(metaname):
                        #open the metafile
                        meta = pd.read_excel(metaname,sheetname="Parameters")
                        if "Chosen Model" in list(meta.keys()):
                            chosen_model = meta["Chosen Model"].iloc[-1]
                        else:
                            #Try to get the model architecture and adjust the combobox
                            try:
                                ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                            except:#No model could be identified
                                chosen_model = "None"
                    else:
                        #Try to get the model architecture and adjust the combobox
                        try:
                            ismlp,chosen_model = model_zoo.mlpconfig_to_str(model_config)
                        except:#No model could be identified
                            chosen_model = "None"

                    if chosen_model is not None:
                        #chosen_model is a string that should be contained in comboBox_ModelSelection
                        index = self.comboBox_ModelSelection.findText(chosen_model, QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_ModelSelection.setCurrentIndex(index)
                    else:
                        index = self.comboBox_ModelSelection.findText('None', QtCore.Qt.MatchFixedString)
                        if index >= 0:
                            self.comboBox_ModelSelection.setCurrentIndex(index)
                    text1 = "Architecture: loaded from .model\nWeights: pretrained weights will be loaded and used when hitting button 'Initialize model!'\n"
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                    msg.setWindowTitle("No valid file was chosen")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    #raise ValueError("No valid file was chosen")

            #In both cases (restart or continue) the input dimensions have to fit
            #The number of output classes should also fit but this is not essential  
            #but most users certainly want the same number of classes (output)->Give Info
            
            try: #Sequential Model
                in_dim = model_config['config'][0]['config']['batch_input_shape']
            except: #Functional Api
                in_dim = model_config['config']["layers"][0]["config"]["batch_input_shape"]
            try: #Sequential Model
                out_dim = model_config['config'][-2]['config']['units']
            except: #Functional Api
                out_dim = model_config['config']["layers"][-2]["config"]["units"]
#            
#            in_dim = model_config['config'][0]['config']['batch_input_shape']
#            out_dim = model_config['config'][-2]['config']['units']

            #Retrieve the color_mode from the model (nr. of channels in last in_dim)
            channels = in_dim[-1] #TensorFlow: channels in last dimension
            if channels==1:
                channel_text = "1 channel (Grayscale)"
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)

            elif channels==3:
                channel_text = "3 channels (RGB)"
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in the ui to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    self.statusbar.showMessage("Color Mode set to RGB",5000)

            text2 = "Model Input: loaded Model takes: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"
            if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                self.spinBox_imagecrop.setValue(in_dim[-2])
                text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
            
            #check that the nr. of classes are equal to the model out put
            SelectedFiles = self.items_clicked()
            indices = [s["class"] for s in SelectedFiles]
            nr_classes = len(set(indices))
            if int(nr_classes)==int(out_dim):
                text3 = "Output: "+str(out_dim)+" classes\n"
            else:
                text3 = "Model Output: The architecture you chose has "+(str(out_dim))+\
                " output nodes (classes) but your selected data has "+str(nr_classes)+\
                " classes. The model could theoretically have more outputs nodes than you have indices but the other way around is not defined\n"                

            text = text1+text2+text3
            self.textBrowser_Info.setText(text)

            if self.radioButton_LoadContinueModel.isChecked():
                #"Load the parameter file of the model that should be continued and apply the same normalization"
                #Make a popup: You are about to continue to train a pretrained model
                #Please select the parameter file of that model to load the normalization method
                #or choose the normalization method manually:
                #this is important
                self.popup_normalization()

    def get_metrics(self,nr_classes):
        Metrics =  []
        f1 = bool(self.checkBox_expertF1.isChecked())
        if f1==True:
            for class_ in range(nr_classes):
                Metrics.append(keras_metrics.categorical_f1_score(label=class_))
        precision = bool(self.checkBox_expertPrecision.isChecked())
        if precision==True:
            for class_ in range(nr_classes):
                Metrics.append(keras_metrics.categorical_precision(label=class_))
        recall = bool(self.checkBox_expertRecall.isChecked())
        if recall==True:
            for class_ in range(nr_classes):
                Metrics.append(keras_metrics.categorical_recall(label=class_))
        
        metrics =  ['accuracy'] + Metrics
        return metrics

    def action_set_modelpath_and_name(self):
        #Get the path and filename for the new model
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save model', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filename = filename[0]
        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        if filename.endswith(".arch"):
            filename = filename.split(".arch")[0]
        #add the suffix .model
        if not filename.endswith(".model"):
            filename = filename +".model"
        self.lineEdit_modelname.setText(filename)
        #Write to Default_dict
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
    def get_dataOverview(self):
        table = self.tableWidget_Info
        cols = table.columnCount()
        header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        rows = table.rowCount()
        tmp_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    tmp_df.iloc[i, j] = table.item(i, j).text()
                except:
                    tmp_df.iloc[i, j] = np.nan
        return tmp_df
        
                 
    def action_initialize_model(self,only_initialize=False):
        #Initialize the model
        #######################Load and restart model##########################
        if self.radioButton_LoadRestartModel.isChecked():
            load_modelname = str(self.lineEdit_LoadModelPath.text())
            text0 = "Loaded model: "+load_modelname
            #load the model and print summary
            if load_modelname.endswith(".arch"):
                json_file = open(load_modelname, 'r')
                model_config = json_file.read()
                json_file.close()
                model_keras = model_from_json(model_config)
                model_config = json.loads(model_config)
                text1 = "\nArchitecture: loaded from .arch\nWeights: randomly initialized\n"

            #Or a .model (FULL model with trained weights) , but for display only load the architecture        
            elif load_modelname.endswith(".model"):
                #Load the model config (this is the architecture)                  
                model_full_h5 = h5py.File(load_modelname, 'r')
                model_config = model_full_h5.attrs['model_config']
                model_full_h5.close() #close the hdf5                
                model_config = json.loads(str(model_config)[2:-1])
                model_keras = model_from_config(model_config)
                text1 = "\nArchitecture: loaded from .model\nWeights: randomly initialized\n"
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                msg.setWindowTitle("No valid file was chosen")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            try:
                metaname = load_modelname.rsplit('_',1)[0]+"_meta.xlsx"
                if os.path.isfile(metaname):
                    #open the metafile
                    meta = pd.read_excel(metaname,sheetname="Parameters")
                    if "Chosen Model" in list(meta.keys()):
                        chosen_model = meta["Chosen Model"].iloc[-1]
            except:
                chosen_model = str(self.comboBox_ModelSelection.currentText())
                
            #In both cases (restart or continue) the input dimensions have to fit
            #The number of output classes should also fit but this is not essential  
            #but most users certainly want the same number of classes (output)->Give Info

            try: #Sequential Model
                in_dim = model_config['config'][0]['config']['batch_input_shape']
            except: #Functional Api
                in_dim = model_config['config']["layers"][0]["config"]["batch_input_shape"]
            try: #Sequential Model
                out_dim = model_config['config'][-2]['config']['units']
            except: #Functional Api
                out_dim = model_config['config']["layers"][-2]["config"]["units"]

#            in_dim = model_config['config'][0]['config']['batch_input_shape']
#            out_dim = model_config['config'][-2]['config']['units']
            channels = in_dim[-1] #TensorFlow: channels in last dimension

            #Compile model (consider user-specific metrics)
            metrics = self.get_metrics(out_dim)            
            model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=metrics)#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled

            if channels==1:
                channel_text = "1 channel (Grayscale)"
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)

            elif channels==3:
                channel_text = "3 channels (RGB)"
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in the ui to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    self.statusbar.showMessage("Color Mode set to RGB",5000)

            text2 = "Model Input: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"

            if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                self.spinBox_imagecrop.setValue(in_dim[-2])
                text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
            
            #check that the nr. of classes are equal to the model out put
            SelectedFiles = self.items_clicked()
            indices = [s["class"] for s in SelectedFiles]
            nr_classes = len(set(indices))
            if int(nr_classes)==int(out_dim):
                text3 = "Output: "+str(out_dim)+" classes\n"
            else:
                text3 = "Model Output: The architecture you chose has "+(str(out_dim))+\
                " output nodes (classes) but your selected data has "+str(nr_classes)+\
                " classes. The model could theoretically have more outputs nodes than you have indices but the other way around is not defined\n"                

        ###############Load and continue training the model####################
        elif self.radioButton_LoadContinueModel.isChecked():
            load_modelname = str(self.lineEdit_LoadModelPath.text())
            text0 = "Loaded model: "+load_modelname

            #User can only choose a .model (FULL model with trained weights) , but for display only load the architecture
            if load_modelname.endswith(".model"):              
                #Load the full model
                try:
                    model_keras = load_model(load_modelname,custom_objects=get_custom_metrics())
                except: 
                    K.clear_session() #On linux It happened that there was an error, if another fitting was run before                  
                    model_keras = load_model(load_modelname,custom_objects=get_custom_metrics())
                #model_config = model_keras.config() #Load the model config (this is the architecture)
                #load_modelname = load_modelname.split(".model")[0]
                text1 = "Architecture: loaded from .model\nWeights: pretrained weights loaded'\n"
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("No valid file was chosen. Please specify a file that was created using AIDeveloper or Keras")
                msg.setWindowTitle("No valid file was chosen")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
                #raise ValueError("No valid file was chosen")

            try:
                metaname = load_modelname.rsplit('_',1)[0]+"_meta.xlsx"
                if os.path.isfile(metaname):
                    #open the metafile
                    meta = pd.read_excel(metaname,sheetname="Parameters")
                    if "Chosen Model" in list(meta.keys()):
                        chosen_model = meta["Chosen Model"].iloc[-1]
                else:
                    chosen_model = str(self.comboBox_ModelSelection.currentText())

            except:
                chosen_model = str(self.comboBox_ModelSelection.currentText())


            #Check input dimensions
            #The number of output classes should also fit but this is not essential  
            #but most users certainly want the same number of classes (output)->Give Info
#            in_dim = model_config['config'][0]['config']['batch_input_shape']
#            out_dim = model_config['config'][-2]['config']['units']
            in_dim = model_keras.get_input_shape_at(0)
            out_dim = model_keras.get_output_shape_at(0)[1]
            channels = in_dim[-1] #TensorFlow: channels in last dimension

            if channels==1:
                channel_text = "1 channel (Grayscale)"
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)

            elif channels==3:
                channel_text = "3 channels (RGB)"
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in the ui to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    self.statusbar.showMessage("Color Mode set to RGB",5000)

            text2 = "Model Input: "+str(in_dim[-3])+" x "+str(in_dim[-2]) + " pixel images and "+channel_text+"\n"
            if int(self.spinBox_imagecrop.value())!=int(in_dim[-2]):
                self.spinBox_imagecrop.setValue(in_dim[-2])
                text2 = text2+ "'Input image size'  in GUI was changed accordingly\n"
            
            #check that the nr. of classes are equal to the model out put
            SelectedFiles = self.items_clicked()
            indices = [s["class"] for s in SelectedFiles]
            nr_classes = len(set(indices))
            if int(nr_classes)==int(out_dim):
                text3 = "Output: "+str(out_dim)+" classes\n"
            else:
                text3 = "Model Output: The architecture you chose has "+(str(out_dim))+\
                " output nodes (classes) but your selected data has "+str(nr_classes)+\
                " classes. The model could theoretically have more outputs nodes than you have indices but the other way around is not defined\n"                

        ###########################New model###################################
        elif self.radioButton_NewModel.isChecked():
            load_modelname = "" #No model is loaded
            text0 = load_modelname
            #Create a new model!
            #Get what the user wants from the dropdown menu!
            chosen_model = str(self.comboBox_ModelSelection.currentText())
            if chosen_model==None:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("No model specified!")
                msg.setWindowTitle("No model specified!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return       
                
            in_dim = int(self.spinBox_imagecrop.value())
            SelectedFiles = self.items_clicked()
            #rtdc_ds = SelectedFiles[0]["rtdc_ds"]

            if str(self.comboBox_GrayOrRGB.currentText())=="Grayscale":
                channels=1
            elif str(self.comboBox_GrayOrRGB.currentText())=="RGB":
                channels=3
                
            indices = [s["class"] for s in SelectedFiles]
            indices_unique = np.unique(np.array(indices))
            if len(indices_unique)<2:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Need at least two classes to fit. Please specify .rtdc files and corresponding indeces")
                msg.setWindowTitle("No valid file was chosen")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return  
            
            if 0 in indices:
                #out_dim = len(set(indices))
                out_dim = np.max(indices)+1
            else:
                out_dim = np.max(indices)           
            nr_classes = out_dim
            
            if chosen_model=="None":
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("No model specified!")
                msg.setWindowTitle("No model specified!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return       
            
            try:                
                model_keras = model_zoo.get_model(chosen_model,in_dim,channels,out_dim)
            except Exception as e: 
                #There is an issue building the model!
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)       
                msg.setText(str(e))
                msg.setWindowTitle("Error occured when building model:")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return       
            
            text1 = "Architecture: created "+chosen_model+" design\nWeights: Initialized random weights\n"
            
            if self.get_color_mode()=="Grayscale":
                channels = 1
                channel_text = "1 channel (Grayscale)"
            elif self.get_color_mode()=="RGB":
                channels = 3
                channel_text = "3 channels (RGB)"
                    
            text2 = "Model Input: "+str(in_dim)+" x "+str(in_dim) + " pixel images and "+channel_text+"\n"

            if int(nr_classes)==int(out_dim):
                text3 = "Output: "+str(out_dim)+" classes\n"
            else:
                text3 = "Model Output: The architecture you chose has "+(str(out_dim))+\
                " output nodes (classes) but your selected data has "+str(nr_classes)+\
                " classes. The model could theoretically have more outputs nodes than you have indices but the other way around is not defined\n"                
        
        else:
            #No radio-button was chosen
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please use the radiobuttons to define the model")
            msg.setWindowTitle("No model defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        model_metrics = self.get_metrics(out_dim)
        if "collection" in chosen_model.lower():
            for m in model_keras[1]: #in a collection, model_keras[0] are the names of the models and model_keras[1] is a list of all models
                m.compile(loss='categorical_crossentropy',optimizer='adam',metrics=self.get_metrics(nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
        if not "collection" in chosen_model.lower():
            model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=model_metrics)#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled

        #If expert mode is on, apply the requested options
        #This affects learning rate, trainability of layers and dropout rate
        expert_mode = bool(self.groupBox_expertMode.isChecked())
        learning_rate_expert = float(self.doubleSpinBox_learningRate.value())
        learning_rate_expert_on = bool(self.checkBox_learningRate.isChecked())   
        train_last_layers = bool(self.checkBox_trainLastNOnly.isChecked())             
        train_last_layers_n = int(self.spinBox_trainLastNOnly.value())              
        train_dense_layers = bool(self.checkBox_trainDenseOnly.isChecked())             
        dropout_expert_on = bool(self.checkBox_dropout.isChecked())
        loss_expert_on = bool(self.checkBox_expt_loss.isChecked())
        loss_expert = str(self.comboBox_expt_loss.currentText()).lower()
        optimizer_expert_on = bool(self.checkBox_optimizer.isChecked())
        optimizer_expert = str(self.comboBox_optimizer.currentText()).lower()
        padding_expert_on = bool(self.checkBox_expt_paddingMode.isChecked())
        padding_expert = str(self.comboBox_expt_paddingMode.currentText()).lower()

        try:
            dropout_expert = str(self.lineEdit_dropout.text()) #due to the validator, there are no squ.brackets
            dropout_expert = "["+dropout_expert+"]"
            dropout_expert = ast.literal_eval(dropout_expert)        
        except:
            dropout_expert = []             

        if type(model_keras)==tuple:#when user chose a Collection of models, a tuple is returned by get_model
            collection = True
        else:
            collection = False
        
        if collection==False: #if there is a single model:
            #Original learning rate (before expert mode is switched on!)
            try:
                self.learning_rate_original = K.eval(model_keras.optimizer.lr)
            except:
                print("Session busy. Try again in fresh session...")
                tf.reset_default_graph() #Make sure to start with a fresh session
                sess = tf.InteractiveSession()
                self.learning_rate_original = K.eval(model_keras.optimizer.lr)
                
            #Get initial trainability states of model
            self.trainable_original, self.layer_names = aid_dl.model_get_trainable_list(model_keras)
            trainable_original, layer_names = self.trainable_original, self.layer_names
            self.do_list_original = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
            do_list_original = self.do_list_original

        if collection==True: #if there is a collection of models:
            #Original learning rate (before expert mode is switched on!)
            self.learning_rate_original = [K.eval(model_keras[1][i].optimizer.lr) for i in range(len(model_keras[1]))]
            #Get initial trainability states of model
            trainable_layerName = [aid_dl.model_get_trainable_list(model_keras[1][i]) for i in range(len(model_keras[1]))]
            self.trainable_original = [trainable_layerName[i][0] for i in range(len(trainable_layerName))]
            self.layer_names = [trainable_layerName[i][1] for i in range(len(trainable_layerName))]
            trainable_original, layer_names = self.trainable_original, self.layer_names
            self.do_list_original = [aid_dl.get_dropout(model_keras[1][i]) for i in range(len(model_keras[1]))]#Get a list of dropout values of the current model
            do_list_original = self.do_list_original

        #TODO add expert mode ability for collection of models. Maybe define self.model_keras as a list in general. So, fitting a single model is just a special case


        if expert_mode==True:
            #Apply the changes to trainable states:
            if train_last_layers==True:#Train only the last n layers
                print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,out_dim,loss_expert,optimizer_expert,learning_rate_expert)

            if train_dense_layers==True:#Train only dense layers
                print("Train only dense layers")
                layer_dense_ind = ["Dense" in x for x in layer_names]
                layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                #create a list of trainable states
                trainable_new = len(trainable_original)*[False]
                for index in layer_dense_ind:
                    trainable_new[index] = True
                aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,out_dim,loss_expert,optimizer_expert,learning_rate_expert)

            if dropout_expert_on==True:
                #The user apparently want to change the dropout rates
                do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                #Compare the dropout values in the model to the dropout values requested by user
                if len(dropout_expert)==1:#if the user gave a float
                    dropout_expert_list=len(do_list)*dropout_expert #convert to list
                elif len(dropout_expert)>1:
                    dropout_expert_list = dropout_expert
                    if not len(dropout_expert_list)==len(do_list):
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Information)       
                        msg.setText("Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers")
                        msg.setWindowTitle("Issue with Expert->Dropout")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        dropout_expert_list = []
                        return
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Could not understand user input at Expert->Dropout")
                    msg.setWindowTitle("Issue with Expert->Dropout")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    
                    dropout_expert_list = []
                 
                if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                    do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)
                    if do_changed==1:
                        text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                    else:
                        text_do = "Dropout rate(s) in model was/were not changed"
                else:
                    text_do = "Dropout rate(s) in model was/were not changed"
                print(text_do)

                
        text_updates = ""

        #Compare current lr and the lr on expert tab:
        if collection == False:
            lr_current = K.eval(model_keras.optimizer.lr)
        else:
            lr_current = K.eval(model_keras[1][0].optimizer.lr)
        print("Current learning rate: "+str(lr_current))
        lr_diff = learning_rate_expert-lr_current
        print("Current learning rate: "+str(lr_current))
        if  abs(lr_diff) > 1e-6:
            K.set_value(model_keras.optimizer.lr, learning_rate_expert)
            text_updates +=  "Changed the learning rate to: "+ str(learning_rate_expert)+"\n"

        recompile = False
        #Compare current optimizer and the optimizer on expert tab:
        if collection==False:
            optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
        else:
            optimizer_current = aid_dl.get_optimizer_name(model_keras[1][0]).lower()#get the current optimizer of the model

        if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
            recompile = True
            text_updates+="Changed the optimizer to: "+optimizer_expert+"\n"
        #Compare current loss function and the loss-function on expert tab:

        if collection==False:
            if model_keras.loss!=loss_expert:
                recompile = True
                text_updates+="Changed the loss function to: "+loss_expert+"\n"
        if collection==True:
            if model_keras[1][0].loss!=loss_expert:
                recompile = True
                text_updates+="Changed the loss function to: "+loss_expert+"\n"

        if recompile==True:
            if collection==True:
                print("Recompiling...")
                aid_dl.model_compile(model_keras,loss_expert,optimizer_expert,learning_rate_expert,model_metrics,nr_classes)
            if collection==False:
                for m in model_keras[1]:
                    print("Recompiling...")
                    aid_dl.model_compile(m,loss_expert,optimizer_expert,learning_rate_expert,model_metrics,nr_classes)
        self.model_keras = model_keras #overwrite the model in self

        if collection == False:
            #Get user-specified filename for the new model
            new_modelname = str(self.lineEdit_modelname.text())
            if len(new_modelname)>0:
                text_new_modelname = "Model will be saved as: "+new_modelname+"\n"
            else:
                text_new_modelname = "Please specify a model path (name for the model to be fitted)\n"

        if collection == True:
            new_modelname = str(self.lineEdit_modelname.text())
            if len(new_modelname)>0:
                new_modelname = os.path.split(new_modelname)
                text_new_modelname = "Collection of Models will be saved into: "+new_modelname[0]+"\n"
            else:
                text_new_modelname = "Please specify a model path (name for the model to be fitted)\n"


        #Info about normalization method
        norm = str(self.comboBox_Normalization.currentText())

        text4 = "Input image normalization method: "+norm+"\n"

        #Check if there are dropout layers:
        #do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
        if len(do_list_original)>0:
            text4 = text4+"Found "+str(len(do_list_original)) +" dropout layers with rates: "+str(do_list_original)+"\n"
        else:
            text4 = text4+"Found no dropout layers\n"
        
        if expert_mode==True:
            if dropout_expert_on:
                text4 = text4+text_do+"\n"
#            if learning_rate_expert_on==True:
#                if K.eval(model_keras.optimizer.lr) != learning_rate_expert: #if the learning rate in UI is NOT equal to the lr of the model...
#                    text_lr = "Changed the learning rate to: "+ str(learning_rate_expert)+"\n"
#                    text4 = text4+text_lr

        text5 = "Model summary:\n"
        summary = []
        if collection==False:
            model_keras.summary(print_fn=summary.append)
            summary = "\n".join(summary)
            text = text_new_modelname+text0+text1+text2+text3+text4+text_updates+text5+summary
            self.textBrowser_Info.setText(text)
                
            #Save the model architecture: serialize to JSON
            model_json = model_keras.to_json()
            with open(new_modelname.split(".model")[0]+".arch", "w") as json_file:
                json_file.write(model_json)

        elif collection==True:
            if self.groupBox_expertMode.isChecked()==True:
                self.groupBox_expertMode.setChecked(False)
                print("Turned off expert mode. Not implemented yet for collections of models. This does not affect user-specified metrics (precision/recall/f1)")
#            print("new_modelname")
#            print(new_modelname)
            
            self.model_keras_arch_path = [new_modelname[0]+os.sep+new_modelname[1].split(".model")[0]+"_"+model_keras[0][i]+".arch" for i in range(len(model_keras[0]))]                
            for i in range(len(model_keras[1])):
                model_keras[1][i].summary(print_fn=summary.append)
                    
                #Save the model architecture: serialize to JSON
                model_json = model_keras[1][i].to_json()
                with open(self.model_keras_arch_path[i], "w") as json_file:
                    json_file.write(model_json)

            summary = "\n".join(summary)
            text = text_new_modelname+text0+text1+text2+text3+text4+text5+summary
            self.textBrowser_Info.setText(text)

        #Save the model to a variable on self
        self.model_keras = model_keras

        #Get the user-defined cropping size
        crop = int(self.spinBox_imagecrop.value())          
        #Make the cropsize a bit larger since the images will later be rotated
        cropsize2 = np.sqrt(crop**2+crop**2)
        cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number

        #Estimate RAM needed
        nr_imgs = np.sum([np.array(list(SelectedFiles)[i]["nr_images"]) for i in range(len(list(SelectedFiles)))])
        ram_needed = np.round(nr_imgs * aid_bin.calc_ram_need(cropsize2),2)
 
        if only_initialize==True:#Stop here if the model just needs to be intialized (for expert mode->partial trainability)
            return
        
        #Tell the user if the data is stored and read from ram or not
        if self.actionDataToRam.isChecked():
            color_mode = self.get_color_mode()
            zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
            zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            zoom_order = int(np.where(np.array(zoom_order)==True)[0])

            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Question)
            text = "<html><head/><body><p>Should the model only be initialized, or do you want to start fitting right after? For fitting, data will be loaded to RAM (since Edit->Data to RAM is enabled), which will require "+str(ram_needed)+"MB of RAM.</p></body></html>"
            msg.setText(text) 
            msg.setWindowTitle("Initialize model or initialize and fit model?")
            msg.addButton(QtGui.QPushButton('Stop after model initialization'), QtGui.QMessageBox.RejectRole)
            msg.addButton(QtGui.QPushButton('Start fitting'), QtGui.QMessageBox.ApplyRole)
            retval = msg.exec_()
            if retval==0: #yes role: Only initialize model
                return
            elif retval == 1:
                dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),cropsize2,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
                self.ram = dic 
                #Finally, activate the 'Fit model' button again
                #self.pushButton_FitModel.setEnabled(True)
                self.action_fit_model()
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Question)
            text = "<html><head/><body><p>Should the model only be initialized, or do you want to start fitting right after? For fitting, data will be loaded to RAM (since Edit->Data to RAM is enabled), which will require "+str(ram_needed)+"MB of RAM.</p></body></html>"
            msg.setText(text) 
            msg.setWindowTitle("Initialize model or initialize and fit model?")
            msg.addButton(QtGui.QPushButton('Stop after model initialization'), QtGui.QMessageBox.RejectRole)
            msg.addButton(QtGui.QPushButton('Start fitting'), QtGui.QMessageBox.ApplyRole)
            retval = msg.exec_()
            if retval==0: #yes role: Only initialize model
                return
            elif retval == 1:
                self.action_fit_model()

        

    def action_fit_model_worker(self,progress_callback,history_callback):
        with tf.Session(graph = tf.Graph(), config=config_gpu) as sess: #Without this multithreading does not work in TensorFlow
            #get an index of the fitting popup
            listindex = self.popupcounter-1
            #Get user-specified filename for the new model
            new_modelname = str(self.lineEdit_modelname.text())
            model_keras_path = self.model_keras_path
            
            if type(model_keras_path)==list:
                collection = True
                #Take the initialized models
                model_keras_path = self.model_keras_path
                model_keras = [load_model(model_keras_path[i],custom_objects=get_custom_metrics()) for i in range(len(model_keras_path)) ]
                model_architecture_names = self.model_keras[0]
                print(model_architecture_names)    
                self.model_keras = None

            else:
                collection = False
                model_keras = load_model(model_keras_path,custom_objects=get_custom_metrics())
                self.model_keras = None

            ##############Main function after hitting FIT MODEL####################
            if self.radioButton_LoadRestartModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_LoadContinueModel.isChecked():
                load_modelname = str(self.lineEdit_LoadModelPath.text())
            elif self.radioButton_NewModel.isChecked():
                load_modelname = "" #No model is loaded

            if collection==False:    
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras.output.shape.dims[1])
            if collection==True:
                #model_config = model_keras.get_config()#["layers"] 
                nr_classes = int(model_keras[0].output.shape.dims[1])
            
            #Metrics to be displayed during fitting (real-time)
            model_metrics = self.get_metrics(nr_classes)
            
            #Compile model
            if collection==False:
                model_keras.compile(loss='categorical_crossentropy',optimizer='adam',metrics=model_metrics)#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
            elif collection==True:
                #Switch off the expert tab!
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(False)
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setEnabled(False)
                for m in model_keras:
                    m.compile(loss='categorical_crossentropy',optimizer='adam',metrics=self.get_metrics(nr_classes))#dont specify loss and optimizer yet...expert stuff will follow and model will be recompiled
                    
            #Original learning rate:
            learning_rate_original = self.learning_rate_original#K.eval(model_keras.optimizer.lr)
            #Original trainable states of layers with parameters
            trainable_original, layer_names = self.trainable_original, self.layer_names
            do_list_original = self.do_list_original
            
            #Collect all information about the fitting routine that was user
            #defined
            if self.actionVerbose.isChecked()==True:
                verbose = 1
            else:
                verbose = 0

            new_model = self.radioButton_NewModel.isChecked()
            chosen_model = str(self.comboBox_ModelSelection.currentText())
                
            crop = int(self.spinBox_imagecrop.value())      
            color_mode = str(self.comboBox_GrayOrRGB.currentText())
            
            loadrestart_model = self.radioButton_LoadRestartModel.isChecked()
            loadcontinue_model = self.radioButton_LoadContinueModel.isChecked()
    
            norm = str(self.comboBox_Normalization.currentText())
    
            nr_epochs = int(self.spinBox_NrEpochs.value())
            keras_refresh_nr_epochs = int(self.spinBox_RefreshAfterEpochs.value())
            h_flip = bool(self.checkBox_HorizFlip.isChecked())
            v_flip = bool(self.checkBox_VertFlip.isChecked())
            rotation = float(self.lineEdit_Rotation.text())
     
            width_shift = float(self.lineEdit_Rotation_2.text())
            height_shift = float(self.lineEdit_Rotation_3.text())
            zoom = float(self.lineEdit_Rotation_4.text())
            shear = float(self.lineEdit_Rotation_5.text())
            
            brightness_refresh_nr_epochs = int(self.spinBox_RefreshAfterNrEpochs.value())
            brightness_add_lower = float(self.spinBox_PlusLower.value())
            brightness_add_upper = float(self.spinBox_PlusUpper.value())
            brightness_mult_lower = float(self.doubleSpinBox_MultLower.value())
            brightness_mult_upper = float(self.doubleSpinBox_MultUpper.value())
            gaussnoise_mean = float(self.doubleSpinBox_GaussianNoiseMean.value())
            gaussnoise_scale = float(self.doubleSpinBox_GaussianNoiseScale.value())

            contrast_on = bool(self.checkBox_contrast.isChecked())        
            contrast_lower = float(self.doubleSpinBox_contrastLower.value())
            contrast_higher = float(self.doubleSpinBox_contrastHigher.value())
            saturation_on = bool(self.checkBox_saturation.isChecked())        
            saturation_lower = float(self.doubleSpinBox_saturationLower.value())
            saturation_higher = float(self.doubleSpinBox_saturationHigher.value())
            hue_on = bool(self.checkBox_hue.isChecked())        
            hue_delta = float(self.doubleSpinBox_hueDelta.value())

            avgBlur_on = bool(self.checkBox_avgBlur.isChecked())        
            avgBlur_min = int(self.spinBox_avgBlurMin.value())
            avgBlur_max = int(self.spinBox_avgBlurMax.value())
            gaussBlur_on = bool(self.checkBox_gaussBlur.isChecked())        
            gaussBlur_min = int(self.spinBox_gaussBlurMin.value())
            gaussBlur_max = int(self.spinBox_gaussBlurMax.value())
            motionBlur_on = bool(self.checkBox_motionBlur.isChecked())        
            motionBlur_kernel = str(self.lineEdit_motionBlurKernel.text())
            motionBlur_angle = str(self.lineEdit_motionBlurAngle.text())
            motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
            motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple


            if collection==False:
                expert_mode = bool(self.groupBox_expertMode.isChecked())
            elif collection==True:
                expert_mode = self.groupBox_expertMode.setChecked(False)
                print("Expert mode was switched off. Not implemented yet for collections")
                expert_mode = False

            batchSize_expert = int(self.spinBox_batchSize.value())
            epochs_expert = int(self.spinBox_epochs.value())
            learning_rate_expert = float(self.doubleSpinBox_learningRate.value())
            learning_rate_expert_on = bool(self.checkBox_learningRate.isChecked()) 
            loss_expert_on = bool(self.checkBox_expt_loss.isChecked())
            loss_expert = str(self.comboBox_expt_loss.currentText()).lower()
            optimizer_expert_on = bool(self.checkBox_optimizer.isChecked())
            optimizer_expert = str(self.comboBox_optimizer.currentText()).lower()
            padding_expert_on = bool(self.checkBox_expt_paddingMode.isChecked())
            padding_expert = str(self.comboBox_expt_paddingMode.currentText()).lower()

            train_last_layers = bool(self.checkBox_trainLastNOnly.isChecked())             
            train_last_layers_n = int(self.spinBox_trainLastNOnly.value())              
            train_dense_layers = bool(self.checkBox_trainDenseOnly.isChecked())             
            dropout_expert_on = bool(self.checkBox_dropout.isChecked())             
            try:
                dropout_expert = str(self.lineEdit_dropout.text()) #due to the validator, there are no squ.brackets
                dropout_expert = "["+dropout_expert+"]"
                dropout_expert = ast.literal_eval(dropout_expert)        
            except:
                dropout_expert = []
            lossW_expert_on = bool(self.checkBox_lossW.isChecked())             
            lossW_expert = str(self.lineEdit_lossW.text())

            #To get the class weights (loss), the SelectedFiles are required 
            SelectedFiles = self.items_clicked()
            self.fittingpopups_ui[listindex].SelectedFiles = SelectedFiles #save to self. to make it accessible for popup showing loss weights
            #Get the class weights. This function runs now the first time in the fitting routine. 
            #It is possible that the user chose Custom weights and then changed the classes. Hence first check if 
            #there is a weight for each class available.
            class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert,custom_check_classes=True)
            if type(class_weight)==list:
                #There has been a mismatch between the classes described in class_weight and the classes available in SelectedFiles!
                lossW_expert = class_weight[0] #overwrite 
                class_weight = class_weight[1]
                print(class_weight)
                print("There has been a mismatch between the classes described in Loss weights and the classes available in the selected files! Hence, the Loss weights are set to Balanced")
#                #Notify user
#                msg = QtWidgets.QMessageBox()
#                msg.setIcon(QtWidgets.QMessageBox.Information)       
#                text = "<html><head/><body><p>There has been a mismatch between the classes described in Loss weights and the classes available in the selected files! Hence, the 'Loss weights' are set to 'Balanced'.</p></body></html>"
#                msg.setText(text) 
#                msg.setWindowTitle("Loss weights set to Balanced")
#                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                msg.exec_()
                
            if collection==False:    
                #Create an excel file
                writer = pd.ExcelWriter(new_modelname.split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                self.writer = writer
                #Used files go to a separate sheet on the MetaFile.xlsx
                SelectedFiles_df = pd.DataFrame(SelectedFiles)
                pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                DataOverview_df = self.get_dataOverview()
                DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet
        
                #create a dictionary for the metafiles (Parameters and History)
                Para_dict = pd.DataFrame()
                Para_dict["AIDeveloper_Version"]=VERSION,
                Para_dict["model_zoo_version"]=model_zoo_version,
                try:
                    Para_dict["OS"]=platform.platform(),
                    Para_dict["CPU"]=platform.processor(),
                except:
                    Para_dict["OS"]="Unknown",
                    Para_dict["CPU"]="Unknown",
                    
                Para_dict["Modelname"]=new_modelname,
                Para_dict["Chosen Model"]=chosen_model,
                Para_dict["Input image size"]=crop,
                Para_dict["Color Mode"]=color_mode,
        
                Para_dict["new_model"]=new_model,
                Para_dict["loadrestart_model"]=loadrestart_model,
                Para_dict["loadcontinue_model"]=loadcontinue_model,
                Para_dict["Continued_Fitting_From"]=load_modelname,
        
                Para_dict["Output Nr. classes"]=nr_classes,
        
                Para_dict["Normalization"]=norm,
                Para_dict["Nr. epochs"]=nr_epochs,
                Para_dict["Keras refresh after nr. epochs"]=keras_refresh_nr_epochs,
                Para_dict["Horz. flip"]=h_flip,
                Para_dict["Vert. flip"]=v_flip,
                Para_dict["rotation"]=rotation,
                Para_dict["width_shift"]=width_shift,
                Para_dict["height_shift"]=height_shift,
                Para_dict["zoom"]=zoom,
                Para_dict["shear"]=shear,
                Para_dict["Brightness refresh after nr. epochs"]=brightness_refresh_nr_epochs,
                Para_dict["Brightness add. lower"]=brightness_add_lower,
                Para_dict["Brightness add. upper"]=brightness_add_upper,
                Para_dict["Brightness mult. lower"]=brightness_mult_lower,  
                Para_dict["Brightness mult. upper"]=brightness_mult_upper,
                Para_dict["Gaussnoise Mean"]=gaussnoise_mean,
                Para_dict["Gaussnoise Scale"]=gaussnoise_scale,
                
                Para_dict["Contrast on"]=contrast_on,                
                Para_dict["Contrast Lower"]=contrast_lower,
                Para_dict["Contrast Higher"]=contrast_higher,
                Para_dict["Saturation on"]=saturation_on,
                Para_dict["Saturation Lower"]=saturation_lower,
                Para_dict["Saturation Higher"]=saturation_higher,
                Para_dict["Hue on"]=hue_on,                
                Para_dict["Hue delta"]=hue_delta,                

                Para_dict["Average blur on"]=avgBlur_on,                
                Para_dict["Average blur Lower"]=avgBlur_min,
                Para_dict["Average blur Higher"]=avgBlur_max,
                Para_dict["Gauss blur on"]=gaussBlur_on,                
                Para_dict["Gauss blur Lower"]=gaussBlur_min,
                Para_dict["Gauss blur Higher"]=gaussBlur_max,
                Para_dict["Motion blur on"]=motionBlur_on,                
                Para_dict["Motion blur Kernel"]=motionBlur_kernel,                
                Para_dict["Motion blur Angle"]=motionBlur_angle,                

                Para_dict["Epoch_Started_Using_These_Settings"]=0,
    
                Para_dict["expert_mode"]=expert_mode,
                Para_dict["batchSize_expert"]=batchSize_expert,
                Para_dict["epochs_expert"]=epochs_expert,
                Para_dict["learning_rate_expert"]=learning_rate_expert,
                Para_dict["learning_rate_expert_on"]=learning_rate_expert_on,

                Para_dict["loss_expert_on"]=loss_expert_on,
                Para_dict["loss_expert"]=loss_expert,
                Para_dict["optimizer_expert_on"]=optimizer_expert_on,
                Para_dict["optimizer_expert"]=optimizer_expert,
                Para_dict["padding_expert_on"]=padding_expert_on,
                Para_dict["padding_expert"]=padding_expert,

                Para_dict["train_last_layers"]=train_last_layers,
                Para_dict["train_last_layers_n"]=train_last_layers_n,
                Para_dict["train_dense_layers"]=train_dense_layers,
                Para_dict["dropout_expert_on"]=dropout_expert_on,
                Para_dict["dropout_expert"]=dropout_expert,
                Para_dict["lossW_expert_on"]=lossW_expert_on,
                Para_dict["lossW_expert"]=lossW_expert,
                Para_dict["class_weight"]=class_weight,
                Para_dict["metrics"]=model_metrics,
                
            elif collection==True: 
                SelectedFiles_df = pd.DataFrame(SelectedFiles)

                Writers = []
                #Create excel files
                for i in range(len(model_keras_path)):
                    writer = pd.ExcelWriter(model_keras_path[i].split(".model")[0]+'_meta.xlsx', engine='openpyxl')
                    Writers.append(writer)
                for writer in Writers:
                    #Used files go to a separate sheet on the MetaFile.xlsx
                    pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
                    SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
                    DataOverview_df = self.get_dataOverview()
                    DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            
                    pd.DataFrame().to_excel(writer,sheet_name='Parameters') #initialize empty Sheet
                    pd.DataFrame().to_excel(writer,sheet_name='History') #initialize empty Sheet
            
                #create a dictionary for the metafiles (Parameters and History)
                Para_dict = pd.DataFrame()
                Para_dict["AIDeveloper_Version"]=VERSION,
                Para_dict["model_zoo_version"]=model_zoo_version,
                try:
                    Para_dict["OS"]=platform.platform(),
                    Para_dict["CPU"]=platform.processor(),
                except:
                    Para_dict["OS"]="Unknown",
                    Para_dict["CPU"]="Unknown",
                
                Para_dict["Modelname"]=model_keras_path[i],
                Para_dict["Chosen Model"]=model_architecture_names[i],
                Para_dict["Input image size"]=crop,
                Para_dict["Color Mode"]=color_mode,
        
                Para_dict["new_model"]=new_model,
                Para_dict["loadrestart_model"]=loadrestart_model,
                Para_dict["loadcontinue_model"]=loadcontinue_model,
                Para_dict["Continued_Fitting_From"]=load_modelname,
        
                Para_dict["Output Nr. classes"]=nr_classes,
        
                Para_dict["Normalization"]=norm,
                Para_dict["Nr. epochs"]=nr_epochs,
                Para_dict["Keras refresh after nr. epochs"]=keras_refresh_nr_epochs,
                Para_dict["Horz. flip"]=h_flip,
                Para_dict["Vert. flip"]=v_flip,
                Para_dict["rotation"]=rotation,
                Para_dict["width_shift"]=width_shift,
                Para_dict["height_shift"]=height_shift,
                Para_dict["zoom"]=zoom,
                Para_dict["shear"]=shear,
                Para_dict["Brightness refresh after nr. epochs"]=brightness_refresh_nr_epochs,
                Para_dict["Brightness add. lower"]=brightness_add_lower,
                Para_dict["Brightness add. upper"]=brightness_add_upper,
                Para_dict["Brightness mult. lower"]=brightness_mult_lower,  
                Para_dict["Brightness mult. upper"]=brightness_mult_upper,
                Para_dict["Gaussnoise Mean"]=gaussnoise_mean,
                Para_dict["Gaussnoise Scale"]=gaussnoise_scale,
                
                Para_dict["Contrast on"]=contrast_on,                
                Para_dict["Contrast Lower"]=contrast_lower,
                Para_dict["Contrast Higher"]=contrast_higher,
                Para_dict["Saturation on"]=saturation_on,
                Para_dict["Saturation Lower"]=saturation_lower,
                Para_dict["Saturation Higher"]=saturation_higher,
                Para_dict["Hue on"]=hue_on,                
                Para_dict["Hue delta"]=hue_delta,                

                Para_dict["Average blur on"]=avgBlur_on,                
                Para_dict["Average blur Lower"]=avgBlur_min,
                Para_dict["Average blur Higher"]=avgBlur_max,
                Para_dict["Gauss blur on"]=gaussBlur_on,                
                Para_dict["Gauss blur Lower"]=gaussBlur_min,
                Para_dict["Gauss blur Higher"]=gaussBlur_max,
                Para_dict["Motion blur on"]=motionBlur_on,                
                Para_dict["Motion blur Kernel"]=motionBlur_kernel,                
                Para_dict["Motion blur Angle"]=motionBlur_angle,                
    
                Para_dict["Epoch_Started_Using_These_Settings"]=0,
    
                Para_dict["expert_mode"]=expert_mode,
                Para_dict["batchSize_expert"]=batchSize_expert,
                Para_dict["epochs_expert"]=epochs_expert,
                Para_dict["learning_rate_expert"]=learning_rate_expert,
                Para_dict["learning_rate_expert_on"]=learning_rate_expert_on,
                Para_dict["loss_expert_on"]=loss_expert_on,
                Para_dict["loss_expert"]=loss_expert,
                Para_dict["optimizer_expert_on"]=optimizer_expert_on,
                Para_dict["optimizer_expert"]=optimizer_expert,
                Para_dict["padding_expert_on"]=padding_expert_on,
                Para_dict["padding_expert"]=padding_expert,
                
                Para_dict["train_last_layers"]=train_last_layers,
                Para_dict["train_last_layers_n"]=train_last_layers_n,
                Para_dict["train_dense_layers"]=train_dense_layers,
                Para_dict["dropout_expert_on"]=dropout_expert_on,
                Para_dict["dropout_expert"]=dropout_expert,
                Para_dict["lossW_expert_on"]=lossW_expert_on,
                Para_dict["lossW_expert"]=lossW_expert,
                Para_dict["class_weight"]=class_weight,
                Para_dict["metrics"]=model_metrics,
                
            ###############################Expert Mode values##################
            expert_mode_before = False #There was no expert mode used before.
            if expert_mode==True:
                #activate groupBox_expertMode_pop
                self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(True)
                expert_mode_before = True
            #Some settings only neet to be changed once, after user clicked apply at next epoch
                        
                #Apply the changes to trainable states:
                if train_last_layers==True:#Train only the last n layers
                    print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                    trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)
                    text1 = "Expert mode: Request for custom trainability states: train only the last "+str(train_last_layers_n)+ " layer(s)\n"
                    #text2 = "\n--------------------\n"
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text1+summary)
                if train_dense_layers==True:#Train only dense layers
                    print("Train only dense layers")
                    layer_dense_ind = ["Dense" in x for x in layer_names]
                    layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                    #create a list of trainable states
                    trainable_new = len(trainable_original)*[False]
                    for index in layer_dense_ind:
                        trainable_new[index] = True
                    summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)                  
                    text1 = "Expert mode: Request for custom trainability states: train only dense layer(s)\n"
                    #text2 = "\n--------------------\n"
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text1+summary)

                if dropout_expert_on==True:
                    #The user apparently want to change the dropout rates
                    do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                    #Compare the dropout values in the model to the dropout values requested by user
                    if len(dropout_expert)==1:#if the user gave a single float
                        dropout_expert_list = len(do_list)*dropout_expert #convert to list
                    elif len(dropout_expert)>1:
                        dropout_expert_list = dropout_expert
                        if not len(dropout_expert_list)==len(do_list):
                            text = "Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers"
                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)
                    else:
                        text = "Could not understand user input at Expert->Dropout"
                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)
                        dropout_expert_list = []
                    if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                        do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)
                        if do_changed==1:
                            text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                        else:
                            text_do = "Dropout rate(s) in model was/were not changed"
                    else:
                        text_do = "Dropout rate(s) in model was/were not changed"
                    print(text_do)
                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text_do)


            text_updates = ""
            #Compare current lr and the lr on expert tab:
            if collection == False:
                lr_current = K.eval(model_keras.optimizer.lr)
            else:
                lr_current = K.eval(model_keras[0].optimizer.lr)

            lr_diff = learning_rate_expert-lr_current
            if  abs(lr_diff) > 1e-6:
                if collection == False:
                    K.set_value(model_keras.optimizer.lr, learning_rate_expert)
                if collection == True:
                    for m in model_keras:
                        K.set_value(m.optimizer.lr, learning_rate_expert)
                text_updates +=  "Changed the learning rate to "+ str(learning_rate_expert)+"\n"
            recompile = False
            #Compare current optimizer and the optimizer on expert tab:
            if collection==False:
                optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
            if collection==True:
                optimizer_current = aid_dl.get_optimizer_name(model_keras[0]).lower()#get the current optimizer of the model

            if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                recompile = True
                text_updates+="Changed the optimizer to "+optimizer_expert+"\n"

            #Compare current loss function and the loss-function on expert tab:
            if collection==False:
                if model_keras.loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"
            if collection==True:
                if model_keras[0].loss!=loss_expert:
                    recompile = True
                    text_updates+="Changed the loss function to "+loss_expert+"\n"





            if recompile==True:
                print("Recompiling...")
                if collection==False:
                    aid_dl.model_compile(model_keras,loss_expert,optimizer_expert,learning_rate_expert,model_metrics,nr_classes)
                if collection==True:
                    for m in model_keras[1]:
                        aid_dl.model_compile(model_keras, loss_expert, optimizer_expert, learning_rate_expert,model_metrics, nr_classes)

            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text_updates)

            self.model_keras = model_keras #overwrite the model on self

            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_train = np.array(SelectedFiles)[ind]
            SelectedFiles_train = list(SelectedFiles_train)
            indices_train = [selectedfile["class"] for selectedfile in SelectedFiles_train]
            nr_events_epoch_train = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_train]
            rtdc_path_train = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_train]
            zoom_factors_train = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_train]
            zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            shuffle_train = [selectedfile["shuffle"] for selectedfile in SelectedFiles_train]
            
            #read self.ram to new variable ; next clear ram. This is required for multitasking (training multiple models with maybe different data)
            DATA = self.ram
            if verbose==1:
                print("Length of DATA (in RAM) = "+str(len(DATA)))
            #clear the ram again 
            self.ram = dict()
            #If the scaling method is "divide by mean and std of the whole training set":
            if norm == "StdScaling using mean and std of all training data":
                mean_trainingdata,std_trainingdata = [],[]
                for i in range(len(SelectedFiles_train)):
                    #if Data_to_RAM was not enabled:
                    #if not self.actionDataToRam.isChecked():
                    if len(DATA)==0: #Here, the entire training set needs to be used! Not only random images!
                        gen_train = aid_img.gen_crop_img(crop,rtdc_path_train[i],random_images=False,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
#                    else: #get a similar generator, using the ram-data
#                        if len(DATA)==0:
#                            gen_train = aid_img.gen_crop_img(crop,rtdc_path_train[i],random_images=False) #Replace true means that individual cells could occur several times
                    else:
                        gen_train = aid_img.gen_crop_img_ram(DATA,rtdc_path_train[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                        
                    images = next(gen_train)[0]
                    mean_trainingdata.append(np.mean(images))
                    std_trainingdata.append(np.std(images))
                mean_trainingdata = np.mean(np.array(mean_trainingdata))
                std_trainingdata = np.mean(np.array(std_trainingdata))
                
                if np.allclose(std_trainingdata,0):
                    std_trainingdata = 0.0001

                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    text = "<html><head/><body><p>The standard deviation of your training data is zero! This would lead to division by zero. To avoid this, I will divide by 0.0001 instead.</p></body></html>"
                    msg.setText(text) 
                    msg.setWindowTitle("Std. is zero")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()

                #This needs to be saved into Para_dict since it will be required for inference
                Para_dict["Mean of training data used for scaling"]=mean_trainingdata,
                Para_dict["Std of training data used for scaling"]=std_trainingdata,
            if collection==False:
                Para_dict.to_excel(writer,sheet_name='Parameters')
                #self.model_meta = writer
                if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH)#change to writable
                writer.save()
                os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)#change to readable, non-writable
            if collection==True:
                for i in range(len(Writers)):
                    Para_dict.to_excel(Writers[i],sheet_name='Parameters')
                    #self.model_meta = writer
                    if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH)
                    Writers[i].save()
                    os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)

            def update_para_dict():
                norm = str(self.fittingpopups_ui[listindex].comboBox_Normalization_pop.currentText())
                nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_NrEpochs_pop.value())
                keras_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterEpochs_pop.value())
                h_flip = bool(self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.isChecked())
                v_flip = bool(self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.isChecked())
                rotation = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.text())
         
                width_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_2.text())
                height_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_3.text())
                zoom = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_4.text())
                shear = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_5.text())
                
                brightness_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterNrEpochs_pop.value())
                brightness_add_lower = float(self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.value())
                brightness_add_upper = float(self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.value())
                brightness_mult_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.value())
                brightness_mult_upper = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.value())
                gaussnoise_mean = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.value())
                gaussnoise_scale = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.value())

                contrast_on = bool(self.fittingpopups_ui[listindex].checkBox_contrast_pop.isChecked())
                contrast_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.value())
                contrast_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.value())
                saturation_on = bool(self.fittingpopups_ui[listindex].checkBox_saturation_pop.isChecked())
                saturation_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.value())
                saturation_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.value())
                hue_on = bool(self.fittingpopups_ui[listindex].checkBox_hue_pop.isChecked())
                hue_delta = float(self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.value())

                avgBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.isChecked())        
                avgBlur_min = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.value())
                avgBlur_max = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.value())
                gaussBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.isChecked())        
                gaussBlur_min = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.value())
                gaussBlur_max = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.value())
                motionBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.isChecked())        
                motionBlur_kernel = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.text())
                motionBlur_angle = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.text())
                motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
                motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple

                expert_mode = bool(self.fittingpopups_ui[listindex].groupBox_expertMode_pop.isChecked())
                batchSize_expert = int(self.fittingpopups_ui[listindex].spinBox_batchSize_pop.value())
                epochs_expert = int(self.fittingpopups_ui[listindex].spinBox_epochs_pop.value())
                learning_rate_expert = float(self.fittingpopups_ui[listindex].doubleSpinBox_learningRate_pop.value())
                
                learning_rate_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_learningRate_pop.isChecked())
                loss_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.isChecked())
                loss_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.currentText())
                optimizer_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_optimizer_pop.isChecked())
                optimizer_expert = str(self.fittingpopups_ui[listindex].comboBox_optimizer_pop.currentText())

                padding_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_expt_paddingMode_pop.isChecked())
                padding_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_paddingMode_pop.currentText())

                train_last_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainLastNOnly_pop.isChecked())             
                train_last_layers_n = int(self.fittingpopups_ui[listindex].spinBox_trainLastNOnly_pop.value())              
                train_dense_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainDenseOnly_pop.isChecked())             
                dropout_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_dropout_pop.isChecked())             
                try:
                    dropout_expert = str(self.fittingpopups_ui[listindex].lineEdit_dropout_pop.text()) #due to the validator, there are no squ.brackets
                    dropout_expert = "["+dropout_expert+"]"
                    dropout_expert = ast.literal_eval(dropout_expert)
                except:
                    dropout_expert = []
                #Issue here! This toggles call of lossweights_popup! Can this be prevented?
                lossW_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_lossW.isChecked())             
                lossW_expert = str(self.fittingpopups_ui[listindex].lineEdit_lossW.text())             
                class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert)
                
                #Document this change in the meta-file
                Para_dict["AIDeveloper_Version"]=VERSION,
                Para_dict["model_zoo_version"]=model_zoo_version,
                try:
                    Para_dict["OS"]=platform.platform(),
                    Para_dict["CPU"]=platform.processor(),
                except:
                    Para_dict["OS"]="Unknown",
                    Para_dict["CPU"]="Unknown",

                Para_dict["Modelname"]=new_modelname,
                Para_dict["Chosen Model"]=chosen_model,
                Para_dict["new_model"]=new_model,
                Para_dict["loadrestart_model"]=loadrestart_model,
                Para_dict["loadcontinue_model"]=loadcontinue_model,
                Para_dict["Continued_Fitting_From"]=load_modelname,                        
                Para_dict["Input image size"]=crop,
                Para_dict["Color Mode"]=color_mode,
                Para_dict["Output Nr. classes"]=nr_classes,
                Para_dict["Normalization"]=norm,
                Para_dict["Nr. epochs"]=nr_epochs,
                Para_dict["Keras refresh after nr. epochs"]=keras_refresh_nr_epochs,
                Para_dict["Horz. flip"]=h_flip,
                Para_dict["Vert. flip"]=v_flip,
                Para_dict["rotation"]=rotation,
                Para_dict["width_shift"]=width_shift,
                Para_dict["height_shift"]=height_shift,
                Para_dict["zoom"]=zoom,
                Para_dict["shear"]=shear,
                Para_dict["Brightness refresh after nr. epochs"]=brightness_refresh_nr_epochs,
                Para_dict["Brightness add. lower"]=brightness_add_lower,
                Para_dict["Brightness add. upper"]=brightness_add_upper,
                Para_dict["Brightness mult. lower"]=brightness_mult_lower,  
                Para_dict["Brightness mult. upper"]=brightness_mult_upper,
                Para_dict["Gaussnoise Mean"]=gaussnoise_mean,
                Para_dict["Gaussnoise Scale"]=gaussnoise_scale,

                Para_dict["Contrast on"]=contrast_on,                
                Para_dict["Contrast Lower"]=contrast_lower,
                Para_dict["Contrast Higher"]=contrast_higher,
                Para_dict["Saturation on"]=saturation_on,
                Para_dict["Saturation Lower"]=saturation_lower,
                Para_dict["Saturation Higher"]=saturation_higher,
                Para_dict["Hue on"]=hue_on,                
                Para_dict["Hue delta"]=hue_delta,                

                Para_dict["Average blur on"]=avgBlur_on,                
                Para_dict["Average blur Lower"]=avgBlur_min,
                Para_dict["Average blur Higher"]=avgBlur_max,
                Para_dict["Gauss blur on"]=gaussBlur_on,                
                Para_dict["Gauss blur Lower"]=gaussBlur_min,
                Para_dict["Gauss blur Higher"]=gaussBlur_max,
                Para_dict["Motion blur on"]=motionBlur_on,                
                Para_dict["Motion blur Kernel"]=motionBlur_kernel,                
                Para_dict["Motion blur Angle"]=motionBlur_angle,                
           
                Para_dict["Epoch_Started_Using_These_Settings"]=counter,

                Para_dict["expert_mode"]=expert_mode,
                Para_dict["batchSize_expert"]=batchSize_expert,
                Para_dict["epochs_expert"]=epochs_expert,
                
                Para_dict["learning_rate_expert"]=learning_rate_expert,
                Para_dict["learning_rate_expert_on"]=learning_rate_expert_on,
                Para_dict["loss_expert_on"]=loss_expert_on,
                Para_dict["loss_expert"]=loss_expert,
                Para_dict["optimizer_expert_on"]=optimizer_expert_on,
                Para_dict["optimizer_expert"]=optimizer_expert,                

                Para_dict["padding_expert_on"]=padding_expert_on,
                Para_dict["padding_expert"]=padding_expert,                
                
                Para_dict["train_last_layers"]=train_last_layers,
                Para_dict["train_last_layers_n"]=train_last_layers_n,
                Para_dict["train_dense_layers"]=train_dense_layers,
                Para_dict["dropout_expert_on"]=dropout_expert_on,
                Para_dict["dropout_expert"]=dropout_expert,
                Para_dict["lossW_expert_on"]=lossW_expert_on,
                Para_dict["lossW_expert"]=lossW_expert,
                Para_dict["class_weight"]=class_weight,
                Para_dict["metrics"]=model_metrics,
                
                #training data cannot be changed during training
                if norm == "StdScaling using mean and std of all training data":                                
                    #This needs to be saved into Para_dict since it will be required for inference
                    Para_dict["Mean of training data used for scaling"]=mean_trainingdata,
                    Para_dict["Std of training data used for scaling"]=std_trainingdata,

                if collection==False:
                    Para_dict.to_excel(self.writer,sheet_name='Parameters',startrow=self.writer.sheets['Parameters'].max_row,header= False)
                    #self.model_meta = writer
                    if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH)#change to read/write
                    try:
                        self.writer.save()
                    except:
                        pass
                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)#change to only readable

                if collection==True:
                    for i in range(len(Writers)):
                        Para_dict["Chosen Model"]=model_architecture_names[i],
                        writer = Writers[i]
                        Para_dict.to_excel(writer,sheet_name='Parameters',startrow=writer.sheets['Parameters'].max_row,header= False)
                        #self.model_meta = writer
                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                        try:
                            writer.save()
                        except:
                            pass
                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH) #read only



            ######################Load the Validation Data################################
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
            ind = np.where(np.array(ind)==True)[0]
            SelectedFiles_valid = np.array(SelectedFiles)[ind]
            SelectedFiles_valid = list(SelectedFiles_valid)
            indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
            nr_events_epoch_valid = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_valid]
            rtdc_path_valid = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_valid]
            zoom_factors_valid = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_valid]
            zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            zoom_order = int(np.where(np.array(zoom_order)==True)[0])
            shuffle_valid = [selectedfile["shuffle"] for selectedfile in SelectedFiles_valid]
        
            ############Cropping#####################
            X_valid,y_valid,Indices = [],[],[]
            for i in range(len(SelectedFiles_valid)):
                if not self.actionDataToRam.isChecked():
                    gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else: #get a similar generator, using the ram-data
                    if len(DATA)==0:
                        gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:
                        gen_valid = aid_img.gen_crop_img_ram(DATA,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                generator_cropped_out = next(gen_valid)
                X_valid.append(generator_cropped_out[0])
                #y_valid.append(np.repeat(indices_valid[i],nr_events_epoch_valid[i]))
                y_valid.append(np.repeat(indices_valid[i],X_valid[-1].shape[0]))
                Indices.append(generator_cropped_out[1])
                del generator_cropped_out
            #Save the validation set (BEFORE normalization!)
            #Write to.rtdc files
            if bool(self.actionExport_Original.isChecked())==True:
                print("Export original images")
                save_cropped = False
                aid_bin.write_rtdc(new_modelname.split(".model")[0]+'_Valid_Data.rtdc',rtdc_path_valid,X_valid,Indices,cropped=save_cropped,color_mode=self.get_color_mode())

            elif bool(self.actionExport_Cropped.isChecked())==True:
                print("Export cropped images")
                save_cropped = True
                aid_bin.write_rtdc(new_modelname.split(".model")[0]+'_Valid_Data.rtdc',rtdc_path_valid,X_valid,Indices,cropped=save_cropped,color_mode=self.get_color_mode())

            elif bool(self.actionExport_Off.isChecked())==True:
                print("Exporting is turned off")
#                msg = QtWidgets.QMessageBox()
#                msg.setIcon(QtWidgets.QMessageBox.Information)       
#                msg.setText("Use a different Exporting option in ->Edit if you want to export the data")
#                msg.setWindowTitle("Export is turned off!")
#                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                msg.exec_()
                
            X_valid = np.concatenate(X_valid)
            y_valid = np.concatenate(y_valid)
            Y_valid = np_utils.to_categorical(y_valid, nr_classes)# * 2 - 1

            if not bool(self.actionExport_Off.isChecked())==True:
                #Save the labels
                np.savetxt(new_modelname.split(".model")[0]+'_Valid_Labels.txt',y_valid.astype(int),fmt='%i')            

            if len(X_valid.shape)==4:
                channels=3
            elif len(X_valid.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X_valid.shape))
            if channels==1:
                #Add the "channels" dimension
                X_valid = np.expand_dims(X_valid,3)

            #get it to theano image format (channels first)
            #X_valid = X_valid.swapaxes(-1,-2).swapaxes(-2,-3)
            if norm == "StdScaling using mean and std of all training data":
                X_valid = aid_img.norm_imgs(X_valid,norm,mean_trainingdata,std_trainingdata)
            else:
                X_valid = aid_img.norm_imgs(X_valid,norm)

            #Validation data can be cropped to final size already since no augmentation
            #will happen on this data set
            dim_val = X_valid.shape
            print("Current dim. of validation set (pixels x pixels)="+str(dim_val[2]))
            if dim_val[2]!=crop:
                print("Change dim. (pixels x pixels) of validation set to ="+str(crop))
                remove = int(dim_val[2]/2.0 - crop/2.0)
                X_valid = X_valid[:,remove:-remove,remove:-remove,:] #crop to crop x crop pixels #TensorFlow


            ####################Update the PopupFitting########################
            self.fittingpopups_ui[listindex].lineEdit_modelname_pop.setText(new_modelname) #set the progress bar to zero
            self.fittingpopups_ui[listindex].spinBox_imagecrop_pop.setValue(crop)
            self.fittingpopups_ui[listindex].spinBox_NrEpochs_pop.setValue(nr_epochs)
            self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.addItems(self.predefined_models)
            chosen_model = str(self.comboBox_ModelSelection.currentText())
            index = self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.findText(chosen_model, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_ModelSelection_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].comboBox_Normalization_pop.addItems(self.norm_methods)            
            index = self.fittingpopups_ui[listindex].comboBox_Normalization_pop.findText(norm, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_Normalization_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].spinBox_RefreshAfterEpochs_pop.setValue(keras_refresh_nr_epochs)
            self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.setChecked(h_flip)
            self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.setChecked(v_flip)
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.setText(str(rotation))
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_2.setText(str(width_shift))
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_3.setText(str(height_shift))
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_4.setText(str(zoom))
            self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_5.setText(str(shear))
            self.fittingpopups_ui[listindex].spinBox_RefreshAfterNrEpochs_pop.setValue(brightness_refresh_nr_epochs)
            self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.setValue(brightness_add_lower)
            self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.setValue(brightness_add_upper)
            self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.setValue(brightness_mult_lower)
            self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.setValue(brightness_mult_upper)
            self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.setValue(gaussnoise_mean)
            self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.setValue(gaussnoise_scale) 

            self.fittingpopups_ui[listindex].checkBox_contrast_pop.setChecked(contrast_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.setValue(contrast_lower) 
            self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.setValue(contrast_higher) 
            self.fittingpopups_ui[listindex].checkBox_saturation_pop.setChecked(saturation_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.setValue(saturation_lower) 
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.setValue(saturation_higher) 
            self.fittingpopups_ui[listindex].checkBox_hue_pop.setChecked(hue_on) 
            self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.setValue(hue_delta) 
            #Special for saturation and hue. Only enabled for RGB:
            saturation_enabled = bool(self.checkBox_saturation.isEnabled())
            self.fittingpopups_ui[listindex].checkBox_saturation_pop.setEnabled(saturation_enabled)
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.setEnabled(saturation_enabled)
            self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.setEnabled(saturation_enabled)
                
            hue_enabled = bool(self.checkBox_hue.isEnabled())
            self.fittingpopups_ui[listindex].checkBox_hue_pop.setEnabled(hue_enabled) 
            self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.setEnabled(hue_enabled)


            self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.setChecked(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].label_avgBlurMin_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.setValue(avgBlur_min) 
            self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].label_avgBlurMax_pop.setEnabled(avgBlur_on)
            self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.setValue(avgBlur_max) 

            self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.setChecked(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].label_gaussBlurMin_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.setValue(gaussBlur_min) 
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].label_gaussBlurMax_pop.setEnabled(gaussBlur_on)
            self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.setValue(gaussBlur_max) 

            self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.setChecked(motionBlur_on)
            self.fittingpopups_ui[listindex].label_motionBlurKernel_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].label_motionBlurAngle_pop.setEnabled(motionBlur_on)
            self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setEnabled(motionBlur_on)
            if len(motionBlur_kernel)==1:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setText(str(motionBlur_kernel[0]))
            if len(motionBlur_kernel)==2:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.setText(str(motionBlur_kernel[0])+","+str(motionBlur_kernel[1]))
            if len(motionBlur_angle)==1:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setText(str(motionBlur_angle[0]))
            if len(motionBlur_kernel)==2:
                self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.setText(str(motionBlur_angle[0])+","+str(motionBlur_angle[1]))

            self.fittingpopups_ui[listindex].groupBox_expertMode_pop.setChecked(expert_mode)
            self.fittingpopups_ui[listindex].spinBox_batchSize_pop.setValue(batchSize_expert)
            self.fittingpopups_ui[listindex].spinBox_epochs_pop.setValue(epochs_expert)

            self.fittingpopups_ui[listindex].checkBox_learningRate_pop.setChecked(learning_rate_expert_on)
            self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.setChecked(loss_expert_on)
            index = self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.findText(loss_expert, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].checkBox_optimizer_pop.setChecked(optimizer_expert_on)
            index = self.fittingpopups_ui[listindex].comboBox_optimizer_pop.findText(optimizer_expert, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_optimizer_pop.setCurrentIndex(index)
            self.fittingpopups_ui[listindex].doubleSpinBox_learningRate_pop.setValue(learning_rate_expert)

            self.fittingpopups_ui[listindex].checkBox_expt_paddingMode_pop.setChecked(padding_expert_on)
            index = self.fittingpopups_ui[listindex].comboBox_expt_paddingMode_pop.findText(padding_expert, QtCore.Qt.MatchFixedString)
            if index >= 0:
                self.fittingpopups_ui[listindex].comboBox_expt_paddingMode_pop.setCurrentIndex(index)

            self.fittingpopups_ui[listindex].checkBox_trainLastNOnly_pop.setChecked(train_last_layers)
            self.fittingpopups_ui[listindex].spinBox_trainLastNOnly_pop.setValue(train_last_layers_n)
            self.fittingpopups_ui[listindex].checkBox_trainDenseOnly_pop.setChecked(train_dense_layers)
            self.fittingpopups_ui[listindex].checkBox_dropout_pop.setChecked(dropout_expert_on)
            do_text = [str(do_i) for do_i in dropout_expert]
            self.fittingpopups_ui[listindex].lineEdit_dropout_pop.setText((', '.join(do_text)))
            self.fittingpopups_ui[listindex].checkBox_lossW.setChecked(lossW_expert_on)
            self.fittingpopups_ui[listindex].pushButton_lossW.setEnabled(lossW_expert_on)
            self.fittingpopups_ui[listindex].lineEdit_lossW.setText(str(lossW_expert))

            if channels==1:
                channel_text = "Grayscale"
            elif channels==3:
                channel_text = "RGB"
            self.fittingpopups_ui[listindex].comboBox_colorMode_pop.addItems([channel_text])

            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
        
            #Dictionary defining affine image augmentation options:
            aug_paras = {"v_flip":v_flip,"h_flip":h_flip,"rotation":rotation,"width_shift":width_shift,"height_shift":height_shift,"zoom":zoom,"shear":shear}
                         
            Histories,Index,Saved,Stopwatch = [],[],[],[]
            if collection==True:
               HISTORIES = [ [] for model in model_keras]
               SAVED = [ [] for model in model_keras]

            counter = 0
#            thresh_acc = 0
#            thresh_loss = 9E20
            
            model_metrics_names = []
            for met in model_metrics:
                if type(met)==str:
                    model_metrics_names.append(met) 
                else:
                    metname = met.name
                    metlabel = met.label
                    if metlabel>0:
                        metname = metname+"_"+str(metlabel)
                    model_metrics_names.append(metname) 
            
            #Dictionary for records in metrics
            model_metrics_records = {}
            model_metrics_records["acc"] = 0 #accuracy  starts at zero and approaches 1 during training         
            model_metrics_records["val_acc"] = 0 #accuracy  starts at zero and approaches 1 during training         
            model_metrics_records["loss"] = 9E20 ##loss starts very high and approaches 0 during training         
            model_metrics_records["val_loss"] = 9E20 ##loss starts very high and approaches 0 during training         
            for key in model_metrics_names:
                if 'precision' in key or 'recall' in key or 'f1_score' in key:
                    model_metrics_records[key] = 0 #those metrics start at zero and approach 1         
                    model_metrics_records["val_"+key] = 0 #those metrics start at zero and approach 1         

            
            time_start = time.time()
            t1 = time.time() #Initialize a timer; this is used to save the meta file every few seconds
            t2 =  time.time() #Initialize a timer; this is used update the fitting parameters
            while counter < nr_epochs:#nr_epochs: #resample nr_epochs times
                #Only keep fitting if the respective window is open:
                isVisible = self.fittingpopups[listindex].isVisible()
                if isVisible:                    
                    ############Keras image augmentation#####################
                    #Start the first iteration:                
                    X_train,y_train = [],[]
                    t3 = time.time()
                    for i in range(len(SelectedFiles_train)):
#                        if not self.actionDataToRam.isChecked():
                        if len(DATA)==0:
                            gen_train = aid_img.gen_crop_img(cropsize2,rtdc_path_train[i],nr_events_epoch_train[i],random_images=shuffle_train[i],replace=True,zoom_factor=zoom_factors_train[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
#                        else:
#                            if len(DATA)==0:
#                                gen_train = aid_img.gen_crop_img(cropsize2,rtdc_path_train[i],nr_events_epoch_train[i],replace=True) #Replace true means that individual cells could occur several times
                        else:
                            gen_train = aid_img.gen_crop_img_ram(DATA,rtdc_path_train[i],nr_events_epoch_train[i],random_images=shuffle_train[i],replace=True) #Replace true means that individual cells could occur several times
                            if self.actionVerbose.isChecked():
                                print("Loaded data from RAM")
                        X_train.append(next(gen_train)[0])
                        #y_train.append(np.repeat(indices_train[i],nr_events_epoch_train[i])) #This does not work if shuffle=False, because actually not all images can be used (the given number in the table (Cells/Epoch) is actually wrong!)
                        y_train.append(np.repeat(indices_train[i],X_train[-1].shape[0]))

                    X_train = np.concatenate(X_train)
                    X_train = X_train.astype(np.uint8)
                        
                    y_train = np.concatenate(y_train)
                    t4 = time.time()
                    if verbose == 1:
                        print("Time to load data (from .rtdc or RAM) and crop="+str(t4-t3))
                    
                    if len(X_train.shape)==4:
                        channels=3
                    elif len(X_train.shape)==3:
                        channels=1
                    else:
                        print("Invalid data dimension:" +str(X_valid.shape))
                    if channels==1:
                        #Add the "channels" dimension
                        X_train = np.expand_dims(X_train,3)

                    t3 = time.time()
                    #Some parallellization: use nr_threads (number of CPUs)
                    nr_threads = 1 #Somehow for MNIST and CIFAR, processing always takes longer. I tried nr_threads=2,4,8,16,24
                    if nr_threads == 1:
                        X_batch = aid_img.affine_augm(X_train,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear) #Affine image augmentation
                        y_batch = np.copy(y_train)
                    else:
                        #Divde data in 4 batches
                        X_train = np.array_split(X_train,nr_threads)
                        y_train = np.array_split(y_train,nr_threads)

                        self.X_batch = [False] * nr_threads
                        self.y_batch = [False] * nr_threads
                        self.counter_aug = 0
                        self.Workers_augm = []
                        
                        def imgaug_worker(aug_paras,progress_callback,history_callback):
                            i = aug_paras["i"]
                            self.X_batch[i] = aid_img.affine_augm(aug_paras["X_train"],v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear)
                            self.y_batch[i] = aug_paras["y_train"]
                            self.counter_aug+=1
    
                        t3_a = time.time()
                        for i in range(nr_threads):
                            aug_paras_ = copy.deepcopy(aug_paras)
                            aug_paras_["i"] = i
                            aug_paras_["X_train"]=X_train[i]#augparas contains rotation and so on. X_train and y_train are overwritten in each iteration (for each worker new X_train)
                            aug_paras_["y_train"]=y_train[i]
                            
                            self.Workers_augm.append(Worker(imgaug_worker,aug_paras_))                            
                            self.threadpool.start(self.Workers_augm[i])
                            
                        while self.counter_aug < nr_threads:
                            time.sleep(0.01)#Wait 0.1s, then check the len again
                        t3_b = time.time()
                        if verbose == 1:
                            print("Time to perform affine augmentation_internal ="+str(t3_b-t3_a))
    
                        X_batch = np.concatenate(self.X_batch)
                        y_batch = np.concatenate(self.y_batch)
       
                    Y_batch = np_utils.to_categorical(y_batch, nr_classes)# * 2 - 1
                    t4 = time.time()
                    if verbose == 1:
                        print("Time to perform affine augmentation ="+str(t4-t3))
                            
                    t3 = time.time()            
                    #Now do the final cropping to the actual size that was set by user
                    dim = X_batch.shape
                    if dim[2]!=crop:
                        remove = int(dim[2]/2.0 - crop/2.0)
                        X_batch = X_batch[:,remove:-remove,remove:-remove,:] #crop to crop x crop pixels #TensorFlow
                    t4 = time.time()
#                    if verbose == 1:
#                        print("Time to crop to final size="+str(t4-t3))

                    X_batch_orig = np.copy(X_batch) #save into new array and do some iterations with varying noise/brightness
                    #reuse this X_batch_orig a few times since this augmentation was costly
                    keras_iter_counter = 0
                    while keras_iter_counter < keras_refresh_nr_epochs and counter < nr_epochs:
                        keras_iter_counter+=1
                        #if t2-t1>5: #check for changed settings every 5 seconds
                        if self.actionVerbose.isChecked()==True:
                            verbose = 1
                        else:
                            verbose = 0                            
                                                
#                        X_batch = np.copy(X_batch_orig)#copy from X_batch_orig, X_batch will be altered without altering X_batch_orig            
#                        X_batch = X_batch.astype(float)
                        
                        #Another while loop if the user wants to reuse the keras-augmented data
                        #several times and only apply brightness augmentation:
                        brightnesss_iter_counter = 0
                        while brightnesss_iter_counter < brightness_refresh_nr_epochs and counter < nr_epochs:
                            #In each iteration, start with non-augmented data
                            X_batch = np.copy(X_batch_orig)#copy from X_batch_orig, X_batch will be altered without altering X_batch_orig            
                            X_batch = X_batch.astype(np.uint8)                            
                            
                            #########X_batch = X_batch.astype(float)########## No floating yet :) !!!
                            
                            brightnesss_iter_counter += 1
                            if self.actionVerbose.isChecked()==True:
                                verbose = 1
                            else:
                                verbose = 0                            

                            if self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.isChecked():
                                nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_NrEpochs_pop.value())
                                #Keras stuff
                                keras_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterEpochs_pop.value())                                
                                h_flip = bool(self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.isChecked())
                                v_flip = bool(self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.isChecked())
                                rotation = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.text())
                                width_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_2.text())
                                height_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_3.text())
                                zoom = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_4.text())
                                shear = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_5.text())
                                #Brightness stuff
                                brightness_refresh_nr_epochs = int(self.fittingpopups_ui[listindex].spinBox_RefreshAfterNrEpochs_pop.value())
                                brightness_add_lower = float(self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.value())
                                brightness_add_upper = float(self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.value())
                                brightness_mult_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.value())
                                brightness_mult_upper = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.value())
                                gaussnoise_mean = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.value())
                                gaussnoise_scale = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.value())

                                contrast_on = bool(self.fittingpopups_ui[listindex].checkBox_contrast_pop.isChecked())
                                contrast_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.value())
                                contrast_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.value())
                                saturation_on = bool(self.fittingpopups_ui[listindex].checkBox_saturation_pop.isChecked())
                                saturation_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.value())
                                saturation_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.value())
                                hue_on = bool(self.fittingpopups_ui[listindex].checkBox_hue_pop.isChecked())
                                hue_delta = float(self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.value())

                                avgBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.isChecked())        
                                avgBlur_min = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.value())
                                avgBlur_max = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.value())
                    
                                gaussBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.isChecked())        
                                gaussBlur_min = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.value())
                                gaussBlur_max = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.value())
                    
                                motionBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.isChecked())        
                                motionBlur_kernel = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.text())
                                motionBlur_angle = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.text())
                                
                                motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
                                motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple


                                #Expert mode stuff
                                expert_mode = bool(self.fittingpopups_ui[listindex].groupBox_expertMode_pop.isChecked())
                                batchSize_expert = int(self.fittingpopups_ui[listindex].spinBox_batchSize_pop.value())
                                epochs_expert = int(self.fittingpopups_ui[listindex].spinBox_epochs_pop.value())
                                
                                learning_rate_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_learningRate_pop.isChecked())
                                learning_rate_expert = float(self.fittingpopups_ui[listindex].doubleSpinBox_learningRate_pop.value())
                                loss_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_expt_loss_pop.isChecked())
                                loss_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_loss_pop.currentText())
                                optimizer_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_optimizer_pop.isChecked())
                                optimizer_expert = str(self.fittingpopups_ui[listindex].comboBox_optimizer_pop.currentText())
                                padding_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_expt_paddingMode_pop.isChecked())
                                padding_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_paddingMode_pop.currentText())

                                train_last_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainLastNOnly_pop.isChecked())             
                                train_last_layers_n = int(self.fittingpopups_ui[listindex].spinBox_trainLastNOnly_pop.value())              
                                train_dense_layers = bool(self.fittingpopups_ui[listindex].checkBox_trainDenseOnly_pop.isChecked())             
                                dropout_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_dropout_pop.isChecked())             
                                try:
                                    dropout_expert = str(self.fittingpopups_ui[listindex].lineEdit_dropout_pop.text()) #due to the validator, there are no squ.brackets
                                    dropout_expert = "["+dropout_expert+"]"
                                    dropout_expert = ast.literal_eval(dropout_expert)
                                except:
                                    dropout_expert = []
                                lossW_expert_on = bool(self.fittingpopups_ui[listindex].checkBox_lossW.isChecked())             
                                lossW_expert = str(self.fittingpopups_ui[listindex].lineEdit_lossW.text())             
                                class_weight = self.get_class_weight(self.fittingpopups_ui[listindex].SelectedFiles,lossW_expert) #
                                
                                print("Updating parameter file (meta.xlsx)!")
                                update_para_dict()

                                #Changes in expert mode can affect the model: apply changed now:
                                if expert_mode==True:
                                    if collection==False: #Expert mode is currently not supported for Collections
                                        expert_mode_before = True
    
                                        #Apply changes to the trainable states:
                                        if train_last_layers==True:#Train only the last n layers
                                            if verbose:
                                                print("Train only the last "+str(train_last_layers_n)+ " layer(s)")
                                            trainable_new = (len(trainable_original)-train_last_layers_n)*[False]+train_last_layers_n*[True]
                                            #Change the trainability states. Model compilation is done inside model_change_trainability
                                            summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)
                                            text1 = "Expert mode: Request for custom trainability states: train only the last "+str(train_last_layers_n)+ " layer(s)\n"
                                            #text2 = "\n--------------------\n"
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text1+summary)
                                        if train_dense_layers==True:#Train only dense layers
                                            if verbose:
                                                print("Train only dense layers")
                                            layer_dense_ind = ["Dense" in x for x in layer_names]
                                            layer_dense_ind = np.where(np.array(layer_dense_ind)==True)[0] #at which indices are dropout layers?
                                            #create a list of trainable states
                                            trainable_new = len(trainable_original)*[False]
                                            for index in layer_dense_ind:
                                                trainable_new[index] = True
                                            #Change the trainability states. Model compilation is done inside model_change_trainability
                                            summary = aid_dl.model_change_trainability(model_keras,trainable_new,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)                 
                                            text1 = "Expert mode: Request for custom trainability states: train only dense layer(s)\n"
                                            #text2 = "\n--------------------\n"
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text1+summary)
    
                                        if dropout_expert_on==True:
                                            #The user apparently want to change the dropout rates
                                            do_list = aid_dl.get_dropout(model_keras)#Get a list of dropout values of the current model
                                            #Compare the dropout values in the model to the dropout values requested by user
                                            if len(dropout_expert)==1:#if the user gave a float
                                                dropout_expert_list = len(do_list)*dropout_expert #convert to list
                                            elif len(dropout_expert)>1:
                                                dropout_expert_list = dropout_expert
                                                if not len(dropout_expert_list)==len(do_list):
                                                    text = "Issue with dropout: you defined "+str(len(dropout_expert_list))+" dropout rates, but model has "+str(len(do_list))+" dropout layers"
                                                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)
                                            else:
                                                text = "Could not understand user input at Expert->Dropout"
                                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)
                                                dropout_expert_list = []

                                            if len(dropout_expert_list)>0 and do_list!=dropout_expert_list:#if the dropout rates of the current model is not equal to the required do_list from user...
                                                #Change dropout. Model .compile happens inside change_dropout function
                                                do_changed = aid_dl.change_dropout(model_keras,dropout_expert_list,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)

                                                if do_changed==1:
                                                    text_do = "Dropout rate(s) in model was/were changed to: "+str(dropout_expert_list)
                                                else:
                                                    text_do = "Dropout rate(s) in model was/were not changed"
                                            else:
                                                text_do = "Dropout rate(s) in model was/were not changed"
                                            if verbose:
                                                print(text_do)
                                            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text_do)


                                ############################Invert 'expert' settings#########################
                                if expert_mode==False and expert_mode_before==True: #if the expert mode was selected before, change the parameters back to original vlaues
                                    if verbose:
                                        print("Expert mode was used before and settings are now inverted")

                                    #Re-set trainable states back to original state                                    
                                    if verbose:
                                        print("Change 'trainable' layers back to original state")
                                    summary = aid_dl.model_change_trainability(model_keras,trainable_original,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)                 
                                    text1 = "Expert mode turns off: Request for orignal trainability states:\n"
                                    #text2 = "\n--------------------\n"
                                    self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text1+summary)
                                    if verbose:
                                        print("Change dropout rates in dropout layers back to original values")
                                    if len(do_list_original)>0:
                                        do_changed = aid_dl.change_dropout(model_keras,do_list_original,model_metrics,nr_classes,loss_expert,optimizer_expert,learning_rate_expert)
                                        if do_changed==1:
                                            text_do = "Dropout rate(s) in model was/were changed to original values: "+str(do_list_original)
                                        else:
                                            text_do = "Dropout rate(s) in model was/were not changed"                                        
                                        self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text_do+"\n")


                                text_updates = ""
                                #Compare current lr and the lr on expert tab:
                                if collection==False:
                                    lr_current = K.eval(model_keras.optimizer.lr)
                                else:
                                    lr_current = K.eval(model_keras[0].optimizer.lr)

                                lr_diff = learning_rate_expert-lr_current
                                if  abs(lr_diff) > 1e-6:
                                    if collection==False:
                                        K.set_value(model_keras.optimizer.lr, learning_rate_expert)
                                    else:
                                        K.set_value(model_keras[0].optimizer.lr, learning_rate_expert)

                                    text_updates +=  "Changed the learning rate to "+ str(learning_rate_expert)+"\n"
                                recompile = False
                                #Compare current optimizer and the optimizer on expert tab:
                                if collection==False:
                                    optimizer_current = aid_dl.get_optimizer_name(model_keras).lower()#get the current optimizer of the model
                                else:
                                    optimizer_current = aid_dl.get_optimizer_name(model_keras[0]).lower()#get the current optimizer of the model

                                if optimizer_current!=optimizer_expert.lower():#if the current model has a different optimizer
                                    recompile = True
                                    text_updates+="Changed the optimizer to "+optimizer_expert+"\n"

                                #Compare current loss function and the loss-function on expert tab:
                                if collection==False:
                                    loss_ = model_keras.loss
                                else:
                                    loss_ = model_keras[0].loss
                                if loss_!=loss_expert:
                                    recompile = True
                                    model_metrics_records["loss"] = 9E20 #Reset the record for loss because new loss function could converge to a different min. value
                                    model_metrics_records["val_loss"] = 9E20 #Reset the record for loss because new loss function could converge to a different min. value
                                    text_updates+="Changed the loss function to "+loss_expert+"\n"

                                if recompile==True and collection==False:
                                    print("Recompiling...")
                                    aid_dl.model_compile(model_keras,loss_expert,optimizer_expert,learning_rate_expert,model_metrics,nr_classes)
                                elif recompile==True and collection==True:
                                    print("Recompiling...")
                                    for m in model_keras:
                                        aid_dl.model_compile(m,loss_expert,optimizer_expert,learning_rate_expert,model_metrics,nr_classes)

                                self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text_updates)

                                self.model_keras = model_keras #overwrite the model in self
                                self.fittingpopups_ui[listindex].checkBox_ApplyNextEpoch.setChecked(False)


                            ##########Contrast/Saturation/Hue augmentation#########
                            #is there any of contrast/saturation/hue augmentation to do?
                            X_batch = X_batch.astype(np.uint8)
                            t5 = time.time()
                            if contrast_on:
                                t_con_aug_1 = time.time()
                                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
                                t_con_aug_2 = time.time()
                                if verbose == 1:
                                    print("Time to augment contrast="+str(t_con_aug_2-t_con_aug_1))

                            if saturation_on or hue_on:
                                t_sat_aug_1 = time.time()
                                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta) #Gray and RGB; both values >0!
                                t_sat_aug_2 = time.time()
                                if verbose == 1:
                                    print("Time to augment saturation/hue="+str(t_sat_aug_2-t_sat_aug_1))

                            ##########Average/Gauss/Motion blurring#########
                            #is there any of blurring to do?
                            
                            if avgBlur_on:
                                t_avgBlur_1 = time.time()
                                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
                                t_avgBlur_2 = time.time()
                                if verbose == 1:
                                    print("Time to perform average blurring="+str(t_avgBlur_2-t_avgBlur_1))

                            if gaussBlur_on:
                                t_gaussBlur_1 = time.time()
                                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
                                t_gaussBlur_2 = time.time()
                                if verbose == 1:
                                    print("Time to perform gaussian blurring="+str(t_gaussBlur_2-t_gaussBlur_1))

                            if motionBlur_on:
                                t_motionBlur_1 = time.time()
                                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)
                                t_motionBlur_2 = time.time()
                                if verbose == 1:
                                    print("Time to perform motion blurring="+str(t_motionBlur_2-t_motionBlur_1))

                            ##########Brightness noise#########
                            t3 = time.time()
                            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)
                            t4 = time.time()
                            if verbose == 1:
                                print("Time to augment brightness="+str(t4-t3))

                            t3 = time.time()
                            if norm == "StdScaling using mean and std of all training data":
                                X_batch = aid_img.norm_imgs(X_batch,norm,mean_trainingdata,std_trainingdata)
                            else:
                                X_batch = aid_img.norm_imgs(X_batch,norm)
                            t4 = time.time()
                            if verbose == 1:
                                print("Time to apply normalization="+str(t4-t3))
                            
                            #Fitting can be paused
                            while str(self.fittingpopups_ui[listindex].pushButton_Pause_pop.text())=="":
                                time.sleep(2) #wait 2 seconds and then check the text on the button again

                            if verbose == 1: 
                                print("X_batch.shape")
                                print(X_batch.shape)
                                print("X_valid.shape")
                                print(X_valid.shape)

                            ###################################################
                            ###############Actual fitting######################
                            ###################################################

                            if collection==False:
                                if expert_mode == False:
                                    history = model_keras.fit(X_batch, Y_batch, batch_size=128, epochs=1,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=None)
                                elif expert_mode==True:
                                    #Here the user can determine batch size epochs and learning rate of the model
                                    #Fit model in with 'expert'-settings
                                    history = model_keras.fit(X_batch, Y_batch, batch_size=batchSize_expert, epochs=epochs_expert,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=class_weight)
                                
                                Histories.append(history.history)
                                Stopwatch.append(time.time()-time_start)

                                #Check if any metric broke a record
                                record_broken = False #initially, assume there is no new record
                                for key in history.history.keys():
                                    value = history.history[key][-1]
                                    record = model_metrics_records[key]
                                    if 'val_acc' in key or 'val_precision' in key or 'val_recall' in key or 'val_f1_score' in key:
                                        #These metrics should go up (towards 1)
                                        if value>record:
                                            model_metrics_records[key] = value
                                            record_broken = True
                                            print(key+" broke record -> Model is saved")

                                            #one could 'break' here, but I want to update all records
                                    elif 'val_loss' in key:
                                        #This metric should go down (towards 0)
                                        if value<record:
                                            model_metrics_records[key] = value
                                            record_broken = True
                                            print(key+" broke record -> Model is saved")

                                if record_broken:
                                    #Save the model
                                    model_keras.save(new_modelname.split(".model")[0]+"_"+str(counter)+".model")
                                    Saved.append(1)
                                
                                #Also save the model upon user-request  
                                elif bool(self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.isChecked())==True:
                                    model_keras.save(new_modelname.split(".model")[0]+"_"+str(counter)+".model")
                                    Saved.append(1)
                                    self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.setChecked(False)
                                else:
                                    Saved.append(0)

                            elif collection==True:
                                for i in range(len(model_keras)):
                                    if expert_mode==False:
                                        history = model_keras[i].fit(X_batch, Y_batch, batch_size=128, epochs=1,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=None)
                                        HISTORIES[i].append(history.history)
                                    elif expert_mode==True:
                                        #Here the user can determine batch size epochs and learning rate of the model
                                        #Fit model in with 'expert'-settings
                                        history = model_keras[i].fit(X_batch, Y_batch, batch_size=batchSize_expert, epochs=epochs_expert,verbose=verbose, validation_data=(X_valid, Y_valid),class_weight=class_weight)
                                        HISTORIES[i].append(history.history)
                                    
                                    print("model_keras_path[i]")
                                    print(model_keras_path[i])

                                    #Check if any metric broke a record
                                    record_broken = False #initially, assume there is no new record
                                    for key in history.history.keys():
                                        value = history.history[key][-1]
                                        record = model_metrics_records[key]
                                        if 'val_acc' in key or 'val_precision' in key or 'val_recall' in key or 'val_f1_score' in key:
                                            #These metrics should go up (towards 1)
                                            if value>record:
                                                model_metrics_records[key] = value
                                                record_broken = True
                                                print(key+" broke record -> Model is saved")
                                                #one could 'break' here, but I want to update all records
                                        elif 'val_loss' in key:
                                            #This metric should go down (towards 0)
                                            if value<record:
                                                model_metrics_records[key] = value
                                                record_broken = True
                                                print(key+" broke record -> Model is saved")
    
                                    if record_broken:
                                        #Save the model
                                        model_keras[i].save(model_keras_path[i].split(".model")[0]+"_"+str(counter)+".model")
                                        SAVED[i].append(1)
                                    elif bool(self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.isChecked())==True:
                                        model_keras[i].save(model_keras_path[i].split(".model")[0]+"_"+str(counter)+".model")
                                        SAVED[i].append(1)
                                        self.fittingpopups_ui[listindex].checkBox_saveEpoch_pop.setChecked(False)
                                    else:
                                        SAVED[i].append(0)


                            callback_progessbar = float(counter)/nr_epochs
                            progress_callback.emit(100.0*callback_progessbar)
                            history_callback.emit(history.history)
                            Index.append(counter)
                            
                            t2 =  time.time()
                            
                            if collection==False:
                                if counter==0:
                                    #If this runs the first time, create the file with header
                                    DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                                    DF1 = pd.concat(DF1)
                                    DF1["Saved"] = Saved
                                    DF1["Time"] = Stopwatch
                                    DF1.index = Index

                                    #If this runs the first time, create the file with header
                                    if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                                    DF1.to_excel(writer,sheet_name='History')
                                    writer.save()
                                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)
                                    print("meta.xlsx was saved")
                                    Index,Histories,Saved,Stopwatch = [],[],[],[]#reset the lists
                                    
                                #Get a sensible frequency for saving the dataframe (every 20s)
                                elif t2-t1>20:                                   
                                #elif counter%50==0:  #otherwise save the history to excel after each n epochs
                                    DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                                    DF1 = pd.concat(DF1)
                                    DF1["Saved"] = Saved
                                    DF1["Time"] = Stopwatch
                                    DF1.index = Index

                                    #in case the history is saved on a server and another PC wants to access 
                                    #the history file (meta.excel) there can be an error
                                    #In such a case try a few times, and wait
                                    tries = 0
#                                    while tries<10:
#                                        try:
                                    #Saving
                                    if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                                        os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                                    DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                                    writer.save()
                                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)  #make read only
                                    print("meta.xlsx was saved")
                                    Index,Histories,Saved,Stopwatch = [],[],[],[]#reset the lists
                                    t1 = time.time()
#                                        except:
#                                            time.sleep(1.5)
#                                            tries+=1
                                        
                                        
                            if collection==True:
                                if counter==0:
                                    for i in range(len(HISTORIES)):
                                        Histories = HISTORIES[i]
                                        Saved = SAVED[i]
                                        #If this runs the first time, create the file with header
                                        DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                                        DF1 = pd.concat(DF1)
                                        DF1["Saved"] = Saved
                                        DF1.index = Index
                                        HISTORIES[i] = []#reset the Histories list
                                        SAVED[i] = []
                                        #If this runs the first time, create the file with header
                                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #read/write
                                        DF1.to_excel(Writers[i],sheet_name='History')
                                        Writers[i].save()
                                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)
                                        print("meta.xlsx was saved")
                                    Index = []#reset the Index list
                                    
                                #Get a sensible frequency for saving the dataframe (every 20s)
                                elif t2-t1>20:                                    
                                    for i in range(len(HISTORIES)):
                                        Histories = HISTORIES[i]
                                        Saved = SAVED[i]
                                        DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                                        DF1 = pd.concat(DF1)
                                        DF1["Saved"] = Saved
                                        DF1.index = Index
                                        HISTORIES[i] = []#reset the Histories list
                                        SAVED[i] = []
                                        #Saving
                                        if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                                            os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                                        DF1.to_excel(Writers[i],sheet_name='History', startrow=Writers[i].sheets['History'].max_row,header= False)
                                        Writers[i].save()
                                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH)  #make read only
                                        print("meta.xlsx was saved")
                                        t1 = time.time()
                                    Index = []#reset the Index list

                            counter+=1
                        
            progress_callback.emit(100.0)

            if collection==False:
                if len(Histories)>0: #if the list for History files is not empty, process it!
                    DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                    DF1 = pd.concat(DF1)
                    DF1["Saved"] = Saved
                    DF1["Time"] = Stopwatch
                    DF1.index = Index
                    Index = []#reset the Index list
                    Histories = []#reset the Histories list
                    Saved = []
                    #does such a file exist already? append! 
                    if not os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                       DF1.to_excel(writer,sheet_name='History')
                    else: # else it exists so append without writing the header
                       DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                if os.path.isfile(new_modelname.split(".model")[0]+'_meta.xlsx'):
                    os.chmod(new_modelname.split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                writer.save()
                writer.close()

            if collection==True:
                for i in range(len(HISTORIES)):
                    Histories = HISTORIES[i]
                    Saved = SAVED[i]
                    if len(Histories)>0: #if the list for History files is not empty, process it!
                        DF1 = [pd.DataFrame(h).iloc[[-1]] for h in Histories] #just in case the nb_epoch in .fit() is >1, only save the last history item, beacuse this would a  model that could be saved 
                        DF1 = pd.concat(DF1)
                        DF1["Saved"] = Saved
                        DF1.index = Index
                        HISTORIES[i] = []#reset the Histories list
                        SAVED[i] = []
                        #does such a file exist already? append! 
                        if not os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                           DF1.to_excel(Writers[i],sheet_name='History')
                        else: # else it exists so append without writing the header
                           DF1.to_excel(writer,sheet_name='History', startrow=writer.sheets['History'].max_row,header= False)
                    if os.path.isfile(model_keras_path[i].split(".model")[0]+'_meta.xlsx'):
                        os.chmod(model_keras_path[i].split(".model")[0]+'_meta.xlsx', S_IREAD|S_IRGRP|S_IROTH|S_IWRITE|S_IWGRP|S_IWOTH) #make read/write
                    Writers[i].save()
                    Writers[i].close()
                    
                Index = []#reset the Index list

                    
                    
                    
                    

    def action_fit_model(self):
        #Take the initialized model
        #Unfortunately, in TensorFlow it is not possile to pass a model from
        #one thread to another. Therefore I have to load and save the models each time :(
        model_keras = self.model_keras
        if type(model_keras)==tuple:
            collection=True
        else:
            collection=False
            
        #Check if there was a model initialized:
        new_modelname = str(self.lineEdit_modelname.text())
        
        if len(new_modelname)==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define a path/filename for the model to be fitted!")
           msg.setWindowTitle("Model path/ filename missing!")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return
       
        if model_keras==None:
            self.action_initialize_model()
            if model_keras==None:
#                msg = QtWidgets.QMessageBox()
#                msg.setIcon(QtWidgets.QMessageBox.Information)       
#                msg.setText("Model could not be initialized")
#                msg.setWindowTitle("Error")
#                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#                msg.exec_()
                return
            if not model_keras==None:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Model is now initialized for you, Please check Model summary window below if everything is correct and then press Fit again!")
                msg.setWindowTitle("No initilized model found!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

        #There should be at least two outputs (index 0 and 1)
        if collection==False:
            #model_config = model_keras.get_config()#["layers"] 
            nr_classes = int(model_keras.output.shape.dims[1])

        if collection==True:
            #model_config = model_keras[1][0].get_config()#["layers"] 
            nr_classes = int(model_keras[1][0].output.shape.dims[1])

        if nr_classes<2:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Please define at least two classes")
           msg.setWindowTitle("Not enough classes")
           msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
           msg.exec_()
           return
        
        if collection==False:
            #define a variable on self which allows the fit_model_worker to load this model and fit
            #(sorry, this is necessary since TensorFlow does not support passing models between threads)
            self.model_keras_path = new_modelname.split(".model")[0]+"_0.model"
            #save a first version of the .model
            model_keras.save(self.model_keras_path)
            #Delete the variable to save RAM
            model_keras = None #Since this uses TensorFlow, I have to reload the model action_fit_model_worker anyway

        if collection==True:
            #define a variable on self which allows the fit_model_worker to load this model and fit
            #(sorry, this is necessary since TensorFlow does not support passing models between threads)
            self.model_keras_path = [new_modelname.split(".model")[0]+"_"+model_keras[0][i]+".model" for i in range(len(model_keras[0]))]
            for i in range(len(self.model_keras_path)):
                #save a first version of the .model
                model_keras[1][i].save(self.model_keras_path[i])

            #Delete the variable to save RAM
            model_keras = None #Since this uses TensorFlow, I have to reload the model action_fit_model_worker anyway
        
        #Check that Data is on RAM
        DATA_len = len(self.ram) #this returns the len of a dictionary. The dictionary is supposed to contain the training/validation data; otherwise the data is read from .rtdc data directly (SLOW unless you have ultra-good SSD)

        def popup_data_to_ram(button):
            yes_or_no = button.text()
            if yes_or_no == "&Yes":
                print("Moving data to ram")
                self.actionDataToRamNow_function()
            elif yes_or_no == "&No":
                pass
            
        if DATA_len==0:
           msg = QtWidgets.QMessageBox()
           msg.setIcon(QtWidgets.QMessageBox.Information)       
           msg.setText("Would you like transfer the Data to RAM now?\n(Currently the data is not in RAM and would be read from .rtdc, which slows down fitting dramatically unless you have a super-fast SSD.)")
           msg.setWindowTitle("Data to RAM now?")
           msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
           msg.buttonClicked.connect(popup_data_to_ram)
           msg.exec_()
            
        ###################Popup Window####################################
        self.fittingpopups.append(MyPopup())
        ui = aid_frontend.Fitting_Ui()

        ui.setupUi(self.fittingpopups[-1]) #append the ui to the last element on the list
        self.fittingpopups_ui.append(ui)
        # Increase the popupcounter by one; this will help to coordinate the data flow between main ui and popup
        self.popupcounter += 1
        listindex=self.popupcounter-1
        
        ##############################Define functions#########################
        self.fittingpopups_ui[listindex].pushButton_UpdatePlot_pop.clicked.connect(lambda: self.update_historyplot_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_Stop_pop.clicked.connect(lambda: self.stop_fitting_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_Pause_pop.clicked.connect(lambda: self.pause_fitting_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_saveTextWindow_pop.clicked.connect(lambda: self.saveTextWindow_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_clearTextWindow_pop.clicked.connect(lambda: self.clearTextWindow_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_showModelSumm_pop.clicked.connect(lambda: self.showModelSumm_pop(listindex))
        self.fittingpopups_ui[listindex].pushButton_saveModelSumm_pop.clicked.connect(lambda: self.saveModelSumm_pop(listindex))
        #Expert mode functions
        #self.fittingpopups_ui[listindex].checkBox_pTr_pop.toggled.connect(lambda on_or_off: self.partialtrainability_activated_pop(on_or_off,listindex))
        self.fittingpopups_ui[listindex].pushButton_lossW.clicked.connect(lambda: self.lossWeights_popup(listindex))
        self.fittingpopups_ui[listindex].checkBox_lossW.clicked.connect(lambda on_or_off: self.lossWeights_activated(on_or_off,listindex))

        self.fittingpopups_ui[listindex].Form.setWindowTitle(os.path.split(new_modelname)[1])
        self.fittingpopups_ui[listindex].progressBar_Fitting_pop.setValue(0) #set the progress bar to zero
        self.fittingpopups_ui[listindex].pushButton_ShowExamleImgs_pop.clicked.connect(lambda: self.action_show_example_imgs_pop(listindex))
        self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.doubleClicked.connect(lambda item: self.tableWidget_HistoryInfo_pop_dclick(item,listindex))






        worker = Worker(self.action_fit_model_worker)
        #Get a signal from the worker to update the progressbar
        worker.signals.progress.connect(self.fittingpopups_ui[listindex].progressBar_Fitting_pop.setValue)
        
        #Define a func which prints information during fitting to textbrowser
        #And furthermore provide option to do real-time plotting
        def real_time_info(dic):
            self.fittingpopups_ui[listindex].Histories.append(dic) #append to a list. Will be used for plotting in the "Update plot" function
#            print("dic")
#            print(dic) 
            OtherMetrics_keys = self.fittingpopups_ui[listindex].RealTime_OtherMetrics.keys()
            #Append to lists for real-time plotting
            self.fittingpopups_ui[listindex].RealTime_Acc.append(dic["acc"][0])
            self.fittingpopups_ui[listindex].RealTime_ValAcc.append(dic["val_acc"][0])
            self.fittingpopups_ui[listindex].RealTime_Loss.append(dic["loss"][0])
            self.fittingpopups_ui[listindex].RealTime_ValLoss.append(dic["val_loss"][0])

            keys = list(dic.keys())            
            #sort keys alphabetically
            keys_ = [l.lower() for l in keys]
            ind_sort = np.argsort(keys_)
            keys = list(np.array(keys)[ind_sort])
            #First keys should always be acc,loss,val_acc,val_loss -in this order
            keys_first = ["acc","loss","val_acc","val_loss"]
            for i in range(len(keys_first)):
                if keys_first[i] in keys:
                    ind = np.where(np.array(keys)==keys_first[i])[0][0]
                    if ind!=i:
                        del keys[ind]
                        keys.insert(i,keys_first[i])
    
            for key in keys:
                if "precision" in key or "f1" in key or "recall" in key:
                    if not key in OtherMetrics_keys: #if this key is missing in self.fittingpopups_ui[listindex].RealTime_OtherMetrics attach it!
                        self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key] = []
                    self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key].append(dic[key])
            dic_text = [("{} {}".format(item, np.round(amount[0],4))) for item, amount in dic.items()]
            text = "Epoch "+str(self.fittingpopups_ui[listindex].epoch_counter)+"\n"+" ".join(dic_text)
            self.fittingpopups_ui[listindex].textBrowser_FittingInfo_pop.append(text)
            self.fittingpopups_ui[listindex].epoch_counter+=1
            if self.fittingpopups_ui[listindex].epoch_counter==1:

                #for each key, put a checkbox on the tableWidget_HistoryInfo_pop
                rowPosition = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.rowCount()
                self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.insertRow(rowPosition)
                self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.setColumnCount(len(keys))

                for columnPosition in range(len(keys)):#(2,4):
                    key = keys[columnPosition]
                    #for each item, also create 2 checkboxes (train/valid)
                    item = QtWidgets.QTableWidgetItem(str(key))#("item {0} {1}".format(rowNumber, columnNumber))
                    item.setBackground(QtGui.QColor(self.colorsQt[columnPosition]))
                    item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
                    item.setCheckState(QtCore.Qt.Unchecked)
                    self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.setItem(rowPosition, columnPosition, item)
            self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.resizeColumnsToContents()
            self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.resizeRowsToContents()


            ########################Real-time plotting#########################
            if self.fittingpopups_ui[listindex].checkBox_realTimePlotting_pop.isChecked():
                #get the range for the real time fitting
                if hasattr(self.fittingpopups_ui[listindex], 'historyscatters'):#if update plot was hit before
                    x = range(len(self.fittingpopups_ui[listindex].Histories))
                    realTimeEpochs = self.fittingpopups_ui[listindex].spinBox_realTimeEpochs.value()
                    if len(x)>realTimeEpochs:
                        x = x[-realTimeEpochs:]
                    #is any metric checked on the table?
                    colcount = int(self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.columnCount())
                    #Collect items that are checked
                    selected_items,Colors = [],[]
                    for colposition in range(colcount):  
                        #is it checked?
                        cb = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(0, colposition)
                        if not cb==None:
                            if cb.checkState() == QtCore.Qt.Checked:
                                selected_items.append(str(cb.text()))
                                Colors.append(cb.background())
                    
                    for i in range(len(self.fittingpopups_ui[listindex].historyscatters)): #iterate over all available plots
                        key = list(self.fittingpopups_ui[listindex].historyscatters.keys())[i]
                        if key in selected_items:
                            if key=="acc":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_Acc).astype(float)
                            elif key=="val_acc":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_ValAcc).astype(float)
                            elif key=="loss":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_Loss).astype(float)
                            elif key=="val_loss":
                                y = np.array(self.fittingpopups_ui[listindex].RealTime_ValLoss).astype(float)
                            elif "precision" in key or "f1" in key or "recall" in key:
                               y = np.array(self.fittingpopups_ui[listindex].RealTime_OtherMetrics[key]).astype(float).reshape(-1,)
                            else:
                                return
                            #Only show the last 250 epochs
                            if y.shape[0]>realTimeEpochs:
                                y = y[-realTimeEpochs:]
                            if y.shape[0]==len(x):
                                self.fittingpopups_ui[listindex].historyscatters[key].setData(x, y)#,pen=None,symbol='o',symbolPen=None,symbolBrush=brush,clear=False)
                            else:
                                print("x and y are not the same size! Omitted plotting. I will try again to plot after the next epoch.")

                        pg.QtGui.QApplication.processEvents()

        self.fittingpopups_ui[listindex].epoch_counter = 0
        worker.signals.history.connect(real_time_info)
        
        #Finally start the worker!
        self.threadpool.start(worker)
        self.fittingpopups[listindex].show()

        
    def action_show_example_imgs(self): #this function is only for the main window
        if self.actionVerbose.isChecked()==True:
            verbose = 1
        else:
            verbose = 0
        #Get state of the comboboxes!
        tr_or_valid = str(self.comboBox_ShowTrainOrValid.currentText())
        w_or_wo_augm = str(self.comboBox_ShowWOrWoAug.currentText())

        #most of it should be similar to action_fit_model_worker
        #Used files go to a separate sheet on the MetaFile.xlsx
        SelectedFiles = self.items_clicked()
        #Collect all information about the fitting routine that was user defined
        crop = int(self.spinBox_imagecrop.value())          
        norm = str(self.comboBox_Normalization.currentText())
        h_flip = bool(self.checkBox_HorizFlip.isChecked())
        v_flip = bool(self.checkBox_VertFlip.isChecked())
        rotation = float(self.lineEdit_Rotation.text())
        width_shift = float(self.lineEdit_Rotation_2.text())
        height_shift = float(self.lineEdit_Rotation_3.text())
        zoom = float(self.lineEdit_Rotation_4.text())
        shear = float(self.lineEdit_Rotation_5.text())
        brightness_add_lower = float(self.spinBox_PlusLower.value())
        brightness_add_upper = float(self.spinBox_PlusUpper.value())
        brightness_mult_lower = float(self.doubleSpinBox_MultLower.value())
        brightness_mult_upper = float(self.doubleSpinBox_MultUpper.value())
        gaussnoise_mean = float(self.doubleSpinBox_GaussianNoiseMean.value())
        gaussnoise_scale = float(self.doubleSpinBox_GaussianNoiseScale.value())

        contrast_on = bool(self.checkBox_contrast.isChecked())        
        contrast_lower = float(self.doubleSpinBox_contrastLower.value())
        contrast_higher = float(self.doubleSpinBox_contrastHigher.value())
        saturation_on = bool(self.checkBox_saturation.isChecked())        
        saturation_lower = float(self.doubleSpinBox_saturationLower.value())
        saturation_higher = float(self.doubleSpinBox_saturationHigher.value())
        hue_on = bool(self.checkBox_hue.isChecked())        
        hue_delta = float(self.doubleSpinBox_hueDelta.value())

        avgBlur_on = bool(self.checkBox_avgBlur.isChecked())        
        avgBlur_min = int(self.spinBox_avgBlurMin.value())
        avgBlur_max = int(self.spinBox_avgBlurMax.value())

        gaussBlur_on = bool(self.checkBox_gaussBlur.isChecked())        
        gaussBlur_min = int(self.spinBox_gaussBlurMin.value())
        gaussBlur_max = int(self.spinBox_gaussBlurMax.value())

        motionBlur_on = bool(self.checkBox_motionBlur.isChecked())        
        motionBlur_kernel = str(self.lineEdit_motionBlurKernel.text())
        motionBlur_angle = str(self.lineEdit_motionBlurAngle.text())
        
        motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
        motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple

        padding_expert = str(self.comboBox_expt_paddingMode.currentText()).lower()

        #which index is requested by user:?
        req_index = int(self.spinBox_ShowIndex.value())
        if tr_or_valid=='Training':
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        elif tr_or_valid=='Validation':
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles = np.array(SelectedFiles)[ind]
        SelectedFiles = list(SelectedFiles)
        indices = [selectedfile["class"] for selectedfile in SelectedFiles]
        ind = np.where(np.array(indices)==req_index)[0]
        if len(ind)<1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no data for this class available")
            msg.setWindowTitle("Class not available")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        indices = list(np.array(indices)[ind])
        SelectedFiles = list(np.array(SelectedFiles)[ind])
        nr_events_epoch = len(indices)*[10] #[selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles]
        rtdc_path = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles]
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        shuffle = [selectedfile["shuffle"] for selectedfile in SelectedFiles]
        #If the scaling method is "divide by mean and std of the whole training set":
        if norm == "StdScaling using mean and std of all training data":
            mean_trainingdata,std_trainingdata = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:    
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                images = next(gen)[0]
                mean_trainingdata.append(np.mean(images))
                std_trainingdata.append(np.std(images))
            mean_trainingdata = np.mean(np.array(mean_trainingdata))
            std_trainingdata = np.mean(np.array(std_trainingdata))
            if np.allclose(std_trainingdata,0):
                std_trainingdata = 0.0001
                print("std_trainingdata was zero and is now set to 0.0001 to avoid div. by zero!")
            if self.actionVerbose.isChecked():
                print("Used all training data to get mean and std for normalization")

        if w_or_wo_augm=='With Augmentation':
            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
            
            ############Cropping and image augmentation#####################
            #Start the first iteration:                
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:   
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=True,replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                try: #When all cells are at the border of the image, the generator will be empty. Avoid program crash by try, except
                    X.append(next(gen)[0])
                except StopIteration:
                    print("All events at border of image and discarded")
                    return
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            X = X.astype(np.uint8) #make sure we stay in uint8
            y = np.concatenate(y)
            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X.shape))
            if channels==1:
                #Add the "channels" dimension
                X = np.expand_dims(X,3)
            
            X_batch, y_batch = aid_img.affine_augm(X,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear), y #Affine image augmentation
            X_batch = X_batch.astype(np.uint8) #make sure we stay in uint8

            #Now do the final cropping to the actual size that was set by user
            dim = X_batch.shape
            if dim[2]!=crop:
                remove = int(dim[2]/2.0 - crop/2.0)
                #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
                X_batch = X_batch[:,remove:-remove,remove:-remove,:] #crop to crop x crop pixels #TensorFlow

            ##########Contrast/Saturation/Hue augmentation#########
            #is there any of contrast/saturation/hue augmentation to do?
            if contrast_on:
                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
            if saturation_on or hue_on:
                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta)

            ##########Average/Gauss/Motion blurring#########
            #is there any of blurring to do?
            if avgBlur_on:
                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
            if gaussBlur_on:
                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
            if motionBlur_on:
                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)

            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)

            if norm == "StdScaling using mean and std of all training data":
                X_batch = aid_img.norm_imgs(X_batch,norm,mean_trainingdata,std_trainingdata)
            else:
                X_batch = aid_img.norm_imgs(X_batch,norm)
            
            X = X_batch
            if verbose: print("Shape of the shown images is:"+str(X.shape))
            
        elif w_or_wo_augm=='Original image':
            ############Cropping#####################
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=True,replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:                        
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=True,replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                try:
                    X.append(next(gen)[0])
                except:
                    return
                y.append(np.repeat(indices[i],X[-1].shape[0]))
            X = np.concatenate(X)
            y = np.concatenate(y)

            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
            else:
                print("Invalid data dimension: " +str(X.shape))
            if channels==1:
                #Add the "channels" dimension
                X = np.expand_dims(X,3)
        
        if norm == "StdScaling using mean and std of all training data":
            X = aid_img.norm_imgs(X,norm,mean_trainingdata,std_trainingdata)
        else:
            X = aid_img.norm_imgs(X,norm)
        
        if verbose: print("Shape of the shown images is: "+str(X.shape))
                
        #Is there already anything shown on the widget?
        children = self.widget_ViewImages.findChildren(QtWidgets.QGridLayout)
        if len(children)>0: #if there is something, delete it!
            for i in reversed(range(self.gridLayout_ViewImages.count())):
                widgetToRemove = self.gridLayout_ViewImages.itemAt(i).widget()
                widgetToRemove.setParent(None)
                widgetToRemove.deleteLater()
        else: #else, create a Gridlayout to put the images
            self.gridLayout_ViewImages = QtWidgets.QGridLayout(self.widget_ViewImages)

        for i in range(5):
            if channels==1:
                img = X[i,:,:,0] #TensorFlow 
            if channels==3:
                img = X[i,:,:,:] #TensorFlow 
            
            #zoom image such that longest side is 64              
            factor = 1#float(64.0/np.max(img.shape))
            #Get the order, specified in Options->Zoom Order
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            order = np.where(np.array(zoom_methods)==True)[0]
            if channels==1:
                img_zoom = ndimage.zoom(img, zoom=factor,order=int(order))
            if channels==3:
                img_zoom = ndimage.zoom(img, zoom=(factor,factor,1),order=int(order))                
                
            img_zoom = np.ascontiguousarray(img_zoom)
            #from 0 to 255
            img_zoom = img_zoom-np.min(img_zoom)
            fac = np.max(img_zoom)
            img_zoom = (img_zoom/fac)*255.0
            img_zoom = img_zoom.astype(np.uint8)
            if channels==1:
                height, width = img_zoom.shape
            if channels==3:
                height, width, _ = img_zoom.shape

#            qi=QtGui.QImage(img_zoom.data, width, height,width, QtGui.QImage.Format_Indexed8)
#            self.label_image_show = QtWidgets.QLabel(self.widget_ViewImages)
#            self.label_image_show.setPixmap(QtGui.QPixmap.fromImage(qi))
#            self.gridLayout_ViewImages.addWidget(self.label_image_show, 1,i)
#            self.label_image_show.show()
            #Use pygtgraph instead, in order to allow for exporting images
            self.image_show = pg.ImageView(self.widget_ViewImages)
            self.image_show.show()
            if verbose: print("Shape of zoomed image: "+str(img_zoom.shape))
            if channels==1:
                self.image_show.setImage(img_zoom.T,autoRange=False)
            if channels==3:
                self.image_show.setImage(np.swapaxes(img_zoom,0,1),autoRange=False)
                
            self.image_show.ui.histogram.hide()
            self.image_show.ui.roiBtn.hide()
            self.image_show.ui.menuBtn.hide()
            self.gridLayout_ViewImages.addWidget(self.image_show, 1,i)
        self.widget_ViewImages.show()

        
        
    def tableWidget_HistoryInfo_pop_dclick(self,item,listindex):
        if item is not None:
            tableitem = self.fittingpopups_ui[listindex].tableWidget_HistoryInfo_pop.item(item.row(), item.column())
            if str(tableitem.text())!="Show saved only":
                color = QtGui.QColorDialog.getColor()
                if color.getRgb()==(0, 0, 0, 255):#no black!
                    return
                else:
                    tableitem.setBackground(color)
                #self.update_historyplot_pop(listindex)
       
        
        
        
        
    def action_show_example_imgs_pop(self,listindex): #this function is only for the main window
        #Get state of the comboboxes!
        tr_or_valid = str(self.fittingpopups_ui[listindex].comboBox_ShowTrainOrValid_pop.currentText())
        w_or_wo_augm = str(self.fittingpopups_ui[listindex].comboBox_ShowWOrWoAug_pop.currentText())

        #most of it should be similar to action_fit_model_worker
        #Used files go to a separate sheet on the MetaFile.xlsx
        SelectedFiles = self.items_clicked()
        #Collect all information about the fitting routine that was user defined
        crop = int(self.fittingpopups_ui[listindex].spinBox_imagecrop_pop.value())          
        norm = str(self.fittingpopups_ui[listindex].comboBox_Normalization_pop.currentText())
        h_flip = bool(self.fittingpopups_ui[listindex].checkBox_HorizFlip_pop.isChecked())
        v_flip = bool(self.fittingpopups_ui[listindex].checkBox_VertFlip_pop.isChecked())
        rotation = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop.text())
        width_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_2.text())
        height_shift = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_3.text())
        zoom = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_4.text())
        shear = float(self.fittingpopups_ui[listindex].lineEdit_Rotation_pop_5.text())
        brightness_add_lower = float(self.fittingpopups_ui[listindex].spinBox_PlusLower_pop.value())
        brightness_add_upper = float(self.fittingpopups_ui[listindex].spinBox_PlusUpper_pop.value())
        brightness_mult_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultLower_pop.value())
        brightness_mult_upper = float(self.fittingpopups_ui[listindex].doubleSpinBox_MultUpper_pop.value())
        gaussnoise_mean = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseMean_pop.value())
        gaussnoise_scale = float(self.fittingpopups_ui[listindex].doubleSpinBox_GaussianNoiseScale_pop.value())
        
        contrast_on = bool(self.fittingpopups_ui[listindex].checkBox_contrast_pop.isChecked())
        contrast_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastLower_pop.value())
        contrast_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_contrastHigher_pop.value())
        saturation_on = bool(self.fittingpopups_ui[listindex].checkBox_saturation_pop.isChecked())
        saturation_lower = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationLower_pop.value())
        saturation_higher = float(self.fittingpopups_ui[listindex].doubleSpinBox_saturationHigher_pop.value())
        hue_on = bool(self.fittingpopups_ui[listindex].checkBox_hue_pop.isChecked())
        hue_delta = float(self.fittingpopups_ui[listindex].doubleSpinBox_hueDelta_pop.value())

        avgBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_avgBlur_pop.isChecked())        
        avgBlur_min = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMin_pop.value())
        avgBlur_max = int(self.fittingpopups_ui[listindex].spinBox_avgBlurMax_pop.value())
    
        gaussBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_gaussBlur_pop.isChecked())        
        gaussBlur_min = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMin_pop.value())
        gaussBlur_max = int(self.fittingpopups_ui[listindex].spinBox_gaussBlurMax_pop.value())
    
        motionBlur_on = bool(self.fittingpopups_ui[listindex].checkBox_motionBlur_pop.isChecked())        
        motionBlur_kernel = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurKernel_pop.text())
        motionBlur_angle = str(self.fittingpopups_ui[listindex].lineEdit_motionBlurAngle_pop.text())
        
        motionBlur_kernel = tuple(ast.literal_eval(motionBlur_kernel)) #translate string in the lineEdits to a tuple
        motionBlur_angle = tuple(ast.literal_eval(motionBlur_angle)) #translate string in the lineEdits to a tuple

        padding_expert = str(self.fittingpopups_ui[listindex].comboBox_expt_paddingMode_pop.currentText()).lower()

        #which index is requested by user:?
        req_index = int(self.fittingpopups_ui[listindex].spinBox_ShowIndex_pop.value())
        if tr_or_valid=='Training':
            ######################Load the Training Data################################
            ind = [selectedfile["TrainOrValid"] == "Train" for selectedfile in SelectedFiles]
        elif tr_or_valid=='Validation':
            ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]
        SelectedFiles = np.array(SelectedFiles)[ind]
        SelectedFiles = list(SelectedFiles)
        indices = [selectedfile["class"] for selectedfile in SelectedFiles]
        ind = np.where(np.array(indices)==req_index)[0]
        if len(ind)<1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no data for this class available")
            msg.setWindowTitle("Class not available")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        indices = list(np.array(indices)[ind])
        SelectedFiles = list(np.array(SelectedFiles)[ind])
        nr_events_epoch = len(indices)*[10] #[selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles]
        rtdc_path = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles]
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        shuffle = [selectedfile["shuffle"] for selectedfile in SelectedFiles]
        #If the scaling method is "divide by mean and std of the whole training set":
        if norm == "StdScaling using mean and std of all training data":
            mean_trainingdata,std_trainingdata = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],random_images=False,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:    
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],random_images=False) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                images = next(gen)[0]
                mean_trainingdata.append(np.mean(images))
                std_trainingdata.append(np.std(images))
            mean_trainingdata = np.mean(np.array(mean_trainingdata))
            std_trainingdata = np.mean(np.array(std_trainingdata))
            if np.allclose(std_trainingdata,0):
                std_trainingdata = 0.0001
                print("std_trainingdata turned out to be zero. I set it to 0.0001, to avoid division by zero!")
            if self.actionVerbose.isChecked():
                print("Used all training data to get mean and std for normalization")

        if w_or_wo_augm=='With Augmentation':
            ###############Continue with training data:augmentation############
            #Rotating could create edge effects. Avoid this by making crop a bit larger for now
            #Worst case would be a 45degree rotation:
            cropsize2 = np.sqrt(crop**2+crop**2)
            cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number

            ############Get cropped images image augmentation#####################
            #Start the first iteration:                
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(cropsize2,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:   
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=shuffle[i],replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")
                    
                X.append(next(gen)[0])
                #y.append(np.repeat(indices[i],nr_events_epoch[i]))
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            y = np.concatenate(y)
            
            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X.shape))
            if channels==1:
                #Add the "channels" dimension
                X = np.expand_dims(X,3)

            X_batch, y_batch = aid_img.affine_augm(X,v_flip,h_flip,rotation,width_shift,height_shift,zoom,shear), y #Affine image augmentation
            X_batch = X_batch.astype(np.uint8) #make sure we stay in uint8

            #Now do the final cropping to the actual size that was set by user
            dim = X_batch.shape
            if dim[2]!=crop:
                remove = int(dim[2]/2.0 - crop/2.0)
                #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
                X_batch = X_batch[:,remove:-remove,remove:-remove,:] #crop to crop x crop pixels #TensorFlow

            ##########Contrast/Saturation/Hue augmentation#########
            #is there any of contrast/saturation/hue augmentation to do?
            if contrast_on:
                X_batch = aid_img.contrast_augm_cv2(X_batch,contrast_lower,contrast_higher) #this function is almost 15 times faster than random_contrast from tf!
            if saturation_on or hue_on:
                X_batch = aid_img.satur_hue_augm_cv2(X_batch.astype(np.uint8),saturation_on,saturation_lower,saturation_higher,hue_on,hue_delta)
           
            ##########Average/Gauss/Motion blurring#########
            #is there any of blurring to do?
            if avgBlur_on:
                X_batch = aid_img.avg_blur_cv2(X_batch,avgBlur_min,avgBlur_max)
            if gaussBlur_on:
                X_batch = aid_img.gauss_blur_cv(X_batch,gaussBlur_min,gaussBlur_max)
            if motionBlur_on:
                X_batch = aid_img.motion_blur_cv(X_batch,motionBlur_kernel,motionBlur_angle)

            X_batch = aid_img.brightn_noise_augm_cv2(X_batch,brightness_add_lower,brightness_add_upper,brightness_mult_lower,brightness_mult_upper,gaussnoise_mean,gaussnoise_scale)
            
            if norm == "StdScaling using mean and std of all training data":
                X_batch = aid_img.norm_imgs(X_batch,norm,mean_trainingdata,std_trainingdata)
            else:
                X_batch = aid_img.norm_imgs(X_batch,norm)
            X = X_batch
            
        elif w_or_wo_augm=='Original image':
            ############Cropping#####################
            X,y = [],[]
            for i in range(len(SelectedFiles)):
                if not self.actionDataToRam.isChecked():
                    gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                else:
                    if len(self.ram)==0:
                        gen = aid_img.gen_crop_img(crop,rtdc_path[i],10,random_images=shuffle[i],replace=True,zoom_factor=zoom_factors[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace true means that individual cells could occur several times
                    else:                        
                        gen = aid_img.gen_crop_img_ram(self.ram,rtdc_path[i],10,random_images=shuffle[i],replace=True) #Replace true means that individual cells could occur several times
                        if self.actionVerbose.isChecked():
                            print("Loaded data from RAM")

                X.append(next(gen)[0])
                #y.append(np.repeat(indices[i],nr_events_epoch[i]))
                y.append(np.repeat(indices[i],X[-1].shape[0]))

            X = np.concatenate(X)
            y = np.concatenate(y)
        
            if len(X.shape)==4:
                channels=3
            elif len(X.shape)==3:
                channels=1
            else:
                print("Invalid data dimension:" +str(X.shape))
            if channels==1:
                #Add the "channels" dimension
                X = np.expand_dims(X,3)

            if norm == "StdScaling using mean and std of all training data":
                X = aid_img.norm_imgs(X,norm,mean_trainingdata,std_trainingdata)
            else:
                X = aid_img.norm_imgs(X,norm)
                
        #Is there already anything shown on the widget?
        children = self.fittingpopups_ui[listindex].widget_ViewImages_pop.findChildren(QtWidgets.QGridLayout)
        if len(children)>0: #if there is something, delete it!
            for i in reversed(range(self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.count())):
                widgetToRemove = self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.itemAt(i).widget()
                widgetToRemove.setParent(None)
                widgetToRemove.deleteLater()
        else: #else, create a Gridlayout to put the images
            self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop = QtWidgets.QGridLayout(self.fittingpopups_ui[listindex].widget_ViewImages_pop)

        for i in range(5):
            if channels==1:
                img = X[i,:,:,0] #TensorFlow 
            if channels==3:
                img = X[i,:,:,:] #TensorFlow 
            #zoom image such that longest side is 64              
            factor = 1#float(64.0/np.max(img.shape))
            #Get the order, specified in Options->Zoom Order
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            order = np.where(np.array(zoom_methods)==True)[0]
            if channels==1:
                img_zoom = ndimage.zoom(img, zoom=factor,order=int(order))
            if channels==3:
                img_zoom = ndimage.zoom(img, zoom=(factor,factor,1),order=int(order))                

            print("Dimension of shown image is: "+str(img_zoom.shape))
            img_zoom = np.ascontiguousarray(img_zoom)
            #from 0 to 255
            img_zoom = img_zoom-np.min(img_zoom)
            fac = np.max(img_zoom)
            img_zoom = (img_zoom/fac)*255.0
            img_zoom = img_zoom.astype(np.uint8)
#            height, width = img_zoom.shape
#            qi=QtGui.QImage(img_zoom.data, width, height,width, QtGui.QImage.Format_Indexed8)
#            self.label_image_show = QtWidgets.QLabel(self.widget_ViewImages)
#            self.label_image_show.setPixmap(QtGui.QPixmap.fromImage(qi))
#            self.gridLayout_ViewImages_pop.addWidget(self.label_image_show, 1,i)
#            self.label_image_show.show()
            #Use pygtgraph instead, in order to allow for exporting images
            self.fittingpopups_ui[listindex].image_show_pop = pg.ImageView(self.fittingpopups_ui[listindex].widget_ViewImages_pop)
            self.fittingpopups_ui[listindex].image_show_pop.show()
            
            if channels==1:
                self.fittingpopups_ui[listindex].image_show_pop.setImage(img_zoom.T,autoRange=False)
            if channels==3:
                self.fittingpopups_ui[listindex].image_show_pop.setImage(np.swapaxes(img_zoom,0,1),autoRange=False)
                
            self.fittingpopups_ui[listindex].image_show_pop.ui.histogram.hide()
            self.fittingpopups_ui[listindex].image_show_pop.ui.roiBtn.hide()
            self.fittingpopups_ui[listindex].image_show_pop.ui.menuBtn.hide()
            self.fittingpopups_ui[listindex].gridLayout_ViewImages_pop.addWidget(self.fittingpopups_ui[listindex].image_show_pop, 1,i)
        self.fittingpopups_ui[listindex].widget_ViewImages_pop.show()

    def get_color_mode(self):
        if str(self.comboBox_GrayOrRGB.currentText())=="Grayscale":
            return "Grayscale"
        elif str(self.comboBox_GrayOrRGB.currentText())=="RGB":
            return "RGB"
        else:
            return None
        
    def checkBox_rollingMedian_statechange(self,item):
        self.horizontalSlider_rollmedi.setEnabled(item)
        
    def update_historyplot(self):
        #After loading a history, there are checkboxes available. Check, if user checked some:
        colcount = self.tableWidget_HistoryItems.columnCount()
        #Collect items that are checked
        selected_items = []
        Colors = []
        for colposition in range(colcount):  
            #get checkbox item and; is it checked?
            cb = self.tableWidget_HistoryItems.item(0, colposition)
            if not cb==None:
                if cb.checkState() == QtCore.Qt.Checked:
                    selected_items.append(str(cb.text()))
                    Colors.append(cb.background())
                 
        #Get a list of the color from the background of the table items
        DF1 = self.loaded_history

        #Clear the plot        
        self.widget_Scatterplot.clear()
            
        #Add plot        
        self.plt1 = self.widget_Scatterplot.addPlot()
        self.plt1.showGrid(x=True,y=True)
        self.plt1.addLegend()
        self.plt1.setLabel('bottom', 'Epoch', units='')
        
        self.plot_rollmedis = [] #list for plots of rolling medians
        
        if "Show saved only" in selected_items:
            #nr_of_selected_items = len(selected_items)-1
            #get the "Saved" column from DF1
            saved = DF1["Saved"]
            saved = np.where(np.array(saved==1))[0]
#        else:
#            nr_of_selected_items = len(selected_items)
            
        self.Colors = Colors
        scatter_x,scatter_y = [],[]
        for i in range(len(selected_items)):
            key = selected_items[i]
            if key!="Show saved only":
                df = DF1[key]  
                epochs = range(len(df))
                win = int(self.horizontalSlider_rollmedi.value())
                rollmedi = df.rolling(window=win).median()
                
                if "Show saved only" in selected_items:
                    df = np.array(df)[saved]
                    epochs = np.array(epochs)[saved]
                    rollmedi = pd.DataFrame(df).rolling(window=win).median()

                scatter_x.append(epochs)
                scatter_y.append(df)
                color = self.Colors[i]
                pen_rollmedi = list(color.color().getRgb())
                pen_rollmedi = pg.mkColor(pen_rollmedi)
                pen_rollmedi = pg.mkPen(color=pen_rollmedi,width=6)
                color = list(color.color().getRgb())
                color[-1] = int(0.6*color[-1])
                color = tuple(color)                
                pencolor = pg.mkColor(color)
                brush = pg.mkBrush(color=pencolor)
                self.plt1.plot(epochs, df,pen=None,symbol='o',symbolPen=None,symbolBrush=brush,name=key,clear=False)
                if bool(self.checkBox_rollingMedian.isChecked()):#Should a rolling median be plotted?
                    try:
                        rollmedi = np.array(rollmedi).reshape(rollmedi.shape[0])
                        rm = self.plt1.plot(np.array(epochs), rollmedi,pen=pen_rollmedi,clear=False)
                        self.plot_rollmedis.append(rm)
                    except Exception as e:
                        #There is an issue for the rolling median plotting!
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Warning)       
                        msg.setText(str(e)+"\n->There are likely too few points to have a rolling median with such a window size ("+str(round(win))+")")
                        msg.setWindowTitle("Error occured when plotting rolling median:")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()                        
                        
        if len(str(self.lineEdit_LoadHistory.text()))==0:
        #if DF1==None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please load History file first (.meta)")
            msg.setWindowTitle("No History file loaded")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        if len(scatter_x)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please select at least one of " +"\n".join(list(DF1.keys())))
            msg.setWindowTitle("No quantity selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #Keep the information as lists available for this function
        self.scatter_x_l, self.scatter_y_l = scatter_x,scatter_y
        if bool(self.checkBox_linearFit.isChecked()):
            #Put a liner region on the plot; cover the last 10% of points
            if np.max(np.concatenate(scatter_x))<12:
                start_x = 0
                end_x = np.max(np.concatenate(scatter_x))+1
            else:
                start_x = int(0.9*np.max(np.concatenate(scatter_x)))
                end_x = int(1.0*np.max(np.concatenate(scatter_x)))
            self.region_linfit = pg.LinearRegionItem([start_x, end_x], bounds=[-np.inf,np.inf], movable=True)
            self.plt1.addItem(self.region_linfit)

            def region_changed():
                try: #clear the plot from other fits if there are any
                    if len(self.plot_fits)>0:
                        for i in range(len(self.plot_fits)):
                            self.plt1.legend.removeItem(self.names[i])                                
                            self.plt1.removeItem(self.plot_fits[i])
                except:
                    pass
                #where did the user drag the region_linfit to?
                new_region = self.region_linfit.getRegion()
                #for each curve, do a linear regression
                self.plot_fits,self.names = [], []
                for i in range(len(self.scatter_x_l)):
                    scatter_x_vals = np.array(self.scatter_x_l[i])
                    ind = np.where( (scatter_x_vals<new_region[1]) & (scatter_x_vals>new_region[0]) )
                    scatter_x_vals = scatter_x_vals[ind]
                    scatter_y_vals = np.array(self.scatter_y_l[i])[ind]
                    if len(scatter_x_vals)>1:
                        fit = np.polyfit(scatter_x_vals,scatter_y_vals,1)
                        fit_y = fit[0]*scatter_x_vals+fit[1]    
                        pencolor = pg.mkColor(self.Colors[i].color())
                        pen = pg.mkPen(color=pencolor,width=6)
                        text = 'y='+("{:.2e}".format(fit[0]))+"x + " +("{:.2e}".format(fit[1]))
                        self.names.append(text)
                        self.plot_fits.append(self.plt1.plot(name=text))
                        self.plot_fits[i].setData(scatter_x_vals,fit_y,pen=pen,clear=False,name=text)

            self.region_linfit.sigRegionChangeFinished.connect(region_changed)

        def slider_changed():
            if bool(self.checkBox_rollingMedian.isChecked()):
                #remove other rolling median lines:
                for i in range(len(self.plot_rollmedis)):
                    self.plt1.removeItem(self.plot_rollmedis[i])
                #Start with fresh list 
                self.plot_rollmedis = []
                win = int(self.horizontalSlider_rollmedi.value())
                for i in range(len(self.scatter_x_l)):
                    epochs = np.array(self.scatter_x_l[i])
                    if type(self.scatter_y_l[i]) == pd.core.frame.DataFrame:
                        rollmedi = self.scatter_y_l[i].rolling(window=win).median()
                    else:
                        rollmedi = pd.DataFrame(self.scatter_y_l[i]).rolling(window=win).median()
                    rollmedi = np.array(rollmedi).reshape(rollmedi.shape[0])
                    pencolor = pg.mkColor(self.Colors[i].color())
                    pen_rollmedi = pg.mkPen(color=pencolor,width=6)
                    rm = self.plt1.plot(np.array(epochs), rollmedi,pen=pen_rollmedi,clear=False)
                    self.plot_rollmedis.append(rm)

        self.horizontalSlider_rollmedi.sliderMoved.connect(slider_changed)



        scatter_x = np.concatenate(scatter_x)
        scatter_y = np.concatenate(scatter_y)
        scatter_x_norm = (scatter_x.astype(float))/float(np.max(scatter_x))
        scatter_y_norm = (scatter_y.astype(float))/float(np.max(scatter_y))

        self.model_was_selected_before = False
        def onClick(event):
            #Get all plotting items
            #if len(self.plt1.listDataItems())==nr_of_selected_items+1:
                #delete the last item if the user selected already one:
            if self.model_was_selected_before:
                self.plt1.removeItem(self.plt1.listDataItems()[-1])

            items = self.widget_Scatterplot.scene().items(event.scenePos())
            #get the index of the viewbox
            isviewbox = [type(item)==pg.graphicsItems.ViewBox.ViewBox for item in items]
            index = np.where(np.array(isviewbox)==True)[0]
            vb = np.array(items)[index]
            try: #when user rescaed the vew and clicks somewhere outside, it could appear an IndexError.
                clicked_x =  float(vb[0].mapSceneToView(event.scenePos()).x())
                clicked_y =  float(vb[0].mapSceneToView(event.scenePos()).y())
            except:
                return
            try:
                a1 = (clicked_x)/float(np.max(scatter_x))            
                a2 = (clicked_y)/float(np.max(scatter_y))
            except Exception as e:
                print(str(e))
                return
            #Which is the closest scatter point?
            dist = np.sqrt(( a1-scatter_x_norm )**2 + ( a2-scatter_y_norm )**2)
            index =  np.argmin(dist)
            clicked_x = scatter_x[index]
            clicked_y = scatter_y[index]
            #Update the spinBox
            #self.spinBox_ModelIndex.setValue(int(clicked_x))
            #Modelindex for textBrowser_SelectedModelInfo
            text_index = "\nModelindex: "+str(clicked_x)
            #Indicate the selected model on the scatter plot
            self.plt1.plot([clicked_x], [clicked_y],pen=None,symbol='o',symbolPen='w',clear=False)

            #Get more information about this model
            Modelname = str(self.loaded_para["Modelname"].iloc[0])
            
            path, filename = os.path.split(Modelname)
            filename = filename.split(".model")[0]+"_"+str(clicked_x)+".model" 
            
            path = os.path.join(path,filename)
            if os.path.isfile(path):
                text_path = "\nFile is located in:"+path
            else:
                text_path = "\nFile not found!:"+path+"\nProbably the .model was deleted or not saved"
            text_acc = str(DF1.iloc[clicked_x])
            self.textBrowser_SelectedModelInfo.setText("Loaded model: "+filename+text_index+text_path+"\nPerformance:\n"+text_acc)
            self.model_was_selected_before = True
            self.model_2_convert = path
        self.widget_Scatterplot.scene().sigMouseClicked.connect(onClick)

    def action_load_history(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta file (*meta.xlsx)")
        filename = filename[0]
        if not filename.endswith("meta.xlsx"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        self.lineEdit_LoadHistory.setText(filename)
        self.action_plot_history(filename)

    def action_load_history_current(self):
        if self.model_keras_path==None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There is no fitting going on")
            msg.setWindowTitle("No current fitting process!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        history_path = self.model_keras_path
        if type(history_path)==list:#collection=True
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Not implemented for collections. Please use 'Load History' button to specify a single .meta file")
            msg.setWindowTitle("Not implemented for collecitons")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        filename = history_path.split("_0.model")[0]+"_meta.xlsx"
        
        if not filename.endswith("meta.xlsx"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        self.lineEdit_LoadHistory.setText(filename)
        self.action_plot_history(filename)
        
    def action_plot_history(self,filename):
        #If there is a file, it can happen that fitting is currently going on
        #and with bad luck AID just tries to write to the file. This would cause a crash.
        #Therfore, first try to copy the file to a temporary folder. If that fails,
        #wait 1 seconds and try again
        
        #There needs to be a "temp" folder. If there os none, create it!
        #does temp exist?
        tries = 0 #during fitting, AID sometimes wants to write to the history file. In this case we cant read
        try:
            while tries<15:#try a few times
                try:
                    temp_path = aid_bin.create_temp_folder()#create a temp folder if it does not already exist
                    #Create a  random filename for a temp. file
                    someletters = list("STERNBURGPILS")
                    temporaryfile = np.random.choice(someletters,5,replace=True)
                    temporaryfile = "".join(temporaryfile)+".xlsx"
                    temporaryfile = os.path.join(temp_path,temporaryfile)
                    shutil.copyfile(filename,temporaryfile) #copy the original excel file there
                    dic = pd.read_excel(temporaryfile,sheetname='History',index_col=0) #open it there
                    self.loaded_history = dic
                    para = pd.read_excel(temporaryfile,sheetname='Parameters')
                    print(temporaryfile)
                    #delete the tempfile
                    os.remove(temporaryfile)
                    self.loaded_para = para    
                    tries = 16
                except:
                    time.sleep(1.5)
                    tries+=1

        except Exception as e:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        #Check if dic exists now
        try:
            keys = list(dic.keys())
        except Exception as e:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
        #sort the list alphabetically
        keys_ = [l.lower() for l in keys]
        ind_sort = np.argsort(keys_)
        keys = list(np.array(keys)[ind_sort])

        #First keys should always be acc,loss,val_acc,val_loss -in this order
        keys_first = ["acc","loss","val_acc","val_loss"]
        for i in range(len(keys_first)):
            if keys_first[i] in keys:
                ind = np.where(np.array(keys)==keys_first[i])[0][0]
                if ind!=i:
                    del keys[ind]
                    keys.insert(i,keys_first[i])
        #Lastly check if there is "Saved" or "Time" present and shift it to the back
        keys_last = ["Saved","Time"]
        for i in range(len(keys_last)):
            if keys_last[i] in keys:
                ind = np.where(np.array(keys)==keys_last[i])[0][0]
                if ind!=len(keys):
                    del keys[ind]
                    keys.append(keys_last[i])


        
        self.tableWidget_HistoryItems.setColumnCount(len(keys)+1) #+1 because of "Show saved only"
        #for each key, put a checkbox on the tableWidget_HistoryInfo_pop
        rowPosition = self.tableWidget_HistoryItems.rowCount()
        if rowPosition==0:
            self.tableWidget_HistoryItems.insertRow(0)
        else:
            rowPosition=0
            
        for columnPosition in range(len(keys)):#(2,4):
            key = keys[columnPosition]
            item = QtWidgets.QTableWidgetItem(str(key))#("item {0} {1}".format(rowNumber, columnNumber))
            item.setBackground(QtGui.QColor(self.colorsQt[columnPosition]))
            item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
            item.setCheckState(QtCore.Qt.Unchecked)
            self.tableWidget_HistoryItems.setItem(rowPosition, columnPosition, item)

        #One checkbox at the end to switch on/of to show only the models that are saved
        columnPosition = len(keys)
        item = QtWidgets.QTableWidgetItem("Show saved only")#("item {0} {1}".format(rowNumber, columnNumber))
        item.setFlags( QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled  )
        item.setCheckState(QtCore.Qt.Unchecked)
        self.tableWidget_HistoryItems.setItem(rowPosition, columnPosition, item)
        
        self.tableWidget_HistoryItems.resizeColumnsToContents()
        self.tableWidget_HistoryItems.resizeRowsToContents()



    def history_tab_get_model_path(self):#Let user define a model he would like to convert
        #pushButton_LoadModel
        #Open a QFileDialog
        filepath = QtWidgets.QFileDialog.getOpenFileName(self, 'Select a trained model you want to convert', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filepath = filepath[0]
        if os.path.isfile(filepath):
            self.model_2_convert = filepath
            path, filename = os.path.split(filepath)
            try:
                modelindex = filename.split(".model")[0]
                modelindex = int(modelindex.split("_")[-1])
            except:
                modelindex = np.nan
                self.textBrowser_SelectedModelInfo.setText("Error loading model")
                return
            text = "Loaded model: "+filename+"\nModelindex: "+str(modelindex)+"\nFile is located in: "+filepath
            self.textBrowser_SelectedModelInfo.setText(text)
                         

    def history_tab_convertModel(self):
        #Check if there is text in textBrowser_SelectedModelInfo
        path = self.model_2_convert
        try:
            os.path.isfile(path)
        except:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No file defined!")
            msg.setWindowTitle("No file defined!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        if not os.path.isfile(path):
            #text_path = "\nFile not found!:"+path+"\nProbably the .model was deleted or not saved"
            #self.pushButton_convertModel.setEnabled(False)
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("\nFile not found!:"+path+"\nProbably the .model was deleted or not saved")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #If the source format is Keras tensforflow:
        source_format = str(self.combobox_initial_format.currentText())                        
        target_format = str(self.comboBox_convertTo.currentText()) #What is the target format?

        ##TODO: All conversion methods to multiprocessing functions!
        def conversion_successful_msg(text):#Enable the Convert to .nnet button
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(text)
            msg.setWindowTitle("Successfully converted model!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()


        ##################Keras TensorFlow -> .nnet############################
        if target_format==".nnet" and source_format=="Keras TensorFlow": 
            ConvertToNnet = 1
            worker = Worker(self.history_tab_convertModel_nnet_worker,ConvertToNnet)
            def get_model_keras_from_worker(dic):
                self.model_keras = dic["model_keras"]
            worker.signals.history.connect(get_model_keras_from_worker)
            def conversion_successful(i):#Enable the Convert to .nnet button
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)
                text = "Conversion Keras TensorFlow -> .nnet done"
                msg.setText(text)
                msg.setWindowTitle("Successfully converted model!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                #self.pushButton_convertModel.setEnabled(True) 
            worker.signals.history.connect(conversion_successful)
            self.threadpool.start(worker)

        ##################Keras TensorFlow -> Frozen .pb#######################
        elif target_format=="Frozen TensorFlow .pb" and source_format=="Keras TensorFlow":
            #target filename should be like source +_frozen.pb
            path_new = os.path.splitext(path)[0] + "_frozen.pb"
            aid_dl.convert_kerastf_2_frozen_pb(path,path_new)
            text = "Conversion Keras TensorFlow -> Frozen .pb is done"
            conversion_successful_msg(text)
        ##################Keras TensorFlow -> Optimized .pb####################
        elif target_format=="Optimized TensorFlow .pb" and source_format=="Keras TensorFlow":
            path_new = os.path.splitext(path)[0] + "_optimized.pb"
            aid_dl.convert_kerastf_2_optimized_pb(path,path_new)
            text = "Conversion Keras TensorFlow -> Optimized .pb is done"
            conversion_successful_msg(text)

        ####################Frozen -> Optimized .pb############################
        elif target_format=="Optimized TensorFlow .pb" and source_format=="Frozen TensorFlow .pb":
            path_new = os.path.splitext(path)[0] + "_optimized.pb"
            aid_dl.convert_frozen_2_optimized_pb(path,path_new)
            text = "Conversion Frozen -> Optimized .pb is done"
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> ONNX####################
        elif target_format=="ONNX (via keras2onnx)" and source_format=="Keras TensorFlow":
            path_new = os.path.splitext(path)[0] + ".onnx"
            aid_dl.convert_kerastf_2_onnx(path,path_new)
            text = "Conversion Keras TensorFlow -> ONNX (via keras2onnx) is done"
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> ONNX via MMdnn####################
        elif target_format=="ONNX (via MMdnn)" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_onnx_mmdnn(path)
            text = "Conversion Keras TensorFlow -> ONNX (via MMdnn) is done"
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> PyTorch Script####################
        elif target_format=="PyTorch Script"  and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"pytorch")
            text = "Conversion Keras TensorFlow -> PyTorch Script is done. You can now use this script and the saved weights to build the model using your PyTorch installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> Caffe Script####################
        elif target_format=="Caffe Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"caffe")
            text = "Conversion Keras TensorFlow -> Caffe Script is done. You can now use this script and the saved weights to build the model using your Caffe installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> CNTK Script####################
        elif target_format=="CNTK Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"cntk")
            text = "Conversion Keras TensorFlow -> CNTK Script is done. You can now use this script and the saved weights to build the model using your CNTK installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> mxnet Script####################
        elif target_format=="MXNet Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"mxnet")
            text = "Conversion Keras TensorFlow -> MXNet Script is done. You can now use this script and the saved weights to build the model using your MXNet installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> onnx Script####################
        elif target_format=="ONNX Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"onnx")
            text = "Conversion Keras TensorFlow -> ONNX Script is done. You can now use this script and the saved weights to build the model using your ONNX installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> TensorFlow Script####################
        elif target_format=="TensorFlow Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"tensorflow")
            text = "Conversion Keras TensorFlow -> TensorFlow Script is done. You can now use this script and the saved weights to build the model using your Tensorflow installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> Keras Script####################
        elif target_format=="Keras Script" and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_script(path,"keras")
            text = "Conversion Keras TensorFlow -> Keras Script is done. You can now use this script and the saved weights to build the model using your Keras installation."
            conversion_successful_msg(text)

        ##################Keras TensorFlow -> CoreML####################
        elif "CoreML" in target_format and source_format=="Keras TensorFlow":
            aid_dl.convert_kerastf_2_coreml(path)
            text = "Conversion Keras TensorFlow -> CoreML is done."
            conversion_successful_msg(text)

        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Not implemeted (yet)")
            msg.setWindowTitle("Not implemeted (yet)")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
                
        #If that worked without error, save the filepath for next time
        Default_dict["Path of last model"] = os.path.split(path)[0]
        aid_bin.save_aid_settings(Default_dict)


        
    def history_tab_convertModel_nnet_worker(self,ConvertToNnet,progress_callback,history_callback):
        #Define a new session -> Necessary for threading in TensorFlow
        with tf.Session(graph = tf.Graph(), config=config_gpu) as sess:
            path = self.model_2_convert
            try:
                model_keras = load_model(path,custom_objects=get_custom_metrics())
            except:
                model_keras = load_model(path)
                
            dic = {"model_keras":model_keras}
            history_callback.emit(dic)
            progress_callback.emit(1)

            if ConvertToNnet==1:
                #Since this happened in a thread, TensorFlow cant access it anywhere else
                #Therefore perform Conversion to nnet right away:
                model_config = model_keras.get_config()#["layers"]
                if type(model_config)==dict:
                    model_config = model_config["layers"]#for keras version>2.2.3, there is a change in the output of get_config()
                #Convert model to theano weights format (Only necesary for CNNs)
                for layer in model_keras.layers:
                   if layer.__class__.__name__ in ['Convolution1D', 'Convolution2D']:
                      original_w = K.get_value(layer.W)
                      converted_w = convert_kernel(original_w)
                      K.set_value(layer.W, converted_w)
                
                nnet_path, nnet_filename = os.path.split(self.model_2_convert)
                nnet_filename = nnet_filename.split(".model")[0]+".nnet" 
                out_path = os.path.join(nnet_path,nnet_filename)
                aid_dl.dump_to_simple_cpp(model_keras=model_keras,model_config=model_config,output=out_path,verbose=False)
                        


    def history_tab_ConvertToNnet(self):
        print("Not used")
#        model_keras = self.model_keras
#        model_config = model_keras.get_config()["layers"]
#        #Convert model to theano weights format (Only necesary for CNNs)
#        for layer in model_keras.layers:
#           if layer.__class__.__name__ in ['Convolution1D', 'Convolution2D']:
#              original_w = K.get_value(layer.W)
#              converted_w = convert_kernel(original_w)
#              K.set_value(layer.W, converted_w)
#        
#        nnet_path, nnet_filename = os.path.split(self.model_2_convert)
#        nnet_filename = nnet_filename.split(".model")[0]+".nnet" 
#        out_path = os.path.join(nnet_path,nnet_filename)
#        aid_dl.dump_to_simple_cpp(model_keras=model_keras,model_config=model_config,output=out_path,verbose=False)
#        msg = QtWidgets.QMessageBox()
#        msg.setIcon(QtWidgets.QMessageBox.Information)       
#        msg.setText("Successfully converted model and saved to\n"+out_path)
#        msg.setWindowTitle("Successfully converted model!")
#        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
#        msg.exec_()
#        self.pushButton_convertModel.setEnabled(False)
#TODO
    def test_nnet(self):
        #I need a function which calls a cpp app that uses the nnet and applies
        #it on a random image.
        #The same image is also used as input the the original .model and 
        #both results are then compared
        print("Not implemented yet")
        print("Placeholder")
        print("Building site")


    def actionDocumentation_function(self):        
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "Currently, there is no detailed written documentation. AIDeveloper instead makes strong use of tooltips."
        text = "<html><head/><body><p>"+text+"</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("Documentation")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def actionSoftware_function(self):
        if sys.platform == "win32":
            plat = "win"
        elif sys.platform=="darwin":
            plat = "mac"      
        elif sys.platform=="linux":
            plat = "linux"
        else:
            print("Unknown Operating system")
            plat = "Win"
            
        dir_deps = os.path.join(dir_root,"aid_dependencies_"+plat+".txt")#dir to aid_dependencies
        f = open(dir_deps, "r")
        text_modules = f.read()
        f.close()
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "<html><head/><body><p>AIDeveloper "+str(VERSION)+"<br>"+sys.version+"<br>Click 'Show Details' to retrieve a list of all used packages.</p></body></html>"
        msg.setText(text)
        msg.setDetailedText(text_modules)
        msg.setWindowTitle("Software")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        
    def actionAbout_function(self):
        icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff))
        icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
        msg = QtWidgets.QMessageBox()
        msg.setIconPixmap(icon)
        text = "AIDeveloper is written and maintained by Maik Herbig. Use maik.herbig@tu-dresden.de to contact the main developer if you find bugs or if you wish a particular feature. Icon theme 2 was mainly designed and created by Konrad Wauer."
        text = "<html><head/><body><p>"+text+"</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("About")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

    def actionLoadSession_function(self):
        #This function should allow to select and load a metafile and
        #Put the GUI the the corresponing state (place the stuff in the table, click Train/Valid)
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open meta-data', Default_dict["Path of last model"],"AIDeveloper Meta or session file (*meta.xlsx *session.xlsx)")
        filename = filename[0]
        if len(filename)==0:
            return
        UsedData = pd.read_excel(filename,sheetname="UsedData")
        Files = list(UsedData["rtdc_path"])

        file_exists = [os.path.exists(url) for url in Files]
        ind_true = np.where(np.array(file_exists)==True)[0]
        UsedData_true = UsedData.iloc[ind_true]
        
        Files_true = list(UsedData_true["rtdc_path"]) #select the indices that are valid

        #Add stuff to table_dragdrop
        rowPosition = int(self.table_dragdrop.rowCount())
        self.dataDropped(Files_true)

        #update the index,  train/valid checkbox and shuffle checkbox
        for i in range(len(Files_true)):
            #set the index (celltype)
            try:
                index = int(np.array(UsedData_true["class"])[i])
            except:
                index = int(np.array(UsedData_true["index"])[i])
                print("You are using an old version of AIDeveloper. Consider upgrading")

            self.table_dragdrop.cellWidget(rowPosition+i, 1).setValue(index)
            #is it checked for train or valid?
            trorvalid = str(np.array(UsedData_true["TrainOrValid"])[i])
            if trorvalid=="Train":
                self.table_dragdrop.item(rowPosition+i, 2).setCheckState(QtCore.Qt.Checked)
            elif trorvalid=="Valid":
                self.table_dragdrop.item(rowPosition+i, 3).setCheckState(QtCore.Qt.Checked)

            #how many cells/epoch during training or validation?
            try:
                nr_events_epoch = str(np.array(UsedData_true["nr_events_epoch"])[i])  
            except:
                nr_events_epoch = str(np.array(UsedData_true["nr_cells_epoch"])[i])  

            self.table_dragdrop.item(rowPosition+i, 6).setText(nr_events_epoch)
            #Shuffle or not?
            shuffle = bool(np.array(UsedData_true["shuffle"])[i])

            if shuffle==False:
                self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Unchecked)
                #Set Cells/Epoch to not editable
                item = self.table_dragdrop.item(rowPosition+i, 6)
                item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                #item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
            else:
                self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Checked)


            #zoom_factor = float(np.array(UsedData_true["zoom_factor"])[i])
            zoom_factor = str(np.array(UsedData_true["zoom_factor"])[i])  
            self.table_dragdrop.item(rowPosition+i, 9).setText(zoom_factor)


#            item = QtWidgets.QTableWidgetItem()
#            item.setData(QtCore.Qt.EditRole,zoom_factor)
#            self.table_dragdrop.setItem(rowPosition+i, 9, item)
#
#            #self.table_dragdrop.cellWidget(rowPosition+i, 9).setValue(zoom_factor)

                
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("This function does only refurbish the window 'Drag and drop .rtdc files here'. To load the corresponding model please use 'Load and continue' in the 'Define Model' -tab. Image augmentation parameters have to be adjusted manually.")
        msg.setWindowTitle("Only Drag and drop table affected!")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()


        #Now take care of missing data
        #Take care of missing files (they might have been moved to a different location)
        ind_false = np.where(np.array(file_exists)==False)[0]  
        #Files_false = list(UsedData_false["rtdc_path"]) #select the indices that are valid
        if len(ind_false)>0:
            UsedData_false = UsedData.iloc[ind_false]
            Files_false = list(UsedData_false["rtdc_path"]) #select the indices that are valid
            self.dataDropped(Files_false)

            self.user_selected_path = None
            #Create popup that informs user that there is missing data and let him specify a location
            #to search for the missing files
            def add_missing_files():
                filename = QtWidgets.QFileDialog.getExistingDirectory(self, 'Select directory', Default_dict["Path of last model"])
                user_selected_path = filename
                if len(filename)==0:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Invalid directory")
                    msg.setWindowTitle("Invalid directory")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                #get the hashes 
                hashes = list(np.array(UsedData_false["hash"])[ind_false])
                paths = list(np.array(UsedData_false["rtdc_path"])[ind_false])

                paths_new,info = aid_bin.find_files(user_selected_path,paths,hashes)
                text = ('\n'.join([str(a) +"\t"+ b for a,b in zip(paths_new,info)]))
                self.textBrowser_Info_pop2.setText(text)

                #Add stuff to table_dragdrop
                rowPosition = int(self.table_dragdrop.rowCount())
                self.dataDropped(paths_new)
                
                for i in range(len(paths_new)):
                    #set the index (celltype)
                    try:
                        index = int(np.array(UsedData_false["class"])[i])
                    except:
                        index = int(np.array(UsedData_false["index"])[i])
                        print("You are using an old version of AIDeveloper. Consider upgrading")

                    self.table_dragdrop.cellWidget(rowPosition+i, 1).setValue(index)
                    #is it checked for train or valid?
                    trorvalid = str(np.array(UsedData_false["TrainOrValid"])[i])
                    if trorvalid=="Train":
                        self.table_dragdrop.item(rowPosition+i, 2).setCheckState(QtCore.Qt.Checked)
                    elif trorvalid=="Valid":
                        self.table_dragdrop.item(rowPosition+i, 3).setCheckState(QtCore.Qt.Checked)
                    #how many cells/epoch during training or validation?
                    nr_events_epoch = str(np.array(UsedData_false["nr_events_epoch"])[i]) 

                    #how many cells/epoch during training or validation?
                    try:
                        nr_events_epoch = str(np.array(UsedData_false["nr_events_epoch"])[i])  
                    except:
                        nr_events_epoch = str(np.array(UsedData_false["nr_cells_epoch"])[i])  
                        print("You are using an old version of AIDeveloper. Consider upgrading")

                    self.table_dragdrop.item(rowPosition+i, 6).setText(nr_events_epoch)
                    #Shuffle or not?
                    shuffle = bool(np.array(UsedData_false["shuffle"])[i])
                    if shuffle==False:
                        self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Unchecked)
                        #Set Cells/Epoch to not editable
                        item = self.table_dragdrop.item(rowPosition+i, 6)
                        item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
                        #item.setFlags(item.flags() |QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable )
                    else:
                        self.table_dragdrop.item(rowPosition+i, 8).setCheckState(QtCore.Qt.Checked)
                    #zoom_factor = float(np.array(UsedData_false["zoom_factor"])[i])
                    zoom_factor = str(np.array(UsedData_false["zoom_factor"])[i])  
                    self.table_dragdrop.item(rowPosition+i, 9).setText(zoom_factor)

            self.w_pop2 = MyPopup()
            self.gridLayout_w_pop2 = QtWidgets.QGridLayout(self.w_pop2)
            self.gridLayout_w_pop2.setObjectName("gridLayout_w_pop2")
            self.verticalLayout_w_pop2 = QtWidgets.QVBoxLayout()
            self.verticalLayout_w_pop2.setObjectName("verticalLayout_w_pop2")
            self.horizontalLayout_w_pop2 = QtWidgets.QHBoxLayout()
            self.horizontalLayout_w_pop2.setObjectName("horizontalLayout_w_pop2")
            self.pushButton_Close_pop2 = QtWidgets.QPushButton(self.centralwidget)
            self.pushButton_Close_pop2.setObjectName("pushButton_Close_pop2")
            self.pushButton_Close_pop2.clicked.connect(self.w_pop2.close)
            self.horizontalLayout_w_pop2.addWidget(self.pushButton_Close_pop2)
            self.pushButton_Search_pop2 = QtWidgets.QPushButton(self.centralwidget)
            self.pushButton_Search_pop2.clicked.connect(add_missing_files)
            self.pushButton_Search_pop2.setObjectName("pushButton_Search")
            self.horizontalLayout_w_pop2.addWidget(self.pushButton_Search_pop2)
            self.verticalLayout_w_pop2.addLayout(self.horizontalLayout_w_pop2)
            self.textBrowser_Info_pop2 = QtWidgets.QTextBrowser(self.centralwidget)
            self.textBrowser_Info_pop2.setObjectName("textBrowser_Info_pop2")
            self.verticalLayout_w_pop2.addWidget(self.textBrowser_Info_pop2)
            self.gridLayout_w_pop2.addLayout(self.verticalLayout_w_pop2, 0, 0, 1, 1)
            self.w_pop2.setWindowTitle("There are missing files. Do you want to search for them?")
            self.pushButton_Close_pop2.setText("No")
            self.pushButton_Search_pop2.setText("Define folder to search files")
            self.w_pop2.show()

#        #update the Define Model tab
#        Parameters = pd.read_excel(filename,sheetname="Parameters")
#        chosen_model = str(Parameters["Chosen Model"])
#        index = self.comboBox_ModelSelection.findText(chosen_model, QtCore.Qt.MatchFixedString)
#        if index >= 0:
#            self.comboBox_ModelSelection.setCurrentIndex(index)
#        #update the Normalization method
#        norm = str(Parameters["Normalization"])
#        index = self.comboBox_Normalization.findText(norm, QtCore.Qt.MatchFixedString)
#        if index >= 0:
#            self.comboBox_Normalization.setCurrentIndex(index)

        #If all this run without error, save the path.
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)
        
        #Update the overview-box
        if self.groupBox_DataOverview.isChecked()==True:
            self.dataOverviewOn()        

    def actionSaveSession_function(self):
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save session', Default_dict["Path of last model"],"AIDeveloper Session file (*_session.xlsx)")
        filename = filename[0]
        path, fname = os.path.split(filename)
                
        if len(fname)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        if fname.endswith(".xlsx"):
            fname = fname.split(".xlsx")[0]            
        if fname.endswith("_session"):
            fname = fname.split("_session")[0]            
        if fname.endswith("_meta"):
            fname = fname.split("_meta")[0]
        if fname.endswith(".model"):
            fname = fname.split(".model")[0]
        if fname.endswith(".arch"):
            fname = fname.split(".arch")[0]

        #add the suffix _session.xlsx
        if not fname.endswith("_session.xlsx"):
            fname = fname +"_session.xlsx"
        filename = os.path.join(path,fname)    
        
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        #Used files go to a separate sheet on the -session.xlsx
        SelectedFiles = self.items_clicked()
        SelectedFiles_df = pd.DataFrame(SelectedFiles)
        pd.DataFrame().to_excel(writer,sheet_name='UsedData') #initialize empty Sheet
        SelectedFiles_df.to_excel(writer,sheet_name='UsedData')
        DataOverview_df = self.get_dataOverview()
        DataOverview_df.to_excel(writer,sheet_name='DataOverview') #write data overview to separate sheet            

        writer.save()
        writer.close()

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("Successfully saved as "+filename)
        msg.setWindowTitle("Successfully saved")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()
        
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]        
        aid_bin.save_aid_settings(Default_dict)
        
    def actionClearList_function(self):
        #Remove all items from dragdrop table
        while (self.table_dragdrop.rowCount() > 0):
            self.table_dragdrop.removeRow(0)
        #reset ram
        self.ram = dict()

        #Remove all items from comboBox_chooseRtdcFile
        self.comboBox_chooseRtdcFile.clear()
        self.comboBox_selectData.clear()
        if self.groupBox_DataOverview.isChecked()==True:
            self.dataOverviewOn()

    def actionRemoveSelected_function(self):
        #Which rows are highlighted?
        rows_selected = np.array([index.row() for index in self.table_dragdrop.selectedIndexes()])
        for row in (rows_selected):
            self.table_dragdrop.removeRow(row)
            self.comboBox_chooseRtdcFile.removeItem(row)
            self.comboBox_selectData.removeItem(row)
            #if there are rows below this row, they will move up one step:
            ind = np.where(np.array(rows_selected)>row)[0]
            rows_selected[ind] -= 1

    def actionSaveToPng_function(self):
        #Which table items are selected?
        rows_selected = np.array([index.row() for index in self.table_dragdrop.selectedIndexes()])
        if len(rows_selected)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)       
            msg.setText("Please first select rows in the table!")
            msg.setWindowTitle("No rows selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Ask user to which folder the images should be written:
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to .png/.jpg', Default_dict["Path of last model"],"Image file format (*.png *.jpg *.bmp *.eps *.gif *.ico *.icns)")
        filename = filename[0]
        if len(filename)==0:
            return
        filename_X, file_extension = os.path.splitext(filename)#divide into path and file_extension if possible
        #Check if the chosen file_extension is valid
        if not file_extension in [".png",".jpg",".bmp",".eps",".gif",".ico",".icns"]:
            print("Invalid file extension detected. Will use .png instead.")
            file_extension = ".png"
            
        #Check the chosen export-options
        if bool(self.actionExport_Original.isChecked())==True:
            print("Export original images")
            save_cropped = False
        elif bool(self.actionExport_Cropped.isChecked())==True:
            print("Export cropped images")
            save_cropped = True
        elif bool(self.actionExport_Off.isChecked())==True:
            print("Exporting is turned off")
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Plase choose a different Export-option in ->Options->Export")
            msg.setWindowTitle("Export is turned off!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        if save_cropped==True:
            #Collect information for image processing
            cropsize = self.spinBox_imagecrop.value()
            color_mode = str(self.comboBox_loadedRGBorGray.currentText())
            zoom_methods = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
            zoom_order = np.where(np.array(zoom_methods)==True)[0]
            
            index = 0
            for row in (rows_selected):
                #get the corresponding rtdc_path
                rtdc_path = str(self.table_dragdrop.cellWidget(row, 0).text())
                nr_events = None #no number needed as we take all images (replace=False in gen_crop_img)
                zoom_factor = float(self.table_dragdrop.item(row, 9).text())            
                gen = aid_img.gen_crop_img(cropsize,rtdc_path,nr_events=nr_events,replace=False,random_images=False,zoom_factor=zoom_factor,zoom_order=zoom_order,color_mode=color_mode,padding_mode='constant')
                images = next(gen)[0]
                #Save the images data to .png/.jpeg...
                for img in images:
                    img = PIL.Image.fromarray(img)
                    img.save(filename_X+"_"+str(index)+file_extension)
                    index+=1

        if save_cropped==False:#save the original images without pre-processing
            index = 0
            for row in (rows_selected):
                rtdc_path = str(self.table_dragdrop.cellWidget(row, 0).text())
                failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                if failed:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)       
                    msg.setText(str(rtdc_ds))
                    msg.setWindowTitle("Error occurred during loading file")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
    
                images = rtdc_ds["image"] #get the images
                #Save the images data to .png/.jpeg...
                for img in images:
                    img = PIL.Image.fromarray(img)
                    img.save(filename_X+"_"+str(index)+file_extension)
                    index+=1
                    
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]
        aid_bin.save_aid_settings(Default_dict)

    def actionRemoveSelectedPeaks_function(self):
        #Which rows are highlighted?
        rows_selected = np.array([index.row() for index in self.tableWidget_showSelectedPeaks.selectedIndexes()])
        #delete each row only once :)
        rows_selected = np.unique(rows_selected)
        for row in (rows_selected):
            self.tableWidget_showSelectedPeaks.removeRow(row)
            #if there are rows below this row, they will move up one step:
            ind = np.where(np.array(rows_selected)>row)[0]
            rows_selected[ind] -=1
        #Update the widget_showSelectedPeaks
        self.update_peak_plot()


    def actionRemoveAllPeaks_function(self):
        #Remove all items from tableWidget_showSelectedPeaks
        while (self.tableWidget_showSelectedPeaks.rowCount() > 0):
            self.tableWidget_showSelectedPeaks.removeRow(0)

    def actionDataToRamNow_function(self):
        self.statusbar.showMessage("Moving data to RAM")
        #check that the nr. of classes are equal to the model out put
        SelectedFiles = self.items_clicked()
        color_mode = self.get_color_mode()
        zoom_factors = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles]
        zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        zoom_order = int(np.where(np.array(zoom_order)==True)[0])

        #Get the user-defined cropping size
        crop = int(self.spinBox_imagecrop.value())          
        #Make the cropsize a bit larger since the images will later be rotated
        cropsize2 = np.sqrt(crop**2+crop**2)
        cropsize2 = np.ceil(cropsize2 / 2.) * 2 #round to the next even number
        
        dic = aid_img.crop_imgs_to_ram(list(SelectedFiles),crop,zoom_factors=zoom_factors,zoom_order=zoom_order,color_mode=color_mode)
        self.ram = dic 
        
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)       
        msg.setText("Successfully moved data to RAM")
        msg.setWindowTitle("Moved Data to RAM")
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        msg.exec_()

        self.statusbar.showMessage("")







    ###########################################################################
    ###########################################################################
    ###########################################################################
    ###########################################################################
    #######################Functions for Assess model tab######################
    def assessmodel_tab_load_model(self):
        #Get the requested model-name from the chosen metafile
        #Open a QFileDialog
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Select a trained model you want to assess', Default_dict["Path of last model"],"Keras Model file (*.model)")
        filename = filename[0]        
        if os.path.isfile(filename):
            #Put this path on the Assess Model tab
            self.lineEdit_LoadModel_2.setText(filename)
            #Save the path to a variable that is then used by history_tab_convertModel_nnet_worker
            self.load_model_path = filename
            
            #Get the modelindex
            path,filename = os.path.split(filename)
            modelindex = filename.split(".model")[0]
            modelindex = int(modelindex.split("_")[-1])
            #Update the modelindex on the Assess Model tab
            self.spinBox_ModelIndex_2.setValue(int(modelindex))

            model_full_h5 = h5py.File(self.load_model_path, 'r')
            model_config = model_full_h5.attrs['model_config']
            model_full_h5.close() #close the hdf5                
            model_config = json.loads(str(model_config)[2:-1])
            
            try: #Sequential Model
                in_dim = model_config['config'][0]['config']['batch_input_shape']
            except: #Functional Api
                in_dim = model_config['config']["layers"][0]["config"]["batch_input_shape"]
            try: #Sequential Model
                out_dim = model_config['config'][-2]['config']['units']
            except: #Functional Api
                out_dim = model_config['config']["layers"][-2]["config"]["units"]
            
            self.spinBox_Crop_2.setValue(int(in_dim[-2]))
            self.spinBox_OutClasses_2.setValue(int(out_dim))
            print("input dimension:"+str(in_dim))
            #Adjust the Color mode in the UI:
            channels = in_dim[-1] #TensorFlow: channels in last dimension
            if channels==1:
                if self.get_color_mode()!="Grayscale":
                    #when model needs Grayscale, set the color mode in comboBox_GrayOrRGB to that
                    index = self.comboBox_GrayOrRGB.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)                                                            
                    index = self.comboBox_loadedRGBorGray.findText("Grayscale", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_loadedRGBorGray.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to Grayscale",5000)
            
            elif channels==3:
                if self.get_color_mode()!="RGB":
                    #when model needs RGB, set the color mode in the ui to that
                    index = self.comboBox_GrayOrRGB.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_GrayOrRGB.setCurrentIndex(index)
                    index = self.comboBox_loadedRGBorGray.findText("RGB", QtCore.Qt.MatchFixedString)
                    if index >= 0:
                        self.comboBox_loadedRGBorGray.setCurrentIndex(index)                                        
                    self.statusbar.showMessage("Color Mode set to RGB",5000)
            else:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Channel dimensions of model ("+str(channels)+" channels) is not supported. Only 1 or 3 channels are allowed.")
                msg.setWindowTitle("Unsupported channel dimension")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
                

            path,fname = os.path.split(self.load_model_path)        
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)

            try:
                parameters = pd.read_excel(metafile_path,sheetname='Parameters')
                norm = str(parameters["Normalization"].values[0])
                model_type = str(parameters["Chosen Model"].values[0])
                
                index = self.comboBox_Normalization_2.findText(norm, QtCore.Qt.MatchFixedString)
                if index >= 0:
                    self.comboBox_Normalization_2.setCurrentIndex(index)
                    
                else:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("Unkown normalization method found in .meta file")
                    msg.setWindowTitle("Unkown normalization method")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()
                    return
                    
                self.lineEdit_ModelSelection_2.setText(model_type)

            except: #there is not such a file, or the file cannot be opened
                #Ask the user to choose the normalization method
                self.lineEdit_ModelSelection_2.setText("Unknown")
                self.comboBox_Normalization_2.setEnabled(True)
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Meta file not found/ Could not be read. Please specify the normalization method manually (dropdown menu)")
                msg.setWindowTitle("No .meta available")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
            
            #If all that run without issue, remember the path for next time
            Default_dict["Path of last model"] = os.path.split(self.load_model_path)[0]
            aid_bin.save_aid_settings(Default_dict)

        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found!:\nProbably the .model was deleted or not saved")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
    
    def inference_time_worker(self,progress_callback,history_callback):
        #Initiate a fresh session
        with tf.Session(graph = tf.Graph(), config=config_gpu) as sess:
            model_keras = load_model(self.load_model_path,custom_objects=get_custom_metrics())
            #Get the model input dimensions
            in_dim = np.array(model_keras.get_input_shape_at(0))
            ind = np.where(in_dim==None)
            in_dim[ind] = 1
            nr_imgs = self.spinBox_inftime_nr_images.value()
            nr_imgs = int(np.round(float(nr_imgs)/10.0))
            
            #Warm up by predicting a single image
            image = (np.random.randint(0,255,size=in_dim)).astype(np.float32)/255.0
            model_keras.predict(image) # warm up

            Times = []
            for k in range(10):
                image = (np.random.randint(0,255,size=in_dim)).astype(np.float32)/255.0
                t1 = time.time()
                for i in range(nr_imgs):#predict 50 times 20 images
                    model_keras.predict(image)
                t2 = time.time()
                dt = (t2-t1)/(nr_imgs) #divide by nr_imgs to get time [s] per image
                dt = dt*1000.0 #multiply by 1000 to change to ms range
                dic = {"outp":str(round(dt,3))+"ms"}
                history_callback.emit(dic)
                Times.append(dt)   
                
        #Send out the Times
        text = " [ms] Mean: "+str(round(np.mean(Times),3))+"; "+"Median: "+str(round(np.median(Times),3))+"; "+"Min: "+str(round(np.min(Times),3))+"; "+"Max: "+str(round(np.max(Times),3))
        dic = {"outp":text}
        history_callback.emit(dic)
        progress_callback.emit(1) #when finished return one
        self.threadpool_single_queue = 0 #reset the thread-counter

    def inference_time(self):
        #Take the model path from the GUI
        self.load_model_path = str(self.lineEdit_LoadModel_2.text())
        if len(self.load_model_path)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        #Increase the thread-counter by one; only after finishing the thread, it will be reset to 0
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.inference_time_worker)
            def get_dt_from_worker(dic):
                outp = dic["outp"]
                self.lineEdit_InferenceTime.setText(outp)
            worker.signals.history.connect(get_dt_from_worker)    
            self.threadpool_single.start(worker)


    def update_check_worker(self,progress_callback,history_callback):
        dic = aid_bin.check_update(VERSION)
        #dic = {"Errors":None,"latest_release":latest_release,"latest_release_url":url,"changelog":changelog}
        history_callback.emit(dic)
        progress_callback.emit(1) #when finished return one
        self.threadpool_single_queue = 0 #reset the thread-counter

    def actionUpdate_check_function(self):
        #Increase the thread-counter by one; only after finishing the thread, it will be reset to 0
        self.threadpool_single_queue += 1
        if self.threadpool_single_queue == 1:
            worker = Worker(self.update_check_worker)
            def get_info_from_worker(dic):

                icon = QtGui.QImage(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_update_01.png"))
                icon = QtGui.QPixmap(icon).scaledToHeight(32, QtCore.Qt.SmoothTransformation)
                msg = QtWidgets.QMessageBox()
                msg.setIconPixmap(icon)

                if dic["Errors"]!=None:
                    msg.setText(str(dic["Errors"]))
                    msg.setWindowTitle("Error")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()

                if dic["Errors"]==None:#No errors! Nice
                    latest_release = dic["latest_release"]
                    
                    if latest_release=="You are up to date":
                        text = "<html><head/><body><p>"+latest_release+"</p></body></html>"
                        msg.setText(text)
                        #msg.setInformativeText("<html><head/><body><p>"+changelog+"</p></body></html>")
                        msg.setWindowTitle(latest_release)
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                    
                    else:
                        text = "Find latest release here: <a href ="+dic["latest_release_url"]+">"+dic["latest_release_url"]+"</a>"
                        text = "<html><head/><body><p>"+text+"</p></body></html>"
                        msg.setText(text)
                        msg.setInformativeText("<html><head/><body><p>"+dic["changelog"]+"</p></body></html>")
                        msg.setWindowTitle("New version detected!")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                
            worker.signals.history.connect(get_info_from_worker)    
            self.threadpool_single.start(worker)


    def get_validation_data_from_clicked(self,get_normalized=True):
        #Check, if files were clicked
        SelectedFiles = self.items_clicked()
        ######################Load the Validation Data################################
        ind = [selectedfile["TrainOrValid"] == "Valid" for selectedfile in SelectedFiles]
        ind = np.where(np.array(ind)==True)[0]

        if len(ind)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No validation data was selected. Please use tab 'Build' and drag/drop to load data")
            msg.setWindowTitle("No validation data selected")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return 0

        SelectedFiles_valid = np.array(SelectedFiles)[ind]
        SelectedFiles_valid = list(SelectedFiles_valid)
        indices_valid = [selectedfile["class"] for selectedfile in SelectedFiles_valid]
        nr_events_epoch_valid = [selectedfile["nr_events_epoch"] for selectedfile in SelectedFiles_valid]
        rtdc_path_valid = [selectedfile["rtdc_path"] for selectedfile in SelectedFiles_valid]
        zoom_factors_valid = [selectedfile["zoom_factor"] for selectedfile in SelectedFiles_valid]
        zoom_order = [self.actionOrder0.isChecked(),self.actionOrder1.isChecked(),self.actionOrder2.isChecked(),self.actionOrder3.isChecked(),self.actionOrder4.isChecked(),self.actionOrder5.isChecked()]
        zoom_order = int(np.where(np.array(zoom_order)==True)[0])
        shuffle_valid = [selectedfile["shuffle"] for selectedfile in SelectedFiles_valid]
        
        #Read other model properties from the Ui
        norm = self.comboBox_Normalization_2.currentText()
        norm = str(norm)
        #if normalization method needs mean/std of training set, the metafile needs to be loaded:
        if norm == "StdScaling using mean and std of all training data":
            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)
            parameters = pd.read_excel(metafile_path,sheetname='Parameters')
            mean_trainingdata = parameters["Mean of training data used for scaling"]
            std_trainingdata = parameters["Std of training data used for scaling"]
        else:
            mean_trainingdata = None
            std_trainingdata = None
            
        crop = int(self.spinBox_Crop_2.value()) 
        padding_expert = str(self.comboBox_expt_paddingMode.currentText()).lower()

        #read self.ram to new variable ; DONT clear ram after since multiple assessments can run on the same data.
        DATA = self.ram
        #self.ram = dict() #DONT clear the ram here! 
         
        ############Cropping#####################
        X_valid,y_valid,Indices = [],[],[]
        for i in range(len(SelectedFiles_valid)):
            if not self.actionDataToRam.isChecked():
                gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace=True means that individual cells could occur several times
            else: #get a similar generator, using the ram-data
                if len(DATA)==0:
                    gen_valid = aid_img.gen_crop_img(crop,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True,zoom_factor=zoom_factors_valid[i],zoom_order=zoom_order,color_mode=self.get_color_mode(),padding_mode=padding_expert) #Replace=True means that individual cells could occur several times
                else:
                    if self.actionVerbose.isChecked():
                        print("Loaded data from RAM")
                    gen_valid = aid_img.gen_crop_img_ram(DATA,rtdc_path_valid[i],nr_events_epoch_valid[i],random_images=shuffle_valid[i],replace=True) #Replace=True means that individual cells could occur several times
            
            gen = next(gen_valid)
            X_valid.append(gen[0])
            y_valid.append(np.repeat(indices_valid[i],X_valid[-1].shape[0]))
            Indices.append(gen[1]) #Cell index to track the event in the data-set(not cell-type!)

            
        X_valid_orig = [X.astype(np.uint8) for X in X_valid]
        X_valid = np.concatenate(X_valid)
        
#        dim = X_valid.shape
#        if dim[2]!=crop:
#            remove = int(dim[2]/2.0 - crop/2.0)
#            #X_batch = X_batch[:,:,remove:-remove,remove:-remove] #crop to crop x crop pixels #Theano
#            X_valid = X_valid[:,remove:-remove,remove:-remove] #crop to crop x crop pixels #TensorFlow

        print("X_valid has following dimension:")
        print(X_valid.shape)
        
        y_valid = np.concatenate(y_valid)

        if len(np.array(X_valid).shape)<3:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Discarded all events because too far at border of image (check zooming/cropping settings!)")
            msg.setWindowTitle("Empty dataset!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()       
            return 0

        if get_normalized == True:
            if norm == "StdScaling using mean and std of all training data":
                X_valid = aid_img.norm_imgs(X_valid,norm,mean_trainingdata,std_trainingdata)
            else:
                X_valid = aid_img.norm_imgs(X_valid,norm)
        else:
            X_valid = None
        dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":rtdc_path_valid,"X_valid_orig":X_valid_orig,"X_valid":X_valid,"y_valid":y_valid,"Indices":Indices}
        self.ValidationSet = dic
        return 1
        
    def export_valid_to_rtdc(self):        
        if not type(self.ValidationSet) is type(None): #If ValidationSet is not none, there has been a ValidationSet loaded already
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Re-used validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM")
            msg.setWindowTitle("Re-Used data")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            worked = 1
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            worked = self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
        if worked==0:
            return
        
        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        X_valid = []
        X_valid.append(self.ValidationSet["X_valid"][:,:,:,0])
#        import pickle
#        with open('X_valid.pickle', 'wb') as handle:
#            pickle.dump(X_valid, handle, protocol=pickle.HIGHEST_PROTOCOL)            
        X_valid_orig = self.ValidationSet["X_valid_orig"]
#        with open('X_valid_orig.pickle', 'wb') as handle:
#            pickle.dump(X_valid_orig, handle, protocol=pickle.HIGHEST_PROTOCOL)            
        
        Indices = self.ValidationSet["Indices"]
        y_valid = self.ValidationSet["y_valid"]
        
        #Get a filename from the user for the new file
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to rtdc', Default_dict["Path of last model"],"rtdc file (*.rtdc)")
        filename = filename[0]
        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
           
        #add the suffix _Valid_Data.avi or _Valid_Labels.npy
        if not filename.endswith(".rtdc"):
            filename = filename +".rtdc"
        filename_X = filename.split(".rtdc")[0]+"_Valid_Data.rtdc"
        filename_y = filename.split(".rtdc")[0]+"_Valid_Labels.txt"

        if bool(self.actionExport_Original.isChecked())==True:
            print("Export original images")
            save_cropped = False
        elif bool(self.actionExport_Cropped.isChecked())==True:
            print("Export cropped images")
            save_cropped = True
        elif bool(self.actionExport_Off.isChecked())==True:
            print("Exporting is turned off")
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("You could choose a different Exporting option in ->Option->Export")
            msg.setWindowTitle("Export is turned off!")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        aid_bin.write_rtdc(filename_X,rtdc_path_valid,X_valid_orig,Indices,cropped=save_cropped,color_mode=self.get_color_mode())
        np.savetxt(filename_y,y_valid.astype(int),fmt='%i')
        
        #If all that run without issue, remember the path for next time
        Default_dict["Path of last model"] = os.path.split(filename)[0]     
        aid_bin.save_aid_settings(Default_dict)
        
    def import_valid_from_rtdc(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Valid_Data.rtdc', Default_dict["Path of last model"],".rtdc file (*_Valid_Data.rtdc)")
        filename = filename[0]
        rtdc_path = filename
        #Load the images:
        if len(filename)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No valid filename was chosen.")
            msg.setWindowTitle("No valid filename was chosen")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #Load the corresponding labels
        filename_labels = filename.split("Data.rtdc")[0]+"Labels.txt"
        if not os.path.isfile(filename_labels):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("No corresponding _Labels.npy file found! Expected it here: "+filename_labels)
            msg.setWindowTitle("No Labels found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        y_valid = np.loadtxt(filename_labels).astype(int)
        #Inform user (statusbar message)
        self.statusbar.showMessage("Loaded labels from "+filename_labels,5000)

        #Read images from .rtdc file
        failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
        if failed:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)       
            msg.setText(str(rtdc_ds))
            msg.setWindowTitle("Error occurred during loading file")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
        #Make the Image dimensions matching the requirements of the model
        model_in = int(self.spinBox_Crop_2.value())
        model_out = int(self.spinBox_OutClasses_2.value())
        color_mode = str(self.comboBox_loadedRGBorGray.currentText())
        if model_in==1 and model_out==1:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please first define a model. The validation data will then be cropped according to the required model-input size")
            msg.setWindowTitle("No model defined")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        
        x_valid = np.array(rtdc_ds["image"])
        dim = x_valid.shape[1]
        channels = x_valid.shape[-1]
        if channels==3 and color_mode=="RGB": #this would fit :)
            pass
        if channels==3 and color_mode=="Grayscale": #we have to convert to grayscale
            print("Use luminosity formula to convert RGB input data to grayscale")
            x_valid = (0.21 * x_valid[:,:,:,:1]) + (0.72 * x_valid[:,:,:,1:2]) + (0.07 * x_valid[:,:,:,-1:])
            x_valid = x_valid[:,:,:,0] 
            x_valid  = x_valid.astype(np.uint8)           
            channels=1
            
        if channels==1 and color_mode=="Grayscale":#this would fit :)
            pass
        if channels==1 and color_mode=="RGB": #this would require conversion from gray to RGB, but this is not trivial!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Loaded data is only Grayscale, while the model needs RGB. There is no conversion possible.")
            msg.setWindowTitle("Grayscale vs RGB error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #The shape of Original images is not squared
        if x_valid.shape[1]!=x_valid.shape[2]:
            print("Images are non-squared->use aid_img.gen_crop_img")
            #Use aid_img.gen_crop_img to crop with respect to centroid
            gen_valid = aid_img.gen_crop_img(model_in,rtdc_path,random_images=False)
            x_valid,index = next(gen_valid)
            #When object is too far at side of image, the frame is dropped.
            #Consider this for y_valid
            y_valid = y_valid[index]
        #Otherwise, if the shape is squared:
        elif x_valid.shape[1]==x_valid.shape[2]:
            #simply grab the images, if the shape is fitting already:
            if model_in==x_valid.shape[1]:
                print("Images already have the correct dimensions")
                x_valid = x_valid        
            #Otherwise, adjust the dimensions by cropping:
            elif model_in<x_valid.shape[1]: #Model input is smaller and images in rtdc_ds are already squared
                print("Squared images are cropped to correct size")
                
                remove = int(dim/2.0 - model_in/2.0)
                print("Crop validation images to " +str(model_in) +"x"+str(model_in) +" pix")
                x_valid = x_valid[:,remove:-remove,remove:-remove] #crop 

            elif model_in>x_valid.shape[1]: #Model needs larger images than what was found in .rtdc! Ups!
                #Inform user!
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("Model input dimension ("+str(model_in)+"x"+str(model_in)+"pix) is larger than validation data dimension ("+str(x_valid.shape)+"). Program will use padding to increase image dimensions. It would be better to get original images that are larger and crop them!")
                msg.setWindowTitle("Wrong image dimension")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()

                diff = int(0.5*(model_in-dim))
                #attach this number of rows and cols to the left,right,up,down
                if color_mode=="Grayscale":
                    x_valid = [np.pad(x_valid[i,:,:], ((diff,diff),(diff,diff)), 'constant') for i in range(len(x_valid))]
                elif color_mode=="RGB":
                    x_valid = [np.pad(x_valid[i,:,:,:], ((diff,diff),(diff,diff),(0,0)), 'constant') for i in range(len(x_valid))]
                    
                x_valid = np.r_[x_valid]

        if not model_in==x_valid.shape[-2]:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Model input dimension ("+str(model_in)+"x"+str(model_in)+"pix) and validation data dimension ("+str(x_valid.shape)+") do not match")
            msg.setWindowTitle("Wrong image dimension")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        #Normalize the images
        norm = self.comboBox_Normalization_2.currentText()
        norm = str(norm)
        #if normalization method needs mean/std of training set, the metafile needs to be loaded:
        if norm == "StdScaling using mean and std of all training data":
            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)
            parameters = pd.read_excel(metafile_path,sheetname='Parameters')
            mean_trainingdata = parameters["Mean of training data used for scaling"]
            std_trainingdata = parameters["Std of training data used for scaling"]
        else:
            mean_trainingdata = None
            std_trainingdata = None

        X_valid_orig = np.copy(x_valid) #copy the cropped but non-normalized images
        if norm == "StdScaling using mean and std of all training data":
            X_valid = aid_img.norm_imgs(x_valid,norm,mean_trainingdata,std_trainingdata)
        else:
            X_valid = aid_img.norm_imgs(x_valid,norm)
        Indices = np.array(range(X_valid.shape[0])) #those are just indices to identify single cells in the file ->not cell-type indices!
        SelectedFiles_valid = None #[].append(rtdc_path)#
        nr_events_epoch_valid = None
        
        dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":[rtdc_path],"X_valid_orig":[X_valid_orig],"X_valid":X_valid,"y_valid":y_valid,"Indices":[Indices]}
        self.ValidationSet = dic

        self.statusbar.showMessage("Validation data loaded to RAM",5000)
        
        #Update the table
        #Prepare a table in tableWidget_Info
        self.tableWidget_Info_2.setColumnCount(0)#Reset table
        self.tableWidget_Info_2.setRowCount(0)#Reset table
        self.tableWidget_Info_2.setColumnCount(4) #Four columns

        nr_ind = len(set(y_valid)) #number of different labels ("indices")
        nr_rows = nr_ind
        self.tableWidget_Info_2.setRowCount(nr_rows)
        #Wich selected file has the most features?
        header_labels = ["Class","Nr of cells","Clr","Name"]
        self.tableWidget_Info_2.setHorizontalHeaderLabels(header_labels) 
        header = self.tableWidget_Info_2.horizontalHeader()
        for i in range(4):
            header.setResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)        
        
        rowPosition = 0      
        #Total nr of cells for each index
        for index in np.unique(y_valid):
            #put the index in column nr. 0
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 0, item)
            #Get the validation files of that index
            ind = np.where(y_valid==index)[0]
            nr_events_epoch = len(ind)
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, str(np.sum(nr_events_epoch)))
            self.tableWidget_Info_2.setItem(rowPosition, 1, item)
            
            #Column for color
            item = QtWidgets.QTableWidgetItem()
            item.setFlags(item.flags() &~QtCore.Qt.ItemIsEnabled &~ QtCore.Qt.ItemIsSelectable )
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole, "")
            item.setBackground(QtGui.QColor(self.colorsQt[index]))            
            self.tableWidget_Info_2.setItem(rowPosition, 2, item)

            #Column for User specified name
            item = QtWidgets.QTableWidgetItem()
            #item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            item.setData(QtCore.Qt.EditRole,str(index))
            self.tableWidget_Info_2.setItem(rowPosition, 3, item)
           
            rowPosition += 1
        self.tableWidget_Info_2.resizeColumnsToContents()            
        self.tableWidget_Info_2.resizeRowsToContents()


                
        
    def export_to_rtdc(self,item):
        """
        Grab validation data of particular class, load the scores (model.predict)
        and save images to .rtdc, or show them (users decision)
        """
        true_label = item.row()
        predicted_label = item.column()

        #If there is X_valid and y_valid on RAM, use it!
        if not type(self.ValidationSet) is type(None): #If X_valid is not none, there has been X_valid loaded already
            self.statusbar.showMessage("Re-used validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM",2000)
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
            self.statusbar.showMessage("Loaded data corresponding to the clicked files on 'Build'-tab",2000)

        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        X_valid_orig = self.ValidationSet["X_valid_orig"] #cropped but non-normalized images
        Indices = self.ValidationSet["Indices"]
        y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
        dic = self.Metrics #gives {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        if len(dic)==0:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Data was altered. Please run 'Update Plots' again")
            msg.setWindowTitle("Data has changed")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        pred = dic["pred"]
        #get the length of each Index-list, 
        lengths = [len(l) for l in Indices]
        starts = np.cumsum(lengths)
        ToSave, y_valid_list, Indices_ = [],[],[] #list; store images remaining to indiv. .rtdc set in there 
        starts = np.array([0]+list(starts))
        for i in range(len(lengths)):
            y_val = y_valid[starts[i]:starts[i+1]]
            pred_ = pred[starts[i]:starts[i+1]]
            #update the indx to prepare for next iteration
            #indx = lengths[i]
            ind = np.where( (y_val==true_label) & (pred_==predicted_label) )[0] #select true_label cells and which of them are clasified as predicted_label
            #Grab the corresponding images
            ToSave.append(X_valid_orig[i][ind,:,:]) #get non-normalized X_valid to new variable
            #np.save("X_valid_orig_"+str(i)+".npy",X_valid_orig[i][ind,:,:])
            y_valid_list.append(y_val[ind])
            Indices_.append(Indices[i][ind]) #get non-normalized X_valid to new variable

        total_number_of_chosen_cells = [len(a) for a in y_valid_list]
        total_number_of_chosen_cells = np.sum(np.array(total_number_of_chosen_cells))

        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Question)
        text = "<html><head/><body><p>Show images or save to .rtdc/.png/.jpg?</p></body></html>"
        msg.setText(text)
        msg.setWindowTitle("Show or save?")
        msg.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.YesToAll | QtGui.QMessageBox.Save | QtGui.QMessageBox.Cancel)
        show = msg.button(QtGui.QMessageBox.Yes)
        show.setText('Show')
        save_rtdc = msg.button(QtGui.QMessageBox.YesToAll)
        save_rtdc.setText('Save to .rtdc')
        save_png = msg.button(QtGui.QMessageBox.Save)
        save_png.setText('Save to .png/.jpg...')
        cancel = msg.button(QtGui.QMessageBox.Cancel)
        cancel.setText('Cancel')
        msg.exec_()        
        
        if msg.clickedButton()==show: #Show video
            #Show ToSave Images as video stream
            if total_number_of_chosen_cells>0:
                self.plt_cm.append(pg.ImageView())
                self.plt_cm[-1].show()
                #append an last frame again to ToSave. Otherwise pyqtgraph does not show all iamges
                viewimgs = np.concatenate(ToSave).swapaxes(1,2)
                viewimgs = np.append(viewimgs,viewimgs[-1:],axis=0)
                self.plt_cm[-1].setImage(viewimgs)
            else:
                return
        
        #For .rtdc saving
        elif msg.clickedButton()==save_rtdc: #Save to .rtdc
            if total_number_of_chosen_cells==0:
                return
            sumlen = np.sum(np.array([len(l) for l in ToSave]))
            self.statusbar.showMessage("Nr. of target cells above threshold = "+str(sumlen),2000)
    
            filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to .rtdc', Default_dict["Path of last model"],".rtdc (*.rtdc)")
            filename = filename[0]
    
            if len(filename)==0:
                return
    
            #add the suffix _Valid_Data.rtdc or _Valid_Labels.txt
            if not filename.endswith(".rtdc"):
                filename = filename +".rtdc"
            filename_X = filename.split(".rtdc")[0]+"_Valid_Data.rtdc"
            filename_y = filename.split(".rtdc")[0]+"_Valid_Labels.txt"
    
            #Save the labels
            y_valid_list = np.concatenate(y_valid_list)
            #Save the .rtdc data (images and all other stuff)
            #Should cropped or original be saved?
            if bool(self.actionExport_Original.isChecked())==True:
                print("Export original images")
                save_cropped = False
            if bool(self.actionExport_Cropped.isChecked())==True:
                print("Export cropped images")
                save_cropped = True
            elif bool(self.actionExport_Off.isChecked())==True:
                print("Exporting is turned off")
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("You may want to choose a different exporting option in ->Options->Export")
                msg.setWindowTitle("Export is turned off!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
    
            np.savetxt(filename_y,y_valid_list.astype(int),fmt='%i')      
            aid_bin.write_rtdc(filename_X,rtdc_path_valid,ToSave,Indices_,cropped=save_cropped,color_mode=self.get_color_mode())

            #If all that run without issue, remember the path for next time
            Default_dict["Path of last model"] = os.path.split(filename)[0]
            aid_bin.save_aid_settings(Default_dict)

        #For .png saving
        elif msg.clickedButton()==save_png: #Save to .png/.jpg
            if total_number_of_chosen_cells==0:
                return
            sumlen = np.sum(np.array([len(l) for l in ToSave]))
            self.statusbar.showMessage("Nr. of target cells above threshold = "+str(sumlen),2000)

            filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save to .png/.jpg', Default_dict["Path of last model"],"Image file format (*.png *.jpg *.bmp *.eps *.gif *.ico *.icns)")
            filename = filename[0]
            if len(filename)==0:
                return
            filename_X, file_extension = os.path.splitext(filename)#divide into path and file_extension if possible
            #Check if chosen file_extension is valid
            if not file_extension in [".png",".jpg",".bmp",".eps",".gif",".ico",".icns"]:
                print("Invalid file extension detected. Will use .png instead.")
                file_extension = ".png"
    
            #Should cropped or original be saved?
            if bool(self.actionExport_Original.isChecked())==True:
                print("Export original images")
                save_cropped = False
            if bool(self.actionExport_Cropped.isChecked())==True:
                print("Export cropped images")
                save_cropped = True
            elif bool(self.actionExport_Off.isChecked())==True:
                print("Exporting is turned off")
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("You may want to choose a different exporting option in ->Options->Export")
                msg.setWindowTitle("Export is turned off!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return

            #Save the images data to .png/.jpeg...
            index = 0
            for imgs in ToSave:
                for img in imgs:
                    img = PIL.Image.fromarray(img)
                    img.save(filename_X+"_"+str(index)+file_extension)
                    index+=1

            #If all that run without issue, remember the path for next time
            Default_dict["Path of last model"] = os.path.split(filename)[0]
            aid_bin.save_aid_settings(Default_dict)


    def copy_cm_to_clipboard(self,cm1_or_cm2):
        if cm1_or_cm2==1:
            table = self.tableWidget_CM1
            cols = table.columnCount()
            header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        elif cm1_or_cm2==2:
            table = self.tableWidget_CM2
            cols = table.columnCount()
            header = [table.horizontalHeaderItem(col).text() for col in range(cols)]
        elif cm1_or_cm2==3: #this is for the classification report table tableWidget_AccPrecSpec
            table = self.tableWidget_AccPrecSpec
            cols = table.columnCount()
            header = list(range(cols))
        
        rows = table.rowCount()

        tmp_df = pd.DataFrame(columns=header,index=range(rows)) 
        for i in range(rows):
            for j in range(cols):
                try:
                    tmp_df.iloc[i, j] = table.item(i, j).text()
                except:
                    tmp_df.iloc[i, j] = np.nan
                    
        tmp_df.to_clipboard()
        if cm1_or_cm2<3:
            self.statusbar.showMessage("Confusion matrix appended to clipboard.",2000)
        if cm1_or_cm2==3:
            self.statusbar.showMessage("Classification report appended to clipboard.",2000)

    def assess_model_plotting(self):
        if self.load_model_path == None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

       #If there is a ValidationSet on RAM-> use it!
        if not type(self.ValidationSet) is type(None): #If ValidationSet is not none, there has been ValidationSet loaded already
            self.statusbar.showMessage("Use validation data (from RAM) loaded earlier. If that is not good, please check and uncheck a file on 'Build' tab. This will delete the validation data from RAM",5000)
        else: #Otherwise get the validation data from the stuff that is clicked on 'Build'-Tab
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist

        #Check if input data is available
        if type(self.ValidationSet)==type(None):
            return
        elif type(self.ValidationSet["X_valid"])==type(None):
            return
        #Check the input dimensions:        
        img_dim = self.ValidationSet["X_valid"].shape[-2]
        model_in = int(self.spinBox_Crop_2.value())
        if model_in!=img_dim:
            self.ValidationSet = None
            self.get_validation_data_from_clicked() #after that, self.ValidationSet will exist
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("New model has different input dimensions (image crop). Validation set is re-loaded (like when you clicked on files on build-tab)")
            msg.setWindowTitle("Automatically re-loaded validation set")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable
        #X_valid = self.X_valid #<--dont do this since it is used only once (.predict) and it would require additional RAM; instad use self.X_valid for .predict
        #Load the model and predict
        with tf.Session(graph = tf.Graph(), config=config_gpu) as sess:
            model_keras = load_model(self.load_model_path,custom_objects=get_custom_metrics())  
            in_dim = model_keras.get_input_shape_at(node_index=0)
            channels_model = in_dim[-1]
            channels_data = self.ValidationSet["X_valid"].shape[-1]
            #Compare channel dimensions of loaded model and validation set
            if channels_model!=channels_data: #Model and validation data have differnt channel dims
                text = "Model expects "+str(int(channels_model))+" channel(s), but data has "+str(int(channels_data))+" channel(s)!"
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText(text)
                msg.setWindowTitle("Model and data channel dimension not equal!")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
            scores = model_keras.predict(self.ValidationSet["X_valid"])
            #I got all I need from the tensorflow session. Close it
            
        #Get settings from the GUI
        threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
        target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for
        thresh_on = bool(self.checkBox_SortingThresh.isChecked())

        #Check that the target index alias "Sorting class" is actually a valid class of the model
        out_dim = int(self.spinBox_OutClasses_2.value())
        if not target_index<out_dim:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("You set the 'Sorting class' to "+str(target_index)+" which is not a valid class of the loaded model. The model only has the following classes: "+str(range(out_dim)))
            msg.setWindowTitle("Class not available in the model")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        dic = aid_bin.metrics_using_threshold(scores,y_valid,threshold,target_index,thresh_on) #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        self.Metrics = dic #write to a variable #     
        
        pred = dic["pred"]
        
        cm = metrics.confusion_matrix(y_valid,pred,labels=range(scores.shape[1]))
        cm_normalized = 100*cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        
        #Show the metrics on tableWidget_CM1 and tableWidget_CM2
        #inds_uni = set(list(set(y_valid))+list(set(pred))) #It could be that a cell-index is not present in the validation data, but, the dimension of the scores tells me, how many indices are supposed to appear
        inds_uni = range(scores.shape[1]) #these indices are explained by model
        
        #look in into tableWidget_Info_2 if there are user defined index names
        
        rowCount = self.tableWidget_Info_2.rowCount()
        #Only counts rows with input
        rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

        try:
            indices_on_table = [int(self.tableWidget_Info_2.item(row, 0).text()) for row in range(rowCount)]
            names_on_table = [str(self.tableWidget_Info_2.item(row, 3).text()) for row in range(rowCount)]
        except Exception as e:
            #There was an error!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        #Check that len(names_on_table) <= len(inds_uni) ->it is impossible that the model for example can predict 2 classes, but there are 3 different classes in the validation set
        if not len(names_on_table) <= len(inds_uni):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Model can only predict "+str(len(inds_uni))+" classes, but validation data contains "+str(len(names_on_table))+" classes")
            msg.setWindowTitle("Too many classes in validation set")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
        
        CellNames = []            
        for ind in inds_uni:
            #check if that index is present on table
            where = np.where(np.array(indices_on_table)==ind)[0]
            if len(where)==1:#if there is exaclty one item...
                CellNames.append(np.array(names_on_table)[where]) #apend the corresponding user defined name to a list
            else:
                CellNames.append(str(ind))

        header_labels = [i[0] for i in CellNames]#list(inds_uni)]

        #Table for CM1 - Total Nr of cells
        self.tableWidget_CM1.setRowCount(len(inds_uni))
        self.tableWidget_CM1.setColumnCount(len(inds_uni))            
        
        self.tableWidget_CM1.setHorizontalHeaderLabels(header_labels) 
        self.tableWidget_CM1.setVerticalHeaderLabels(header_labels) 
        for i in inds_uni:
            for j in inds_uni:
                rowPosition = i
                colPosition = j
                #Total nr of cells for each index
                #put the index in column nr. 0
                item = QtWidgets.QTableWidgetItem()
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                item.setData(QtCore.Qt.EditRole,str(cm[i,j]))
                self.tableWidget_CM1.setItem(rowPosition, colPosition, item)
        self.tableWidget_CM1.resizeColumnsToContents()            
        self.tableWidget_CM1.resizeRowsToContents()


        #Table for CM2 - Normalized Confusion matrix
        self.tableWidget_CM2.setRowCount(len(inds_uni))
        self.tableWidget_CM2.setColumnCount(len(inds_uni))
        self.tableWidget_CM2.setHorizontalHeaderLabels(header_labels) 
        self.tableWidget_CM2.setVerticalHeaderLabels(header_labels) 
        for i in range(len(inds_uni)):
            for j in range(len(inds_uni)):
                rowPosition = i
                colPosition = j
                #Total nr of cells for each index
                #put the index in column nr. 0
                item = QtWidgets.QTableWidgetItem()
                item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                item.setData(QtCore.Qt.EditRole,str(cm_normalized[i,j]))
                self.tableWidget_CM2.setItem(rowPosition, colPosition, item)
        self.tableWidget_CM2.resizeColumnsToContents()            
        self.tableWidget_CM2.resizeRowsToContents()
        
        ############Fill tableWidget_AccPrecSpec with information##########          
                
        #Compute more metrics and put them on the table below                    
        nr_target_init = float(len(np.where(y_valid==target_index)[0])) #number of target cells in the initial sample
        conc_init = nr_target_init/float(len(y_valid)) #concentration of the target cells in the initial sample
                    
        acc = metrics.accuracy_score(y_valid,pred)
        
        #Reset the table
        self.tableWidget_AccPrecSpec.setColumnCount(0)#Reset table
        self.tableWidget_AccPrecSpec.setRowCount(0)#Reset table
        nr_cols = np.max([5,len(inds_uni)+1])
        self.tableWidget_AccPrecSpec.setColumnCount(nr_cols) #Five columns
        self.tableWidget_AccPrecSpec.setRowCount(7+len(inds_uni)+2) #Nr. of rows

        #Put lots and lots of Info on tableWidget_AccPrecSpec
        text_conc_init = "Init. conc. of cells from class/name "+header_labels[target_index]
        self.tableWidget_AccPrecSpec.setItem(0 , 0, QtGui.QTableWidgetItem(text_conc_init))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(100*conc_init,4)))
        self.tableWidget_AccPrecSpec.setItem(0, 1, item)

        text_conc_final = "Final conc. in target region"
        self.tableWidget_AccPrecSpec.setItem(1 , 0, QtGui.QTableWidgetItem(text_conc_final))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["conc_target_cell"],4)))
        self.tableWidget_AccPrecSpec.setItem(1, 1, item)

        text_enrich = "Enrichment"
        self.tableWidget_AccPrecSpec.setItem(2 , 0, QtGui.QTableWidgetItem(text_enrich))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["enrichment"],4)))
        self.tableWidget_AccPrecSpec.setItem(2, 1, item)

        text_yield = "Yield"
        self.tableWidget_AccPrecSpec.setItem(3 , 0, QtGui.QTableWidgetItem(text_yield))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(dic["yield_"],4)))
        self.tableWidget_AccPrecSpec.setItem(3, 1, item)

        text_acc = "Accuracy"#+str(round(acc,4))+"\n"
        self.tableWidget_AccPrecSpec.setItem(4 , 0, QtGui.QTableWidgetItem(text_acc))
        item = QtWidgets.QTableWidgetItem()
        item.setData(QtCore.Qt.EditRole, float(np.round(acc,4)))
        self.tableWidget_AccPrecSpec.setItem(4, 1, item)

        text_classification_report = "Classification Report"#+metrics.classification_report(y_valid, pred, target_names=header_labels)
        self.tableWidget_AccPrecSpec.setItem(5 , 0, QtGui.QTableWidgetItem(text_classification_report))
        class_rep = metrics.classification_report(y_valid, pred,labels=inds_uni, target_names=header_labels,output_dict =True)

        try:
            df = pd.DataFrame(class_rep)
            df = df.T
            ax_left = df.axes[0]
            for row in range(len(ax_left)):
                self.tableWidget_AccPrecSpec.setItem(7+row, 0, QtGui.QTableWidgetItem(str(ax_left[row])))
            ax_up = df.axes[1]
            for col in range(len(ax_up)):
                self.tableWidget_AccPrecSpec.setItem(6, 1+col, QtGui.QTableWidgetItem(str(ax_up[col])))
                
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    val = df.iloc[row,col]
                    val = float(np.round(val,4))
                    item = QtWidgets.QTableWidgetItem()
                    item.setData(QtCore.Qt.EditRole, val)
                    self.tableWidget_AccPrecSpec.setItem(7+row, 1+col, item)
        except Exception as e:
            #There is an issue loading the files!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()

        self.tableWidget_AccPrecSpec.resizeColumnsToContents()            
        self.tableWidget_AccPrecSpec.resizeRowsToContents()

        #AFTER the table is resized to the contents, fill in also information 
        #about the used data:
        rowPosition = self.tableWidget_AccPrecSpec.rowCount()
        self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert a new row
        self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem("Used Files"))
        rowPosition = self.tableWidget_AccPrecSpec.rowCount()
        self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert another row!
        self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem("File"))
        
        #dic = {"SelectedFiles_valid":SelectedFiles_valid,"nr_events_epoch_valid":nr_events_epoch_valid,"rtdc_path_valid":[rtdc_path],"X_valid_orig":[X_valid_orig],"X_valid":X_valid,"y_valid":y_valid,"Indices":[Indices]}
        
        rtdc_path_valid = self.ValidationSet["rtdc_path_valid"]
        #nr_events_epoch_valid = self.ValidationSet["nr_events_epoch_valid"]
        y_valid = self.ValidationSet["y_valid"] #y_valid is a long array containing the label of all cell (of all clicked files)
        Indices = self.ValidationSet["Indices"] #Index is a list with arrays containing cell-indices (to track events in a data-set)
        
        
        y_valid_uni = np.unique(np.array(y_valid),return_counts=True)
        #set the column count the at least match the amount of different cell-types available
        if self.tableWidget_AccPrecSpec.columnCount() < len(y_valid_uni[0]):
            diff = len(y_valid_uni[0])-self.tableWidget_AccPrecSpec.columnCount()
            for col_ind in range(diff):
                colPosition = self.tableWidget_AccPrecSpec.columnCount()
                self.tableWidget_AccPrecSpec.insertColumn(colPosition) #Insert a new col for each cell-type
        
        #Create a column for each cell-type
        for col_ind in range(len(y_valid_uni[0])):
            #how_many = y_valid_uni[1][col_ind]
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+col_ind, QtGui.QTableWidgetItem(float(how_many)))
            content = "Class "+str(y_valid_uni[0][col_ind])
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, content)
            self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+col_ind, item)
                
        loc = 0
        for row in range(len(rtdc_path_valid)):
            rowPosition = self.tableWidget_AccPrecSpec.rowCount()
            self.tableWidget_AccPrecSpec.insertRow(rowPosition) #Insert a new row for each entry
            self.tableWidget_AccPrecSpec.setItem(rowPosition , 0, QtGui.QTableWidgetItem(rtdc_path_valid[row]))
            #y_valid_uni = np.unique(y_valid[row])
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 1, QtGui.QTableWidgetItem(np.array(y_valid_uni)))
            #self.tableWidget_AccPrecSpec.setItem(rowPosition , 2, QtGui.QTableWidgetItem(float(nr_events_epoch_valid[row])))
            index = Indices[row] #get the array of indices of a single measurement
            y_valid_i = y_valid[loc:loc+len(index)]
            loc = loc+len(index)
            y_valid_i_uni = np.unique(y_valid_i,return_counts=True)
            for col_ind in range(len(y_valid_i_uni[0])):
                #what is the cell-type
                cell_type = int(y_valid_i_uni[0][col_ind])#cell-type index alway starts with 0. Nr. of cells of cell-type 0 remain to column 1
                how_many = y_valid_i_uni[1][col_ind]
                item = QtWidgets.QTableWidgetItem()
                item.setData(QtCore.Qt.EditRole, int(how_many))
                self.tableWidget_AccPrecSpec.setItem(rowPosition , 1+cell_type, item)

        #Draw the probability histogram
        self.probability_histogram()
        #Finally, also update the third plot
        self.thirdplot()
 
    def create_random_table(self):
        print("def create_random_table only useful for development")
#        matrix = np.random.randint(0,100,size=(3,3))
#        self.tableWidget_CM1.setRowCount(matrix.shape[0])
#        self.tableWidget_CM1.setColumnCount(matrix.shape[1])
#
#        for i in range(matrix.shape[0]):
#            for j in range(matrix.shape[1]):
#                item = QtWidgets.QTableWidgetItem()
#                item.setData(QtCore.Qt.EditRole,str(matrix[i,j]))
#                self.tableWidget_CM1.setItem(i, j, item)
#
#        self.tableWidget_CM1.resizeColumnsToContents()            
#        self.tableWidget_CM1.resizeRowsToContents()

      
    def probability_histogram(self): #    def probability_histogram(self):   
        """
        Grab the scores of each class and show it in histogram
        """
        if len(self.Metrics) ==0: #but if not give message and return
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("There are no Metrics determined yet (use ->'Update Plots' first)")
            msg.setWindowTitle("No Metrics found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        dic = self.Metrics #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}
        scores = dic["scores"]

        #Get the available cell indices (cell-type identifier)
        inds_uni = range(scores.shape[1]) #these indices are explained by model

        threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
        target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for
       
        try:
            #What is the probability of cell with y_valid=i that it belongs to class target_index?
            scores_i = []
            y_valid = self.ValidationSet["y_valid"]
            for i in inds_uni:
                ind = np.where(y_valid==i)[0]
                if len(ind)>0: #if there are no cells available, dont append. In this case there will also be no color defined
                    scores_i.append(scores[ind,target_index])
        except Exception as e:
            #There is an issue loading the files!
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText(str(e))
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
            
        rowCount = self.tableWidget_Info_2.rowCount()
        #only count rows with content
        rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])
        names_on_table = [str(self.tableWidget_Info_2.item(row, 3).text()) for row in range(rowCount)]
        index_on_table = [int(self.tableWidget_Info_2.item(row, 0).text()) for row in range(rowCount)]
        #On which row is the target_index?
        ind = np.where(np.array(index_on_table)==target_index)[0]
        if len(ind) == 1:
            target_name = str(np.array(names_on_table)[ind][0])          
        else:
            target_name = str(target_index)
        
        #Get the user-defined colors from table
        colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]
        #it can be that the table was not updated and there are more scores than table-items
        if len(colors_on_table)!=len(scores_i):
            #update table
            SelectedFiles = self.items_clicked_no_rtdc_ds()
            self.update_data_overview(SelectedFiles)
            self.update_data_overview_2(SelectedFiles)
            #update colors on table
            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]
            
        #Clear the plot        
        self.widget_probHistPlot.clear()
        #Add plot        
        hist = self.widget_probHistPlot.addPlot()
        hist.showGrid(x=True,y=True)
        hist.setLabel('bottom', "p("+target_name+")", units='')
        hist.setLabel('left', "#", units='')
        
        #Get the user defined histogram style from the combobox
        style = str(self.comboBox_probability_histogram.currentText())
        for i in range(len(scores_i)): # I had previously range(len(scores_i)); but this causes an error if there is a cell-type missing in the validation set
            hist_i = hist.plot()
            if len(scores_i[i])>1:#only continue of there multiple events (histogram does not make sense otherwise)
                range_hist = (scores_i[i].min(), scores_i[i].max())
                first_edge, last_edge = np.lib.histograms._get_outer_edges(scores_i[i], range=range_hist)
                try: #numpy 1.15
                    width = np.lib.histograms._hist_bin_selectors['auto'](scores_i[i])
                except:#numpy >1.15
                    width = np.lib.histograms._hist_bin_selectors['auto'](scores_i[i],(np.min(scores_i[i]),np.min(scores_i[i])))
                    
                n_equal_bins = int(np.ceil(np.lib.histograms._unsigned_subtract(last_edge, first_edge) / width))
                if n_equal_bins>1E4: #Who needs more than 10k bins?!:
                    n_equal_bins = int(1E4)
                else:
                    n_equal_bins='auto'
    
                y,x = np.histogram(scores_i[i], bins=n_equal_bins)
                if style=="Style1":
                    pencolor = pg.mkColor(colors_on_table[i].color())
                    pen = pg.mkPen(color=pencolor,width=5)
                    hist_i.setData(x, y, stepMode=True, pen=pen,clear=False)
                elif style=="Style2":
                    pencolor = pg.mkColor(colors_on_table[i].color())
                    pen = pg.mkPen(color=pencolor,width=10)
                    hist_i.setData(x, y, stepMode=True, pen=pen,clear=False)
                elif style=="Style3":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.6*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)
                elif style=="Style4":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.7*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)
                elif style=="Style5":
                    color = colors_on_table[i].color()
                    color.setAlpha(0.8*255.0) 
                    pencolor = pg.mkColor(color)
                    brush = pg.mkBrush(color=pencolor)
                    hist_i.setData(x, y, stepMode=True, fillLevel=0, brush=brush,clear=False)

        #Add a vertical line indicating the threshold
        self.line = pg.InfiniteLine(pos=threshold, angle=90, pen='w', movable=False)
        hist.addItem(self.line)
        hist.setXRange(0, 1, padding=0)
        
    def thirdplot(self):
        target_index =self.spinBox_indexOfInterest.value()
        cb_text = self.comboBox_3rdPlot.currentText()
        
        if cb_text=='None':
            return
        if cb_text=='ROC-AUC':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0:
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #If no Metrics available yet...
                dic = self.Metrics
            if len(dic)==0:
                return
            #Get the ValidationSet
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
            scores = dic["scores"]
            
            inds_uni = list(range(scores.shape[1])) #these indices are explained by model
            #ROC-curve is only available for binary problems:
            Y_valid = np_utils.to_categorical(y_valid,num_classes=len(inds_uni))
            
            # Compute ROC curve and ROC area for each class
            fpr,tpr,roc_auc = dict(),dict(),dict()
            for i in range(len(inds_uni)):                
                fpr[i], tpr[i], _ = metrics.roc_curve(Y_valid[:, i], scores[:, i])
                roc_auc[i] = metrics.auc(fpr[i], tpr[i])
            
            # Compute micro-average ROC curve and ROC area
            fpr["micro"], tpr["micro"], _ = metrics.roc_curve(Y_valid.ravel(), scores.ravel())
            roc_auc["micro"] = metrics.auc(fpr["micro"], tpr["micro"])

            #Get the user-defined colors from table
            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "False Positive Rate", units='')
            hist.setLabel('left', "True Positive Rate", units='')
            for i, color in zip(range(len(inds_uni)), colors_on_table):
                text = 'Class '+str(i)+', AUC='+str(round(roc_auc[i],2))
                hist.plot(fpr[i], tpr[i], pen=None,symbol='o',symbolPen=None,symbolBrush=color,name=text,clear=False)
                clr = color.color()
                hist.plot(fpr[i],tpr[i],pen=clr)
            hist.setXRange(0, 1, padding=0)

        if cb_text=='Precision-Recall':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0: 
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #Otherwise, there are Metrics available already :) Use them
                dic = self.Metrics
            if len(dic)==0:
                return
              
            #Get the ValidationSet
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable      
            scores = dic["scores"]#[:,target_index]
            
            inds_uni = list(range(scores.shape[1])) #these indices are explained by model
            #ROC-curve is only available for binary problems:
            Y_valid = np_utils.to_categorical(y_valid,num_classes=len(inds_uni))
            
            # Compute Precision Recall curve and P-R area for each class
            precision,recall,precision_recall_auc = dict(),dict(),dict()
            for i in range(len(inds_uni)):                
                precision[i], recall[i], _ = metrics.precision_recall_curve(Y_valid[:, i], scores[:, i])
                precision_recall_auc[i] = metrics.auc(recall[i], precision[i])
            
            # Compute micro-average ROC curve and ROC area
            precision["micro"], recall["micro"], _ = metrics.roc_curve(Y_valid.ravel(), scores.ravel())
            precision_recall_auc["micro"] = metrics.auc(recall["micro"],precision["micro"])

            #Get the user-defined colors from table
            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "Recall", units='')
            hist.setLabel('left', "Precision", units='')
            for i, color in zip(range(len(inds_uni)), colors_on_table):
                text = 'Class '+str(i)+', AUC='+str(round(precision_recall_auc[i],2))
                hist.plot(recall[i],precision[i], pen=None,symbol='o',symbolPen=None,symbolBrush=color,name=text,clear=False)
                clr = color.color()
                hist.plot(recall[i],precision[i],pen=clr)
            hist.setXRange(0, 1, padding=0)

        if cb_text=='Enrichment vs. Threshold' or cb_text=='Yield vs. Threshold' or cb_text=='Conc. vs. Threshold':
            #Check if self.Metrics are available
            if len(self.Metrics) == 0: #if not,
                self.assess_model_plotting() #run this function to create self.Metrics
                dic = self.Metrics
            else: #If Metrics are already available, use it. Load it
                dic = self.Metrics
            if len(dic)==0:
                return
            
            scores = dic["scores"]
            y_valid = self.ValidationSet["y_valid"] #load the validation labels to a new variable

            #The dic only contains metrics for a single threshold, which is not enough
            #call aid_bin.metrics_using_threshold with a range of thresholds:
            #(it might makes sense to evaluate this for each possible target_index. Now only perform the measurement for the user defined target index)
            Dics,Threshs = [],[]
            for thresh in np.linspace(0,1,25):
                dic_ = aid_bin.metrics_using_threshold(scores,y_valid,thresh,target_index) #returns dic = {"scores":scores,"pred":pred,"conc_target_cell":conc_target_cell,"enrichment":enrichment,"yield_":yield_}                    
                Dics.append(dic_)
                Threshs.append(thresh)
            #Collect information in arrays
            enrichment_ = np.array([d["enrichment"] for d in Dics])
            yield__ = np.array([d["yield_"] for d in Dics])
            conc_target_cell = np.array([d["conc_target_cell"] for d in Dics])
            Threshs = np.array(Threshs)

            rowCount = self.tableWidget_Info_2.rowCount()
            #only count rows with content
            rowCount = sum([self.tableWidget_Info_2.item(row, 0)!=None for row in range(rowCount)])

            colors_on_table = [self.tableWidget_Info_2.item(row, 2).background() for row in range(rowCount)]

            #Clear the plot        
            self.widget_3rdPlot.clear()
            #Add plot        
            hist = self.widget_3rdPlot.addPlot()
            hist.showGrid(x=True,y=True)
            hist.addLegend()
            hist.setLabel('bottom', "Threshold", units='')
            color = '#0000ff'
            if cb_text=='Enrichment vs. Threshold':
                hist.setLabel('left', "Enrichment", units='')
                hist.plot(Threshs,enrichment_, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,enrichment_,pen=color)
            if cb_text=='Yield vs. Threshold':
                hist.setLabel('left', "Yield", units='')
                hist.plot(Threshs,yield__, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,yield__,pen=color)
            if cb_text=='Conc. vs. Threshold':
                hist.setLabel('left', "Conc. of target cell in target region", units='')
                hist.plot(Threshs,conc_target_cell, pen=None,symbol='o',symbolPen=None,symbolBrush=color,name='',clear=False)
                hist.plot(Threshs,conc_target_cell,pen=color)

            hist.setXRange(0, 1, padding=0)
            #Add indicator for the currently used threshold
            threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
            self.line = pg.InfiniteLine(pos=threshold, angle=90, pen='w', movable=False)
            hist.addItem(self.line)



    def classify(self):
        #Very similar function to "Update Plot". But here, no graphs are produced
        #Resulting scores/predictions etc are simply stored to excel file
        #This function does NOT take labels.

        #Check if a model was defined
        if self.load_model_path == None:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please define a model path first")
            msg.setWindowTitle("No model path found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        #Classify all datasets or just one?
        Files = []
        if self.radioButton_selectAll.isChecked():
            #Get the urls of ALL files on the "build"-tab
            rowCount = self.table_dragdrop.rowCount()
            for rowPosition in range(rowCount):  
                #get the filename/path
                rtdc_path = self.table_dragdrop.cellWidget(rowPosition, 0).text()
                rtdc_path = str(rtdc_path)
                Files.append(rtdc_path)
            if len(Files)==0:
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)       
                msg.setText("There are no files on the 'Build'-Tab")
                msg.setWindowTitle("No files found")
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.exec_()
                return
        elif self.radioButton_selectDataSet.isChecked():
            rtdc_path = self.comboBox_selectData.currentText()
            Files.append(rtdc_path)
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("Please use the radiobuttons (left) to indicate if all or only a selected file should be classified.")
            msg.setWindowTitle("No file(s) specified")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
          
        print("Chosen file(s):")
        print(Files)

        #what input size is required by loaded model?
        crop = int(self.spinBox_Crop_2.value())
        norm = str(self.comboBox_Normalization_2.currentText())
        padding_expert = str(self.comboBox_expt_paddingMode.currentText())

        #if normalization method needs mean/std of training set, the metafile needs to be loaded:
        if norm == "StdScaling using mean and std of all training data":
            modelindex = int(self.spinBox_ModelIndex_2.value())
            path,fname = os.path.split(self.load_model_path)    
            fname = fname.split(str(modelindex)+".model")[0]+"meta.xlsx"
            metafile_path = os.path.join(path,fname)
            parameters = pd.read_excel(metafile_path,sheetname='Parameters')
            mean_trainingdata = parameters["Mean of training data used for scaling"]
            std_trainingdata = parameters["Std of training data used for scaling"]
        else:
            mean_trainingdata = None
            std_trainingdata = None
        
        with tf.Session(graph = tf.Graph(), config=config_gpu) as sess:#load the model since its the same for each file            
            model_keras = load_model(self.load_model_path,custom_objects=get_custom_metrics())
            in_dim = model_keras.get_input_shape_at(node_index=0)
            #Get the color mode of the model
            channels_model = in_dim[-1]
            if channels_model==1:
                color_mode='Grayscale'
            elif channels_model==3:
                color_mode='RGB'
            else:
                print("Invalid number of channels. AID only supports grayscale (1 channel) and RGB (3 channels) images.")


            #Get the user-set export option (Excel or to 'userdef0' in .rtdc file)
            export_option = str(self.comboBox_scoresOrPrediction.currentText())
        
            if export_option == "Add predictions to .rtdc file (userdef0)" or export_option=="Add pred&scores to .rtdc file (userdef0 to 9)":
            #Users sometimes need to have Donor-ID (Parent foldername) added to the .rtdc file
            #Ask the user: Do you want to get a specific fixed addon to filename, OR do you want to have the parent-foldername added?
                msg = QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Question)
                text = "Do you want to get a specific fixed addon to filename, <b>or do you want to have the parent-foldername added for each file individually?"
                text = "<html><head/><body><p>"+text+"</p></body></html>"
                msg.setText(text)
                msg.setWindowTitle("Filename-addon for created files")
                msg.addButton(QtGui.QPushButton('Specific fixed addon...'), QtGui.QMessageBox.YesRole)
                msg.addButton(QtGui.QPushButton('Parent foldername'), QtGui.QMessageBox.NoRole)
                msg.addButton(QtGui.QPushButton('Cancel'), QtGui.QMessageBox.RejectRole)
                retval = msg.exec_()
    
                if retval==0: 
                    #Get some user input:
                    fname_addon, ok = QtWidgets.QInputDialog.getText(self, 'Specific fixed addon...', 'Enter filname addon:')
                    if ok:
                        fname_addon = str(fname_addon)
                    else:
                        return
                elif retval==1:
                    fname_addon = "Action:GetParentFolderName!"
                else:
                    return

            #Iterate over all Files
            for rtdc_path in Files:
                #get all images, cropped correcetly
                gen_train = aid_img.gen_crop_img(crop,rtdc_path,random_images=False,color_mode=color_mode,padding_mode=padding_expert)    
                x_train,index = next(gen_train) #x_train-images of all cells, index-original index of all cells           
                
                if norm == "StdScaling using mean and std of all training data":
                    x_train = aid_img.norm_imgs(x_train,norm,mean_trainingdata,std_trainingdata)
                else:
                    x_train = aid_img.norm_imgs(x_train,norm)
                            
                #Check the input dimensions:
                img_dim = x_train.shape[-2]
                model_in = int(self.spinBox_Crop_2.value())
                if model_in!=img_dim:
                    msg = QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)       
                    msg.setText("New model has different input dimensions (image crop). Validation set is re-loaded (clicked files on build-tab)")
                    msg.setWindowTitle("Input dimensions not fitting")
                    msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msg.exec_()

                scores = model_keras.predict(x_train)
                scores_normal = np.copy(scores)
                pred_normal = np.argmax(scores_normal,axis=1)
                
                #Get settings from the GUI
                threshold = float(self.doubleSpinBox_sortingThresh.value())#threshold probability obove which a cell is sorted
                target_index = int(self.spinBox_indexOfInterest.value())#index of the cell type that should be sorted for

                #Use argmax for prediction (threshold can only be applied to one index)
                pred_normal = np.argmax(scores,axis=1)

                #First: check the scores_in_function of the sorting index and adjust them using the threshold
                pred_thresh = np.array([1 if p>threshold else 0 for p in scores[:,target_index]])
                #replace the corresponding column in the scores_in_function
                scores[:,target_index] = pred_thresh
                #Determine the prediction again, considering the threshold for the target index
                pred_thresh = np.argmax(scores,axis=1)
                
                normal_or_thresh = bool(self.checkBox_SortingThresh.isChecked())
                if normal_or_thresh==True: #(if its true means its normal means p=0.5)
                    prediction_to_rtdc_ds = pred_normal
                if normal_or_thresh==False: #(if its false means its thresholded for some class)
                    prediction_to_rtdc_ds = pred_thresh
                
                if export_option == "Scores and predictions to Excel sheet":
                    info = np.array([[self.load_model_path],[rtdc_path],[target_index],[threshold]]).T
                    info = pd.DataFrame(info,columns=["load_model_path","rtdc_path","target_class","threshold"])
                    #Combine all information in nice excel sheet
                    filename = rtdc_path.split(".rtdc")[0]+"_Prediction.xlsx"
                    writer = pd.ExcelWriter(filename, engine='openpyxl')
                    #Used files go to a separate sheet on the -session.xlsx
                    pd.DataFrame().to_excel(writer,sheet_name='Info') #initialize empty Sheet
                    info.to_excel(writer,sheet_name='Info')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Scores_normal') #initialize empty Sheet
                    pd.DataFrame(scores_normal).to_excel(writer,sheet_name='Scores_normal')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Prediction_normal') #initialize empty Sheet
                    pd.DataFrame(pred_normal).to_excel(writer,sheet_name='Prediction_normal')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Scores_thresholded') #initialize empty Sheet
                    pd.DataFrame(scores).to_excel(writer,sheet_name='Scores_thresholded')
    
                    pd.DataFrame().to_excel(writer,sheet_name='Prediction_thresholded') #initialize empty Sheet
                    pd.DataFrame(pred_thresh).to_excel(writer,sheet_name='Prediction_thresholded')
    
                    writer.save()
                    writer.close()

                if export_option == "Add predictions to .rtdc file (userdef0)" or export_option==         "Add pred&scores to .rtdc file (userdef0 to 9)":           
                    #Get initial dataset
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    if failed:
                        msg = QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Critical)       
                        msg.setText(str(rtdc_ds))
                        msg.setWindowTitle("Error occurred during loading file")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                        msg.exec_()
                        return
                    
                    failed,rtdc_ds = aid_bin.load_rtdc(rtdc_path)
                    rtdc_ds_len = rtdc_ds["image"].shape[0] #this way is actually faster than asking any other feature for its len :)
                    prediction_fillnan = np.full([rtdc_ds_len], np.nan)#put initially np.nan for all cells

                    if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                        classes = scores_normal.shape[1]
                        if classes>9:
                            classes = 9#set the max number of classes to 9. It cannot saved more to .rtdc due to limitation of userdef
                        scores_fillnan = np.full([rtdc_ds_len,classes], np.nan)

                    #Make sure the predictions get again to the same length as the initial data set
                    #Fill array with corresponding predictions
                    for i in range(len(prediction_to_rtdc_ds)):
                        indx = index[i]
                        prediction_fillnan[indx] = prediction_to_rtdc_ds[i]
                        if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                            #for class_ in range(classes):
                            scores_fillnan[indx,0:classes] = scores_normal[i,0:classes]

                    #Get savename
                    path, rtdc_file = os.path.split(rtdc_path)

                    if fname_addon!="Action:GetParentFolderName!":#dont get the parentfoldername, instead used user defined addon!
                        savename = rtdc_path.split(".rtdc")[0]
                        savename = savename+"_"+str(fname_addon)+".rtdc"
                        
                    elif fname_addon=="Action:GetParentFolderName!":                        
                        savename = rtdc_path.split(".rtdc")[0]
                        parentfolder = aid_bin.splitall(rtdc_path)[-2]
                        savename = savename+"_"+str(parentfolder)+".rtdc"
                    else:
                        return
                    
                    if not os.path.isfile(savename):#if such a file does not yet exist...
                        savename = savename
                    else:#such a file already exists!!!
                        #Avoid to overwriting an existing file:
                        print("Adding additional number since file exists!")
                        addon = 1
                        while os.path.isfile(savename):
                            savename = savename.split(".rtdc")[0]
                            if addon>1:
                                savename = savename.split("_"+str(addon-1))[0]
                            savename = savename+"_"+str(addon)+".rtdc"
                            addon += 1        
                    
                    print(savename)                    
                    shutil.copy(rtdc_path, savename) #copy original file
                    #append to hdf5 file
                    with h5py.File(savename, mode="a") as h5:
                        h5["events/userdef0"] = prediction_fillnan
                        if export_option == "Add pred&scores to .rtdc file (userdef0 to 9)":
                            #add the scores to userdef1...9
                            userdef_ind = 1
                            for class_ in range(classes):
                                scores_i = scores_fillnan[:,class_]
                                h5["events/userdef"+str(userdef_ind)] = scores_i
                                userdef_ind += 1
                                



    #####################Python Editor/Console#################################
    def pythonInRun(self):
        code = self.plainTextEdit_pythonIn.toPlainText()
        
        codeOut = io.StringIO()
        out,error = "",""
        # capture output
        sys.stdout = codeOut
        try:
            exec(code,globals())
        except Exception as e:
            error = str(e)
        # restore stdout
        sys.stdout = sys.__stdout__
        out = codeOut.getvalue()
        codeOut.close()
        
        text_out = "Out:\n"+out
        text_error = "Error:\n"+error
        
        #Print both to textBrowser_pythonOut
        self.textBrowser_pythonOut.append(text_out)
        if len(error)>0:
            self.textBrowser_pythonOut.append(text_error)

    def pythonInClear(self):
        self.plainTextEdit_pythonIn.clear()
        self.lineEdit_pythonCurrentFile.clear()

    def pythonInSaveAs(self):
        filename = QtWidgets.QFileDialog.getSaveFileName(self, 'Save code', Default_dict["Path of last model"],"Python file (*.py)")
        filename = filename[0]
        if len(filename)==0:
            return
        #add the suffix .csv
        if not filename.endswith(".py"):
            filename = filename +".py"               

        code = self.plainTextEdit_pythonIn.toPlainText()

        myfile = open(filename,'w')#Open the file with writing permission
        myfile.write(code)        
        myfile.close()
        self.lineEdit_pythonCurrentFile.setText(filename)

    def pythonInOpen(self):
        filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open Python file', Default_dict["Path of last model"],"Python file (*.py)")
        filename = filename[0]
        if not filename.endswith(".py"):
            return
        if not os.path.isfile(filename):
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)       
            msg.setText("File not found")
            msg.setWindowTitle("File not found")
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return
        with open(filename, 'r') as myfile:
            data = myfile.read()
        
        self.plainTextEdit_pythonIn.clear()
        self.plainTextEdit_pythonIn.insertPlainText(data)
        self.lineEdit_pythonCurrentFile.setText(filename)

    def pythonOutClear(self):
        self.textBrowser_pythonOut.clear()

    #Show cpu and ram usage on the status bar
    def cpu_ram_worker(self,progress_callback,history_callback):
        while True:
            cpu,ram = psutil.cpu_percent(),psutil.virtual_memory().percent
            self.statusbar_cpuRam.setText("CPU: "+str(cpu)+"%  RAM: "+str(ram)+"%")
            time.sleep(2)

    def delete_ram(self):
        if self.actionVerbose.isChecked()==True:
            print("Deleting Ram")
        self.ram = dict()
        self.ValidationSet = None
        self.Metrics = dict()

    def quit_app(self):
        sys.exit()

    def closeEvent(self, event):
        sys.exit()


def main():
    global app
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon(os.path.join(dir_root,"art",Default_dict["Icon theme"],"main_icon_simple_04_256"+icon_suff)))

    if Default_dict["Layout"] == "Dark":
        dir_layout = os.path.join(dir_root,"layout_dark.txt")#dir to settings
        f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/ColinDuquesnoy/QDarkStyleSheet/blob/master/qdarkstyle/style.qss
        f = f.read()
        app.setStyleSheet(f)
    elif Default_dict["Layout"] == "DarkOrange":
        dir_layout = os.path.join(dir_root,"layout_darkorange.txt")#dir to settings
        f = open(dir_layout, "r") #I obtained the layout file from: https://github.com/nphase/qt-ping-grapher/blob/master/resources/darkorange.stylesheet
        f = f.read()
        app.setStyleSheet(f)
    else:
        app.setStyleSheet("")
    
    ui = MainWindow()
    ui.show()
    try:
        splash.finish(ui)
    except:
        pass

    ret = app.exec_()
    sys.exit(ret)

if __name__ == '__main__':
    main()
